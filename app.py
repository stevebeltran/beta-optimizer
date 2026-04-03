import streamlit as st
import pandas as pd
import geopandas as gpd
import numpy as np
import plotly.graph_objects as go
from shapely.geometry import Point, Polygon, MultiPolygon, box, shape
from shapely.ops import unary_union
import os, itertools, glob, math, simplekml, heapq, re, random, json, io, datetime, base64, smtplib, uuid, traceback
from concurrent.futures import ThreadPoolExecutor
import pulp
import urllib.request
import urllib.parse
import zipfile
import streamlit.components.v1 as components
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import gspread
from google.oauth2.service_account import Credentials
import pyproj
from PIL import Image

# --- PAGE CONFIG & INITIALIZE SESSION STATE ---
st.set_page_config(page_title="BRINC COS Drone Optimizer", layout="wide", initial_sidebar_state="expanded")

# This MUST run before any st.session_state checks to prevent KeyError
defaults = {
    'csvs_ready': False, 'df_calls': None, 'df_calls_full': None, 'df_stations': None,
    'active_city': "Rockford", 'active_state': "IL", 'estimated_pop': 65000,
    'k_resp': 2, 'k_guard': 0, 'r_resp': 2.0, 'r_guard': 8.0,
    'dfr_rate': 12, 'deflect_rate': 25, 'total_original_calls': 0, 'total_modeled_calls': 0,
    'onboarding_done': False, 'trigger_sim': False, 'city_count': 1,
    'brinc_user': 'steven.beltran',
    'pd_chief_name': '', 'pd_dept_name': '', 'pd_dept_email': '', 'pd_dept_phone': '',
    'session_start': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    'session_id': str(uuid.uuid4())[:8],
    'data_source': 'unknown',   # 'cad_upload' | 'simulation' | 'demo' | 'brinc_file'
    'map_build_logged': False,  # prevent duplicate map-build rows per session
    'boundary_kind': 'place',
    'boundary_source_path': '',
    'location_detection_source': '',
    'boundary_detection_mode': '',
    'master_gdf_override': None,  # GeoDataFrame from coordinate-based jurisdiction lookup
    # ── NEW: file ingestion metadata & engagement tracking ──────────────────
    'file_meta': {},            # populated by aggressive_parse_calls; see _extract_file_meta()
    'export_event_log': [],     # ordered list of export types clicked this session
    'export_count': 0,          # total download button clicks this session
    'demo_mode_used': False,    # True if any demo city was loaded
    'sim_mode_used': False,     # True if simulation (not real upload) was run
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v


if 'target_cities' not in st.session_state:
    st.session_state['target_cities'] = [{"city": "", "state": st.session_state.get('active_state', 'IL')}]


GUARDIAN_FLIGHT_HOURS_PER_DAY = 23.5

SIMULATOR_DISCLAIMER_SHORT = (
    "Simulation output only. Coverage, station placement, response time, and ROI figures are model estimates based on uploaded data and configuration settings. "
    "They are not guarantees of real-world performance, legal compliance, FAA approval, procurement outcome, or financial results."
)

def _extract_file_meta(raw_df, res_df, filename=""):
    """
    Compute and return a dict of data-matrix statistics from a parsed CAD upload.
    Call this once per file inside aggressive_parse_calls() and store the result
    in st.session_state['file_meta'].  All values are JSON-safe scalars or strings.
    """
    meta = {}
    try:
        meta['uploaded_filename']   = str(filename)
        meta['file_row_count']      = int(len(raw_df))
        meta['file_col_count']      = int(len(raw_df.columns))
        meta['file_col_names']      = json.dumps(list(raw_df.columns))

        # ── City / state inferred from the file ──────────────────────────────
        meta['file_inferred_city']  = str(res_df['_csv_city'].iloc[0])  if '_csv_city'  in res_df.columns and len(res_df) > 0 else ''
        meta['file_inferred_state'] = str(res_df['_csv_state'].iloc[0]) if '_csv_state' in res_df.columns and len(res_df) > 0 else ''

        # ── Date range ───────────────────────────────────────────────────────
        if 'date' in res_df.columns:
            _dates = pd.to_datetime(res_df['date'], errors='coerce').dropna()
            if not _dates.empty:
                meta['file_date_range_start'] = _dates.min().strftime('%Y-%m-%d')
                meta['file_date_range_end']   = _dates.max().strftime('%Y-%m-%d')
                meta['file_date_span_days']   = int((_dates.max() - _dates.min()).days)
                meta['peak_month']            = int(_dates.dt.month.value_counts().idxmax())
                meta['peak_day_of_week']      = int(_dates.dt.dayofweek.value_counts().idxmax())
            else:
                meta['file_date_range_start'] = ''
                meta['file_date_range_end']   = ''
                meta['file_date_span_days']   = 0
                meta['peak_month']            = 0
                meta['peak_day_of_week']      = 0
        else:
            meta['file_date_range_start'] = ''
            meta['file_date_range_end']   = ''
            meta['file_date_span_days']   = 0
            meta['peak_month']            = 0
            meta['peak_day_of_week']      = 0

        # ── Peak hour ────────────────────────────────────────────────────────
        if 'time' in res_df.columns:
            _times = pd.to_datetime(res_df['time'], format='%H:%M:%S', errors='coerce').dropna()
            meta['peak_hour'] = int(_times.dt.hour.value_counts().idxmax()) if not _times.empty else -1
        else:
            meta['peak_hour'] = -1

        # ── Null rate across key CAD fields ──────────────────────────────────
        _key_fields = [c for c in ['lat', 'lon', 'date', 'time', 'priority', 'call_type_desc'] if c in res_df.columns]
        if _key_fields:
            _null_pct = res_df[_key_fields].isnull().values.mean()
            meta['file_null_rate_pct'] = round(float(_null_pct) * 100, 1)
        else:
            meta['file_null_rate_pct'] = 0.0

        # ── Coordinate detection ─────────────────────────────────────────────
        meta['file_has_lat_lon']  = bool('lat' in res_df.columns and 'lon' in res_df.columns and res_df[['lat','lon']].dropna().shape[0] > 0)
        meta['file_has_priority'] = bool('priority' in res_df.columns and res_df['priority'].dropna().shape[0] > 0)

        # ── Call-type breakdown (top 10) ─────────────────────────────────────
        _type_col = next((c for c in ['call_type_desc','agencyeventtypecodedesc','calldesc','description','nature'] if c in res_df.columns), None)
        if _type_col:
            _tc = res_df[_type_col].dropna().str.strip().value_counts().head(10)
            meta['call_type_breakdown'] = json.dumps({str(k): int(v) for k, v in _tc.items()})
        else:
            meta['call_type_breakdown'] = ''

        # ── Priority distribution ─────────────────────────────────────────────
        if 'priority' in res_df.columns:
            _pc = res_df['priority'].dropna().astype(str).value_counts().sort_index()
            meta['priority_distribution'] = json.dumps({str(k): int(v) for k, v in _pc.items()})
        else:
            meta['priority_distribution'] = ''

    except Exception:
        pass
    return meta


def _build_details_html(details):
    """Shared HTML block for deployment details used in email notifications."""
    if not details: return ""
    drone_list = "".join([
        f"<li><b>{d['name']}</b> ({d['type']}) @ {d['lat']:.4f}, {d['lon']:.4f}</li>"
        for d in details.get('active_drones', [])
    ])
    pop   = details.get('population', 0)
    calls = details.get('total_calls', 0)
    daily = details.get('daily_calls', 0)
    area  = details.get('area_sq_mi', 0)
    be    = details.get('break_even', 'N/A')
    src   = details.get('data_source', '—')
    sid   = details.get('session_id', '—')
    stime = details.get('session_start', '—')
    dur   = details.get('session_duration_min', '—')
    pd_c  = details.get('pd_chief', '—')
    pd_d  = details.get('pd_dept', '—')
    avg_t = details.get('avg_response_min', 0)
    time_saved = details.get('avg_time_saved_min', 0)
    area_cov = details.get('area_covered_pct', 0)
    return f"""
    <div style="margin-top:20px; padding-top:20px; border-top:1px solid #f0f0f0;">
        <h4 style="color:#555; margin-bottom:10px;">Session Info</h4>
        <table style="width:100%; border-collapse:collapse; font-size:12px; margin-bottom:15px;">
            <tr><td style="padding:4px; color:#888; width:50%;">Session ID</td><td style="padding:4px;">{sid}</td></tr>
            <tr><td style="padding:4px; color:#888;">Session Start</td><td style="padding:4px;">{stime}</td></tr>
            <tr><td style="padding:4px; color:#888;">Session Duration</td><td style="padding:4px;">{dur} min</td></tr>
            <tr><td style="padding:4px; color:#888;">Data Source</td><td style="padding:4px;">{src}</td></tr>
        </table>
        <h4 style="color:#555; margin-bottom:10px;">Jurisdiction</h4>
        <table style="width:100%; border-collapse:collapse; font-size:12px; margin-bottom:15px;">
            <tr><td style="padding:4px; color:#888; width:50%;">Population</td><td style="padding:4px;">{pop:,}</td></tr>
            <tr><td style="padding:4px; color:#888;">Total Annual Calls</td><td style="padding:4px;">{calls:,}</td></tr>
            <tr><td style="padding:4px; color:#888;">Daily Calls</td><td style="padding:4px;">{daily:,}</td></tr>
            <tr><td style="padding:4px; color:#888;">Coverage Area</td><td style="padding:4px;">{area:,} sq mi</td></tr>
        </table>
        <h4 style="color:#555; margin-bottom:10px;">Deployment Settings</h4>
        <table style="width:100%; border-collapse:collapse; font-size:12px; margin-bottom:15px;">
            <tr><td style="padding:4px; color:#888; width:50%;">Strategy</td><td style="padding:4px;">{details.get('opt_strategy', '')}</td></tr>
            <tr><td style="padding:4px; color:#888;">Incremental Build</td><td style="padding:4px;">{details.get('incremental_build', False)}</td></tr>
            <tr><td style="padding:4px; color:#888;">Allow Overlap</td><td style="padding:4px;">{details.get('allow_redundancy', False)}</td></tr>
            <tr><td style="padding:4px; color:#888;">DFR Dispatch Rate</td><td style="padding:4px;">{details.get('dfr_rate', 0)}%</td></tr>
            <tr><td style="padding:4px; color:#888;">Deflection Rate</td><td style="padding:4px;">{details.get('deflect_rate', 0)}%</td></tr>
            <tr><td style="padding:4px; color:#888;">Total CapEx</td><td style="padding:4px;">${details.get('fleet_capex', 0):,.0f}</td></tr>
            <tr><td style="padding:4px; color:#888;">Annual Savings</td><td style="padding:4px;">${details.get('annual_savings', 0):,.0f}</td></tr>
            <tr><td style="padding:4px; color:#888;">Thermal Upside</td><td style="padding:4px;">${details.get('thermal_savings', 0):,.0f}</td></tr>
            <tr><td style="padding:4px; color:#888;">K-9 Upside</td><td style="padding:4px;">${details.get('k9_savings', 0):,.0f}</td></tr>
            <tr><td style="padding:4px; color:#888;">Break-Even</td><td style="padding:4px;">{be}</td></tr>
            <tr><td style="padding:4px; color:#888;">Avg Response Time</td><td style="padding:4px;">{avg_t:.1f} min</td></tr>
            <tr><td style="padding:4px; color:#888;">Time Saved vs Patrol</td><td style="padding:4px;">{time_saved:.1f} min</td></tr>
            <tr><td style="padding:4px; color:#888;">Geographic Coverage</td><td style="padding:4px;">{area_cov:.1f}%</td></tr>
        </table>
        <h4 style="color:#555; margin-bottom:10px;">Police Dept Contact</h4>
        <table style="width:100%; border-collapse:collapse; font-size:12px; margin-bottom:15px;">
            <tr><td style="padding:4px; color:#888; width:50%;">Signatory</td><td style="padding:4px;">{pd_c}</td></tr>
            <tr><td style="padding:4px; color:#888;">Department</td><td style="padding:4px;">{pd_d}</td></tr>
        </table>
        <h4 style="color:#555; margin-bottom:10px;">Active Drones Placed</h4>
        <ul style="font-size:12px; color:#444; padding-left:20px;">{drone_list}</ul>
    </div>
    """

def _build_sheets_row(city, state, event_type, k_resp, k_guard, coverage, name, email, details=None):
    """Build the flat list of values for a Google Sheets row — single source of truth."""
    d = details or {}
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    session_start = d.get('session_start', now)
    dur = d.get('session_duration_min', '')
    try:
        if dur == '':
            start_dt = datetime.datetime.strptime(session_start, "%Y-%m-%d %H:%M:%S")
            dur = round((datetime.datetime.now() - start_dt).total_seconds() / 60, 1)
    except Exception:
        dur = ''
    fm = d.get('file_meta', {})
    return [
        # ── Identity ────────────────────────────────────────────────────────
        now,                                    # A: Timestamp
        d.get('session_id', ''),               # B: Session ID
        session_start,                          # C: Session Start
        dur,                                    # D: Session Duration (min)
        d.get('data_source', ''),              # E: Data Source
        # ── Who ─────────────────────────────────────────────────────────────
        name,                                   # F: BRINC Rep Name
        email,                                  # G: BRINC Rep Email
        d.get('pd_chief', ''),                 # H: PD Chief / Signatory
        d.get('pd_dept', ''),                  # I: Department Name
        d.get('pd_dept_email', ''),            # J: Department Email
        d.get('pd_dept_phone', ''),            # K: Department Phone
        # ── Where (user-selected) ────────────────────────────────────────────
        city,                                   # L: City (user-selected)
        state,                                  # M: State (user-selected)
        d.get('population', ''),               # N: Population
        d.get('area_sq_mi', ''),               # O: Area (sq mi)
        # ── Where (file-inferred) ────────────────────────────────────────────
        fm.get('file_inferred_city', ''),      # P: City inferred from uploaded file
        fm.get('file_inferred_state', ''),     # Q: State inferred from uploaded file
        d.get('city_confirmed_match', ''),     # R: File city matched user selection (True/False)
        d.get('multi_city_targets', ''),       # S: All target cities JSON
        d.get('num_cities_targeted', ''),      # T: Count of cities analyzed
        # ── Calls ───────────────────────────────────────────────────────────
        d.get('total_calls', ''),              # U: Total Annual Calls
        d.get('daily_calls', ''),              # V: Daily Calls
        d.get('calls_per_capita', ''),         # W: Calls per Capita
        # ── Fleet ───────────────────────────────────────────────────────────
        event_type,                             # X: Event Type
        k_resp,                                 # Y: Responders
        k_guard,                                # Z: Guardians
        round(coverage, 1) if coverage else '', # AA: Call Coverage %
        d.get('area_covered_pct', ''),         # AB: Area Coverage %
        d.get('avg_response_min', ''),         # AC: Avg Response (min)
        d.get('avg_time_saved_min', ''),       # AD: Time Saved vs Patrol (min)
        # ── Financials ──────────────────────────────────────────────────────
        d.get('fleet_capex', ''),              # AE: Fleet CapEx
        d.get('annual_savings', ''),           # AF: Annual Savings
        d.get('break_even', ''),               # AG: Break-Even
        # ── Settings ────────────────────────────────────────────────────────
        d.get('opt_strategy', ''),             # AH: Opt Strategy
        d.get('dfr_rate', ''),                 # AI: DFR Rate %
        d.get('deflect_rate', ''),             # AJ: Deflection Rate %
        d.get('incremental_build', ''),        # AK: Incremental Build
        d.get('allow_redundancy', ''),         # AL: Allow Overlap
        d.get('r_resp_radius', ''),            # AM: Responder Radius (mi)
        d.get('r_guard_radius', ''),           # AN: Guardian Radius (mi)
        d.get('estimated_pop_input', ''),      # AO: Population input by user
        # ── File Data Matrix ─────────────────────────────────────────────────
        fm.get('uploaded_filename', ''),       # AP: Uploaded filename(s)
        fm.get('file_row_count', ''),          # AQ: Raw file row count
        fm.get('file_col_count', ''),          # AR: Column count
        fm.get('file_col_names', ''),          # AS: Column names JSON
        fm.get('file_date_range_start', ''),   # AT: Earliest date in data
        fm.get('file_date_range_end', ''),     # AU: Latest date in data
        fm.get('file_date_span_days', ''),     # AV: Days of history in file
        fm.get('file_null_rate_pct', ''),      # AW: Null rate % across key fields
        fm.get('file_has_lat_lon', ''),        # AX: Lat/lon detected (True/False)
        fm.get('file_has_priority', ''),       # AY: Priority col detected (True/False)
        fm.get('call_type_breakdown', ''),     # AZ: Top call types JSON
        fm.get('priority_distribution', ''),   # BA: Priority counts JSON
        fm.get('peak_hour', ''),               # BB: Peak hour of day (0-23)
        fm.get('peak_day_of_week', ''),        # BC: Peak day of week (0=Mon)
        fm.get('peak_month', ''),              # BD: Peak month (1-12)
        # ── User Interaction Signals ─────────────────────────────────────────
        d.get('boundary_kind', ''),            # BE: Boundary type (place/county)
        d.get('boundary_source_path', ''),     # BF: Shapefile path used
        d.get('sim_or_upload', ''),            # BG: simulation vs cad_upload
        d.get('onboarding_completed', ''),     # BH: Onboarding finished (True/False)
        d.get('demo_mode_used', ''),           # BI: Demo city loaded (True/False)
        d.get('export_type_sequence', ''),     # BJ: Ordered export clicks (JSON list)
        d.get('total_exports_in_session', ''),# BK: Total export clicks
        d.get('map_viewed', ''),               # BL: Map rendered this session
        # ── Drones detail (JSON) ─────────────────────────────────────────────
        json.dumps([{"name": dr.get("name"), "type": dr.get("type"),
                     "lat": dr.get("lat"), "lon": dr.get("lon"),
                     "avg_time_min": dr.get("avg_time_min"),
                     "faa_ceiling": dr.get("faa_ceiling"),
                     "annual_savings": dr.get("annual_savings")}
                    for dr in d.get('active_drones', [])]),  # BM: Drone JSON
    ]

def _notify_email(city, state, file_type, k_resp, k_guard, coverage, name, email, details=None):
    try:
        gmail_address  = st.secrets.get("GMAIL_ADDRESS", "")
        app_password   = st.secrets.get("GMAIL_APP_PASSWORD", "")
        notify_address = st.secrets.get("NOTIFY_EMAIL", gmail_address)
        if not gmail_address or not app_password: return
        emoji = {"HTML": "📄", "KML": "🌏", "BRINC": "💾", "MAP_BUILD": "🗺️"}.get(file_type, "📥")
        subject = f"{emoji} BRINC {file_type.replace('_',' ').title()} — {city}, {state}"
        details_html = _build_details_html(details)
        d = details or {}
        pop  = d.get('population', 0)
        body = f"""
        <html><body style="font-family:Arial,sans-serif;color:#333;padding:20px;">
        <div style="max-width:560px;margin:0 auto;border:1px solid #ddd;border-radius:8px;overflow:hidden;">
            <div style="background:#000;padding:16px 20px;border-bottom:3px solid #00D2FF;">
                <span style="color:#00D2FF;font-size:18px;font-weight:900;letter-spacing:2px;">BRINC</span>
                <span style="color:#888;font-size:12px;margin-left:8px;">{file_type.replace('_',' ').title()} Notification</span>
            </div>
            <div style="padding:20px;">
                <table style="width:100%;border-collapse:collapse;font-size:14px;">
                    <tr style="border-bottom:1px solid #f0f0f0;"><td style="padding:8px 4px;color:#888;width:40%;">Event</td><td style="padding:8px 4px;font-weight:bold;">{emoji} {file_type.replace('_',' ').title()}</td></tr>
                    <tr style="border-bottom:1px solid #f0f0f0;"><td style="padding:8px 4px;color:#888;">Jurisdiction</td><td style="padding:8px 4px;font-weight:bold;">{city}, {state}</td></tr>
                    <tr style="border-bottom:1px solid #f0f0f0;"><td style="padding:8px 4px;color:#888;">Population</td><td style="padding:8px 4px;">{pop:,}</td></tr>
                    <tr style="border-bottom:1px solid #f0f0f0;"><td style="padding:8px 4px;color:#888;">Fleet</td><td style="padding:8px 4px;">{k_resp} Responder · {k_guard} Guardian</td></tr>
                    <tr style="border-bottom:1px solid #f0f0f0;"><td style="padding:8px 4px;color:#888;">Call Coverage</td><td style="padding:8px 4px;">{coverage:.1f}%</td></tr>
                    <tr style="border-bottom:1px solid #f0f0f0;"><td style="padding:8px 4px;color:#888;">BRINC Rep</td><td style="padding:8px 4px;">{name if name else '—'}</td></tr>
                    <tr><td style="padding:8px 4px;color:#888;">Rep Email</td><td style="padding:8px 4px;">{f'<a href="mailto:{email}">{email}</a>' if email else '—'}</td></tr>
                </table>
                {details_html}
                <div style="margin-top:16px;font-size:11px;color:#bbb;">{datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")} UTC</div>
            </div>
        </div>
        </body></html>
        """
        msg = MIMEMultipart("alternative")
        msg["Subject"], msg["From"], msg["To"] = subject, gmail_address, notify_address
        msg.attach(MIMEText(body, "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=8) as server:
            server.login(gmail_address, app_password)
            server.sendmail(gmail_address, notify_address, msg.as_string())
    except: pass

def _log_to_sheets(city, state, file_type, k_resp, k_guard, coverage, name, email, details=None):
    try:
        sheet_id = st.secrets.get("GOOGLE_SHEET_ID", "")
        creds_dict = st.secrets.get("gcp_service_account", {})
        if not sheet_id or not creds_dict: return
        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_info(dict(creds_dict), scopes=scopes)
        client = gspread.authorize(creds)
        sheet = client.open_by_key(sheet_id).sheet1
        row = _build_sheets_row(city, state, file_type, k_resp, k_guard, coverage, name, email, details)
        sheet.append_row(row)
    except: pass

# --- GLOBAL CONFIGURATION ---
CONFIG = {
    "RESPONDER_COST": 80000, "GUARDIAN_COST": 160000, "RESPONDER_RANGE_MI": 2.0,
    "OFFICER_COST_PER_CALL": 82, "DRONE_COST_PER_CALL": 6,
    # Specialty-response upside (conservative defaults; see helper below)
    "THERMAL_DEFAULT_APPLICABLE_RATE": 0.12,
    "THERMAL_SAVINGS_PER_CALL": 38,
    "K9_DEFAULT_APPLICABLE_RATE": 0.03,
    "K9_SAVINGS_PER_CALL": 155,
    # Fire department value: aerial recon/scene size-up + overhaul hotspot detection
    # ~5% of addressable calls are fire-related; blended savings $450/call assisted
    # (15% aerial ladder avoidance at $4,500/deploy + $90 overhaul crew time saved)
    "FIRE_DEFAULT_APPLICABLE_RATE": 0.05,
    "FIRE_SAVINGS_PER_CALL": 450,
    "DEFAULT_TRAFFIC_SPEED": 35.0, "RESPONDER_SPEED": 42.0, "GUARDIAN_SPEED": 60.0,
    # Guardian duty cycle: 60 min flight + 3 min charge = 63 min cycle
    # Daily airtime = (24*60) / 63 * 60 = 1371.4 min = 22.86 hrs
    "GUARDIAN_FLIGHT_MIN":  60,   # flight minutes per cycle
    "GUARDIAN_CHARGE_MIN":   3,   # charge minutes per cycle
    # Responder duty cycle: 30-min max flight per sortie, 11.6hr shift
    "RESPONDER_FLIGHT_MIN":   30,    # max flight minutes per sortie
    "RESPONDER_PATROL_HOURS": 11.6,
    # Officer / drone cost model
    "OFFICER_HOURLY_WAGE": 37.0,     # baseline hourly wage for overtime estimates
    # Outcome rates (modeled estimates; adjust per-agency as needed)
    "OUTCOME_ARREST_RATE":      0.043,
    "OUTCOME_RESCUE_RATE":      0.021,
    "OUTCOME_DEESCALATION_RATE":0.11,
    "OUTCOME_MISSING_RATE":     0.008,
    # drone_wins_pct formula: pct = clamp(calls_covered * WINS_MULTIPLIER, WINS_FLOOR, 99)
    "DRONE_WINS_MULTIPLIER":    0.72,
    "DRONE_WINS_FLOOR":         60,
}
# Derived: compute Guardian daily airtime from duty cycle
CONFIG["GUARDIAN_DAILY_FLIGHT_MIN"] = (
    (24 * 60) / (CONFIG["GUARDIAN_FLIGHT_MIN"] + CONFIG["GUARDIAN_CHARGE_MIN"])
) * CONFIG["GUARDIAN_FLIGHT_MIN"]
CONFIG["GUARDIAN_PATROL_HOURS"] = CONFIG["GUARDIAN_DAILY_FLIGHT_MIN"] / 60
STATE_FIPS = {"AL": "01", "AK": "02", "AZ": "04", "AR": "05", "CA": "06", "CO": "08", "CT": "09", "DE": "10", "FL": "12", "GA": "13", "HI": "15", "ID": "16", "IL": "17", "IN": "18", "IA": "19", "KS": "20", "KY": "21", "LA": "22", "ME": "23", "MD": "24", "MA": "25", "MI": "26", "MN": "27", "MS": "28", "MO": "29", "MT": "30", "NE": "31", "NV": "32", "NH": "33", "NJ": "34", "NM": "35", "NY": "36", "NC": "37", "ND": "38", "OH": "39", "OK": "40", "OR": "41", "PA": "42", "RI": "44", "SC": "45", "SD": "46", "TN": "47", "TX": "48", "UT": "49", "VT": "50", "VA": "51", "WA": "53", "WV": "54", "WI": "55", "WY": "56"}
US_STATES_ABBR = {"Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR", "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY", "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT", "Vermont": "VT", "Virginia": "VA", "Washington": "WA", "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY"}
KNOWN_POPULATIONS = {"Victoria": 65534, "New York": 8336817, "Los Angeles": 3822238, "Chicago": 2665039, "Houston": 1304379, "Phoenix": 1644409, "Philadelphia": 1567258, "San Antonio": 2302878, "San Diego": 1472530, "Dallas": 1299544, "San Jose": 1381162, "Austin": 974447, "Jacksonville": 971319, "Fort Worth": 956709, "Columbus": 907971, "Indianapolis": 880621, "Charlotte": 897720, "San Francisco": 971233, "Seattle": 749256, "Denver": 713252, "Washington": 678972, "Nashville": 683622, "Oklahoma City": 694800, "El Paso": 694553, "Boston": 650706, "Portland": 635067, "Las Vegas": 656274, "Detroit": 620376, "Memphis": 633104, "Louisville": 628594, "Baltimore": 620961, "Milwaukee": 620251, "Albuquerque": 677122, "Tucson": 564559, "Fresno": 677102, "Sacramento": 808418, "Kansas City": 697738, "Mesa": 504258, "Atlanta": 499127, "Omaha": 508901, "Colorado Springs": 483956, "Raleigh": 476587, "Miami": 449514, "Virginia Beach": 455369, "Oakland": 530763, "Minneapolis": 563332, "Tulsa": 547239, "Arlington": 398654, "New Orleans": 562503, "Wichita": 402263, "Cleveland": 900000, "Tampa": 449514, "Orlando": 316081}
DEMO_CITIES = [("Las Vegas", "NV"), ("Austin", "TX"), ("Seattle", "WA"), ("Denver", "CO"), ("Nashville", "TN"), ("Columbus", "OH"), ("Detroit", "MI"), ("San Diego", "CA"), ("Charlotte", "NC"), ("Portland", "OR"), ("Memphis", "TN"), ("Louisville", "KY"), ("Baltimore", "MD"), ("Milwaukee", "WI"), ("Albuquerque", "NM"), ("Tucson", "AZ"), ("Fresno", "CA"), ("Sacramento", "CA"), ("Kansas City", "MO"), ("Mesa", "AZ"), ("Atlanta", "GA"), ("Omaha", "NE"), ("Colorado Springs", "CO"), ("Raleigh", "NC"), ("Miami", "FL"), ("Minneapolis", "MN"), ("Tulsa", "OK"), ("Arlington", "TX"), ("Tampa", "FL"), ("New Orleans", "LA"), ("Wichita", "KS"), ("Cleveland", "OH"), ("Virginia Beach", "VA"), ("Oakland", "CA"), ("Indianapolis", "IN"), ("Jacksonville", "FL"), ("Fort Worth", "TX"), ("Boston", "MA"), ("El Paso", "TX"), ("Oklahoma City", "OK"), ("Boise", "ID"), ("Richmond", "VA"), ("Spokane", "WA"), ("Tacoma", "WA"), ("Aurora", "CO"), ("Anaheim", "CA"), ("Bakersfield", "CA"), ("Riverside", "CA"), ("Stockton", "CA"), ("Corpus Christi", "TX"), ("Lexington", "KY"), ("Henderson", "NV"), ("Saint Paul", "MN"), ("Anchorage", "AK"), ("Plano", "TX"), ("Lincoln", "NE"), ("Buffalo", "NY"), ("Fort Wayne", "IN"), ("Jersey City", "NJ"), ("Chula Vista", "CA"), ("Orlando", "FL"), ("St. Louis", "MO"), ("Madison", "WI"), ("Durham", "NC"), ("Lubbock", "TX"), ("Winston-Salem", "NC"), ("Garland", "TX"), ("Glendale", "AZ"), ("Hialeah", "FL"), ("Scottsdale", "AZ"), ("Irving", "TX"), ("Fremont", "CA"), ("Baton Rouge", "LA"), ("Birmingham", "AL"), ("Rochester", "NY"), ("Des Moines", "IA"), ("Montgomery", "AL"), ("Modesto", "CA"), ("Fayetteville", "NC"), ("Shreveport", "LA"), ("Akron", "OH"), ("Grand Rapids", "MI"), ("Huntington Beach", "CA"), ("Little Rock", "AR")]
FAST_DEMO_CITIES = [("Henderson", "NV"), ("Lincoln", "NE"), ("Boise", "ID"), ("Des Moines", "IA"), ("Madison", "WI"), ("Colorado Springs", "CO"), ("Richmond", "VA"), ("Raleigh", "NC"), ("Durham", "NC"), ("Fort Wayne", "IN"), ("Omaha", "NE"), ("Wichita", "KS"), ("Tulsa", "OK"), ("Spokane", "WA"), ("Tacoma", "WA"), ("Aurora", "CO"), ("Las Vegas", "NV"), ("Nashville", "TN"), ("Columbus", "OH"), ("Charlotte", "NC"), ("Louisville", "KY"), ("Indianapolis", "IN"), ("Memphis", "TN"), ("Detroit", "MI"), ("Milwaukee", "WI"), ("Minneapolis", "MN"), ("Seattle", "WA"), ("Denver", "CO"), ("Portland", "OR"), ("Austin", "TX")]
FAA_CEILING_COLORS = {0: {"line": "rgba(255,  20,  20, 0.95)", "fill": "rgba(255,  20,  20, 0.20)"}, 50: {"line": "rgba(255, 120,   0, 0.95)", "fill": "rgba(255, 120,   0, 0.18)"}, 100: {"line": "rgba(255, 210,   0, 0.95)", "fill": "rgba(255, 210,   0, 0.18)"}, 200: {"line": "rgba(180, 230,   0, 0.95)", "fill": "rgba(180, 230,   0, 0.16)"}, 300: {"line": "rgba( 80, 200,  50, 0.95)", "fill": "rgba( 80, 200,  50, 0.16)"}, 400: {"line": "rgba(  0, 180, 100, 0.95)", "fill": "rgba(  0, 180, 100, 0.15)"}}
FAA_DEFAULT_COLOR = {"line": "rgba(150,150,150,0.8)", "fill": "rgba(150,150,150,0.10)"}
STATION_COLORS = ["#00D2FF", "#39FF14", "#FFD700", "#FF007F", "#FF4500", "#00FFCC", "#FF3333", "#7FFF00", "#00FFFF", "#FF9900"]

# --- THEME VARIABLES ---
bg_main = "#000000"
bg_sidebar = "#111111"
text_main = "#ffffff"
text_muted = "#aaaaaa"
accent_color = "#00D2FF"
card_bg = "#111111"
card_border = "#333333"
card_text = "#eeeeee"
card_title = "#ffffff"
budget_box_bg = "#0a0a0a"
budget_box_border = "#00D2FF"
budget_box_shadow = "rgba(0, 210, 255, 0.15)"
map_style = "carto-darkmatter"
map_boundary_color = "#ffffff"
map_incident_color = "#00D2FF"
legend_bg = "rgba(0, 0, 0, 0.7)"
legend_text = "#ffffff"

HERO_MESSAGES = ["🚔 Building safer communities, one drone at a time…", "🛡️ Loading data because your officers deserve better tools…", "🫡 Honoring the men and women who answer the call every day…", "💙 Officers run toward danger so the rest of us don't have to…", "🚁 Optimizing so your team gets there first — every time…", "🌟 Every second we save is a life better protected…", "🤝 Technology in service of the community's greatest heroes…", "💪 Your officers deserve every advantage we can give them…", "🙏 Dedicated to the families who wait at home while heroes serve…", "🏅 Processing data worthy of those who wear the badge with pride…", "🌃 Mapping the city your officers protect through every shift…", "🔵 Building a network as reliable as the officers who depend on it…", "❤️ Because faster response means more lives saved…", "🌅 Creating tools that let officers come home safely every night…", "🦅 Guardian drones — always watching, always ready to assist…", "🏘️ Modeling coverage for the neighborhoods they protect and serve…", "📡 Connecting technology to the courage already on the streets…", "🧠 Smart systems for smarter, safer law enforcement…", "🌟 Every data point represents a community worth protecting…", "🚨 Fewer false alarms. More real backup. Better outcomes for all…"]
FAA_MESSAGES = ["✈️ Checking FAA airspace — keeping your drones and your pilots safe…", "🛫 Loading LAANC data — because safe skies mean more missions completed…", "🗺️ Mapping controlled airspace — so every flight is a legal, safe one…", "✈️ FAA compliance check in progress — protecting officers on the ground and drones in the air…", "🛡️ Pulling airspace boundaries — safe operations start before takeoff…", "🌐 Verifying flight corridors — your pilots deserve a clear path forward…", "📡 Syncing with FAA LAANC — because your department deserves zero surprises in the sky…", "🛩️ Loading aviation data — the same skies your officers look up to every night…"]
AIRFIELD_MESSAGES = ["🏗️ Locating nearby airfields — coordinating with the aviation community that shares your skies…", "📍 Mapping airports near each station — great neighbors make great operators…", "🛬 Finding local airfields — because your team coordinates with everyone keeping the community safe…", "✈️ Scanning for nearby aviation assets — your drones respect every aircraft they share the sky with…", "🗺️ Identifying airport proximity — so your officers always know what's overhead…", "🤝 Locating nearby airfields — collaboration between aviation and law enforcement saves lives…", "📡 Querying aviation infrastructure — the sky belongs to everyone who protects this community…"]
JURISDICTION_MESSAGES = ["🗺️ Identifying jurisdictions — every boundary represents a community counting on you…", "📐 Loading geographic boundaries — the lines officers cross every shift to keep people safe…", "🏙️ Mapping your jurisdiction — the streets your officers know better than anyone…", "🌆 Matching data to boundaries — every block is someone's home, someone's neighborhood…", "📍 Finding your coverage area — the community that trusts you with their safety…", "🗺️ Resolving jurisdictions — where every call for help deserves an answer…"]
SPATIAL_MESSAGES = ["⚡ Crunching coverage geometry — because your officers deserve precision, not guesswork…", "🧮 Computing spatial matrices — doing the math so your team can focus on what matters…", "📊 Building coverage model — every calculation brings faster response one step closer…", "🔬 Analyzing incident patterns — understanding the city so your officers can better protect it…", "💡 Optimizing station geometry — smart placement means no neighborhood is left behind…", "🧠 Modeling response zones — technology standing behind the officers who stand for us…"]

def get_hero_message(): return random.choice(HERO_MESSAGES)
def get_faa_message(): return random.choice(FAA_MESSAGES)
def get_airfield_message(): return random.choice(AIRFIELD_MESSAGES)
def get_jurisdiction_message(): return random.choice(JURISDICTION_MESSAGES)
def get_spatial_message(): return random.choice(SPATIAL_MESSAGES)


def _count_points_within_boundary(df_calls, boundary_gdf):
    try:
        if df_calls is None or len(df_calls) == 0 or boundary_gdf is None or getattr(boundary_gdf, "empty", True):
            return 0
        work = df_calls.copy()
        work['lat'] = pd.to_numeric(work['lat'], errors='coerce')
        work['lon'] = pd.to_numeric(work['lon'], errors='coerce')
        work = work.dropna(subset=['lat', 'lon'])
        if work.empty:
            return 0
        pts = gpd.GeoDataFrame(work, geometry=gpd.points_from_xy(work.lon, work.lat), crs="EPSG:4326")
        try:
            boundary_4326 = boundary_gdf.to_crs(epsg=4326)
        except Exception:
            boundary_4326 = boundary_gdf
        poly = unary_union(boundary_4326.geometry)
        inside = pts.within(poly)
        return int(inside.sum())
    except Exception:
        return 0


def find_jurisdictions_by_coordinates(df_calls, min_call_share=0.001, min_call_count=3):
    """
    Purely coordinate-driven jurisdiction lookup.

    Spatially joins call points against places_lite.parquet (and
    counties_lite.parquet as fallback) to find every jurisdiction that
    contains at least `min_call_share` of the uploaded calls OR at least
    `min_call_count` absolute calls.  Returns a GeoDataFrame with columns
    [DISPLAY_NAME, data_count, geometry] sorted descending by data_count,
    or None if nothing is found.
    """
    import traceback as _tb
    _debug_msgs = []
    try:
        if df_calls is None or df_calls.empty:
            _debug_msgs.append("df_calls is None or empty")
            st.session_state['_jur_debug'] = _debug_msgs
            return None

        pts = df_calls.copy()
        pts['lat'] = pd.to_numeric(pts['lat'], errors='coerce')
        pts['lon'] = pd.to_numeric(pts['lon'], errors='coerce')
        pts = pts.dropna(subset=['lat', 'lon'])
        if pts.empty:
            _debug_msgs.append("no valid lat/lon after cleaning")
            st.session_state['_jur_debug'] = _debug_msgs
            return None

        _debug_msgs.append(f"pts count: {len(pts)}, lat range: {pts['lat'].min():.3f}–{pts['lat'].max():.3f}, lon range: {pts['lon'].min():.3f}–{pts['lon'].max():.3f}")

        sample = pts.sample(min(len(pts), 5000), random_state=42) if len(pts) > 5000 else pts
        pts_gdf = gpd.GeoDataFrame(sample, geometry=gpd.points_from_xy(sample.lon, sample.lat), crs='EPSG:4326')
        bbox = tuple(pts_gdf.total_bounds)
        _debug_msgs.append(f"bbox: {bbox}")

        results = []

        for parquet_file, name_col, kind in [
            ('places_lite.parquet',  'NAME', 'place'),
            ('counties_lite.parquet','NAME', 'county'),
        ]:
            _exists = os.path.exists(parquet_file)
            _debug_msgs.append(f"{parquet_file} exists: {_exists}")
            if not _exists:
                continue
            try:
                # Full read then cx filter — bbox kwarg not supported by this parquet build
                poly_gdf = gpd.read_parquet(parquet_file)
                poly_gdf = poly_gdf.cx[bbox[0]:bbox[2], bbox[1]:bbox[3]]

                _debug_msgs.append(f"{parquet_file} rows in bbox: {len(poly_gdf)}")
                if poly_gdf is None or poly_gdf.empty:
                    continue
                if poly_gdf.crs is None:
                    poly_gdf = poly_gdf.set_crs(epsg=4326)
                poly_gdf = poly_gdf.to_crs(epsg=4326)

                joined = gpd.sjoin(pts_gdf[['geometry']], poly_gdf[[name_col, 'geometry']], how='left', predicate='within')
                hit_counts = joined[name_col].value_counts().dropna()
                _debug_msgs.append(f"sjoin hits: {dict(list(hit_counts.items())[:10])}")
                if hit_counts.empty:
                    continue

                total = hit_counts.sum()
                for jname, cnt in hit_counts.items():
                    # Include if meets share threshold OR absolute minimum count
                    if cnt / total < min_call_share and cnt < min_call_count:
                        continue
                    row = poly_gdf[poly_gdf[name_col] == jname].copy()
                    if row.empty:
                        continue
                    display = str(jname)
                    if kind == 'county' and not display.lower().endswith('county'):
                        display = display + ' County'
                    already = any(r['DISPLAY_NAME'] == display for r in results)
                    if not already:
                        results.append({
                            'DISPLAY_NAME': display,
                            'data_count': int(cnt),
                            'geometry': row.geometry.iloc[0],
                        })
            except Exception as _e:
                _debug_msgs.append(f"{parquet_file} ERROR: {_e}\n{_tb.format_exc()[-300:]}")
                continue

            if results and parquet_file.startswith('places'):
                break

        _debug_msgs.append(f"total results: {len(results)}, names: {[r['DISPLAY_NAME'] for r in results]}")
        st.session_state['_jur_debug'] = _debug_msgs

        if not results:
            return None

        out = gpd.GeoDataFrame(results, crs='EPSG:4326')
        out = out.sort_values('data_count', ascending=False).reset_index(drop=True)
        return out

    except Exception as _e:
        _debug_msgs.append(f"OUTER ERROR: {_e}\n{_tb.format_exc()[-400:]}")
        st.session_state['_jur_debug'] = _debug_msgs
        return None
    """
    Purely coordinate-driven jurisdiction lookup.

    Spatially joins call points against places_lite.parquet (and
    counties_lite.parquet as fallback) to find every jurisdiction that
    contains at least `min_call_share` of the uploaded calls.  Returns a
    GeoDataFrame with columns [DISPLAY_NAME, data_count, geometry] sorted
    descending by data_count, or None if nothing is found.

    This is the primary boundary-detection path for uploaded CAD files.
    It requires no city/state name inference — it works purely from lat/lon.
    """
    import traceback as _tb
    _debug_msgs = []
    try:
        if df_calls is None or df_calls.empty:
            _debug_msgs.append("df_calls is None or empty")
            st.session_state['_jur_debug'] = _debug_msgs
            return None

        pts = df_calls.copy()
        pts['lat'] = pd.to_numeric(pts['lat'], errors='coerce')
        pts['lon'] = pd.to_numeric(pts['lon'], errors='coerce')
        pts = pts.dropna(subset=['lat', 'lon'])
        if pts.empty:
            _debug_msgs.append("no valid lat/lon after cleaning")
            st.session_state['_jur_debug'] = _debug_msgs
            return None

        _debug_msgs.append(f"pts count: {len(pts)}, lat range: {pts['lat'].min():.3f}–{pts['lat'].max():.3f}, lon range: {pts['lon'].min():.3f}–{pts['lon'].max():.3f}")

        # Sample for speed on large files (spatial join is O(n·m))
        sample = pts.sample(min(len(pts), 5000), random_state=42) if len(pts) > 5000 else pts
        pts_gdf = gpd.GeoDataFrame(sample, geometry=gpd.points_from_xy(sample.lon, sample.lat), crs='EPSG:4326')
        bbox = tuple(pts_gdf.total_bounds)  # (minx, miny, maxx, maxy)
        _debug_msgs.append(f"bbox: {bbox}")

        results = []

        for parquet_file, name_col, kind in [
            ('places_lite.parquet',  'NAME', 'place'),
            ('counties_lite.parquet','NAME', 'county'),
        ]:
            _exists = os.path.exists(parquet_file)
            _debug_msgs.append(f"{parquet_file} exists: {_exists}")
            if not _exists:
                continue
            try:
                try:
                    poly_gdf = gpd.read_parquet(parquet_file, bbox=bbox)
                except Exception as _be:
                    _debug_msgs.append(f"bbox read failed ({_be}), falling back to full read")
                    poly_gdf = gpd.read_parquet(parquet_file)
                    poly_gdf = poly_gdf.cx[bbox[0]:bbox[2], bbox[1]:bbox[3]]

                _debug_msgs.append(f"{parquet_file} rows in bbox: {len(poly_gdf)}")
                if poly_gdf is None or poly_gdf.empty:
                    continue
                if poly_gdf.crs is None:
                    poly_gdf = poly_gdf.set_crs(epsg=4326)
                poly_gdf = poly_gdf.to_crs(epsg=4326)

                joined = gpd.sjoin(pts_gdf[['geometry']], poly_gdf[[name_col, 'geometry']], how='left', predicate='within')
                hit_counts = joined[name_col].value_counts().dropna()
                _debug_msgs.append(f"sjoin hits: {dict(list(hit_counts.items())[:5])}")
                if hit_counts.empty:
                    continue

                total = hit_counts.sum()
                for jname, cnt in hit_counts.items():
                    if cnt / total < min_call_share:
                        continue
                    row = poly_gdf[poly_gdf[name_col] == jname].copy()
                    if row.empty:
                        continue
                    display = str(jname)
                    if kind == 'county' and not display.lower().endswith('county'):
                        display = display + ' County'
                    already = any(r['DISPLAY_NAME'] == display for r in results)
                    if not already:
                        results.append({
                            'DISPLAY_NAME': display,
                            'data_count': int(cnt),
                            'geometry': row.geometry.iloc[0],
                        })
            except Exception as _e:
                _debug_msgs.append(f"{parquet_file} ERROR: {_e}\n{_tb.format_exc()[-300:]}")
                continue

            if results and parquet_file.startswith('places'):
                break

        _debug_msgs.append(f"total results: {len(results)}, names: {[r['DISPLAY_NAME'] for r in results]}")
        st.session_state['_jur_debug'] = _debug_msgs

        if not results:
            return None

        out = gpd.GeoDataFrame(results, crs='EPSG:4326')
        out = out.sort_values('data_count', ascending=False).reset_index(drop=True)
        return out

    except Exception as _e:
        _debug_msgs.append(f"OUTER ERROR: {_e}\n{_tb.format_exc()[-400:]}")
        st.session_state['_jur_debug'] = _debug_msgs
        return None
    """
    Purely coordinate-driven jurisdiction lookup.

    Spatially joins call points against places_lite.parquet (and
    counties_lite.parquet as fallback) to find every jurisdiction that
    contains at least `min_call_share` of the uploaded calls.  Returns a
    GeoDataFrame with columns [DISPLAY_NAME, data_count, geometry] sorted
    descending by data_count, or None if nothing is found.

    This is the primary boundary-detection path for uploaded CAD files.
    It requires no city/state name inference — it works purely from lat/lon.
    """
    try:
        if df_calls is None or df_calls.empty:
            return None

        pts = df_calls.copy()
        pts['lat'] = pd.to_numeric(pts['lat'], errors='coerce')
        pts['lon'] = pd.to_numeric(pts['lon'], errors='coerce')
        pts = pts.dropna(subset=['lat', 'lon'])
        if pts.empty:
            return None

        # Sample for speed on large files (spatial join is O(n·m))
        sample = pts.sample(min(len(pts), 5000), random_state=42) if len(pts) > 5000 else pts
        pts_gdf = gpd.GeoDataFrame(sample, geometry=gpd.points_from_xy(sample.lon, sample.lat), crs='EPSG:4326')
        bbox = tuple(pts_gdf.total_bounds)  # (minx, miny, maxx, maxy)

        results = []

        for parquet_file, name_col, kind in [
            ('places_lite.parquet',  'NAME', 'place'),
            ('counties_lite.parquet','NAME', 'county'),
        ]:
            if not os.path.exists(parquet_file):
                continue
            try:
                poly_gdf = gpd.read_parquet(parquet_file, bbox=bbox)
                if poly_gdf is None or poly_gdf.empty:
                    # bbox kwarg not supported by all versions — fall back to full read + filter
                    poly_gdf = gpd.read_parquet(parquet_file)
                    poly_gdf = poly_gdf.cx[bbox[0]:bbox[2], bbox[1]:bbox[3]]
                if poly_gdf.empty:
                    continue
                if poly_gdf.crs is None:
                    poly_gdf = poly_gdf.set_crs(epsg=4326)
                poly_gdf = poly_gdf.to_crs(epsg=4326)

                joined = gpd.sjoin(pts_gdf[['geometry']], poly_gdf[[name_col, 'geometry']], how='left', predicate='within')
                hit_counts = joined[name_col].value_counts().dropna()
                if hit_counts.empty:
                    continue

                total = hit_counts.sum()
                for jname, cnt in hit_counts.items():
                    if cnt / total < min_call_share:
                        continue
                    row = poly_gdf[poly_gdf[name_col] == jname].copy()
                    if row.empty:
                        continue
                    display = str(jname)
                    if kind == 'county' and not display.lower().endswith('county'):
                        display = display + ' County'
                    # Avoid duplicating a place already captured from places_lite
                    already = any(r['DISPLAY_NAME'] == display for r in results)
                    if not already:
                        results.append({
                            'DISPLAY_NAME': display,
                            'data_count': int(cnt),
                            'geometry': row.geometry.iloc[0],
                        })
            except Exception:
                continue

            # If places gave us results, don't also add county duplicates
            if results and parquet_file.startswith('places'):
                break

        if not results:
            return None

        out = gpd.GeoDataFrame(results, crs='EPSG:4326')
        out = out.sort_values('data_count', ascending=False).reset_index(drop=True)
        return out

    except Exception:
        return None


def _select_best_boundary_for_calls(df_calls, city_text, state_abbr, prefer_county=False):
    """Try place and county boundaries and keep the candidate containing the most uploaded calls."""
    candidates = []

    try:
        place_success, place_gdf = fetch_place_boundary_local(state_abbr, city_text)
        if place_success and place_gdf is not None and not place_gdf.empty:
            candidates.append(('place', place_gdf, _count_points_within_boundary(df_calls, place_gdf)))
    except Exception:
        pass

    county_names = [city_text]
    if not str(city_text).lower().endswith(" county"):
        county_names.append(f"{city_text} County")

    for cname in county_names:
        try:
            county_success, county_gdf = fetch_county_boundary_local(state_abbr, cname)
            if county_success and county_gdf is not None and not county_gdf.empty:
                candidates.append(('county', county_gdf, _count_points_within_boundary(df_calls, county_gdf)))
                break
        except Exception:
            pass

    if not candidates:
        # ── TIGER fallback: parquet not present or city not found — download from Census ──
        state_fips = STATE_FIPS.get(state_abbr)
        if state_fips:
            try:
                tiger_success, tiger_gdf = fetch_tiger_city_shapefile(state_fips, city_text, SHAPEFILE_DIR)
                if tiger_success and tiger_gdf is not None and not tiger_gdf.empty:
                    tiger_gdf = tiger_gdf.copy()
                    if 'NAME' not in tiger_gdf.columns:
                        tiger_gdf['NAME'] = city_text
                    hits = _count_points_within_boundary(df_calls, tiger_gdf)
                    candidates.append(('place', tiger_gdf, hits))
            except Exception:
                pass

    if not candidates:
        return False, None, 'place', 0

    if prefer_county:
        candidates.sort(key=lambda x: (x[2], 1 if x[0] == 'county' else 0), reverse=True)
    else:
        candidates.sort(key=lambda x: (x[2], 1 if x[0] == 'place' else 0), reverse=True)

    best_kind, best_gdf, best_hits = candidates[0]
    return True, best_gdf, best_kind, int(best_hits)

def get_base64_of_bin_file(bin_file):
    try:
        with open(bin_file, 'rb') as f: return base64.b64encode(f.read()).decode()
    except Exception: return None


def get_themed_logo_base64(logo_file="logo.png", theme="dark"):
    """Return a recolored transparent PNG logo as base64.

    theme='dark'  -> white logo on transparent background
    theme='light' -> black logo on transparent background
    """
    try:
        target_rgb = (255, 255, 255) if str(theme).lower() == 'dark' else (0, 0, 0)
        with Image.open(logo_file).convert('RGBA') as img:
            alpha = img.getchannel('A')
            recolored = Image.new('RGBA', img.size, target_rgb + (0,))
            recolored.putalpha(alpha)
            buf = io.BytesIO()
            recolored.save(buf, format='PNG')
            return base64.b64encode(buf.getvalue()).decode()
    except Exception:
        return get_base64_of_bin_file(logo_file)


def get_transparent_product_base64(image_file="gigs.png", threshold=32):
    """Return product image as transparent PNG by removing near-black background."""
    try:
        with Image.open(image_file).convert('RGBA') as img:
            arr = np.array(img)
            r, g, b, a = arr[:, :, 0], arr[:, :, 1], arr[:, :, 2], arr[:, :, 3]
            mask = (r <= threshold) & (g <= threshold) & (b <= threshold)
            arr[:, :, 3] = np.where(mask, 0, a)
            result = Image.fromarray(arr, 'RGBA')
            buf = io.BytesIO()
            result.save(buf, format='PNG')
            return base64.b64encode(buf.getvalue()).decode()
    except Exception:
        return get_base64_of_bin_file(image_file)

# ============================================================
# COMMAND CENTER ANALYTICS GENERATOR
# ============================================================

def _detect_datetime_series_for_labels(df):
    """Return a best-effort parsed datetime series from common CAD field patterns."""
    if df is None or len(df) == 0:
        return None
    try:
        if 'date' in df.columns and 'time' in df.columns:
            s = pd.to_datetime(df['date'].astype(str).fillna('') + ' ' + df['time'].astype(str).fillna(''), errors='coerce')
            if s.notna().sum() > 0:
                return s
        if 'date' in df.columns:
            s = pd.to_datetime(df['date'], errors='coerce')
            if s.notna().sum() > 0:
                return s
        candidates = [
            'createdtime_central', 'created time', 'createdtime', 'call datetime', 'calldatetime',
            'timestamp', 'datetime', 'incident datetime', 'received time', 'time received',
            'dispatch datetime', 'event time', 'event datetime'
        ]
        col_map = {str(c).strip().lower(): c for c in df.columns}
        for cand in candidates:
            col = cand if cand in df.columns else col_map.get(cand)
            if col is not None:
                s = pd.to_datetime(df[col], errors='coerce')
                if s.notna().sum() > 0:
                    return s
    except Exception:
        return None
    return None



def estimate_high_activity_overtime(df_calls_full, state_abbr, calls_covered_perc, dfr_dispatch_rate, deflection_rate):
    """Estimate high-activity monthly staffing pressure and officer overtime replacement cost."""
    if df_calls_full is None or len(df_calls_full) == 0:
        return None
    try:
        dt = _detect_datetime_series_for_labels(df_calls_full)
        if dt is None:
            return None
        work = df_calls_full.copy()
        work['_dt'] = pd.to_datetime(dt, errors='coerce')
        work = work.dropna(subset=['_dt'])
        if work.empty:
            return None

        work['_month'] = work['_dt'].dt.to_period('M').astype(str)
        work['_hour_key'] = work['_dt'].dt.floor('H')
        hourly = work.groupby('_hour_key').size().rename('calls').reset_index()
        hourly['_month'] = hourly['_hour_key'].dt.to_period('M').astype(str)

        monthly_rows = []
        for month, grp in hourly.groupby('_month'):
            if grp.empty:
                continue
            threshold = grp['calls'].quantile(0.75)
            busy = grp[grp['calls'] >= threshold].copy()
            if busy.empty:
                busy = grp.nlargest(max(1, int(len(grp) * 0.25)), 'calls')
            total_busy_calls = float(busy['calls'].sum())
            busy_hours = int(busy['_hour_key'].nunique())
            if busy_hours <= 0:
                continue

            officer_hourly, wage_source = CONFIG['OFFICER_HOURLY_WAGE'], 'estimate'
            overtime_hourly = officer_hourly * 1.5

            drone_relief_share = (calls_covered_perc / 100.0) * dfr_dispatch_rate * deflection_rate
            residual_busy_calls = total_busy_calls * max(0.0, 1.0 - drone_relief_share)
            overtime_cost = residual_busy_calls * CONFIG['OFFICER_COST_PER_CALL']
            overtime_hours = overtime_cost / overtime_hourly if overtime_hourly > 0 else 0.0

            monthly_rows.append({
                'month': month,
                'busy_hours': busy_hours,
                'busy_calls': total_busy_calls,
                'residual_calls': residual_busy_calls,
                'ot_hourly': overtime_hourly,
                'ot_cost': overtime_cost,
                'ot_hours': overtime_hours,
                'wage_source': wage_source,
            })

        if not monthly_rows:
            return None
        monthly_df = pd.DataFrame(monthly_rows).sort_values('month')
        return {
            'monthly': monthly_df,
            'avg_busy_hours': float(monthly_df['busy_hours'].mean()),
            'avg_ot_hours': float(monthly_df['ot_hours'].mean()),
            'avg_ot_cost': float(monthly_df['ot_cost'].mean()),
            'ot_hourly': float(monthly_df['ot_hourly'].median()),
            'wage_source': monthly_df['wage_source'].iloc[0],
            'peak_month': monthly_df.loc[monthly_df['ot_cost'].idxmax(), 'month'],
            'peak_ot_cost': float(monthly_df['ot_cost'].max()),
        }
    except Exception:
        return None



def estimate_specialty_response_savings(df_calls_full, total_calls_annual, calls_covered_perc=100.0):
    """Estimate additional annual savings from thermal-enabled search efficiency and avoided K-9 deployments.

    The model intentionally stays conservative:
    - Thermal value is applied to a subset of addressable calls that often benefit from search / locate / perimeter support.
    - K-9 value is applied to a smaller subset of addressable calls that are likely to require tracking or perimeter work.
    - If CAD call types are available, the function uses them; otherwise it falls back to conservative defaults.
    """
    addressable_calls = max(0.0, float(total_calls_annual or 0) * max(0.0, min(1.0, float(calls_covered_perc or 0) / 100.0)))
    out = {
        'addressable_calls_annual': addressable_calls,
        'thermal_rate': float(CONFIG["THERMAL_DEFAULT_APPLICABLE_RATE"]),
        'k9_rate': float(CONFIG["K9_DEFAULT_APPLICABLE_RATE"]),
        'fire_rate': float(CONFIG["FIRE_DEFAULT_APPLICABLE_RATE"]),
        'thermal_calls_annual': 0.0,
        'k9_calls_annual': 0.0,
        'fire_calls_annual': 0.0,
        'thermal_savings': 0.0,
        'k9_savings': 0.0,
        'fire_savings': 0.0,
        'additional_savings_total': 0.0,
        'source': 'default_model',
    }
    if addressable_calls <= 0:
        return out

    call_type_col = None
    if df_calls_full is not None and len(df_calls_full) > 0:
        for c in ['call_type_desc','agencyeventtypecodedesc','calldesc','description','nature','event_desc']:
            if c in df_calls_full.columns and df_calls_full[c].dropna().shape[0] > 0:
                call_type_col = c
                break

    if call_type_col is not None:
        s = df_calls_full[call_type_col].fillna('').astype(str).str.lower().str.strip()
        if not s.empty:
            thermal_pattern = (
                r'suspicious|prowler|alarm|burglar|robbery|theft|assault|shots|gun|weapon|person search|search|perimeter|'
                r'missing|welfare|suicid|disturbance|fight|domestic|trespass|subject check|unknown trouble|wanted'
            )
            k9_pattern = (
                r'k-?9|canine|track|tracking|perimeter|search|search warrant|manhunt|flee|fled|foot pursuit|'
                r'missing|burglary|robbery|woods|field|suspect search'
            )
            thermal_rate_raw = float(s.str.contains(thermal_pattern, regex=True, na=False).mean())
            k9_rate_raw = float(s.str.contains(k9_pattern, regex=True, na=False).mean())
            fire_pattern = (
                r'fire|structure fire|building fire|fire alarm|alarm fire|brush fire|grass fire|'
                r'wildfire|vegetation fire|vehicle fire|dumpster fire|smoke|smoke investigation|'
                r'odor of smoke|fire investigation|carbon monoxide|co alarm|gas leak|hazmat'
            )
            fire_rate_raw = float(s.str.contains(fire_pattern, regex=True, na=False).mean())
            out['thermal_rate'] = min(0.25, max(CONFIG["THERMAL_DEFAULT_APPLICABLE_RATE"] * 0.5, thermal_rate_raw if thermal_rate_raw > 0 else CONFIG["THERMAL_DEFAULT_APPLICABLE_RATE"]))
            out['k9_rate'] = min(0.08, max(CONFIG["K9_DEFAULT_APPLICABLE_RATE"] * 0.5, k9_rate_raw if k9_rate_raw > 0 else CONFIG["K9_DEFAULT_APPLICABLE_RATE"]))
            out['fire_rate'] = min(0.20, max(CONFIG["FIRE_DEFAULT_APPLICABLE_RATE"] * 0.5, fire_rate_raw if fire_rate_raw > 0 else CONFIG["FIRE_DEFAULT_APPLICABLE_RATE"]))
            out['source'] = f'cad_call_types:{call_type_col}'

    out['thermal_calls_annual'] = addressable_calls * out['thermal_rate']
    out['k9_calls_annual'] = addressable_calls * out['k9_rate']
    out['fire_calls_annual'] = addressable_calls * out['fire_rate']
    out['thermal_savings'] = out['thermal_calls_annual'] * float(CONFIG["THERMAL_SAVINGS_PER_CALL"])
    out['k9_savings'] = out['k9_calls_annual'] * float(CONFIG["K9_SAVINGS_PER_CALL"])
    out['fire_savings'] = out['fire_calls_annual'] * float(CONFIG["FIRE_SAVINGS_PER_CALL"])
    out['additional_savings_total'] = out['thermal_savings'] + out['k9_savings'] + out['fire_savings']
    return out

def build_high_activity_staffing_html(overtime_stats, dark=True, compact=False):
    """Return an HTML block for the High-Activity Staffing Pressure section."""
    if overtime_stats is None:
        return ""
    bg = "#06060a" if dark else "#ffffff"
    card = "#0c0c12" if dark else "#ffffff"
    border = "#1a1a26" if dark else "#e5e7eb"
    text_main = "#e8e8f2" if dark else "#111118"
    text_muted = "#7777a0" if dark else "#6b7280"
    accent = "#00D2FF"
    title_size = "13px" if compact else "0.95rem"
    body_size = "11px" if compact else "0.72rem"
    metric_size = "24px" if compact else "1.45rem"
    monthly_rows_html = "".join([
        f"<tr><td style='padding:6px 8px; border-top:1px solid {border}; color:{text_main};'>{row.month}</td>"
        f"<td style='padding:6px 8px; border-top:1px solid {border}; text-align:right; color:{text_main};'>{int(row.busy_hours):,}</td>"
        f"<td style='padding:6px 8px; border-top:1px solid {border}; text-align:right; color:{text_main};'>{row.ot_hours:,.0f}</td>"
        f"<td style='padding:6px 8px; border-top:1px solid {border}; text-align:right; color:{accent};'>${row.ot_cost:,.0f}</td></tr>"
        for row in overtime_stats['monthly'].itertuples(index=False)
    ])
    return f"""
    <div style="background:{bg}; border:1px solid {border}; border-radius:8px; padding:16px 18px; margin:14px 0 14px 0;">
        <div style="display:flex; justify-content:space-between; align-items:flex-end; gap:12px; flex-wrap:wrap; margin-bottom:10px;">
            <div>
                <div style="font-size:10px; color:{text_muted}; text-transform:uppercase; letter-spacing:1px; margin-bottom:4px;">High-Activity Staffing Pressure</div>
                <div style="font-size:{title_size}; color:{text_main}; font-weight:800;">Estimated officer overtime needed to cover residual peak demand</div>
            </div>
            <div style="font-size:10px; color:{text_muted};">Officer wage basis: <span style="color:{text_main};">{overtime_stats['wage_source']}</span></div>
        </div>
        <div style="display:grid; grid-template-columns:repeat(4,1fr); gap:10px; margin-bottom:12px;">
            <div style="background:{card}; border:1px solid {border}; border-radius:6px; padding:10px; text-align:center;">
                <div style="font-size:10px; color:{text_muted}; text-transform:uppercase;">Avg High-Activity Hours / Mo</div>
                <div style="font-size:{metric_size}; font-weight:800; color:{text_main}; font-family:'IBM Plex Mono', monospace;">{overtime_stats['avg_busy_hours']:.0f}</div>
            </div>
            <div style="background:{card}; border:1px solid {border}; border-radius:6px; padding:10px; text-align:center;">
                <div style="font-size:10px; color:{text_muted}; text-transform:uppercase;">Avg OT Hours Needed / Mo</div>
                <div style="font-size:{metric_size}; font-weight:800; color:{text_main}; font-family:'IBM Plex Mono', monospace;">{overtime_stats['avg_ot_hours']:.0f}</div>
            </div>
            <div style="background:{card}; border:1px solid {border}; border-radius:6px; padding:10px; text-align:center;">
                <div style="font-size:10px; color:{text_muted}; text-transform:uppercase;">Avg OT Cost / Mo</div>
                <div style="font-size:{metric_size}; font-weight:800; color:{accent}; font-family:'IBM Plex Mono', monospace;">${overtime_stats['avg_ot_cost']:,.0f}</div>
            </div>
            <div style="background:{card}; border:1px solid {border}; border-radius:6px; padding:10px; text-align:center;">
                <div style="font-size:10px; color:{text_muted}; text-transform:uppercase;">Avg OT Hourly Rate</div>
                <div style="font-size:{metric_size}; font-weight:800; color:{text_main}; font-family:'IBM Plex Mono', monospace;">${overtime_stats['ot_hourly']:.2f}</div>
            </div>
        </div>
        <div style="font-size:{body_size}; color:{text_muted}; margin-bottom:10px;">Peak month: <span style="color:{text_main}; font-weight:700;">{overtime_stats['peak_month']}</span> · estimated OT spend <span style="color:{accent}; font-weight:700;">${overtime_stats['peak_ot_cost']:,.0f}</span></div>
        <div style="overflow-x:auto;">
            <table style="width:100%; border-collapse:collapse; font-size:{body_size};">
                <thead>
                    <tr>
                        <th style="text-align:left; padding:6px 8px; color:{text_muted}; font-weight:700; border-bottom:1px solid {border};">Month</th>
                        <th style="text-align:right; padding:6px 8px; color:{text_muted}; font-weight:700; border-bottom:1px solid {border};">High-Activity Hours</th>
                        <th style="text-align:right; padding:6px 8px; color:{text_muted}; font-weight:700; border-bottom:1px solid {border};">OT Hours</th>
                        <th style="text-align:right; padding:6px 8px; color:{text_muted}; font-weight:700; border-bottom:1px solid {border};">OT Cost</th>
                    </tr>
                </thead>
                <tbody>{monthly_rows_html}</tbody>
            </table>
        </div>
    </div>
    """

def generate_command_center_html(df, total_orig_calls, export_mode=False):
    """Generates the full Command Center visual suite with interactive Javascript filtering."""
    if df is None or df.empty:
        return "<div style='color:gray; padding:20px;'>Analytics unavailable. No incident records loaded.</div>"

    import calendar as _cal
    import json

    df_ana = df.copy()

    dt_obj = None
    if 'date' in df_ana.columns:
        _date_str = df_ana['date'].astype(str).fillna('')
        if 'time' in df_ana.columns:
            _time_str = df_ana['time'].astype(str).fillna('')
            _combined = _date_str + ' ' + _time_str
        else:
            _combined = _date_str
        # Try ISO format first (what our parser stores), then fall back
        for _fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d']:
            try:
                _trial = pd.to_datetime(_combined, format=_fmt, errors='coerce')
                if _trial.notna().sum() > len(df_ana) * 0.5:
                    dt_obj = _trial
                    break
            except Exception:
                continue
        if dt_obj is None or dt_obj.dropna().empty:
            dt_obj = pd.to_datetime(_combined, errors='coerce')

    if dt_obj is None or dt_obj.dropna().empty:
        dt_candidates = [
            'createdtime_central', 'created time', 'createdtime', 'call datetime', 'calldatetime',
            'timestamp', 'datetime', 'incident datetime', 'received time', 'time received',
            'dispatch datetime', 'event time', 'event datetime'
        ]
        for cand in dt_candidates:
            if cand in df_ana.columns:
                trial = pd.to_datetime(df_ana[cand], errors='coerce')
                if trial.dropna().shape[0] > 0:
                    dt_obj = trial
                    break

    if dt_obj is None or dt_obj.dropna().empty:
        col_map = {str(c).strip().lower(): c for c in df_ana.columns}
        for cand in [
            'createdtime_central', 'created time', 'createdtime', 'call datetime', 'calldatetime',
            'timestamp', 'datetime', 'incident datetime', 'received time', 'time received',
            'dispatch datetime', 'event time', 'event datetime'
        ]:
            if cand in col_map:
                _col_real = col_map[cand]
                _fmt_try = ['%m/%d/%Y %I:%M %p','%m/%d/%Y %H:%M:%S','%Y-%m-%d %H:%M:%S',
                            '%Y-%m-%dT%H:%M:%S','%m/%d/%Y %H:%M','%Y-%m-%d %H:%M']
                for _fmt in _fmt_try:
                    try:
                        trial = pd.to_datetime(df_ana[_col_real], format=_fmt, errors='coerce')
                        if trial.notna().sum() > 0:
                            dt_obj = trial
                            break
                    except Exception:
                        continue
                if dt_obj is not None and dt_obj.dropna().shape[0] > 0:
                    break
                # Last resort: dateutil inference
                if dt_obj is None or dt_obj.dropna().empty:
                    trial = pd.to_datetime(df_ana[_col_real], errors='coerce')
                    if trial.dropna().shape[0] > 0:
                        dt_obj = trial
                        break

    # Universal scan: try every object column as a datetime source
    if dt_obj is None or dt_obj.dropna().empty:
        for _col in df_ana.select_dtypes(include='object').columns:
            try:
                _samp = df_ana[_col].dropna().head(20)
                _trial = pd.to_datetime(_samp, errors='coerce')
                if _trial.notna().sum() >= 10 and _trial.dt.year.between(2000, 2035).mean() > 0.8:
                    dt_obj = pd.to_datetime(df_ana[_col], errors='coerce')
                    break
            except Exception:
                continue

    if dt_obj is None or dt_obj.dropna().empty:
        return "<div style='color:gray; padding:20px;'>Analytics unavailable. Missing date/time fields.</div>"

    df_ana['dt_obj'] = dt_obj
    df_ana = df_ana.dropna(subset=['dt_obj'])
    if df_ana.empty: return "<div>No valid dates found in data.</div>"

    # 1. Parse dates and build a lightweight records array for JavaScript
    records = []
    for _, r in df_ana.iterrows():
        dt = r['dt_obj']
        p_val = str(r['priority']).upper().strip() if 'priority' in r else 'UNKNOWN'
        if p_val == 'NAN' or not p_val: p_val = 'UNKNOWN'
        records.append({
            'd': dt.strftime('%Y-%m-%d'),
            'h': dt.hour,
            'dow': dt.dayofweek, # Mon=0, Sun=6
            'p': p_val
        })
        
    # 2. Dynamically identify all priority types
    unique_pris = sorted(list(set(r['p'] for r in records)))
    options_html = '<option value="ALL">ALL PRIORITIES</option>'
    for p in unique_pris:
        # Clean up the display name slightly
        display_p = f"PRIORITY {p}" if len(p) <= 2 else p
        options_html += f'<option value="{p}">{display_p}</option>'

    # 3. Build the initial HTML shell (JS will populate the numbers)
    month_keys = sorted(list(set(r['d'][:7] for r in records)))
    cal_html = "<div style='display:grid; grid-template-columns:repeat(auto-fill, minmax(250px, 1fr)); gap:15px; margin-top:20px;'>"
    
    for mk in month_keys[:12]:
        yr, mo = int(mk.split('-')[0]), int(mk.split('-')[1])
        cal_html += f"<div style='background:#0c0c12; border:1px solid #1a1a26; border-radius:6px; padding:12px;'>"
        cal_html += f"<div style='display:flex; justify-content:space-between; align-items:baseline; border-bottom:1px solid #252535; padding-bottom:6px; margin-bottom:8px;'><span style='color:#00D2FF; font-weight:800; font-size:12px; text-transform:uppercase; letter-spacing:1px;'>{_cal.month_name[mo]} {yr}</span><span id='month-total-{mk}' style='color:#7777a0; font-size:10px; font-family:monospace;'>0 calls</span></div>"
        
        cal_html += "<div style='display:grid; grid-template-columns:repeat(7, 1fr); gap:2px; margin-bottom:4px;'>"
        for i, dname in enumerate(['Su','Mo','Tu','We','Th','Fr','Sa']):
            c = ['#FF6B6B','#4ECDC4','#45B7D1','#F0B429','#96CEB4','#DDA0DD','#FF9A8B'][i]
            cal_html += f"<div style='font-size:9px; text-align:center; color:{c}; font-weight:600;'>{dname}</div>"
        cal_html += "</div>"
        
        cal_html += "<div style='display:grid; grid-template-columns:repeat(7, 1fr); gap:2px;'>"
        first_dow_sun = (_cal.weekday(yr, mo, 1) + 1) % 7
        last_day = _cal.monthrange(yr, mo)[1]
        
        for _ in range(first_dow_sun): cal_html += "<div></div>"
            
        for d in range(1, last_day + 1):
            dk = f"{yr}-{mo:02d}-{d:02d}"
            dow_idx = (_cal.weekday(yr, mo, d) + 1) % 7 
            cal_html += f"<div class='day-cell' data-date='{dk}' data-mkey='{mk}' data-month='{_cal.month_name[mo]}' data-d='{d}' data-y='{yr}' data-dow='{dow_idx}' style='aspect-ratio:1; border-radius:2px; display:flex; flex-direction:column; align-items:center; justify-content:center; position:relative; font-family:monospace; cursor:default; border:1px solid transparent; transition:transform 0.1s;' onmouseover='showTooltip(this, event)' onmouseout='hideTooltip()'></div>"
            
        cal_html += "</div></div>"
    cal_html += "</div>"

    controls_html = f"""
    <div style="display:flex; gap:20px; align-items:center; background:#0c0c12; border:1px solid #1a1a26; padding:12px 18px; border-radius:6px; margin-bottom:20px;">
        <div style="display:flex; align-items:center; gap:10px;">
            <span style="font-size:10px; color:#7777a0; font-weight:bold; letter-spacing:1px; text-transform:uppercase;">Priority Filter:</span>
            <select id="pri-select" onchange="currentPri=this.value; updateDashboard();" style="background:#1a1a26; color:#00D2FF; border:1px solid #252535; padding:6px 12px; border-radius:4px; font-weight:bold; cursor:pointer;">
                {options_html}
            </select>
        </div>
        <div style="display:flex; align-items:center; gap:10px;">
            <span style="font-size:10px; color:#7777a0; font-weight:bold; letter-spacing:1px; text-transform:uppercase;">Shift Length:</span>
            <select id="shift-select" onchange="currentShift=parseInt(this.value); updateDashboard();" style="background:#1a1a26; color:#00D2FF; border:1px solid #252535; padding:6px 12px; border-radius:4px; font-weight:bold; cursor:pointer;">
                <option value="8" selected>8 HOURS</option>
                <option value="10">10 HOURS</option>
                <option value="12">12 HOURS</option>
            </select>
        </div>
    </div>
    """

    full_html = f"""
    <div style="background:#000; color:#e8e8f2; font-family: 'Barlow', sans-serif; padding:15px; border-radius:8px;">
        <style>
            .day-cell:hover {{ transform: scale(1.15); z-index: 10; box-shadow: 0 4px 12px rgba(0,0,0,0.5); }}
            .day-peak {{ border-color: #cc0000 !important; font-weight: 700; }}
            #dfr-tooltip {{ position: fixed; z-index: 9999; background: #09090f; border: 1px solid #252535; border-radius: 6px; padding: 12px 16px; font-family: monospace; font-size: 11px; color: #e8e8f2; pointer-events: none; box-shadow: 0 6px 24px rgba(0,0,0,0.8); display: none; min-width: 220px; }}
        </style>
        
        <div id="dfr-tooltip"></div>
        <div style="color:#00D2FF; font-weight:900; letter-spacing:3px; font-size:14px; text-transform:uppercase; margin-bottom:20px; border-bottom:1px solid #1a1a26; padding-bottom:10px;">Data Ingestion Analytics</div>
        
        {controls_html}
        
        <div style="display:grid; grid-template-columns: 1fr 1fr; gap:15px; margin-bottom:20px;">
            <div style="background:#0c0c12; border-left:4px solid #00D2FF; padding:15px; border-radius:4px; border-top:1px solid #1a1a26; border-right:1px solid #1a1a26; border-bottom:1px solid #1a1a26;">
                <div style="color:#00D2FF; font-size:26px; font-weight:900; font-family:monospace;" id="kpi-total-val">0</div>
                <div style="color:#7777a0; font-size:10px; text-transform:uppercase; letter-spacing:1px; margin-top:4px;">Displayed Incidents</div>
            </div>
            <div style="background:#0c0c12; border-left:4px solid #F0B429; padding:15px; border-radius:4px; border-top:1px solid #1a1a26; border-right:1px solid #1a1a26; border-bottom:1px solid #1a1a26;">
                <div style="color:#F0B429; font-size:26px; font-weight:900; font-family:monospace;" id="kpi-peak-val">0:00</div>
                <div style="color:#7777a0; font-size:10px; text-transform:uppercase; letter-spacing:1px; margin-top:4px;">Peak Activity Hour</div>
            </div>
        </div>
        
        <div style="display:grid; grid-template-columns: 3fr 2fr; gap:15px; margin-bottom:25px;">
            <div style="background:#06060a; border:1px solid #1a1a26; border-radius:6px; padding:15px;">
                <div style="margin-bottom:12px; font-size:10px; color:#7777a0; text-transform:uppercase; letter-spacing:1px; font-weight:bold;">Optimized DFR Shift Windows</div>
                <div id="shift-container"></div>
            </div>
            <div style="background:#06060a; border:1px solid #1a1a26; border-radius:6px; padding:15px; display:flex; flex-direction:column;">
                <div style="margin-bottom:12px; font-size:10px; color:#7777a0; text-transform:uppercase; letter-spacing:1px; font-weight:bold;">Call Volume by Day of Week</div>
                <div style="display:flex; justify-content:space-between; align-items:flex-end; flex-grow:1; padding:10px 5px 0;" id="dow-container"></div>
            </div>
        </div>
        
        <div style="display:flex; align-items:center; justify-content:space-between; margin-bottom:5px; padding-top:15px; border-top:1px solid #1a1a26;">
            <div style="font-size:14px; font-weight:800; color:#fff; letter-spacing:1px; text-transform:uppercase;">DFR Deployment Calendar</div>
            <div style="display:flex; gap:12px; font-family:monospace; font-size:9px; color:#8d93b8; flex-wrap:wrap;">
                <div style="display:flex; align-items:center; gap:5px;"><div style="width:9px; height:9px; background:#59B7FF; border:1px solid #1E5D91; border-radius:2px; box-shadow:0 0 6px rgba(89,183,255,0.18);"></div>VERY LOW</div>
                <div style="display:flex; align-items:center; gap:5px;"><div style="width:9px; height:9px; background:#45E28A; border:1px solid #1D7D49; border-radius:2px; box-shadow:0 0 6px rgba(69,226,138,0.18);"></div>LOW</div>
                <div style="display:flex; align-items:center; gap:5px;"><div style="width:9px; height:9px; background:#FFD84D; border:1px solid #8A6F00; border-radius:2px; box-shadow:0 0 6px rgba(255,216,77,0.18);"></div>MEDIUM</div>
                <div style="display:flex; align-items:center; gap:5px;"><div style="width:9px; height:9px; background:#FF9F43; border:1px solid #9A4F00; border-radius:2px; box-shadow:0 0 6px rgba(255,159,67,0.18);"></div>HIGH</div>
                <div style="display:flex; align-items:center; gap:5px;"><div style="width:9px; height:9px; background:#FF5B6E; border:1px solid #A51F2D; border-radius:2px; box-shadow:0 0 8px rgba(255,91,110,0.24);"></div>PEAK</div>
            </div>
        </div>
        
        {cal_html}
        
        <script>
            const rawData = {json.dumps(records)};
            const totalOrigCalls = {int(total_orig_calls) if total_orig_calls else len(records)};
            let currentShift = 8;
            let currentPri = 'ALL';
            window.dateHourly = {{}};
            const dowNames = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'];
            const dowColors = ['#4ECDC4','#45B7D1','#F0B429','#96CEB4','#DDA0DD','#FF9A8B','#FF6B6B'];
            const stripeColors = ['#FF6B6B','#4ECDC4','#45B7D1','#F0B429','#96CEB4','#DDA0DD','#FF9A8B'];

            function updateDashboard() {{
                let filtered = rawData;
                if(currentPri !== 'ALL') {{
                    filtered = rawData.filter(d => String(d.p) === currentPri);
                }}

                let total = filtered.length;
                if(currentPri === 'ALL') {{ total = totalOrigCalls; }}
                document.getElementById('kpi-total-val').innerHTML = total.toLocaleString();

                let hourly = new Array(24).fill(0);
                let daily = {{}};
                let dowCounts = new Array(7).fill(0);
                window.dateHourly = {{}};
                let mTotal = {{}};
                let mMax = {{}};

                filtered.forEach(r => {{
                    hourly[r.h]++;
                    daily[r.d] = (daily[r.d] || 0) + 1;
                    dowCounts[r.dow]++;
                    
                    if(!window.dateHourly[r.d]) window.dateHourly[r.d] = new Array(24).fill(0);
                    window.dateHourly[r.d][r.h]++;
                    
                    let m = r.d.substring(0,7);
                    mTotal[m] = (mTotal[m] || 0) + 1;
                    if(!mMax[m] || daily[r.d] > mMax[m]) mMax[m] = daily[r.d];
                }});

                // Peak Hour
                let peakHr = hourly.indexOf(Math.max(...hourly));
                if(total === 0) peakHr = 0;
                document.getElementById('kpi-peak-val').innerText = peakHr + ':00';

                // Shifts
                let shiftHtml = "";
                [8, 10, 12].forEach(win => {{
                    let bestV = 0, bestS = 0;
                    for(let s=0; s<24; s++) {{
                        let v = 0;
                        for(let h=0; h<win; h++) v += hourly[(s+h)%24];
                        if(v > bestV) {{ bestV = v; bestS = s; }}
                    }}
                    let pct = total > 0 ? (bestV/total)*100 : 0;
                    let isActive = (win === currentShift);
                    let bgColor = isActive ? "rgba(0,210,255,0.08)" : "#0c0c12";
                    let brColor = isActive ? "#00D2FF" : "#252535";
                    let badge = isActive ? "<div style='font-size:8px; color:#00D2FF; margin-left:10px; border:1px solid #00D2FF; padding:1px 4px; border-radius:2px;'>SELECTED</div>" : "";
                    let eHr = (bestS + win) % 24;
                    let pS = String(bestS).padStart(2,'0');
                    let pE = String(eHr).padStart(2,'0');
                    
                    let shiftSegments = '';
                    if (bestS + win <= 24) {{
                        shiftSegments = `<div style="position:absolute; left:${{(bestS/24)*100}}%; width:${{(win/24)*100}}%; background:#00D2FF; height:100%; border-radius:4px; opacity:0.6;"></div>`;
                    }} else {{
                        const firstWidth = ((24 - bestS) / 24) * 100;
                        const secondWidth = (((bestS + win) % 24) / 24) * 100;
                        shiftSegments = `
                            <div style="position:absolute; left:${{(bestS/24)*100}}%; width:${{firstWidth}}%; background:#00D2FF; height:100%; border-radius:4px; opacity:0.6;"></div>
                            <div style="position:absolute; left:0%; width:${{secondWidth}}%; background:#00D2FF; height:100%; border-radius:4px; opacity:0.6;"></div>`;
                    }}
                    shiftHtml += `<div style="display:flex; align-items:center; background:${{bgColor}}; border:1px solid ${{brColor}}; padding:8px; margin-bottom:5px; border-radius:4px; transition:all 0.2s;">
                        <div style="width:50px; font-weight:800; color:#fff; font-size:13px;">${{win}}hr</div>
                        <div style="width:110px; font-family:monospace; color:#00D2FF; font-size:12px;">${{pS}}:00 - ${{pE}}:00</div>
                        <div style="flex-grow:1; background:#1a1a26; height:8px; border-radius:4px; margin:0 15px; position:relative; overflow:hidden;">
                            ${{shiftSegments}}
                        </div>
                        <div style="width:50px; text-align:right; font-family:monospace; color:#00D2FF; font-size:13px;">${{pct.toFixed(1)}}%</div>
                        ${{badge}}
                    </div>`;
                }});
                document.getElementById('shift-container').innerHTML = shiftHtml;

                // DOW
                let maxDow = Math.max(...dowCounts, 1);
                let dowHtml = "";
                for(let i=0; i<7; i++) {{
                    let hPct = (dowCounts[i]/maxDow)*100;
                    dowHtml += `<div style="flex:1; display:flex; flex-direction:column; align-items:center;">
                        <div style="background:#1a1a26; width:22px; height:80px; position:relative; border-radius:2px;">
                            <div style="position:absolute; bottom:0; width:100%; height:${{hPct}}%; background:${{dowColors[i]}}; border-radius:2px; transition:height 0.3s;"></div>
                        </div>
                        <span style="font-size:10px; color:#7777a0; margin-top:6px; font-family:monospace;">${{dowNames[i]}}</span>
                    </div>`;
                }}
                document.getElementById('dow-container').innerHTML = dowHtml;

                // Month Headers
                document.querySelectorAll('[id^="month-total-"]').forEach(el => {{
                    let m = el.id.replace('month-total-','');
                    let cnt = mTotal[m] || 0;
                    el.innerText = cnt.toLocaleString() + ' calls';
                }});

                // Calendar Cells
                document.querySelectorAll('.day-cell').forEach(cell => {{
                    if(!cell.hasAttribute('data-date')) return;
                    let d = cell.getAttribute('data-date');
                    let mkey = cell.getAttribute('data-mkey');
                    let cnt = daily[d] || 0;
                    let max = mMax[mkey] || 1;
                    let ratio = max > 0 ? cnt / max : 0;
                    
                    let bg, fc, cls;
                    if (cnt === 0) {{ bg='#08080f'; fc='#333'; cls='day-zero'; }}
                    else if (ratio >= 0.85) {{ bg='#3d0a0a'; fc='#ff4444'; cls='day-peak'; }}
                    else if (ratio >= 0.55) {{ bg='#3d1a00'; fc='#ff8c00'; cls='day-high'; }}
                    else if (ratio >= 0.25) {{ bg='#2d2d00'; fc='#d4c000'; cls='day-med'; }}
                    else {{ bg='#0d3320'; fc='#2ecc71'; cls='day-low'; }}

                    cell.className = 'day-cell ' + cls;
                    cell.style.background = bg;
                    cell.style.color = fc;
                    cell.setAttribute('data-count', cnt);
                    cell.setAttribute('data-ratio', ratio);
                    
                    let domD = cell.getAttribute('data-d');
                    let html = `<span style='font-size:11px; z-index:1; font-weight:bold;'>${{domD}}</span>`;
                    if(cnt > 0) {{
                        html += `<span style='font-size:8px; opacity:0.7; margin-top:1px;'>${{cnt}}</span>`;
                        let dowIdx = parseInt(cell.getAttribute('data-dow'));
                        html += `<div style='position:absolute; bottom:0; left:0; right:0; height:2px; background:${{stripeColors[dowIdx]}}; opacity:0.7; border-radius:0 0 2px 2px;'></div>`;
                    }}
                    cell.innerHTML = html;
                }});
            }}

            function showTooltip(el, ev) {{
                const cnt = parseInt(el.getAttribute('data-count'));
                if (cnt === 0) return;
                
                const dk = el.getAttribute('data-date');
                const ratio = parseFloat(el.getAttribute('data-ratio'));
                const mName = el.getAttribute('data-month');
                const d = el.getAttribute('data-d');
                const y = el.getAttribute('data-y');
                const dow = parseInt(el.getAttribute('data-dow'));
                
                let loadText = '';
                if (ratio >= 0.85) loadText = '<span style="color:#FF5B6E">■ PEAK</span> — Full crew';
                else if (ratio >= 0.65) loadText = '<span style="color:#FF9F43">■ HIGH</span> — Priority deploy';
                else if (ratio >= 0.45) loadText = '<span style="color:#FFD84D">■ MEDIUM</span> — Standard ops';
                else if (ratio >= 0.25) loadText = '<span style="color:#45E28A">■ LOW</span> — Light staffing';
                else loadText = '<span style="color:#59B7FF">■ VERY LOW</span> — Opportunistic coverage';
                
                const hrArr = window.dateHourly[dk] || Array(24).fill(0);
                let bestV = 0, bestS = 0;
                for (let s=0; s<24; s++) {{
                    let v = 0;
                    for (let h=0; h<currentShift; h++) v += hrArr[(s+h)%24];
                    if (v > bestV) {{ bestV = v; bestS = s; }}
                }}
                const dayPct = cnt > 0 ? Math.round((bestV / cnt) * 100) : 0;
                const eHr = (bestS + currentShift) % 24;
                const fmt = (h) => (h%12 || 12) + (h<12 ? 'AM' : 'PM');
                
                const tt = document.getElementById('dfr-tooltip');
                tt.innerHTML = `
                    <div style="color:#00D2FF; margin-bottom:6px; font-size:12px; font-weight:bold; border-bottom:1px solid #252535; padding-bottom:4px;">${{mName}} ${{d}}, ${{y}} · ${{dowNames[dow]}}</div>
                    <div style="margin-bottom:8px; font-size:13px;">Calls: <span style="color:#fff; font-weight:bold;">${{cnt}}</span>  ·  ${{loadText}}</div>
                    <div style="background:#1a1a26; padding:8px; border-radius:4px;">
                        <div style="color:#7777a0; font-size:9px; letter-spacing:1px; text-transform:uppercase; margin-bottom:4px;">Best ${{currentShift}}hr Shift</div>
                        <div style="color:#00D2FF; font-size:14px; font-weight:bold; margin-bottom:2px;">${{fmt(bestS)}} – ${{fmt(eHr)}}</div>
                        <div style="color:#aaa; font-size:10px;">Covers <span style="color:#fff;">${{dayPct}}%</span> of daily volume</div>
                    </div>
                `;
                
                tt.style.display = 'block';
                
                let left = ev.clientX + 15;
                let top = ev.clientY - 20;
                if (left + 220 > window.innerWidth) left = ev.clientX - 235;
                if (top + 100 > window.innerHeight) top = ev.clientY - 110;
                
                tt.style.left = left + 'px';
                tt.style.top = top + 'px';
            }}
            
            function hideTooltip() {{
                document.getElementById('dfr-tooltip').style.display = 'none';
            }}

            // Run once to populate the dashboard on load
            updateDashboard();
        </script>
    </div>
    """
    return full_html

# ============================================================
# AGGRESSIVE DATA PARSER
# ============================================================
def aggressive_parse_calls(uploaded_files):
    all_calls_list = []
    CV = {
        'date': ['received date','incident date','call date','call creation date','calldatetime','call datetime','calltime','timestamp','date','datetime','date time','dispatch date','time received','incdate','date_rept','date_occu','createdtime','created_time','receivedtime','received_time','eventtime','event_time','incidenttime','incident_time','reportedtime','reported_time','entrytime','entry_time','time_central','time_stamp','created'],
        'time': ['call creation time','call time','dispatch time','received time','time', 'hour', 'hour_rept','hour_occu'],
        'priority': ['call priority', 'priority level', 'priority', 'pri', 'urgency'],
        'lat': ['latitude','lat','y coord','ycoord','ycoor','addressy','geoy','y_coord','map_y',
                'point_y','gps_lat','gps_latitude','ylat','coord_y','northing','y_wgs','lat_wgs',
                'incident_lat','inc_lat','event_lat','y_coordinate','address_y','ylocation'],
        'lon': ['longitude','lon','long','x coord','xcoord','xcoor','addressx','geox','x_coord',
                'map_x','point_x','gps_lon','gps_long','gps_longitude','xlon','coord_x','easting',
                'x_wgs','lon_wgs','incident_lon','inc_lon','event_lon','x_coordinate','address_x','xlocation']
    }


    def _looks_like_headerless_geocoder_export(df):
        try:
            cols = [str(c).strip() for c in df.columns]
            coord_pat = re.compile(r'^-?\d+(?:\.\d+)?\s*,\s*-?\d+(?:\.\d+)?$')
            zip_pat = re.compile(r'\b[A-Z]{2}\b\s*,\s*\d{5}(?:-\d{4})?$', re.I)
            has_coord_col = any(coord_pat.match(c) for c in cols)
            has_address_col = any(',' in c and zip_pat.search(c) for c in cols)
            has_matchish = any(str(c).strip().lower() in {'match', 'no_match', 'exact', 'non_exact'} for c in cols)
            return has_coord_col and has_address_col and has_matchish
        except Exception:
            return False

    def _normalize_headerless_geocoder_export(df):
        rows = [list(df.columns)] + df.astype(str).fillna('').values.tolist()
        width = max(len(r) for r in rows)
        padded = [r + [''] * (width - len(r)) for r in rows]
        norm = pd.DataFrame(padded)
        base_cols = ['source_id', 'input_address', 'match_status', 'match_type', 'matched_address', 'lonlat', 'external_id', 'side']
        if width > len(base_cols):
            base_cols += [f'extra_{i}' for i in range(1, width - len(base_cols) + 1)]
        norm.columns = base_cols[:width]
        return norm

    def _extract_lonlat_pair(series):
        s = series.astype(str).str.strip()
        pair = s.str.extract(r'^\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*$')
        lon = pd.to_numeric(pair[0], errors='coerce')
        lat = pd.to_numeric(pair[1], errors='coerce')
        valid = ((lat.between(-90, 90)) & (lon.between(-180, 180))).mean()
        return lon, lat, float(valid)

    def _infer_city_from_location_text(raw_df):
        text_cols = [c for c in raw_df.columns if c in ['location', 'address', 'incident_location', 'addr', 'street', 'input_address', 'matched_address']]
        if not text_cols:
            return None

        s = raw_df[text_cols[0]].dropna().astype(str).str.upper().str.strip()
        if s.empty:
            return None

        s = s.str.replace(r':.*$', '', regex=True)
        s = s.str.replace(r'\bCNTY\b', 'COUNTY', regex=True)
        s = s.str.replace(r'[^A-Z0-9 /,-]', ' ', regex=True)
        s = s.str.replace(r'\s+', ' ', regex=True).str.strip()

        candidates = []
        for val in s:
            padded = f' {val} '
            if ' MOBILE ' in padded:
                candidates.append('Mobile')
                continue

            parts = [p.strip() for p in val.split(',') if p and p.strip()]
            if len(parts) >= 2:
                locality = parts[-2] if re.match(r'^[A-Z]{2}$', parts[-1]) else parts[-1]
                locality = locality.strip()
                if locality and locality not in {'COUNTY', 'CITY', 'TOWN', 'VILLAGE', 'HAMLET'}:
                    candidates.append(locality.title())
                    continue

            m = re.search(r'\b([A-Z]{3,}(?:\s+[A-Z]{3,}){0,2})$', val)
            if m:
                city = m.group(1).title()
                if city not in {'County', 'City', 'Town', 'Village', 'Hamlet'}:
                    candidates.append(city)

        if not candidates:
            return None

        vc = pd.Series(candidates).value_counts()
        return vc.index[0] if not vc.empty else None

    def _infer_state_from_text(raw_df, inferred_city=None):
        for col in ['state', 'state_name']:
            if col in raw_df.columns:
                top = raw_df[col].dropna().astype(str).str.strip().value_counts()
                if not top.empty:
                    state_val = top.index[0]
                    state_up = str(state_val).upper()
                    if state_up in STATE_FIPS:
                        return state_up
                    state_title = str(state_val).title()
                    if state_title in US_STATES_ABBR:
                        return US_STATES_ABBR[state_title]

        if inferred_city == 'Mobile':
            return 'AL'
        return None


    def _choose_priority_column(raw_df):
        exact_names = ['priority', 'call priority', 'priority level', 'pri']
        exact = [c for c in raw_df.columns if c.strip().lower() in exact_names]
        if exact:
            exact.sort(key=lambda c: (
                pd.to_numeric(raw_df[c], errors='coerce').dropna().isin([1,2,3,4,5,6,7,8,9]).mean(),
                -raw_df[c].dropna().nunique()
            ), reverse=True)
            return exact[0]

        loose_names = ['priority', 'call priority', 'priority level', 'pri', 'urgency']
        loose = [c for c in raw_df.columns if any(k in c for k in loose_names)]
        if loose:
            loose.sort(key=lambda c: (
                pd.to_numeric(raw_df[c], errors='coerce').dropna().isin([1,2,3,4,5,6,7,8,9]).mean(),
                -raw_df[c].dropna().nunique()
            ), reverse=True)
            return loose[0]
        return None

    def parse_priority(raw):
        s = str(raw).strip().upper()
        if not s or s == 'NAN': return None
        # Smart inference for PD offenses if priority column is missing
        if any(w in s for w in ['ROBBERY','BURGLARY','ASSAULT','SHOOTING','STABBING','CRITICAL','EMERG']): return 1
        if any(w in s for w in ['ACCIDENT','DISTURBANCE','THEFT','MED','ALARM']): return 2
        if any(w in s for w in ['NON REPORTABLE','FOUND PROPERTY','INFO','ROUTINE','MISC']): return 4
        
        m = re.search(r'^(\d+)', s)
        if m: return int(m.group(1))
        return 3

    for cfile in uploaded_files:
        try:
            fname = cfile.name.lower()
            excel_exts = ('.xlsx', '.xls', '.xlsb', '.xlsm')

            if fname.endswith(excel_exts):
                # ── Excel path ────────────────────────────────────────────────
                raw_bytes = cfile.getvalue()
                engine = 'openpyxl'
                if fname.endswith('.xls'):
                    engine = 'xlrd'
                elif fname.endswith('.xlsb'):
                    engine = 'pyxlsb'

                def _sheet_score(ws):
                    score = 0
                    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
                    if not rows:
                        return -1
                    header = rows[0] or []
                    header_norm = [str(h).strip().lower() for h in header if h is not None]
                    if not header_norm:
                        return -1
                    hints = ['latitude', 'longitude', 'lat', 'lon', 'priority', 'location', 'date', 'time']
                    score += sum(10 for h in header_norm if any(k == h or k in h for k in hints))
                    score += sum(1 for h in header_norm if h and not re.match(r'^column\d+$', h))
                    if len(rows) > 1 and rows[1] and any(v is not None and str(v).strip() != '' for v in rows[1]):
                        score += 25
                    # Penalize external-data placeholder sheets
                    if len(header_norm) == 1 and header_norm[0].startswith('externaldata_'):
                        score -= 100
                    return score

                try:
                    import openpyxl as _oxl
                    _wb = _oxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
                    _sheet_name = max(_wb.sheetnames, key=lambda sn: _sheet_score(_wb[sn]))
                    _ws = _wb[_sheet_name]
                    _row_iter = _ws.iter_rows(values_only=True)
                    _headers_raw = next(_row_iter)
                    if _headers_raw is None:
                        raise ValueError("Selected Excel sheet has no header row.")
                    _real_idx = [
                        i for i, h in enumerate(_headers_raw)
                        if h is not None and not (str(h).startswith('Column') and str(h)[6:].isdigit())
                    ]
                    if not _real_idx:
                        _real_idx = [i for i, h in enumerate(_headers_raw) if h is not None]
                    _real_headers = [str(_headers_raw[i]).lower().strip() for i in _real_idx]
                    _rows_data = []
                    for _row in _row_iter:
                        if _row is None:
                            continue
                        _trimmed = [_row[i] if i < len(_row) else None for i in _real_idx]
                        if any(v is not None and str(v).strip() != '' for v in _trimmed):
                            _rows_data.append(_trimmed)
                    _wb.close()
                    raw_df = pd.DataFrame(_rows_data, columns=_real_headers)
                    raw_df = raw_df.dropna(how='all')
                    raw_df.columns = [str(c).lower().strip() for c in raw_df.columns]
                except Exception as _xe:
                    raw_df = None
                    # Try all sheets with pandas and pick the one that looks most like CAD data
                    try:
                        _all = pd.read_excel(io.BytesIO(raw_bytes), engine=engine, sheet_name=None)
                        best_score = -10**9
                        best_df = None
                        for _sn, _df in _all.items():
                            _df.columns = [str(c).lower().strip() for c in _df.columns]
                            _score = 0
                            for _c in _df.columns:
                                if _c in ('latitude', 'longitude', 'priority', 'location'):
                                    _score += 20
                                elif any(k in _c for k in ['lat', 'lon', 'priority', 'location', 'date', 'time']):
                                    _score += 5
                            _score += min(len(_df), 100)
                            if len(_df.columns) == 1 and str(_df.columns[0]).startswith('externaldata_'):
                                _score -= 100
                            if _score > best_score:
                                best_score = _score
                                best_df = _df
                        if best_df is not None:
                            raw_df = best_df
                    except Exception:
                        pass
                    if raw_df is None:
                        raw_df = pd.read_excel(io.BytesIO(raw_bytes), engine=engine, dtype=str)
                        raw_df.columns = [str(c).lower().strip() for c in raw_df.columns]
            else:
                # ── CSV / TXT path ────────────────────────────────────────────
                content = cfile.getvalue().decode('utf-8', errors='ignore')
                first_line = content.split('\n')[0]
                delim = ',' if first_line.count(',') > first_line.count('\t') else '\t'
                raw_df = pd.read_csv(io.StringIO(content), sep=delim, dtype=str)
                if _looks_like_headerless_geocoder_export(raw_df):
                    raw_df = _normalize_headerless_geocoder_export(raw_df)
                raw_df.columns = [str(c).lower().strip() for c in raw_df.columns]

                # ── Census Geocoder: split combined 'lonlat' column ──────────
                # After normalization the file has a 'lonlat' column storing
                # "lon,lat" pairs (e.g. "-93.283,36.601").  The generic
                # coord-name scanner below matches 'lonlat' for BOTH lat and
                # lon but pd.to_numeric returns all-NaN on comma-pair strings.
                # Handle it here explicitly before the scanner runs, then drop
                # No_Match rows whose lonlat is empty.
                if 'lonlat' in raw_df.columns and 'lat' not in raw_df.columns:
                    _pair = raw_df['lonlat'].astype(str).str.strip().str.extract(
                        r'^\s*(-?[\d.]+)\s*,\s*(-?[\d.]+)\s*$'
                    )
                    _lon_cand = pd.to_numeric(_pair[0], errors='coerce')
                    _lat_cand = pd.to_numeric(_pair[1], errors='coerce')
                    # Census geocoder stores lon first, lat second — verify US range
                    if _lon_cand.between(-180, -50).mean() > 0.3 and _lat_cand.between(18, 72).mean() > 0.3:
                        raw_df['lon'] = _lon_cand
                        raw_df['lat'] = _lat_cand
                    # Drop rows with no geocoded location (No_Match rows)
                    if 'lat' in raw_df.columns:
                        raw_df = raw_df[raw_df['lat'].notna()].copy()

                # ── Census Geocoder: extract city & state from matched_address ──
                # matched_address format: "32 GOLFSHORES DR, BRANSON, MO, 65616"
                if 'matched_address' in raw_df.columns and '_csv_city' not in raw_df.columns:
                    try:
                        _ma = raw_df['matched_address'].dropna().astype(str)
                        _ma_parts = _ma.str.split(',')
                        _ma_cities = _ma_parts.apply(
                            lambda p: p[-3].strip().title() if len(p) >= 4 else (p[-2].strip().title() if len(p) >= 2 else None)
                        ).dropna()
                        _ma_states = _ma_parts.apply(
                            lambda p: p[-2].strip().upper() if len(p) >= 2 else None
                        ).dropna()
                        if not _ma_cities.empty:
                            raw_df['_csv_city'] = _ma_cities.value_counts().index[0]
                        if not _ma_states.empty:
                            _top_st = _ma_states.value_counts().index[0]
                            if _top_st in STATE_FIPS:
                                raw_df['_csv_state'] = _top_st
                    except Exception:
                        pass

            res = pd.DataFrame()
            exact_coord_names = {
                'lat': ['latitude', 'lat', 'gps_lat', 'gps_latitude'],
                'lon': ['longitude', 'lon', 'long', 'gps_lon', 'gps_longitude']
            }
            for field in ['lat', 'lon']:
                found_exact = [c for c in raw_df.columns if c.strip().lower() in exact_coord_names[field]]
                # Exclude bare 'lonlat' from the loose scan — it's a combined field,
                # not a plain numeric column, and will produce all-NaN via pd.to_numeric.
                found_loose = [c for c in raw_df.columns
                               if c != 'lonlat' and any(s in c for s in CV[field])]
                found = found_exact or found_loose
                if found:
                    res[field] = pd.to_numeric(raw_df[found[0]], errors='coerce')

            if 'lat' not in res.columns or 'lon' not in res.columns:
                for c in raw_df.columns:
                    lon_series, lat_series, valid_rate = _extract_lonlat_pair(raw_df[c])
                    if valid_rate >= 0.50:
                        res['lon'] = lon_series
                        res['lat'] = lat_series
                        break

            # ── Fallback: no column name matched — scan numeric columns by value range ──
            # Lat: -90 to 90, Lon: -180 to 180. Pick best candidate for each.
            if 'lat' not in res.columns or 'lon' not in res.columns:
                numeric_cols = []
                for c in raw_df.columns:
                    series = pd.to_numeric(raw_df[c], errors='coerce').dropna()
                    if len(series) > 10:
                        numeric_cols.append((c, series))

                lat_candidates, lon_candidates = [], []
                for c, series in numeric_cols:
                    mn, mx = series.min(), series.max()
                    # Reject if already assigned
                    if c in (res.get('_lat_col',''), res.get('_lon_col','')):
                        continue
                    # Large integer coords (State Plane) — treat as potential coord pair
                    if mx > 1000:
                        lat_candidates.append((c, series))
                        lon_candidates.append((c, series))
                        continue
                    if -90 <= mn and mx <= 90 and mn < -1:
                        lat_candidates.append((c, series))
                    if -180 <= mn and mx <= 180 and (mn < -90 or mx > 90):
                        lon_candidates.append((c, series))

                # Prefer candidate whose name hints at lat/lon
                def _score(name, hints):
                    return sum(1 for h in hints if h in name)

                if 'lat' not in res.columns and lat_candidates:
                    lat_candidates.sort(key=lambda x: -_score(x[0], ['lat','y','north']))
                    best_lat_col = lat_candidates[0][0]
                    res['lat'] = pd.to_numeric(raw_df[best_lat_col], errors='coerce')

                if 'lon' not in res.columns and lon_candidates:
                    # Don't reuse the lat column
                    used = res.get('lat', pd.Series()).name if 'lat' in res.columns else None
                    lon_candidates = [(c, s) for c, s in lon_candidates if c != used]
                    if lon_candidates:
                        lon_candidates.sort(key=lambda x: -_score(x[0], ['lon','long','x','east']))
                        best_lon_col = lon_candidates[0][0]
                        res['lon'] = pd.to_numeric(raw_df[best_lon_col], errors='coerce')
            
            _p_col = _choose_priority_column(raw_df)
            p_found = [_p_col] if _p_col else []
            if _p_col:
                parsed_priority = raw_df[_p_col].apply(parse_priority)
                parsed_priority = pd.to_numeric(parsed_priority, errors='coerce')
                parsed_priority = parsed_priority.where(parsed_priority.isin([1, 2, 3, 4, 5, 6, 7, 8, 9]))
                if parsed_priority.dropna().empty:
                    res['priority'] = 3
                else:
                    res['priority'] = parsed_priority.fillna(3).astype(int)
            else:
                # No trustworthy priority field — keep the app usable with a neutral default
                res['priority'] = 3
            
            # ── Event type description — carried through for CAD analytics charts ──
            _desc_hints = ['desc','type','nature','offense','calltype','call_type','event_type',
                           'eventtype','calldesc','incident_type','agencyeventtype','violation','call_nature','cfs_type']
            _desc_found = [c for c in raw_df.columns
                           if any(h in c for h in _desc_hints)
                           and c not in (p_found[:1] if p_found else [])]
            if _desc_found:
                # Pick the column with the most unique text values (most descriptive)
                _best_desc = max(_desc_found, key=lambda c: raw_df[c].dropna().nunique())
                if raw_df[_best_desc].dropna().nunique() > 2:
                    res['call_type_desc'] = raw_df[_best_desc].astype(str).str.strip()

            d_found = [c for c in raw_df.columns if any(s in c for s in CV['date'])]
            t_found = [c for c in raw_df.columns if any(s in c for s in CV['time'])]

            # Fallback: if no date column found by name hint, scan all string columns
            # for any that successfully parse as datetime (catches columns like
            # 'createdtime_central', 'call_ts', 'event_dttm', etc.)
            if not d_found:
                for _col in raw_df.columns:
                    if _col in (t_found or []):
                        continue
                    try:
                        _test = pd.to_datetime(raw_df[_col].dropna().head(50), errors='coerce')
                        _valid = _test.dropna()
                        if len(_valid) >= 10 and _valid.dt.year.between(2000, 2035).mean() > 0.8:
                            d_found = [_col]
                            break
                    except Exception:
                        continue

            if d_found:
                # Build the raw string series to parse — combine date+time cols if separate
                if t_found and d_found[0] != t_found[0]:
                    _raw_dt_str = raw_df[d_found[0]].fillna('') + ' ' + raw_df[t_found[0]].fillna('')
                else:
                    _raw_dt_str = raw_df[d_found[0]]

                # Try explicit common formats first (orders of magnitude faster than
                # dateutil fallback on large files, and avoids NaT on ghost rows).
                # Format detection: sample the first non-null value.
                _sample_vals = _raw_dt_str.dropna().str.strip()
                _sample_vals = _sample_vals[_sample_vals != ''].head(5)
                _fmt_candidates = [
                    '%m/%d/%Y %I:%M %p',   # 2/14/2025 6:03 PM  (Mobile AL)
                    '%m/%d/%Y %H:%M:%S',   # 2/14/2025 18:03:00
                    '%m/%d/%Y %H:%M',      # 2/14/2025 18:03
                    '%Y-%m-%d %H:%M:%S',   # 2025-02-14 18:03:00
                    '%Y-%m-%dT%H:%M:%S',   # ISO 8601
                    '%Y-%m-%d %H:%M',      # 2025-02-14 18:03
                    '%Y/%m/%d %H:%M:%S',
                    '%d/%m/%Y %H:%M:%S',
                    '%m-%d-%Y %H:%M:%S',
                ]
                dt_series = None
                if not _sample_vals.empty:
                    for _fmt in _fmt_candidates:
                        try:
                            _trial = pd.to_datetime(_sample_vals.iloc[0], format=_fmt, errors='raise')
                            # Format matched — apply to full series
                            dt_series = pd.to_datetime(_raw_dt_str, format=_fmt, errors='coerce')
                            break
                        except Exception:
                            continue
                if dt_series is None:
                    # Final fallback: let pandas infer (slow but handles edge cases)
                    dt_series = pd.to_datetime(_raw_dt_str, errors='coerce')

                res['date'] = dt_series.dt.strftime('%Y-%m-%d')
                res['time'] = dt_series.dt.strftime('%H:%M:%S')

            # --- COORDINATE CLEANUP: sentinel values & sign errors ---
            if not res.empty and 'lat' in res.columns and 'lon' in res.columns:
                # Drop obvious sentinel/null-coordinate rows before any further processing
                # lat=0 and lon=0 are common CAD null sentinels (no valid location on the equator/prime meridian)
                # lon=-179.99999 is another common sentinel used by some CAD vendors
                _sentinel_mask = (
                    (res['lat'] == 0) | (res['lon'] == 0) |
                    (res['lat'].abs() < 0.001) | (res['lon'].abs() < 0.001) |
                    (res['lon'] < -179.9)
                )
                if _sentinel_mask.any():
                    res = res[~_sentinel_mask].copy()

                # Fix wrong-sign longitudes: some CAD exports omit the minus sign for
                # western-hemisphere longitudes (e.g. 81.31 instead of -81.31).
                # Detect by checking if the majority of lons are negative (correct for US)
                # while a small minority are positive with the same absolute magnitude.
                if not res.empty and 'lon' in res.columns:
                    _neg_count = (res['lon'] < 0).sum()
                    _pos_count = (res['lon'] > 0).sum()
                    _total = len(res)
                    # If >90% are negative but some are positive AND the median negative lon
                    # matches -(positive lon range), flip the positive ones
                    if _neg_count > 0 and _pos_count > 0 and (_neg_count / _total) > 0.90:
                        _median_neg = res.loc[res['lon'] < 0, 'lon'].median()
                        _pos_vals = res.loc[res['lon'] > 0, 'lon']
                        # Check if flipping would land near the median negative cluster
                        _would_match = ((-_pos_vals).between(_median_neg - 2, _median_neg + 2)).mean()
                        if _would_match > 0.5:
                            res.loc[res['lon'] > 0, 'lon'] = -res.loc[res['lon'] > 0, 'lon']

            # --- COORDINATE CONVERSION (STATE PLANE / LARGE-INTEGER DETECTOR) ---
            if not res.empty and 'lat' in res.columns and 'lon' in res.columns:
                res = res[(res['lat'] != 0) & (res['lon'] != 0)].dropna(subset=['lat', 'lon'])
                if not res.empty:
                    max_val = max(res['lat'].abs().max(), res['lon'].abs().max())
                    if max_val > 1000:
                        converted = False
                        # Strategy 1: Try common State Plane CRS at /100 and /1 scales
                        candidate_crs = [
                            "EPSG:2278",  # TX South Central (ftUS)
                            "EPSG:2277",  # TX Central (ftUS)
                            "EPSG:2276",  # TX North Central (ftUS)
                            "EPSG:2279",  # TX South (ftUS)
                            "EPSG:32140", # TX South Central (m)
                        ]
                        for scale in [100.0, 1.0]:
                            for crs in candidate_crs:
                                try:
                                    transformer = pyproj.Transformer.from_crs(crs, "EPSG:4326", always_xy=True)
                                    test_lons, test_lats = transformer.transform(
                                        res['lon'].values[:20] / scale,
                                        res['lat'].values[:20] / scale
                                    )
                                    if (24 < float(test_lats.mean()) < 50 and
                                            -130 < float(test_lons.mean()) < -60 and
                                            float(test_lats.std()) < 5 and
                                            float(test_lons.std()) < 5):
                                        lons, lats = transformer.transform(
                                            res['lon'].values / scale, res['lat'].values / scale
                                        )
                                        res['lon'], res['lat'] = lons, lats
                                        converted = True
                                        break
                                except Exception:
                                    continue
                            if converted:
                                break

                        # Strategy 2: If CRS conversion failed, anchor to city column geocode
                        if not converted:
                            try:
                                city_name = None
                                for col in ['city', 'city_name', 'municipality', 'jurisdiction']:
                                    if col in raw_df.columns:
                                        top = raw_df[col].dropna().str.strip().value_counts()
                                        if not top.empty:
                                            city_name = top.index[0]
                                            break
                                state_name = None
                                for col in ['state', 'state_name']:
                                    if col in raw_df.columns:
                                        top = raw_df[col].dropna().str.strip().value_counts()
                                        if not top.empty:
                                            state_name = top.index[0]
                                            break
                                if city_name:
                                    query_str = f"{city_name}, {state_name}" if state_name else city_name
                                    geo_url = f"https://nominatim.openstreetmap.org/search?format=json&q={urllib.parse.quote(query_str)}&limit=1"
                                    req = urllib.request.Request(geo_url, headers={"User-Agent": "BRINC_COS_Optimizer/1.0"})
                                    with urllib.request.urlopen(req, timeout=10) as resp:
                                        geo_data = json.loads(resp.read().decode("utf-8"))
                                    if geo_data:
                                        anchor_lat = float(geo_data[0]["lat"])
                                        anchor_lon = float(geo_data[0]["lon"])
                                        raw_cx = res["lon"].median()
                                        raw_cy = res["lat"].median()
                                        city_radius_deg = 0.35
                                        raw_spread = max(res["lon"].std(), res["lat"].std(), 1)
                                        deg_per_unit = city_radius_deg / raw_spread
                                        res["lon"] = anchor_lon + (res["lon"] - raw_cx) * deg_per_unit
                                        res["lat"] = anchor_lat + (res["lat"] - raw_cy) * deg_per_unit
                                        converted = True
                            except Exception:
                                pass

                        if converted:
                            res = res[
                                (res["lat"] > 18) & (res["lat"] < 72) &
                                (res["lon"] > -170) & (res["lon"] < -60)
                            ]

            # ── Agency / source tagging (Fire vs Police) ─────────────────────
            # Look for a column named 'agency', 'department', or 'dept' and
            # carry it through as a lowercase 'agency' column so the map
            # renderer can colour fire calls red and police calls the default colour.
            _agency_col = next(
                (c for c in raw_df.columns if c.strip().lower() in ('agency', 'department', 'dept')),
                None
            )
            if _agency_col:
                res['agency'] = raw_df[_agency_col].astype(str).str.strip().str.lower()
            else:
                res['agency'] = 'police'   # safe default for single-agency files

            # ── Agency / source tagging (Fire vs Police) ─────────────────────
            # Look for a column named 'agency', 'department', or 'dept' and
            # carry it through as a lowercase 'agency' column so the map
            # renderer can colour fire calls red and police calls the default colour.
            _agency_col = next(
                (c for c in raw_df.columns if c.strip().lower() in ('agency', 'department', 'dept')),
                None
            )
            if _agency_col:
                res['agency'] = raw_df[_agency_col].astype(str).str.strip().str.lower()
            else:
                res['agency'] = 'police'   # safe default for single-agency files

            # City/state detection: store top values on rows for location detection
            top_city_name = None
            for col in ["city", "city_name", "municipality", "jurisdiction"]:
                if col in raw_df.columns:
                    top_city = raw_df[col].dropna().astype(str).str.strip().value_counts()
                    if not top_city.empty:
                        top_city_name = str(top_city.index[0]).title()
                        res["_csv_city"] = top_city_name
                        break

            if "_csv_city" not in res.columns:
                inferred_city = _infer_city_from_location_text(raw_df)
                if inferred_city:
                    top_city_name = inferred_city
                    res["_csv_city"] = inferred_city

            inferred_state = _infer_state_from_text(raw_df, top_city_name)
            if not inferred_state:
                for _addr_col in ['input_address', 'matched_address', 'address', 'location']:
                    if _addr_col in raw_df.columns:
                        _addr_series = raw_df[_addr_col].astype(str)
                        # Pattern 1: "..., ST, ZIPCODE" (comma-separated state and zip)
                        _states = _addr_series.str.extract(r',\s*([A-Z]{2})\s*,\s*\d{5}(?:-\d{4})?')[0].dropna()
                        if _states.empty:
                            # Pattern 2: "..., ST ZIPCODE" (state and zip in same segment, common format)
                            _states = _addr_series.str.extract(r',\s*([A-Z]{2})\s+\d{5}(?:-\d{4})?')[0].dropna()
                        if not _states.empty:
                            inferred_state = _states.value_counts().idxmax()
                            break
            if inferred_state:
                res["_csv_state"] = inferred_state

            # ── Capture file data matrix for Sheets/email logging ────────────
            try:
                _meta = _extract_file_meta(raw_df, res, filename=cfile.name)
                # Merge into session-level file_meta (last file wins for per-field values;
                # accumulate filenames if multiple files are uploaded at once)
                _existing = st.session_state.get('file_meta', {})
                _existing_names = _existing.get('uploaded_filename', '')
                if _existing_names and _meta.get('uploaded_filename','') and _meta['uploaded_filename'] not in _existing_names:
                    _meta['uploaded_filename'] = _existing_names + ' | ' + _meta['uploaded_filename']
                st.session_state['file_meta'] = {**_existing, **_meta}
            except Exception:
                pass

            all_calls_list.append(res)
        except: continue
        
    if not all_calls_list: return pd.DataFrame()
    # Only keep frames that actually have lat/lon columns — Excel sheets
    # without coordinate data should not crash the concat
    valid = [df for df in all_calls_list if 'lat' in df.columns and 'lon' in df.columns]
    if not valid: return pd.DataFrame()
    combined = pd.concat(valid, ignore_index=True)
    # Safe dropna — columns guaranteed to exist now
    combined = combined.dropna(subset=['lat', 'lon'])
    combined['lat'] = pd.to_numeric(combined['lat'], errors='coerce')
    combined['lon'] = pd.to_numeric(combined['lon'], errors='coerce')
    combined = combined[(combined['lat'].between(-90, 90)) & (combined['lon'].between(-180, 180))]
    # IMPORTANT: keep the full parsed CAD dataset here.
    #
    # The optimizer is sampled later (after upload) for performance, but the
    # parsed dataframe itself must preserve every incident so:
    #   1) Total Incidents shows the true uploaded count
    #   2) the stations map can render a much denser full-history call cloud
    #   3) export/reporting math stays tied to the source file, not a k-means
    #      surrogate created during parsing
    return combined

def _build_cad_charts_html(df_calls):
    """Generate a self-contained HTML block for the PDF/HTML export.
    Includes the Drone Apprehension Impact Value table and the Top Call Types chart.
    Returns an empty string if no real CAD data is available."""
    if df_calls is None or df_calls.empty:
        return ""
    try:
        total_calls = len(df_calls)

        # ── Apprehension metric calculations ─────────────────────────────────
        import streamlit as _st
        dfr_rate        = float(_st.session_state.get('dfr_rate', 12)) / 100.0
        pursuit_rate    = 0.18
        pursuit_calls   = round(total_calls * pursuit_rate)
        dfr_pursuit     = round(pursuit_calls * dfr_rate)
        arr_lift        = 0.20   # +20 pp
        additional_arr  = round(dfr_pursuit * arr_lift)
        coverage_pct    = float(_st.session_state.get('calls_covered_perc', 70) or 70)
        time_saved      = float(_st.session_state.get('avg_time_saved_min', 6) or 6)
        score = round(
            0.40 * min(coverage_pct, 100) +
            0.35 * min(time_saved / 10.0 * 100, 100) +
            0.25 * min(dfr_rate * 100 / 30.0 * 100, 100)
        )
        score = max(0, min(score, 100))
        if score >= 75:
            score_label = "HIGH"
            score_color = "#008060"
        elif score >= 50:
            score_label = "MODERATE"
            score_color = "#b06000"
        else:
            score_label = "LOW"
            score_color = "#b00020"

        rows = [
            ("Average officer response time",          "8 – 12 min",             "2 – 4 min (DFR first on scene)",    "BRINC field deployments"),
            ("Suspect located before officer arrival", "~18% of pursuits",        "~62% of pursuits",                  "Aerial ID + thermal"),
            ("Apprehension rate per pursuit incident", "34%",                     "54%  (+20 pp)",                     "Perimeter intel, real-time relay"),
            ("Additional arrests per 100 calls",       "—",                       "+20 apprehensions",                 "Net lift on DFR-covered incidents"),
            ("Thermal imaging (nighttime pursuits)",   "Unavailable",             "100% of flight hours",              "Eliminates blind foot searches"),
            ("Perimeter containment",                  "4 – 6 officers required", "Drone in < 90 sec",                 "Officers freed for contact"),
            ("DFR-dispatched pursuit calls / year",    "—",                       f"{dfr_pursuit:,}",                  f"{int(dfr_rate*100)}% DFR × {pursuit_calls:,} pursuit calls"),
            ("Est. additional arrests / year",         "—",                       f"+ {additional_arr:,} arrests",     "DFR pursuit calls × +20 pp lift"),
        ]

        rows_html = ""
        for i, (factor, base, drone, source) in enumerate(rows):
            bg = "#f9fafb" if i % 2 == 0 else "#ffffff"
            rows_html += f"""
  <tr style="background:{bg};">
    <td style="padding:9px 12px; font-size:13px; color:#333; border-bottom:1px solid #e5e7eb; width:34%;">{factor}</td>
    <td style="padding:9px 12px; font-size:13px; color:#666; border-bottom:1px solid #e5e7eb; width:20%; text-align:right;">{base}</td>
    <td style="padding:9px 12px; font-size:13px; color:#00695c; font-weight:700; border-bottom:1px solid #e5e7eb; width:20%; text-align:right;">{drone}</td>
    <td style="padding:9px 12px; font-size:11px; color:#888; border-bottom:1px solid #e5e7eb; width:26%;">{source}</td>
  </tr>"""

        # ── Top event types ───────────────────────────────────────────────────
        type_labels, type_vals = [], []
        for _c in ['call_type_desc','agencyeventtypecodedesc','eventdesc','calldesc','description','nature','event_desc']:
            if _c in df_calls.columns and df_calls[_c].dropna().nunique() > 2:
                tc = df_calls[_c].dropna().str.strip().value_counts().head(10)
                type_labels = tc.index.tolist()
                type_vals   = tc.values.tolist()
                break

        import json
        type_labels_js = json.dumps(type_labels)
        type_vals_js   = json.dumps(type_vals)
        has_types      = "true" if type_vals else "false"
        bar_height     = max(260, len(type_labels) * 28 + 60) if type_labels else 260

        return f"""
<h2 style="color:#111; font-size:22px; font-weight:800; margin-top:40px; margin-bottom:20px;
           padding-bottom:10px; border-bottom:2px solid #eee;">Incident Data Analysis</h2>
<p style="font-size:13px; color:#666; margin-bottom:20px;">
  Summary of <strong>{total_calls:,}</strong> calls for service used to optimise drone placement.
</p>

<p style="font-size:12px; font-weight:700; color:#333; text-transform:uppercase;
          letter-spacing:0.6px; margin:0 0 8px;">🎯 Drone Apprehension Impact Value</p>
<p style="font-size:12px; color:#666; margin:0 0 12px 0;">
  How drone deployment improves suspect apprehension — derived from your call volume and DFR
  dispatch rate. Baseline figures from national law enforcement benchmarks.
</p>
<div style="overflow-x:auto; border-radius:8px; border:1px solid #e5e7eb; margin-bottom:10px;">
<table style="width:100%; border-collapse:collapse; font-family:inherit;">
  <thead>
    <tr style="background:#f0faf8;">
      <th style="padding:10px 12px; font-size:11px; font-weight:700; text-transform:uppercase;
                 letter-spacing:0.6px; color:#555; border-bottom:1px solid #d1d5db; text-align:left;">Factor</th>
      <th style="padding:10px 12px; font-size:11px; font-weight:700; text-transform:uppercase;
                 letter-spacing:0.6px; color:#555; border-bottom:1px solid #d1d5db; text-align:right;">Without Drone</th>
      <th style="padding:10px 12px; font-size:11px; font-weight:700; text-transform:uppercase;
                 letter-spacing:0.6px; color:#555; border-bottom:1px solid #d1d5db; text-align:right;">With Drone</th>
      <th style="padding:10px 12px; font-size:11px; font-weight:700; text-transform:uppercase;
                 letter-spacing:0.6px; color:#555; border-bottom:1px solid #d1d5db; text-align:left;">Basis</th>
    </tr>
  </thead>
  <tbody>
{rows_html}
    <tr style="background:#e6f4f1;">
      <td style="padding:10px 12px; font-size:14px; font-weight:700; color:#111; border-bottom:1px solid #d1d5db;">
        Apprehension Value Score
      </td>
      <td style="padding:10px 12px; color:#888; border-bottom:1px solid #d1d5db; text-align:right;">—</td>
      <td colspan="2" style="padding:10px 12px; font-size:18px; font-weight:800;
          color:{score_color}; border-bottom:1px solid #d1d5db;">
        {score_label} &nbsp;<span style="font-size:12px; font-weight:400; color:#888;">({score}/100 composite)</span>
      </td>
    </tr>
  </tbody>
</table>
</div>
<p style="font-size:10px; color:#aaa; margin:4px 0 28px 0;">
  Score weighted: 40% geographic coverage · 35% time saved vs patrol · 25% DFR dispatch rate.
  Arrest estimates are model projections; actual results vary by deployment, terrain, and incident type.
</p>

<p style="font-size:12px; font-weight:700; color:#555; text-transform:uppercase;
          letter-spacing:0.5px; margin:0 0 8px;">Top Call Types</p>
<div style="position:relative; height:{bar_height}px; margin-bottom:24px;">
  <canvas id="expTypeChart"></canvas>
</div>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<script>
(function(){{
  var typL={type_labels_js}, typV={type_vals_js};
  var hasTypes={has_types};
  if(hasTypes && typL.length) {{
    new Chart(document.getElementById('expTypeChart'), {{
      type:'bar',
      data:{{
        labels:typL,
        datasets:[{{data:typV,backgroundColor:'#00D2FF',borderRadius:3,borderSkipped:false}}]
      }},
      options:{{responsive:true,maintainAspectRatio:false,indexAxis:'y',
        plugins:{{legend:{{display:false}}}},
        scales:{{
          x:{{ticks:{{callback:function(v){{return v>=1000?Math.round(v/1000)+'k':v}}}}}},
          y:{{ticks:{{font:{{size:11}}}}}}
        }}
      }}
    }});
  }}
}})();
</script>
"""
    except Exception as e:
        print(f"[BRINC] _build_cad_charts_html failed: {e}\n{traceback.format_exc()}")
        return "<div style='color:#888;padding:20px;text-align:center;font-size:13px;'>Chart unavailable — data could not be rendered.</div>"


def _get_annualized_calls(raw_count: int) -> int:
    """Return raw_count scaled to a full year using the uploaded file's date span.

    If the file covers less than a full year (and at least 14 days), the raw
    count is extrapolated to 365 days.  Falls back to raw_count when no date
    span is available (simulated data, unknown span, or span ≥ 330 days).
    """
    span_days = int(st.session_state.get('file_meta', {}).get('file_date_span_days', 0) or 0)
    if 14 <= span_days < 330:
        return round(raw_count * 365 / span_days)
    return raw_count


def _build_apprehension_table(df_calls, text_main, text_muted, card_bg, card_border, accent_color):
    """Compute and render the Drone Apprehension Impact Value table.

    Derived metrics use call volume, DFR dispatch rate, and coverage percentage
    stored in session state — no static placeholders.
    """
    if df_calls is None or df_calls.empty:
        return

    # ── Pull session values ───────────────────────────────────────────────────
    total_calls      = int(st.session_state.get('total_original_calls', len(df_calls)) or len(df_calls))
    dfr_rate         = float(st.session_state.get('dfr_rate', 12)) / 100.0   # fraction dispatched by drone
    calls_per_year   = _get_annualized_calls(total_calls)

    # Pursuit-eligible calls: incidents where a suspect is potentially fleeing
    # — conservatively 18% of all calls (PERF national average for patrol pursuits)
    pursuit_rate     = 0.18
    pursuit_calls    = round(calls_per_year * pursuit_rate)

    # Apprehension lift: drone raises locate-before-arrival from 18 % → 62 %
    # (+20 pp net apprehension rate lift per BRINC field deployments)
    baseline_arr_rate  = 0.34   # officer-only apprehension rate per pursuit incident
    drone_arr_rate     = 0.54   # with drone aerial ID + perimeter intel
    arr_lift_pp        = round((drone_arr_rate - baseline_arr_rate) * 100, 0)

    # Annual additional arrests from DFR-dispatched pursuit calls
    dfr_pursuit_calls   = round(pursuit_calls * dfr_rate)
    additional_arrests  = round(dfr_pursuit_calls * (drone_arr_rate - baseline_arr_rate))

    # Apprehension Value Score: composite of speed + coverage + thermal (0–100)
    coverage_pct  = float(st.session_state.get('calls_covered_perc', 70) or 70)
    time_saved    = float(st.session_state.get('avg_time_saved_min', 6) or 6)
    # Weighted: 40% coverage, 35% time saved (normalized to 10-min max), 25% DFR rate
    score = round(
        0.40 * min(coverage_pct, 100) +
        0.35 * min(time_saved / 10.0 * 100, 100) +
        0.25 * min(dfr_rate * 100 / 30.0 * 100, 100)
    )
    score = max(0, min(score, 100))
    if score >= 75:
        score_label = "🟢 HIGH"
        score_color = "#00D2FF"
    elif score >= 50:
        score_label = "🟡 MODERATE"
        score_color = "#EF9F27"
    else:
        score_label = "🔴 LOW"
        score_color = "#E24B4A"

    # ── HTML table ────────────────────────────────────────────────────────────
    row_style_a = f"background:{card_bg};"
    row_style_b = f"background:rgba(0,210,255,0.04);"
    th_style    = (f"padding:10px 14px; text-align:left; font-size:11px; font-weight:700; "
                   f"text-transform:uppercase; letter-spacing:0.6px; color:{text_muted}; "
                   f"border-bottom:1px solid {card_border};")
    td_l_style  = (f"padding:10px 14px; font-size:13px; color:{text_muted}; "
                   f"border-bottom:1px solid {card_border}; width:36%;")
    td_b_style  = (f"padding:10px 14px; font-size:13px; color:{text_main}; "
                   f"border-bottom:1px solid {card_border}; width:19%; text-align:right;")
    td_d_style  = (f"padding:10px 14px; font-size:13px; color:{accent_color}; font-weight:700; "
                   f"border-bottom:1px solid {card_border}; width:19%; text-align:right;")
    td_s_style  = (f"padding:10px 14px; font-size:11px; color:{text_muted}; "
                   f"border-bottom:1px solid {card_border}; width:26%;")

    rows = [
        ("row_a", "Average officer response time",
         "8 – 12 min", "2 – 4 min (DFR first on scene)",
         "BRINC field deployments; avg aerial ETA"),
        ("row_b", "Suspect located before officer arrival",
         "~18% of pursuits", "~62% of pursuits",
         "Drone situational awareness + thermal"),
        ("row_a", "Apprehension rate per pursuit incident",
         f"{int(baseline_arr_rate*100)}%", f"{int(drone_arr_rate*100)}%  (+{int(arr_lift_pp)} pp)",
         "Aerial ID, perimeter intel, real-time relay"),
        ("row_b", "Additional arrests per 100 pursuit calls",
         "—", f"+{int(arr_lift_pp)} apprehensions",
         "Net lift applied to DFR-covered incidents"),
        ("row_a", "Thermal imaging (nighttime pursuits)",
         "Unavailable", "100% of flight hours",
         "Eliminates blind foot searches in darkness"),
        ("row_b", "Perimeter containment established",
         "4 – 6 officers required", "Drone in < 90 sec",
         "Officers freed for contact; drone holds perimeter"),
        ("row_a", "DFR-dispatched pursuit calls / year",
         "—", f"{dfr_pursuit_calls:,}",
         f"{int(dfr_rate*100)}% DFR rate × {pursuit_calls:,} pursuit-eligible calls"),
        ("row_b", "Est. additional arrests / year",
         "—", f"+ {additional_arrests:,} arrests",
         "DFR pursuit calls × +20 pp apprehension lift"),
    ]

    table_html = f"""
<div style="margin-top:4px; margin-bottom:20px;">
  <p style="font-size:13px; font-weight:700; color:{text_main}; text-transform:uppercase;
            letter-spacing:0.6px; margin:0 0 10px 0;">🎯 Drone Apprehension Impact Value</p>
  <p style="font-size:12px; color:{text_muted}; margin:0 0 14px 0;">
    How drone deployment improves suspect apprehension — derived from your call volume,
    DFR dispatch rate, and coverage. Baseline figures from national law enforcement benchmarks.
  </p>
  <div style="overflow-x:auto; border-radius:8px; border:1px solid {card_border};">
  <table style="width:100%; border-collapse:collapse; font-family:inherit;">
    <thead>
      <tr style="background:rgba(0,210,255,0.08);">
        <th style="{th_style}">Factor</th>
        <th style="{th_style} text-align:right;">Without Drone</th>
        <th style="{th_style} text-align:right;">With Drone</th>
        <th style="{th_style}">Basis</th>
      </tr>
    </thead>
    <tbody>
"""
    for i, (variant, factor, base, drone, source) in enumerate(rows):
        bg = row_style_b if i % 2 else row_style_a
        table_html += f"""
      <tr style="{bg}">
        <td style="{td_l_style}">{factor}</td>
        <td style="{td_b_style}">{base}</td>
        <td style="{td_d_style}">{drone}</td>
        <td style="{td_s_style}">{source}</td>
      </tr>"""

    # Final composite score row
    table_html += f"""
      <tr style="background:rgba(0,210,255,0.10);">
        <td style="{td_l_style} font-weight:700; color:{text_main}; font-size:14px;">
          Apprehension Value Score
        </td>
        <td style="{td_b_style}">—</td>
        <td colspan="2" style="padding:10px 14px; font-size:16px; font-weight:800;
            color:{score_color}; border-bottom:1px solid {card_border};">
          {score_label} &nbsp;<span style="font-size:12px; font-weight:400;
          color:{text_muted};">({score}/100 composite)</span>
        </td>
      </tr>
    </tbody>
  </table>
  </div>
  <p style="font-size:10px; color:{text_muted}; margin:6px 0 0 0;">
    Score weighted: 40% geographic coverage · 35% time saved vs patrol · 25% DFR dispatch rate.
    Arrest estimates are model projections — actual results depend on deployment, terrain, and incident type.
  </p>
</div>
"""
    # Wrap in a full HTML document so components.html renders the table faithfully.
    # st.markdown strips <table> tags in recent Streamlit versions.
    full_html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  body {{
    margin: 0; padding: 0;
    background: transparent;
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
  }}
</style>
</head>
<body>
{table_html}
</body>
</html>"""
    # Height: header ~60px + description ~40px + 9 rows × 44px + score row 54px + footnote 30px
    _table_height = 60 + 40 + (len(rows) * 44) + 54 + 44 + 30
    import streamlit.components.v1 as _comp
    _comp.html(full_html, height=_table_height, scrolling=False)


def _build_cad_charts(df_calls, text_main, text_muted, card_bg, card_border, accent_color):
    """Render apprehension impact table + top call types chart."""
    import plotly.graph_objects as go

    if df_calls is None or df_calls.empty:
        return

    layout_base = dict(
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(color=text_muted, size=11),
        margin=dict(l=10, r=10, t=32, b=10),
        hoverlabel=dict(bgcolor=card_bg, font_size=12, font_color=text_main, bordercolor=accent_color),
    )
    grid_color = card_border

    # ── Apprehension Impact Value table (replaces priority donut + density curve) ──
    _build_apprehension_table(df_calls, text_main, text_muted, card_bg, card_border, accent_color)

    # ── Top event types (horizontal bar) ──────────────────────────────────────
    desc_col = None
    for _c in ['call_type_desc','agencyeventtypecodedesc','eventdesc','calldesc','description','nature','event_desc']:
        if _c in df_calls.columns and df_calls[_c].dropna().nunique() > 2:
            desc_col = _c
            break

    if desc_col:
        top_types = df_calls[desc_col].dropna().str.strip().value_counts().head(12)
        if not top_types.empty:
            fig_types = go.Figure(go.Bar(
                x=top_types.values, y=top_types.index,
                orientation='h',
                marker_color=accent_color,
                text=[f'{v:,}' for v in top_types.values],
                textposition='outside',
                hovertemplate='<b>%{y}</b><br>%{x:,} calls<extra></extra>',
            ))
            fig_types.update_layout(**layout_base,
                height=max(280, len(top_types) * 30 + 60),
                title=dict(text='Top Call Types', font=dict(size=13, color=text_main), x=0),
                xaxis=dict(showgrid=True, gridcolor=grid_color, title='Calls'),
                yaxis=dict(showgrid=False, autorange='reversed'),
                showlegend=False,
            )
            st.plotly_chart(fig_types, use_container_width=True, config={'displayModeBar': False})



def _safe_df_to_records(df):
    """Safely serialize a DataFrame to a JSON-safe list of records.
    Returns an empty list if df is None, empty, or serialization fails."""
    if df is None:
        return []
    try:
        if hasattr(df, 'empty') and df.empty:
            return []
        return json.loads(df.replace({float('nan'): None}).to_json(orient='records'))
    except Exception:
        try:
            return json.loads(df.fillna('').to_json(orient='records'))
        except Exception:
            return []


def _make_random_stations(df_calls, n=40, boundary_geom=None, epsg_code=None):
    """Fallback station generator based on call-density hotspots.

    If a city boundary is supplied, only incidents inside that boundary are used and
    final station coordinates are snapped to the nearest in-boundary incident so every
    suggested site remains inside the geographic area.
    """
    if df_calls is None or df_calls.empty:
        return pd.DataFrame()

    work = df_calls.copy()
    work['lat'] = pd.to_numeric(work['lat'], errors='coerce')
    work['lon'] = pd.to_numeric(work['lon'], errors='coerce')
    work = work.dropna(subset=['lat', 'lon']).reset_index(drop=True)
    if work.empty:
        return pd.DataFrame()

    if boundary_geom is not None and epsg_code is not None:
        try:
            work_gdf = gpd.GeoDataFrame(work, geometry=gpd.points_from_xy(work.lon, work.lat), crs="EPSG:4326").to_crs(epsg=int(epsg_code))
            inside_mask = work_gdf.within(boundary_geom)
            if inside_mask.any():
                work = work.loc[inside_mask.values].reset_index(drop=True)
        except Exception:
            pass
        if work.empty:
            return pd.DataFrame()

    lats = work['lat'].dropna().values
    lons = work['lon'].dropna().values
    if len(lats) == 0:
        return pd.DataFrame()

    q1_la, q3_la = np.percentile(lats, 5), np.percentile(lats, 95)
    q1_lo, q3_lo = np.percentile(lons, 5), np.percentile(lons, 95)
    iqr_la, iqr_lo = q3_la - q1_la, q3_lo - q1_lo
    buf_la, buf_lo = max(iqr_la * 0.5, 0.01), max(iqr_lo * 0.5, 0.01)
    mask = ((lats >= q1_la - buf_la) & (lats <= q3_la + buf_la) &
            (lons >= q1_lo - buf_lo) & (lons <= q3_lo + buf_lo))
    clean_lats, clean_lons = lats[mask], lons[mask]
    if len(clean_lats) == 0:
        clean_lats, clean_lons = lats, lons

    base_coords = np.column_stack([clean_lats, clean_lons])
    if len(base_coords) == 0:
        return pd.DataFrame()

    try:
        from sklearn.cluster import MiniBatchKMeans as _KM
        k = min(n, len(base_coords))
        km = _KM(n_clusters=k, random_state=42, batch_size=1024, n_init=3)
        km.fit(base_coords)
        centroids = km.cluster_centers_
    except Exception:
        np.random.seed(42)
        idx = np.random.choice(len(base_coords), min(n, len(base_coords)), replace=False)
        centroids = base_coords[idx]

    # Snap every centroid to the nearest actual in-boundary call to guarantee the
    # proposed station remains inside the jurisdiction geometry.
    snapped = []
    for cen_lat, cen_lon in centroids:
        d2 = (base_coords[:, 0] - cen_lat) ** 2 + (base_coords[:, 1] - cen_lon) ** 2
        nearest = base_coords[int(np.argmin(d2))]
        snapped.append((float(nearest[0]), float(nearest[1])))

    if not snapped:
        return pd.DataFrame()

    deduped = list(dict.fromkeys((round(lat, 6), round(lon, 6)) for lat, lon in snapped))
    station_lats = np.array([lat for lat, _ in deduped])
    station_lons = np.array([lon for _, lon in deduped])

    k_actual = len(station_lats)
    types = (['Police'] * max(1, math.ceil(k_actual * 0.5)) +
             ['Fire']   * max(1, math.ceil(k_actual * 0.3)) +
             ['School'] * max(1, math.ceil(k_actual * 0.2)))[:k_actual]
    return pd.DataFrame({
        'name': [f"{types[i]} Station {i+1}" for i in range(k_actual)],
        'lat':  station_lats,
        'lon':  station_lons,
        'type': types,
    })

@st.cache_data(show_spinner=False)
def _fetch_osm_stations_cached(cen_lat_r: float, cen_lon_r: float, max_stations: int = 100):
    """Cache-friendly OSM query keyed on rounded centroid (2 dp ≈ 1 km grid).
    Returns (list_of_dicts | None, note_str).  Timeout is kept short so the
    UI falls back to random stations quickly if Overpass is unresponsive.
    """
    osm_urls = [
        'https://overpass-api.de/api/interpreter',
        'https://overpass.kumi.systems/api/interpreter',
        'https://overpass.openstreetmap.ru/api/interpreter',
    ]
    for R in [0.25, 0.45]:
        bbox = f"{cen_lat_r - R},{cen_lon_r - R},{cen_lat_r + R},{cen_lon_r + R}"
        query = (
            f'[out:json][timeout:20];'
            f'(node["amenity"="fire_station"]({bbox});'
            f'node["amenity"="police"]({bbox});'
            f'node["amenity"="school"]({bbox});'
            f'way["amenity"="fire_station"]({bbox});'
            f'way["amenity"="police"]({bbox});'
            f'way["amenity"="school"]({bbox});'
            f');out center;'
        )
        data = None
        for osm_url in osm_urls:
            for _attempt in range(2):  # one retry per mirror before moving on
                try:
                    req = urllib.request.Request(
                        f"{osm_url}?data={urllib.parse.quote(query)}",
                        headers={'User-Agent': 'BRINC_COS_Optimizer/1.0'}
                    )
                    with urllib.request.urlopen(req, timeout=8) as resp:  # fail fast
                        data = json.loads(resp.read().decode('utf-8'))
                    break
                except Exception:
                    if _attempt == 0:
                        import time as _t; _t.sleep(1)
                    continue
            if data is not None:
                break

        if data is None:
            continue

        rows = []
        for el in data.get('elements', []):
            tags = el.get('tags', {})
            lat = el.get('lat') or (el.get('center') or {}).get('lat')
            lon = el.get('lon') or (el.get('center') or {}).get('lon')
            if lat is None or lon is None:
                continue
            amenity = tags.get('amenity', '')
            type_label = 'Fire' if amenity == 'fire_station' else 'Police' if amenity == 'police' else 'School'
            rows.append({'name': tags.get('name', f"{type_label} Station"),
                         'lat': round(lat, 6), 'lon': round(lon, 6), 'type': type_label})

        if rows:
            df_s = pd.DataFrame(rows).drop_duplicates(subset=['lat', 'lon']).reset_index(drop=True)
            counts, new_names = {}, []
            for n in df_s['name']:
                if n in counts:
                    counts[n] += 1
                    new_names.append(f"{n} ({counts[n]})")
                else:
                    counts[n] = 0
                    new_names.append(n)
            df_s['name'] = new_names
            if len(df_s) > max_stations:
                pri = {'Police': 0, 'Fire': 1, 'School': 2}
                df_s['_pri'] = df_s['type'].map(pri).fillna(3)
                df_s = df_s.sort_values('_pri').head(max_stations).drop(columns='_pri').reset_index(drop=True)
            return df_s.to_dict('records'), f"Found {len(df_s)} stations from OpenStreetMap."

    return None, "OSM unavailable"


def generate_stations_from_calls(df_calls, max_stations=100):
    """Query OpenStreetMap for real stations; fall back gracefully if unavailable."""
    lats = df_calls['lat'].dropna().values
    lons = df_calls['lon'].dropna().values
    if len(lats) == 0:
        return None, "No coordinates available to generate stations."

    q1_la, q3_la = np.percentile(lats, 25), np.percentile(lats, 75)
    q1_lo, q3_lo = np.percentile(lons, 25), np.percentile(lons, 75)
    iqr_la, iqr_lo = q3_la - q1_la, q3_lo - q1_lo
    mask = (
        (lats >= q1_la - 2.5 * iqr_la) & (lats <= q3_la + 2.5 * iqr_la) &
        (lons >= q1_lo - 2.5 * iqr_lo) & (lons <= q3_lo + 2.5 * iqr_lo)
    )
    # Round to 2 dp (~1 km) so the same city always hits the cache
    cen_lat_r = round(float(lats[mask].mean()), 2)
    cen_lon_r = round(float(lons[mask].mean()), 2)

    rows, note = _fetch_osm_stations_cached(cen_lat_r, cen_lon_r, max_stations)
    if rows:
        return pd.DataFrame(rows), note

    # All OSM attempts failed — derive candidate sites from call density
    df_fallback = _make_random_stations(df_calls, n=40)
    if not df_fallback.empty:
        return df_fallback, "⚠️ OpenStreetMap unavailable — using estimated station locations from call data. Upload a stations CSV for accuracy."
    return None, "Could not generate stations — no valid call coordinates."

# ============================================================
# CACHED DATA FUNCTIONS
# ============================================================
@st.cache_data
def get_address_from_latlon(lat, lon):
    url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat}&lon={lon}&zoom=18&addressdetails=1"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'BRINC_DFR_Optimizer_App/2.0'})
        with urllib.request.urlopen(req, timeout=2) as response:
            data = json.loads(response.read().decode('utf-8'))
            if 'address' in data:
                addr = data['address']
                road = addr.get('road', '')
                house_number = addr.get('house_number', '')
                city = addr.get('city', addr.get('town', addr.get('village', '')))
                if road:
                    return f"{house_number} {road}, {city}".strip(', ')
    except Exception:
        pass
    # Fallback to coordinates if an exact street address isn't found
    return f"{lat:.5f}, {lon:.5f}"

@st.cache_data
def forward_geocode(address_str):
    import urllib.parse
    url = f"https://nominatim.openstreetmap.org/search?format=json&q={urllib.parse.quote(address_str)}&limit=1"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'BRINC_COS_Optimizer/1.0'})
        with urllib.request.urlopen(req, timeout=5) as response:
            data = json.loads(response.read().decode('utf-8'))
            if data:
                return float(data[0]['lat']), float(data[0]['lon'])
    except Exception: pass
    return None, None

@st.cache_data(show_spinner=False)
def lookup_zip_code(zip_code: str):
    """
    Look up a US ZIP code and return (city, state_abbr, county) using the free
    Zippopotam.us API.  Returns (None, None, None) on failure.
    """
    zip_code = zip_code.strip()
    if not re.match(r'^\d{5}$', zip_code):
        return None, None, None
    try:
        url = f"https://api.zippopotam.us/us/{zip_code}"
        req = urllib.request.Request(url, headers={'User-Agent': 'BRINC_COS_Optimizer/1.0'})
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode('utf-8'))
        place = data['places'][0]
        city  = place['place name']
        state = place['state abbreviation']
        return city, state, place.get('state', '')
    except Exception:
        return None, None, None

@st.cache_data
def normalize_jurisdiction_name(name):
    if not name:
        return ""
    name = str(name).lower().strip()
    name = re.sub(r'\bst\b\.?', 'saint', name)
    name = re.sub(r'[^a-z0-9\s-]', ' ', name)
    for suffix in [' city', ' town', ' village', ' borough', ' township', ' cdp', ' municipality', ' county', ' parish']:
        if name.endswith(suffix):
            name = name[:-len(suffix)].strip()
            break
    name = re.sub(r'\s+', ' ', name).strip()
    return name

def lookup_county_for_city(city_name, state_abbr):
    """Use Nominatim reverse-geocode to find the county name for a city that
    doesn't directly match a county name in the local parquet."""
    try:
        lat, lon = forward_geocode(f"{city_name}, {state_abbr}, USA")
        if lat is None: return None
        url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat}&lon={lon}&zoom=8&addressdetails=1"
        req = urllib.request.Request(url, headers={'User-Agent': 'BRINC_COS_Optimizer/1.0'})
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode('utf-8'))
        county_raw = data.get('address', {}).get('county', '')
        # Nominatim returns "Winnebago County" — strip the suffix
        county_name = county_raw.replace(' County', '').replace(' Parish', '').replace(' Borough', '').strip()
        return county_name if county_name else None
    except Exception:
        return None

def fetch_county_by_centroid(df_calls, state_abbr):
    """Find the county boundary that contains the median centroid of the call data.

    Uses a pure spatial lookup against counties_lite.parquet — no network calls,
    no name-matching.  Returns (True, GeoDataFrame) or (False, None).
    """
    local_file = "counties_lite.parquet"
    if not os.path.exists(local_file):
        return False, None

    state_fips = STATE_FIPS.get(state_abbr)
    if not state_fips:
        return False, None

    try:
        lat = float(df_calls['lat'].dropna().median())
        lon = float(df_calls['lon'].dropna().median())
    except Exception:
        return False, None

    try:
        gdf = gpd.read_parquet(local_file)
        state_rows = gdf[gdf['STATEFP'] == state_fips].copy()
        if state_rows.empty:
            return False, None

        from shapely.geometry import Point
        pt = Point(lon, lat)  # geographic order: (x=lon, y=lat)

        containing = state_rows[state_rows.geometry.contains(pt)]
        if containing.empty:
            # Fall back to nearest centroid in case the point lands on a boundary
            state_rows = state_rows.copy()
            state_rows['_dist'] = state_rows.geometry.distance(pt)
            containing = state_rows.nsmallest(1, '_dist')

        if not containing.empty:
            result = containing[['NAME', 'geometry']].copy()
            result['NAME'] = result['NAME'].astype(str) + " County"
            return True, result
    except Exception as e:
        print(f"[BRINC] fetch_county_by_centroid failed: {e}")

    return False, None


@st.cache_data
def fetch_county_boundary_local(state_abbr, county_name_input):
    # 1. Clean the input
    search_name = normalize_jurisdiction_name(county_name_input)
        
    state_fips = STATE_FIPS.get(state_abbr)
    if not state_fips: return False, None
    
    # 2. Look for our new ultra-compressed parquet file
    local_file = "counties_lite.parquet"
    if not os.path.exists(local_file):
        st.error(f"Missing {local_file}! Please ensure it is uploaded to your repository.")
        return False, None
                
    # 3. Read directly from the Parquet file instantly
    try:
        # Geopandas reads Parquet files in milliseconds!
        gdf = gpd.read_parquet(local_file)
        
        # Filter for the exact State FIPS code and County Name
        match = gdf[(gdf['STATEFP'] == state_fips) & (gdf['NAME'].str.lower() == search_name)]
        
        if not match.empty:
            # Put the word "County" back on for the UI displays
            match = match.copy()
            match['NAME'] = match['NAME'] + " County"
            return True, match[['NAME', 'geometry']]
    except Exception as e:
        st.error(f"Error reading local database: {e}")
        pass
        
    return False, None

@st.cache_data
def fetch_place_boundary_local(state_abbr, place_name_input):
    """Look up a city/town/CDP boundary from the local places_lite.parquet.
    Returns (True, GeoDataFrame) on success, (False, None) if not found or
    the file doesn't exist yet (falls back to county lookup in caller)."""
    local_file = "places_lite.parquet"
    if not os.path.exists(local_file):
        return False, None   # file not yet added — caller falls back to county

    state_fips = STATE_FIPS.get(state_abbr)
    if not state_fips: return False, None

    search_name = normalize_jurisdiction_name(place_name_input)

    try:
        gdf = gpd.read_parquet(local_file)
        state_rows = gdf[gdf["STATEFP"] == state_fips]

        state_rows = state_rows.copy()
        state_rows['_norm_name'] = state_rows['NAME'].astype(str).apply(normalize_jurisdiction_name)
        if 'NAMELSAD' in state_rows.columns:
            state_rows['_norm_lsad'] = state_rows['NAMELSAD'].astype(str).apply(normalize_jurisdiction_name)
        else:
            state_rows['_norm_lsad'] = state_rows['_norm_name']

        # Exact normalized match first
        match = state_rows[(state_rows['_norm_name'] == search_name) | (state_rows['_norm_lsad'] == search_name)]

        # Partial normalized match fallback (e.g. Fort Worth / Fort Worth city)
        if match.empty:
            match = state_rows[
                state_rows['_norm_name'].str.startswith(search_name) |
                state_rows['_norm_lsad'].str.startswith(search_name)
            ]
            if not match.empty:
                match = match.copy()
                match['_diff'] = match['NAME'].astype(str).str.len() - len(search_name)
                match = match.sort_values('_diff').head(1)

        if match.empty:
            return False, None

        result = match.copy()
        # Use NAMELSAD for display if available (e.g. "Rockford city"), else NAME
        name_col = "NAMELSAD" if "NAMELSAD" in result.columns else "NAME"
        result["NAME"] = result[name_col].astype(str)
        return True, result[["NAME", "geometry"]]

    except Exception:
        return False, None

@st.cache_data
def reverse_geocode_state(lat, lon):
    url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat}&lon={lon}&zoom=10&addressdetails=1"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'BRINC_COS_Optimizer/1.0'})
        with urllib.request.urlopen(req, timeout=8) as response:
            data = json.loads(response.read().decode('utf-8'))
            address = data.get('address', {})
            state = address.get('state', '')
            city = (
                address.get('city')
                or address.get('town')
                or address.get('village')
                or address.get('municipality')
                or address.get('hamlet')
            )
            return state, city
    except Exception:
        return None, None

@st.cache_data
def fetch_census_population(state_fips, place_name, is_county=False):
    if is_county:
        url = f"https://api.census.gov/data/2020/dec/pl?get=P1_001N,NAME&for=county:*&in=state:{state_fips}"
    else:
        url = f"https://api.census.gov/data/2020/dec/pl?get=P1_001N,NAME&for=place:*&in=state:{state_fips}"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as response:
            data = json.loads(response.read().decode('utf-8'))
            search_name = str(place_name).lower().strip()
            for suffix in [' city', ' town', ' village', ' borough', ' township', ' cdp', ' municipality']:
                if search_name.endswith(suffix):
                    search_name = search_name[:-len(suffix)].strip()
                    break

            exact_match = None
            prefix_match = None
            for row in data[1:]:
                place_full = str(row[1]).lower().split(',')[0].strip()
                place_core = place_full
                for suffix in [' city', ' town', ' village', ' borough', ' township', ' cdp', ' municipality', ' county']:
                    if place_core.endswith(suffix):
                        place_core = place_core[:-len(suffix)].strip()
                        break
                if place_core == search_name or place_full == search_name:
                    exact_match = int(row[0])
                    break
                if is_county and (place_full.startswith(search_name + ' ') or place_core.startswith(search_name + ' ')):
                    prefix_match = int(row[0])
            if exact_match is not None:
                return exact_match
            if prefix_match is not None:
                return prefix_match
    except Exception:
        pass
    return KNOWN_POPULATIONS.get(place_name)

SHAPEFILE_DIR = "jurisdiction_data"
if not os.path.exists(SHAPEFILE_DIR): os.makedirs(SHAPEFILE_DIR)

def _sanitize_boundary_token(value):
    return str(value or "").strip().replace(" ", "_").replace("/", "_")

def _boundary_shp_base(kind, name, state_abbr):
    return os.path.join(SHAPEFILE_DIR, f"{kind}__{_sanitize_boundary_token(name)}_{state_abbr}")

def save_boundary_gdf(boundary_gdf, kind, name, state_abbr):
    """Save boundary to a type-specific shapefile base so place/county do not overwrite each other."""
    try:
        base = _boundary_shp_base(kind, name, state_abbr)
        # Remove older files for this exact base so a fresh write wins cleanly
        for ext in [".shp", ".shx", ".dbf", ".prj", ".cpg"]:
            fp = base + ext
            if os.path.exists(fp):
                try:
                    os.remove(fp)
                except Exception:
                    pass
        boundary_gdf.to_file(base + ".shp")
        return base + ".shp"
    except Exception:
        return None

def load_saved_boundary(kind, name, state_abbr):
    """Load a previously saved boundary, preferring the exact typed name."""
    try:
        exact = _boundary_shp_base(kind, name, state_abbr) + ".shp"
        if os.path.exists(exact):
            gdf = gpd.read_file(exact)
            if gdf.crs is None:
                gdf = gdf.set_crs(epsg=4269)
            return gdf.to_crs(epsg=4326)
    except Exception:
        pass
    return None

@st.cache_data
def fetch_tiger_city_shapefile(state_fips, city_name, output_dir):
    # Check if we already downloaded and cached this state's places file
    temp_dir = os.path.join(output_dir, f"temp_tiger_{state_fips}")
    cached_shp = os.path.join(temp_dir, f"tl_2023_{state_fips}_place.shp")
    gdf = None

    if os.path.exists(cached_shp):
        try:
            gdf = gpd.read_file(cached_shp)
        except Exception:
            gdf = None

    if gdf is None:
        # Download from Census TIGER — try 2023 then 2022 as fallback
        for year in ["2023", "2022"]:
            url = f"https://www2.census.gov/geo/tiger/TIGER{year}/PLACE/tl_{year}_{state_fips}_place.zip"
            try:
                req = urllib.request.Request(url, headers={"User-Agent": "BRINC_COS_Optimizer/1.0"})
                with urllib.request.urlopen(req, timeout=45) as resp:
                    zip_data = resp.read()
                zip_file = zipfile.ZipFile(io.BytesIO(zip_data))
                os.makedirs(temp_dir, exist_ok=True)
                zip_file.extractall(temp_dir)
                shp_files = glob.glob(os.path.join(temp_dir, "*.shp"))
                if shp_files:
                    gdf = gpd.read_file(shp_files[0])
                    break
            except Exception:
                continue

    if gdf is None:
        return False, None

    try:
        search_name = city_name.lower().strip()
        exact_mask = gdf['NAME'].str.lower().str.strip() == search_name
        if exact_mask.any():
            city_gdf = gdf[exact_mask].copy()
        else:
            # Partial match — prefer the longest name match to avoid tiny place with same substring
            partial = gdf[gdf['NAME'].str.lower().str.contains(search_name, case=False, na=False)].copy()
            if partial.empty:
                return False, None
            # Pick the row whose NAME most closely matches (shortest extra chars)
            partial['_diff'] = partial['NAME'].str.len() - len(search_name)
            city_gdf = partial.sort_values('_diff').head(1)

        if city_gdf.empty:
            return False, None

        city_gdf = city_gdf.dissolve(by='NAME').reset_index()
        if city_gdf.crs is None:
            city_gdf = city_gdf.set_crs(epsg=4269)
        city_gdf = city_gdf.to_crs(epsg=4326)
        save_path = os.path.join(output_dir, f"{city_name.replace(' ', '_')}_{state_fips}.shp")
        city_gdf.to_file(save_path)
        return True, city_gdf
    except Exception:
        return False, None

def generate_mock_faa_grid(minx, miny, maxx, maxy):
    features = []
    x_steps = np.linspace(minx, maxx, 20)
    y_steps = np.linspace(miny, maxy, 20)
    mock_airports = [{"lon": minx + 0.3 * (maxx - minx), "lat": miny + 0.3 * (maxy - miny), "radius": 0.15, "name": "Mock Intl (MCK)"}]
    for i in range(len(x_steps) - 1):
        for j in range(len(y_steps) - 1):
            cell_poly = [[x_steps[i], y_steps[j]], [x_steps[i+1], y_steps[j]], [x_steps[i+1], y_steps[j+1]], [x_steps[i], y_steps[j+1]], [x_steps[i], y_steps[j]]]
            cell_center = Point((x_steps[i] + x_steps[i+1]) / 2, (y_steps[j] + y_steps[j+1]) / 2)
            ceiling, arpt_name = None, ""
            for ap in mock_airports:
                dist_ratio = cell_center.distance(Point(ap["lon"], ap["lat"])) / ap["radius"]
                if dist_ratio < 1.0:
                    if   dist_ratio < 0.15: ceiling, arpt_name = 0,   ap["name"]
                    elif dist_ratio < 0.35: ceiling, arpt_name = 50,  ap["name"]
                    elif dist_ratio < 0.55: ceiling, arpt_name = 100, ap["name"]
                    else:                   ceiling, arpt_name = 200, ap["name"]
                    break
            if ceiling is not None:
                features.append({"type": "Feature", "geometry": {"type": "Polygon", "coordinates": [cell_poly]}, "properties": {"CEILING": ceiling, "ARPT_Name": arpt_name}})
    return {"type": "FeatureCollection", "features": features}

@st.cache_data
def load_faa_parquet(minx, miny, maxx, maxy):
    if not os.path.exists("faa_uasfm.parquet"): return generate_mock_faa_grid(minx, miny, maxx, maxy)
    try:
        gdf = gpd.read_parquet("faa_uasfm.parquet")
        pad = 0.05
        filtered = gdf.cx[minx-pad:maxx+pad, miny-pad:maxy+pad]
        if filtered.empty: return {"type": "FeatureCollection", "features": []}
        return json.loads(filtered.to_json())
    except Exception as e: return generate_mock_faa_grid(minx, miny, maxx, maxy)

def add_faa_laanc_layer_to_plotly(fig, faa_geojson, is_dark=True):
    if not faa_geojson or not faa_geojson.get("features"): return
    text_lons, text_lats, text_strings, text_hovers = [], [], [], []
    for feature in faa_geojson.get("features", []):
        geom = feature.get("geometry")
        props = feature.get("properties", {})
        ceiling = props.get("CEILING")
        arpt = props.get("ARPT_Name") or props.get("ARPT_NAME") or "Unknown Airport"
        if ceiling is None or geom is None or geom.get("type") != "Polygon": continue
        snapped = min(FAA_CEILING_COLORS.keys(), key=lambda v: abs(v - ceiling))
        colors = FAA_CEILING_COLORS.get(snapped, FAA_DEFAULT_COLOR)
        coords = geom["coordinates"][0]
        bx, by = zip(*coords)
        fig.add_trace(go.Scattermap(mode="lines", lon=list(bx), lat=list(by), fill="toself", fillcolor=colors["fill"], line=dict(color=colors["line"], width=1.5), hoverinfo="text", text=f"<b>{ceiling} ft AGL</b><br>{arpt}", name=f"LAANC {ceiling}ft", showlegend=False))
        try:
            centroid = shape(geom).centroid
            text_lons.append(centroid.x); text_lats.append(centroid.y); text_strings.append(str(ceiling)); text_hovers.append(f"{ceiling} ft — {arpt}")
        except Exception: pass
    if text_lons:
        fig.add_trace(go.Scattermap(mode="text", lon=text_lons, lat=text_lats, text=text_strings, hovertext=text_hovers, hoverinfo="text", textfont=dict(size=10, color="#ffffff" if is_dark else "#000000"), showlegend=False, name="LAANC Labels"))

def get_station_faa_ceiling(lat, lon, faa_geojson):
    if not faa_geojson or 'features' not in faa_geojson: return "400 ft (Class G)"
    pt = Point(lon, lat)
    for feature in faa_geojson['features']:
        if 'geometry' in feature and feature['geometry']:
            try:
                s = shape(feature['geometry'])
                if s.contains(pt):
                    val = feature['properties'].get('CEILING')
                    if val is not None: return f"{val} ft (Controlled)"
            except Exception: pass
    return "400 ft (Class G)"

@st.cache_data
def fetch_airfields(minx, miny, maxx, maxy):
    pad = 0.2
    query = f"""[out:json];(node["aeroway"~"aerodrome|heliport"]({miny-pad},{minx-pad},{maxy+pad},{maxx+pad});way["aeroway"~"aerodrome|heliport"]({miny-pad},{minx-pad},{maxy+pad},{maxx+pad}););out center;"""
    try:
        req = urllib.request.Request("https://overpass-api.de/api/interpreter", data=query.encode('utf-8'), headers={'User-Agent': 'BRINC_Optimizer'})
        with urllib.request.urlopen(req, timeout=10) as response:
            data = json.loads(response.read().decode('utf-8'))
            airfields = []
            for el in data.get('elements', []):
                lat = el.get('lat') or el.get('center', {}).get('lat')
                lon = el.get('lon') or el.get('center', {}).get('lon')
                name = el.get('tags', {}).get('name', 'Unknown Airfield')
                if lat and lon: airfields.append({'name': name, 'lat': lat, 'lon': lon})
            return airfields
    except Exception: return []

def get_nearest_airfield(lat, lon, airfields):
    if not airfields: return "No data"
    min_dist = float('inf')
    best = None
    for af in airfields:
        lat1, lon1, lat2, lon2 = map(math.radians, [lat, lon, af['lat'], af['lon']])
        a = math.sin((lat2-lat1)/2)**2 + math.cos(lat1)*math.cos(lat2)*math.sin((lon2-lon1)/2)**2
        dist = 3958.8 * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
        if dist < min_dist:
            y = math.sin(lon2-lon1)*math.cos(lat2)
            x = math.cos(lat1)*math.sin(lat2) - math.sin(lat1)*math.cos(lat2)*math.cos(lon2-lon1)
            bearing = (math.degrees(math.atan2(y, x)) + 360) % 360
            dirs = ['N','NNE','NE','ENE','E','ESE','SE','SSE','S','SSW','SW','WSW','W','WNW','NW','NNW']
            min_dist = dist
            best = (af['name'], dist, dirs[int((bearing+11.25)/22.5) % 16])
    if best:
        n = best[0][:18] + ("..." if len(best[0]) > 18 else "")
        return f"{best[1]:.1f}mi {best[2]} ({n})"
    return "No data"

def generate_random_points_in_polygon(polygon, num_points):
    # Flatten MultiPolygon to its largest component so bbox sampling stays efficient
    if isinstance(polygon, MultiPolygon):
        polygon = max(polygon.geoms, key=lambda p: p.area)
    points = []
    minx, miny, maxx, maxy = polygon.bounds
    while len(points) < num_points:
        x_coords = np.random.uniform(minx, maxx, 1000)
        y_coords = np.random.uniform(miny, maxy, 1000)
        for x, y in zip(x_coords, y_coords):
            if len(points) >= num_points: break
            if polygon.contains(Point(x, y)): points.append((y, x))
    return points

def generate_clustered_calls(polygon, num_points):
    if isinstance(polygon, MultiPolygon):
        polygon = max(polygon.geoms, key=lambda p: p.area)
    points = []
    minx, miny, maxx, maxy = polygon.bounds
    hotspots = []
    while len(hotspots) < random.randint(5, 15):
        hx, hy = random.uniform(minx, maxx), random.uniform(miny, maxy)
        if polygon.contains(Point(hx, hy)): hotspots.append((hx, hy))
    target_clustered = int(num_points * 0.75)
    while len(points) < target_clustered:
        hx, hy = random.choice(hotspots)
        px, py = np.random.normal(hx, 0.02), np.random.normal(hy, 0.02)
        if polygon.contains(Point(px, py)): points.append((py, px))
    while len(points) < num_points:
        px, py = random.uniform(minx, maxx), random.uniform(miny, maxy)
        if polygon.contains(Point(px, py)): points.append((py, px))
    np.random.shuffle(points)
    return points

def estimate_grants(population):
    if population > 1000000: return "$1.5M - $3.0M+"
    elif population > 500000: return "$500k - $1.5M"
    elif population > 250000: return "$250k - $500k"
    elif population > 100000: return "$100k - $250k"
    else: return "$25k - $100k"

def get_circle_coords(lat, lon, r_mi=2.0):
    angles = np.linspace(0, 2*np.pi, 100)
    c_lats = lat + (r_mi/69.172) * np.sin(angles)
    c_lons = lon + (r_mi/(69.172 * np.cos(np.radians(lat)))) * np.cos(angles)
    return c_lats, c_lons

def format_3_lines(name_str):
    match = re.search(r'\s(\d{1,5}\s+[A-Za-z])', name_str)
    if match:
        idx = match.start()
        line1 = name_str[:idx].strip()
        rest = name_str[idx:].strip()
        if ',' in rest:
            parts = rest.split(',', 1)
            return f"{line1}<br>{parts[0].strip()},<br>{parts[1].strip()}"
        return f"{line1}<br>{rest}"
    if ',' in name_str:
        parts = name_str.split(',')
        if len(parts) >= 3:
            return f"{parts[0].strip()},<br>{parts[1].strip()},<br>{','.join(parts[2:]).strip()}"
    return name_str

def _build_unit_cards_html(active_drones, text_main, text_muted, card_bg, card_border, card_title, accent_color, columns_per_row=2, simple=False, deflection_rate=0.25):
    if not active_drones:
        return ""
    # Per-type daily airtime budgets derived from CONFIG duty cycles:
    #   Guardian: 60 min flight + 3 min charge → (24*60/63)*60 = 1371.4 min = 22.86 hr
    #   Responder: patrol-unit model, 11.6 hr shift equivalent
    _GUARDIAN_DAILY_MINS  = CONFIG["GUARDIAN_DAILY_FLIGHT_MIN"]   # ~1371.4
    _GUARDIAN_DAILY_HOURS = CONFIG["GUARDIAN_PATROL_HOURS"]        # ~22.86
    _RESPONDER_DAILY_MINS  = CONFIG["RESPONDER_PATROL_HOURS"] * 60 # 11.6 * 60 = 696
    _RESPONDER_DAILY_HOURS = CONFIG["RESPONDER_PATROL_HOURS"]      # 11.6
    columns_per_row = max(1, int(columns_per_row))

    # Specialty-response values are independent from Annual Capacity Value.
    # They are modeled per station from that station's own calls-in-range and
    # resulting drone flights, not allocated as a share of fleet totals.
    _THERMAL_RATE = float(CONFIG.get("THERMAL_DEFAULT_APPLICABLE_RATE", 0.12) or 0)
    _THERMAL_PER_CALL = float(CONFIG.get("THERMAL_SAVINGS_PER_CALL", 38) or 0)
    _K9_RATE = float(CONFIG.get("K9_DEFAULT_APPLICABLE_RATE", 0.03) or 0)
    _K9_PER_CALL = float(CONFIG.get("K9_SAVINGS_PER_CALL", 155) or 0)
    _FIRE_RATE = float(CONFIG.get("FIRE_DEFAULT_APPLICABLE_RATE", 0.05) or 0)
    _FIRE_PER_CALL = float(CONFIG.get("FIRE_SAVINGS_PER_CALL", 450) or 0)

    cards_html = []
    for d in active_drones:
        short_name  = format_3_lines(d["name"])
        d_color     = d["color"]
        d_type      = d["type"]
        d_step      = d["deploy_step"]
        d_savings   = d["annual_savings"]
        d_flights   = d["marginal_flights"]
        d_shared    = d["shared_flights"]
        # Resolved/day = total station flights (exclusive + shared) × deflection rate.
        d_deflected  = (d_flights + d_shared) * deflection_rate
        d_time      = d["avg_time_min"]
        d_faa       = d["faa_ceiling"]
        d_airport   = d["nearest_airport"]
        d_cost      = d["cost"]
        d_be        = d["be_text"]
        d_lat       = d["lat"]
        d_lon       = d["lon"]
        d_address   = get_address_from_latlon(d_lat, d_lon)
        gmaps_url   = f"https://www.google.com/maps/search/?api=1&query={d_lat},{d_lon}"

        # Pick duty-cycle values for this drone type
        is_guardian = (d_type == "GUARDIAN")
        max_patrol_mins  = _GUARDIAN_DAILY_MINS  if is_guardian else _RESPONDER_DAILY_MINS
        max_patrol_hours = _GUARDIAN_DAILY_HOURS if is_guardian else _RESPONDER_DAILY_HOURS

        # Uptime tooltip: show the duty-cycle breakdown for Guardians
        if is_guardian:
            _g_fl  = CONFIG["GUARDIAN_FLIGHT_MIN"]
            _g_ch  = CONFIG["GUARDIAN_CHARGE_MIN"]
            _g_cyc = _g_fl + _g_ch
            _cycles_per_day = (24 * 60) / _g_cyc
            uptime_tooltip = (
                f"{_g_fl}min flight + {_g_ch}min charge = {_g_cyc}min cycle · "
                f"{_cycles_per_day:.1f} cycles/day · "
                f"{max_patrol_hours:.2f}hr airtime"
            )
        else:
            uptime_tooltip = f"{max_patrol_hours}hr patrol shift"

        total_daily_flights = d_flights + d_shared
        d_zone_calls = float(d.get("zone_calls_annual", 0) or 0)
        d_zone_flights_annual = float(d.get("zone_flights_annual", total_daily_flights * 365.0) or 0)
        # Cap thermal/K9 base to physically serviceable flights (max_flights_cap * 365)
        # zone_flights_annual is raw DEMANDED flights — thermal/K9 assists can only
        # happen on flights actually flown within the 10-min scene-floor capacity.
        _serviceable_annual = float(d.get("max_flights_cap", 0) or 0) * 365.0
        _flight_base = min(d_zone_flights_annual, _serviceable_annual) if _serviceable_annual > 0 else d_zone_flights_annual
        # Further cap: assists cannot exceed total zone calls in range
        _flight_base = min(_flight_base, d_zone_calls) if d_zone_calls > 0 else _flight_base
        d_thermal_calls = _flight_base * _THERMAL_RATE
        d_k9_calls      = _flight_base * _K9_RATE
        d_fire_calls    = _flight_base * _FIRE_RATE
        d_thermal = d_thermal_calls * _THERMAL_PER_CALL
        d_k9      = d_k9_calls      * _K9_PER_CALL
        d_fire    = d_fire_calls    * _FIRE_PER_CALL
        patrol_time_line = ""
        if total_daily_flights > 0:
            # Raw calculation: patrol budget / flights = available min per flight
            raw_mins_per_flight = max_patrol_mins / total_daily_flights
            # Cap at drone's physical max flight time
            max_single_flight = CONFIG["GUARDIAN_FLIGHT_MIN"] if is_guardian else CONFIG["RESPONDER_FLIGHT_MIN"]
            mins_per_flight = min(raw_mins_per_flight, max_single_flight)
            capped = raw_mins_per_flight > max_single_flight
            # Always show the line so low-volume Responders display correctly
            patrol_color = "#F0B429" if mins_per_flight < 15 else "#2ecc71" if mins_per_flight >= max_single_flight * 0.9 else "#00D2FF"
            cap_note = f" (max {max_single_flight}min)" if capped else ""
            _annual_flights = total_daily_flights * 365
            patrol_time_line = (
                f'<div style="font-size:0.65rem; color:{text_muted}; text-align:right; line-height:1.2;" '
                f'title="{uptime_tooltip}">'
                f'<span style="font-weight:800; color:{patrol_color};">{total_daily_flights:.1f} flights/day</span> '
                f'<span style="font-weight:400; color:{text_muted}; font-size:0.60rem;">({_annual_flights:,.0f}/yr)</span><br>'
                f'<span style="font-weight:600; color:{patrol_color};">{mins_per_flight:.1f} min/flight{cap_note}</span></div>'
            )

        # Concurrency / value breakdown
        d_util         = d.get('utilization', 0)
        d_true_util    = d.get('true_util', d_util)
        d_on_scene     = d.get('on_scene_min', 99.0)
        d_max_cap      = d.get('max_flights_cap', 0)
        d_has_deficit  = d.get('has_deficit', False)
        d_deficit_f    = d.get('deficit_flights', 0)
        d_unserv_day   = d.get('unserv_calls_day', 0)
        d_unserv_yr    = d.get('unserv_calls_yr', 0)
        d_total_flights_possible_yr = max(0.0, float(d_max_cap or 0) * 365.0)
        d_total_uncovered_flights_yr = max(0.0, float(d_zone_flights_annual or 0) - d_total_flights_possible_yr)
        d_extra_same   = d.get('extra_same', 0)
        d_extra_alt    = d.get('extra_alt', 0)
        d_extra_same_capex = d.get('extra_same_capex', 0)
        d_extra_alt_capex  = d.get('extra_alt_capex', 0)
        d_same_lbl     = d.get('same_type_label', d_type.title())
        d_alt_lbl      = d.get('alt_type_label', 'Guardian' if d_type == 'RESPONDER' else 'Responder')
        d_blocked      = d.get('blocked_per_day', 0)
        d_base_annual  = d.get('base_annual', d_savings)
        d_conc_annual  = d.get('concurrent_annual', 0)
        d_best         = d.get('best_case_annual', d_savings)
        d_best_be      = d.get('best_be_text', d_be)
        # True utilization display — uncapped so it can show > 100% in deficit
        util_pct       = f"{d_true_util*100:.1f}%"
        util_color     = "#dc3545" if d_has_deficit or d_true_util > 0.75 else "#F0B429" if d_true_util > 0.4 else "#2ecc71"
        # On-scene time color coding
        if d_on_scene < 10.0:
            scene_color = "#dc3545"
        elif d_on_scene < 20.0:
            scene_color = "#F0B429"
        else:
            scene_color = "#2ecc71"
        has_concurrent = d_shared > 0.1 and d_conc_annual > 0
        if has_concurrent:
            _excl_str = f"${d_base_annual:,.0f} exclusive"
            _conc_str = f"+ ${d_conc_annual:,.0f} concurrent"
        else:
            _excl_str = "exclusive zone coverage"
            _conc_str = ""

        # ── DEFICIT FOOTER (compact strip at card bottom) ─────────────────────────────
        _sc_fmt = f"${d_extra_same_capex:,}" if d_has_deficit else ""
        _ac_fmt = f"${d_extra_alt_capex:,}" if d_has_deficit else ""

        # ── SIMPLE CARD (toggled from Display Options) ───────────────────────
        if simple:
            _pin_badge = (
                f'<span style="font-size:0.55rem;background:rgba(255,215,0,0.15);color:#FFD700;border:1px solid rgba(255,215,0,0.4);border-radius:3px;padding:1px 5px;margin-left:4px;">🔒 Guardian</span>'
                if (d.get("pinned") and d_type == "GUARDIAN") else
                f'<span style="font-size:0.55rem;background:rgba(0,210,255,0.15);color:#00D2FF;border:1px solid rgba(0,210,255,0.4);border-radius:3px;padding:1px 5px;margin-left:4px;">🔒 Responder</span>'
                if (d.get("pinned") and d_type == "RESPONDER") else ""
            )
            _deficit_strip = (
                f'<div style="font-size:0.60rem;color:#dc3545;font-weight:700;margin-top:6px;padding-top:5px;border-top:1px solid rgba(220,53,69,0.3);">⚠️ Capacity deficit — {d_on_scene:.1f} min on-scene</div>'
                if d_has_deficit else ""
            )
            _specialty_total = d_thermal + d_k9 + d_fire
            _cap_strip = (
                f'<div style="display:flex;align-items:center;gap:5px;margin-top:6px;padding-top:5px;border-top:1px solid rgba(220,53,69,0.3);">'
                f'<span style="font-size:0.60rem;color:#dc3545;font-weight:700;">⚠️ Over capacity</span>'
                f'<span style="font-size:0.59rem;color:{text_muted};">· {d_unserv_day:.0f} calls/day unserviceable</span></div>'
                if d_has_deficit else
                f'<div style="display:flex;align-items:center;gap:5px;margin-top:6px;padding-top:5px;border-top:1px solid rgba(34,197,94,0.2);">'
                f'<span style="font-size:0.60rem;color:#2ecc71;font-weight:700;">✓ Within capacity</span>'
                f'<span style="font-size:0.59rem;color:{text_muted};">· {d_on_scene:.1f} min on-scene</span></div>'
            )
            cards_html.append(f'''
<div class="unit-card" style="background:{card_bg};border:1px solid {"#dc3545" if d_has_deficit else card_border};border-top:3px solid {d_color};border-radius:8px;padding:10px 12px;box-sizing:border-box;">
  <div style="display:flex;align-items:baseline;justify-content:space-between;gap:4px;margin-bottom:2px;">
    <span style="font-weight:700;font-size:0.78rem;color:{card_title};overflow:hidden;text-overflow:ellipsis;white-space:nowrap;flex:1;min-width:0;">{"🔒 " if d.get("pinned") else ""}{d["name"]}</span>
    <span style="font-size:0.58rem;color:#666;text-transform:uppercase;white-space:nowrap;flex-shrink:0;">{d_type} · #{d_step}</span>
  </div>
  <div style="font-size:0.62rem;margin-bottom:6px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">
    <a href="{gmaps_url}" target="_blank" style="color:{accent_color};text-decoration:none;opacity:0.85;">📍 {d_address} ↗</a>
  </div>
  <!-- Response time hero -->
  <div style="background:rgba(255,255,255,0.05);border:1px solid {card_border};border-radius:6px;padding:7px 10px;text-align:center;margin-bottom:6px;">
    <div style="font-size:0.58rem;color:{text_muted};text-transform:uppercase;letter-spacing:0.3px;margin-bottom:1px;">Avg Aerial Response<span class="tip" data-tip="Average travel time from this station to incidents in its zone, based on drone speed and straight-line distance with a 1.4x routing factor.">?</span></div>
    <div style="font-size:1.6rem;font-weight:900;color:{card_title};line-height:1.1;">{d_time:.1f}<span style="font-size:0.9rem;font-weight:600;"> min</span></div>
  </div>
  <!-- 2x2 key metrics -->
  <div style="display:grid;grid-template-columns:1fr 1fr;gap:5px;margin-bottom:6px;">
    <div style="background:rgba(0,210,255,0.07);border:1px solid rgba(0,210,255,0.18);border-radius:5px;padding:6px 8px;text-align:center;">
      <div style="font-size:0.57rem;color:{text_muted};text-transform:uppercase;letter-spacing:0.3px;">Annual Value<span class="tip" data-tip="Best-case annual savings from calls this drone resolves without sending a ground unit. Includes both exclusive and concurrent zone coverage.">?</span></div>
      <div style="font-size:0.95rem;font-weight:900;color:{accent_color};">${d_best:,.0f}</div>
    </div>
    <div style="background:rgba(0,210,255,0.07);border:1px solid rgba(0,210,255,0.18);border-radius:5px;padding:6px 8px;text-align:center;">
      <div style="font-size:0.57rem;color:{text_muted};text-transform:uppercase;letter-spacing:0.3px;">Break-Even<span class="tip" data-tip="Months to recover the unit CapEx from annual capacity savings at current DFR and deflection rates.">?</span></div>
      <div style="font-size:0.95rem;font-weight:900;color:{accent_color};">{d_be}</div>
    </div>
    <div style="background:rgba(255,255,255,0.04);border:1px solid {card_border};border-radius:5px;padding:6px 8px;text-align:center;">
      <div style="font-size:0.57rem;color:{text_muted};text-transform:uppercase;letter-spacing:0.3px;">Resolved/day<span class="tip" data-tip="Calls per day closed without dispatching an officer: total station flights/day (exclusive + shared zone) × deflection rate. Drone arrived, assessed, and dispatch stood down.">?</span></div>
      <div style="font-size:0.88rem;font-weight:800;color:{card_title};">{d_deflected:.1f} calls</div>
    </div>
    <div style="background:rgba(255,255,255,0.04);border:1px solid {card_border};border-radius:5px;padding:6px 8px;text-align:center;">
      <div style="font-size:0.57rem;color:{text_muted};text-transform:uppercase;letter-spacing:0.3px;">Zone Calls/yr<span class="tip" data-tip="Total historical calls for service within this drone's patrol radius. Used as the ceiling for thermal, K-9, and fire assist estimates.">?</span></div>
      <div style="font-size:0.88rem;font-weight:800;color:{card_title};">{int(d_zone_calls):,}</div>
    </div>
  </div>
  <!-- Specialty total -->
  <div style="background:rgba(251,191,36,0.06);border:1px solid rgba(251,191,36,0.18);border-radius:5px;padding:5px 10px;display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;">
    <span style="font-size:0.60rem;color:{text_muted};text-transform:uppercase;letter-spacing:0.3px;">🔥🐕🚒 Specialty Value<span class="tip" data-tip="Combined value from thermal imaging assists (12% of flights), K-9 replacement (3% of flights), and fire scene support (5% of flights). Separate from Annual Capacity Value.">?</span></span>
    <span style="font-size:0.85rem;font-weight:800;color:#fbbf24;">${_specialty_total:,.0f}/yr</span>
  </div>
  <!-- CapEx footer -->
  <div style="display:flex;justify-content:space-between;align-items:center;padding-top:5px;border-top:1px solid {card_border};font-size:0.65rem;">
    <span style="color:{text_muted};">CapEx<span class="tip" data-tip="One-time hardware cost for this unit. Responder: $80,000. Guardian: $160,000.">?</span></span>
    <span style="font-weight:700;color:{card_title};">${d_cost:,.0f}</span>
  </div>
  {_cap_strip}
</div>''')
            continue

        cards_html.append(f'''
<div class="unit-card" style="background:{card_bg}; border:1px solid {"#dc3545" if d_has_deficit else card_border}; border-top:3px solid {d_color}; border-radius:8px; padding:10px 12px; display:flex; flex-direction:column; box-sizing:border-box;">
  <!-- Header: single compact row -->
  <div style="margin-bottom:5px; flex-shrink:0;">
    <div style="display:flex; align-items:baseline; gap:5px; overflow:hidden;">
      <span style="font-weight:700; font-size:0.78rem; color:{card_title}; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; flex:1; min-width:0;">{"🔒 " if d.get("pinned") else ""}{d["name"]}</span>
      <span style="font-size:0.58rem; color:#666; text-transform:uppercase; letter-spacing:0.3px; white-space:nowrap; flex-shrink:0;">{d_type} · #{d_step}</span>
    </div>
    <div style="font-size:0.65rem; margin-top:1px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">
      <a href="{gmaps_url}" target="_blank" style="color:{accent_color}; text-decoration:none; font-weight:500; opacity:0.85;">📍 {d_address} ↗</a>
    </div>
  </div>
  <div style="background:rgba(0,210,255,0.07); border:1px solid rgba(0,210,255,0.15); border-radius:6px; padding:8px 10px; margin-bottom:6px;"
       title="Annual Capacity Value is based on calls handled without sending a squad.">
    <div style="font-size:0.68rem; color:{text_muted}; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:2px;">Annual Capacity Value<span class="tip" data-tip="Estimated annual savings from calls this drone resolves without sending a ground unit. Capped at physical flight capacity.">?</span>{"  ⚠️ capped at physical max" if d_has_deficit else ""}</div>
    <div style="display:flex; align-items:baseline; justify-content:space-between; gap:6px;">
      <div>
        <div style="font-size:1.3rem; font-weight:900; color:{accent_color}; line-height:1.1;">${d_best:,.0f}</div>
        <div style="font-size:0.60rem; color:{text_muted}; margin-top:2px;">calls handled without sending a squad</div>
      </div>
      {patrol_time_line}
    </div>
    <div style="display:grid; grid-template-columns:1fr 1fr; gap:6px; margin-top:7px;">
      <div title="Calls assisted by thermal are modeled from this station's own flights and calls for service in range." style="background:rgba(251,191,36,0.08); border:1px solid rgba(251,191,36,0.22); border-radius:6px; padding:6px 8px;">
        <div style="font-size:0.60rem; color:{text_muted}; text-transform:uppercase; letter-spacing:0.4px; margin-bottom:1px;">🔥 Thermal Response Value<span class="tip" data-tip="Flights where thermal imaging aided the response - est. 12% of serviceable flights. Capped to actual zone calls in range.">?</span></div>
        <div style="font-size:0.72rem; font-weight:700; color:{card_title}; line-height:1.15;">{d_thermal_calls:,.0f} calls assisted</div>
        <div style="font-size:0.85rem; font-weight:800; color:#fbbf24; line-height:1.1; margin-top:2px;">${d_thermal:,.0f}/yr</div>
      </div>
      <div title="K-9 calls assisted are modeled from this station's own flights and calls for service in range." style="background:rgba(57,255,20,0.06); border:1px solid rgba(57,255,20,0.18); border-radius:6px; padding:6px 8px;">
        <div style="font-size:0.60rem; color:{text_muted}; text-transform:uppercase; letter-spacing:0.4px; margin-bottom:1px;">🐕 K-9 Replacement Value<span class="tip" data-tip="Flights where drone thermal support replaced a K-9 deployment - est. 3% of serviceable flights. Capped to zone calls in range.">?</span></div>
        <div style="font-size:0.72rem; font-weight:700; color:{card_title}; line-height:1.15;">{d_k9_calls:,.0f} calls assisted</div>
        <div style="font-size:0.85rem; font-weight:800; color:#39FF14; line-height:1.1; margin-top:2px;">${d_k9:,.0f}/yr</div>
      </div>
    </div>
    <div title="Fire response value is modeled from fire/smoke/alarm calls in this station's zone." style="background:rgba(251,113,33,0.07); border:1px solid rgba(251,113,33,0.25); border-radius:6px; padding:6px 10px; margin-top:6px;">
      <div style="font-size:0.60rem; color:{text_muted}; text-transform:uppercase; letter-spacing:0.4px; margin-bottom:3px;">🚒 Fire Department Value<span class="tip" data-tip="Est. 5% of serviceable flights assist fire calls. Value = aerial scene size-up (avoids premature aerial ladder deployment at ~$4,500/deploy) + overhaul hotspot detection (~$90/call crew time saved). Blended: $450/fire call assisted.">?</span></div>
      <div style="display:flex; justify-content:space-between; align-items:baseline; margin-top:2px;">
        <div>
          <div style="font-size:0.62rem; color:{text_muted};">Fire calls assisted</div>
          <div style="font-size:0.78rem; font-weight:700; color:{card_title};">{d_fire_calls:,.0f}</div>
        </div>
        <div style="font-size:0.85rem; font-weight:800; color:#fb7121;">${d_fire:,.0f}/yr</div>
      </div>
    </div>
  </div>

  <!-- Value breakdown box -->
  <div style="border:1px solid rgba(57,255,20,0.18); border-radius:6px; padding:6px 10px; margin-bottom:8px; background:rgba(57,255,20,0.04);">
    <div style="font-size:0.60rem; color:{text_muted}; text-transform:uppercase; letter-spacing:0.3px; margin-bottom:4px;">Value Breakdown<span class="tip" data-tip="EXCLUSIVE: savings from calls only this drone can reach. Formula: exclusive flights/day x deflection rate x $76/call x 365. Guaranteed regardless of other drones. CONCURRENT: savings from shared-zone calls captured only when the partner drone is already busy. Same formula but conditional — if partner is free it handles those calls instead. Combined = best-case Annual Capacity Value.">?</span></div>
    <div style="display:grid; grid-template-columns:1fr auto 1fr; gap:4px; align-items:center; margin-bottom:4px;">
      <div style="text-align:center;">
        <div style="color:{accent_color}; font-weight:700; font-size:0.78rem;">${d_base_annual:,.0f}</div>
        <div style="color:{text_muted}; font-size:0.63rem;">exclusive</div>
      </div>
      <div style="color:{text_muted}; font-size:0.75rem; opacity:0.5; text-align:center;">+</div>
      <div style="text-align:center;">
        <div style="color:#39FF14; font-weight:700; font-size:0.78rem;">${d_conc_annual:,.0f}</div>
        <div style="color:{text_muted}; font-size:0.63rem;">concurrent</div>
      </div>
    </div>
    <div style="font-size:0.65rem; color:{text_muted}; opacity:0.8; border-top:1px dashed rgba(255,255,255,0.1); padding-top:4px; text-align:center;">{util_pct} utilization{"  ·  ⚠️ over capacity" if d_has_deficit else ""} · ROI {d_best_be}</div>
  </div>

  <div style="display:grid; grid-template-columns:1fr 1fr; gap:4px; font-size:0.68rem; flex:1; margin-bottom:8px; align-content:start;">
    <div style="background:rgba(255,255,255,0.04); border:1px solid {card_border}; border-radius:5px; padding:5px 7px;">
      <div style="color:{text_muted}; font-size:0.60rem; text-transform:uppercase; letter-spacing:0.3px; margin-bottom:1px;">Zone Flights/day<span class="tip" data-tip="Flights demanded per day in this zone: zone_pct x daily_calls x DFR_rate. This is demand, not actual flights flown.">?</span></div>
      <div style="font-weight:800; color:{accent_color}; font-size:0.82rem;">{d.get("zone_flights",d_flights):.1f}/day</div>
      <div style="font-size:0.59rem; color:{text_muted};">({d.get("zone_flights",d_flights)*365:,.0f}/yr)</div>
    </div>
    <div style="background:{"rgba(220,53,69,0.08)" if d_has_deficit else "rgba(255,255,255,0.04)"}; border:1px solid {"#dc3545" if d_has_deficit else card_border}; border-radius:5px; padding:5px 7px;">
      <div style="color:{text_muted}; font-size:0.60rem; text-transform:uppercase; letter-spacing:0.3px; margin-bottom:1px;">{"Effective" if d.get("effective_dfr_rate", dfr_dispatch_rate) < dfr_dispatch_rate - 0.001 else "Max"} DFR Rate<span class="tip" data-tip="The DFR dispatch rate at which this drone hits exactly 100% utilization. Enable Auto-cap in Deployment Strategy to apply this limit per-station without affecting others. Formula: max_flights_cap / (zone_perc × daily_calls).">?</span></div>
      <div style="font-weight:800; color:{"#dc3545" if d_has_deficit else "#2ecc71"}; font-size:0.82rem;">{(d_max_cap / max(d.get("zone_flights", d_flights+d_shared), 0.001)) * dfr_dispatch_rate * 100:.1f}%</div>
      <div style="font-size:0.59rem; color:{text_muted};">{"▼ reduce to clear" if d_has_deficit else ("⚡ auto-capped" if d.get("effective_dfr_rate", dfr_dispatch_rate) < dfr_dispatch_rate - 0.001 else "✓ current rate ok")}</div>
    </div>
    <div style="background:{"rgba(220,53,69,0.08)" if d_has_deficit else "rgba(255,255,255,0.04)"}; border:1px solid {"#dc3545" if d_has_deficit else card_border}; border-radius:5px; padding:5px 7px;">
      <div style="color:{text_muted}; font-size:0.60rem; text-transform:uppercase; letter-spacing:0.3px; margin-bottom:1px;">Utilization<span class="tip" data-tip="Flight time demanded as % of daily capacity using the 10-min on-scene floor model. Over 100% means this drone cannot serve all calls in its zone.">?</span></div>
      <div style="font-weight:800; color:{util_color}; font-size:0.82rem;">{util_pct}</div>
    </div>
    <div style="background:rgba(255,255,255,0.04); border:1px solid {card_border}; border-radius:5px; padding:5px 7px;">
      <div style="color:{text_muted}; font-size:0.60rem; text-transform:uppercase; letter-spacing:0.3px; margin-bottom:1px;">Resolved/day<span class="tip" data-tip="Calls per day closed without dispatching an officer: total station flights/day (exclusive + shared zone) × deflection rate. Drone arrived, assessed, and dispatch stood down.">?</span></div>
      <div style="font-weight:800; color:{card_title}; font-size:0.82rem;">{d_deflected:.1f}</div>
    </div>
    <div style="background:rgba(255,255,255,0.04); border:1px solid {card_border}; border-radius:5px; padding:5px 7px;">
      <div style="color:{text_muted}; font-size:0.60rem; text-transform:uppercase; letter-spacing:0.3px; margin-bottom:1px;">Avg Response<span class="tip" data-tip="Average travel time from this station to incidents in its zone, based on drone speed and straight-line distance with a 1.4x routing factor.">?</span></div>
      <div style="font-weight:800; color:{card_title}; font-size:0.82rem;">{d_time:.1f} min</div>
    </div>
    <div style="background:rgba(255,255,255,0.04); border:1px solid {card_border}; border-radius:5px; padding:5px 7px;">
      <div style="color:{text_muted}; font-size:0.60rem; text-transform:uppercase; letter-spacing:0.3px; margin-bottom:1px;">Zone Calls/yr<span class="tip" data-tip="Total historical calls for service within this drone's patrol radius. Ceiling for thermal and K-9 assist estimates.">?</span></div>
      <div style="font-weight:800; color:{card_title}; font-size:0.82rem;">{int(d_zone_calls):,}</div>
    </div>
    <div style="background:rgba(255,255,255,0.04); border:1px solid {card_border}; border-radius:5px; padding:5px 7px;">
      <div style="color:{text_muted}; font-size:0.60rem; text-transform:uppercase; letter-spacing:0.3px; margin-bottom:1px;">FAA Ceiling<span class="tip" data-tip="Max authorized altitude from FAA LAANC maps at this location. 0 ft = controlled airspace requiring coordination before flight.">?</span></div>
      <div style="font-weight:700; color:{card_title}; font-size:0.72rem; line-height:1.2;">{d_faa}</div>
    </div>
    <div style="background:rgba(255,255,255,0.04); border:1px solid {card_border}; border-radius:5px; padding:5px 7px;">
      <div style="color:{text_muted}; font-size:0.60rem; text-transform:uppercase; letter-spacing:0.3px; margin-bottom:1px;">Nearest Airfield<span class="tip" data-tip="Closest airport or airfield. Affects LAANC authorization altitude and Part 107 waiver requirements.">?</span></div>
      <div style="font-weight:600; color:{card_title}; font-size:0.68rem; line-height:1.2; word-break:break-word;">{d_airport}</div>
    </div>

  </div>

  <!-- CapEx + ROI footer -->
  <div style="border-top:1px solid {card_border}; padding-top:6px; display:grid; grid-template-columns:1fr 1fr; gap:4px 8px; font-size:0.68rem; margin-bottom:8px;">
    <div style="color:{text_muted};">CapEx<span class="tip" data-tip="One-time hardware cost for this unit. Responder: $80,000. Guardian: $160,000.">?</span></div>
    <div style="text-align:right; font-weight:700; color:{card_title};">${d_cost:,.0f}</div>
    <div style="color:{text_muted};">Base ROI<span class="tip" data-tip="Months to recover unit CapEx from exclusive-zone savings alone at current DFR and deflection rates.">?</span></div>
    <div style="text-align:right; font-weight:800; color:{accent_color};">{d_be}</div>
  </div>

  { (f'<div style="border-top:1px solid rgba(220,53,69,0.35);margin-top:4px;padding-top:5px;">'  
       f'<div style="font-size:0.62rem;font-weight:800;color:#dc3545;margin-bottom:3px;">⚠️ CAPACITY DEFICIT · {d_on_scene:.1f} min on-scene (min 10)</div>'  
       f'<div style="font-size:0.59rem;color:{text_muted};margin-bottom:4px;">{d_unserv_day:.0f} calls/day unserviceable · {d_unserv_yr:,.0f}/yr</div>'  
       f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:3px;margin-bottom:4px;">'  
       f'<div style="background:rgba(220,53,69,0.08);border:1px solid rgba(220,53,69,0.2);border-radius:4px;padding:3px 6px;font-size:0.59rem;">'  
       f'<div style="color:{text_muted};">Total Flights Possible</div>'  
       f'<div style="font-weight:700;color:{card_title};">{d_total_flights_possible_yr:,.0f}/yr</div></div>'  
       f'<div style="background:rgba(220,53,69,0.08);border:1px solid rgba(220,53,69,0.2);border-radius:4px;padding:3px 6px;font-size:0.59rem;">'  
       f'<div style="color:{text_muted};">Uncovered Flights</div>'  
       f'<div style="font-weight:700;color:#dc3545;">{d_total_uncovered_flights_yr:,.0f}/yr</div></div>'  
       f'</div>'  
       f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:3px;">'  
       f'<div style="background:rgba(220,53,69,0.08);border:1px solid rgba(220,53,69,0.2);border-radius:4px;padding:3px 6px;font-size:0.59rem;">'  
       f'<div style="color:{text_muted};">+{d_extra_same} {d_same_lbl}</div>'  
       f'<div style="font-weight:700;color:#F0B429;">{_sc_fmt}</div></div>'  
       f'<div style="background:rgba(220,53,69,0.08);border:1px solid rgba(220,53,69,0.2);border-radius:4px;padding:3px 6px;font-size:0.59rem;">'  
       f'<div style="color:{text_muted};">+{d_extra_alt} {d_alt_lbl}</div>'  
       f'<div style="font-weight:700;color:#F0B429;">{_ac_fmt}</div></div>'  
       f'</div></div>')  
    if d_has_deficit else  
    (f'<div style="border-top:1px solid rgba(34,197,94,0.2);margin-top:4px;padding-top:4px;display:flex;align-items:center;gap:5px;">'  
     f'<span style="font-size:0.60rem;color:#2ecc71;font-weight:700;">✓ WITHIN CAPACITY</span>'  
     f'<span style="font-size:0.60rem;color:{scene_color};font-weight:600;">· {d_on_scene:.1f} min on-scene</span>'  
     f'</div>') }
  <!-- Inline lock status indicators -->
  <div style="display:grid; grid-template-columns:1fr 1fr; gap:3px; margin-top:auto; padding-top:4px; flex-shrink:0;">
    <div style="{"background:rgba(255,215,0,0.15);border:1px solid rgba(255,215,0,0.5);" if (d.get("pinned") and d_type=="GUARDIAN") else "background:rgba(255,255,255,0.03);border:1px dashed rgba(255,215,0,0.18);"} border-radius:4px; padding:3px 6px; font-size:0.57rem; color:{"#FFD700" if (d.get("pinned") and d_type=="GUARDIAN") else "rgba(255,215,0,0.35)"}; text-align:center; line-height:1.5; white-space:nowrap;">{"🔒 Guardian" if (d.get("pinned") and d_type=="GUARDIAN") else "🦅 lock guard"}</div>
    <div style="{"background:rgba(0,210,255,0.15);border:1px solid rgba(0,210,255,0.5);" if (d.get("pinned") and d_type=="RESPONDER") else "background:rgba(255,255,255,0.03);border:1px dashed rgba(0,210,255,0.18);"} border-radius:4px; padding:3px 6px; font-size:0.57rem; color:{"#00D2FF" if (d.get("pinned") and d_type=="RESPONDER") else "rgba(0,210,255,0.35)"}; text-align:center; line-height:1.5; white-space:nowrap;">{"🔒 Responder" if (d.get("pinned") and d_type=="RESPONDER") else "🚁 lock resp"}</div>
  </div>
</div>''')

    grid = (
        '<div style="display:grid; grid-template-columns:repeat(' + str(columns_per_row) + ', minmax(0, 1fr));'
        ' gap:10px; align-items:start; margin-bottom:12px; width:100%; box-sizing:border-box;">'
        + "".join(cards_html)
        + '</div>'
    )
    # Wrap in a style-scoped div to prevent Streamlit container from collapsing width.
    # overflow:visible is required so the 1.5x hover scale isn't clipped by the grid container.
    grid_id = f"ucg_{abs(hash(str([d.get('name') for d in active_drones]))) % 100000}"
    return (
        '<style>'
        '.unit-card-grid { display:grid; gap:10px; width:100%; box-sizing:border-box; overflow:visible; }'
        '.unit-card-grid > .unit-card { min-width:0; box-sizing:border-box; overflow:visible; }'
        '.tip { display:inline-flex; align-items:center; justify-content:center; width:11px; height:11px; border-radius:50%; background:rgba(255,255,255,0.12); color:#888; font-size:8px; font-weight:700; cursor:default; margin-left:3px; vertical-align:middle; position:relative; flex-shrink:0; }'
        '.tip:hover::after { content:attr(data-tip); position:absolute; bottom:130%; left:50%; transform:translateX(-50%); background:#1a1a2e; color:#e0e0e0; font-size:10px; font-weight:400; padding:5px 8px; border-radius:5px; white-space:normal; width:200px; line-height:1.4; z-index:9999; border:1px solid #333; box-shadow:0 4px 12px rgba(0,0,0,0.5); pointer-events:none; text-transform:none; letter-spacing:normal; }'
        '</style>'
        f'<div id="{grid_id}" class="unit-card-grid" style="grid-template-columns:repeat({columns_per_row}, minmax(0,1fr)); overflow:visible;">'
        + "".join(cards_html)
        + '</div>'
        f'<script>'
        f'(function(){{'
        f'  function eq(){{'
        f'    var g=document.getElementById("{grid_id}");'
        f'    if(!g)return;'
        f'    var cards=g.querySelectorAll(".unit-card");'
        f'    if(!cards.length)return;'
        f'    cards.forEach(function(c){{c.style.height="auto";}});'
        f'    var maxH=0;'
        f'    cards.forEach(function(c){{maxH=Math.max(maxH,c.getBoundingClientRect().height);}});'
        f'    cards.forEach(function(c){{c.style.height=maxH+"px";}});'
        f'  }}'
        f'  if(document.readyState==="complete"){{eq();}}else{{window.addEventListener("load",eq);}}'
        f'  setTimeout(eq,150);'
        f'  setTimeout(eq,600);'
        f'}})();'
        f'</script>'
    )

def to_kml_color(hex_str):
    h = hex_str.lstrip('#')
    return f"ff{h[4:6]}{h[2:4]}{h[0:2]}" if len(h) == 6 else "ff0000ff"

def calculate_zoom(min_lon, max_lon, min_lat, max_lat):
    lon_diff = max_lon - min_lon
    lat_diff = max_lat - min_lat
    if lon_diff <= 0 or lat_diff <= 0: return 12
    return min(max(min(np.log2(360/lon_diff), np.log2(180/lat_diff)) + 1.6, 5), 18)

def generate_kml(active_gdf, active_drones, calls_gdf):
    kml = simplekml.Kml()
    kml.document.name = "BRINC DFR Deployment Plan"
    kml.document.description = (
        "SIMULATOR DISCLAIMER: This file was generated by the BRINC Drones Coverage Optimization Simulator. "
        "All coverage zones, station locations, and incident data are model estimates based on user-provided inputs. "
        "Real-world results will vary. This file does not constitute a legal recommendation, binding proposal, "
        "contract, or guarantee of any product, service, or financial outcome. "
        "All deployments require FAA authorization, local ordinances review, and formal procurement."
    )
    fol_bounds = kml.newfolder(name="Jurisdictions")
    for _, row in active_gdf.iterrows():
        geoms = [row.geometry] if isinstance(row.geometry, Polygon) else row.geometry.geoms
        for geom in geoms:
            pol = fol_bounds.newpolygon(name=row.get('DISPLAY_NAME', 'Boundary'))
            pol.outerboundaryis = list(geom.exterior.coords)
            pol.style.linestyle.color = simplekml.Color.red
            pol.style.linestyle.width = 3
            pol.style.polystyle.color = simplekml.Color.changealphaint(30, simplekml.Color.red)
    fol_stations = kml.newfolder(name="Station Points")
    fol_rings = kml.newfolder(name="Coverage Rings")
    for d in active_drones:
        kml_c = to_kml_color(d['color'])
        pnt = fol_stations.newpoint(name=f"[{d['type'][:3]}] {d['name']}")
        pnt.coords = [(d['lon'], d['lat'])]
        pnt.style.iconstyle.icon.href = 'http://maps.google.com/mapfiles/kml/paddle/blu-blank.png'
        lats, lons = get_circle_coords(d['lat'], d['lon'], r_mi=d['radius_m']/1609.34)
        ring_coords = list(zip(lons, lats))
        ring_coords.append(ring_coords[0])
        pol = fol_rings.newpolygon(name=f"Range: {d['name']}")
        pol.outerboundaryis = ring_coords
        pol.style.linestyle.color = kml_c
        pol.style.linestyle.width = 2
        pol.style.polystyle.color = simplekml.Color.changealphaint(60, kml_c)
    fol_calls = kml.newfolder(name="Incident Data (Sample)")
    calls_export = calls_gdf.to_crs(epsg=4326)
    if len(calls_export) > 2000:
        calls_export = calls_export.sample(2000, random_state=42)
    for _, row in calls_export.iterrows():
        pnt = fol_calls.newpoint()
        pnt.coords = [(row.geometry.x, row.geometry.y)]
        pnt.style.iconstyle.scale = 0.5
        pnt.style.iconstyle.icon.href = 'http://maps.google.com/mapfiles/kml/shapes/placemark_circle.png'
    return kml.kml()

@st.cache_data
def find_relevant_jurisdictions(calls_df, stations_df, shapefile_dir, preferred_shp=None):
    points_list = []
    if calls_df is not None: points_list.append(calls_df[['lat', 'lon']])
    if stations_df is not None: points_list.append(stations_df[['lat', 'lon']])
    if not points_list: return None
    full_points = pd.concat(points_list)
    full_points = full_points[(full_points.lat.abs() > 1) & (full_points.lon.abs() > 1)]
    scan_points = full_points.sample(50000, random_state=42) if len(full_points) > 50000 else full_points
    points_gdf = gpd.GeoDataFrame(scan_points, geometry=gpd.points_from_xy(scan_points.lon, scan_points.lat), crs="EPSG:4326")
    total_bounds = points_gdf.total_bounds

    # Always scan all saved shapefiles in the directory so multi-jurisdiction
    # uploads show every boundary, not just the first one saved.
    shp_files = glob.glob(os.path.join(shapefile_dir, "*.shp"))
    # If no shapefiles exist at all and a preferred path was given, use just that
    if not shp_files and preferred_shp and os.path.exists(preferred_shp):
        shp_files = [preferred_shp]

    relevant_polys = []
    _calls_minx, _calls_miny, _calls_maxx, _calls_maxy = total_bounds
    for shp_path in shp_files:
        try:
            import fiona
            with fiona.open(shp_path) as _shp_src:
                _shp_bounds = _shp_src.bounds
            _no_overlap = (
                _shp_bounds[2] < _calls_minx or _shp_bounds[0] > _calls_maxx or
                _shp_bounds[3] < _calls_miny or _shp_bounds[1] > _calls_maxy
            )
            if _no_overlap:
                continue
        except Exception:
            pass

        try:
            gdf_chunk = gpd.read_file(shp_path, bbox=tuple(total_bounds))
            if not gdf_chunk.empty:
                if gdf_chunk.crs is None: gdf_chunk.set_crs(epsg=4269, inplace=True)
                gdf_chunk = gdf_chunk.to_crs(epsg=4326)
                hits = gpd.sjoin(gdf_chunk, points_gdf, how="inner", predicate="intersects")
                if not hits.empty:
                    subset = gdf_chunk.loc[hits.index.unique()].copy()
                    subset['data_count'] = hits.index.value_counts()
                    name_col = next((c for c in ['NAME','DISTRICT','NAMELSAD'] if c in subset.columns), subset.columns[0])
                    subset['DISPLAY_NAME'] = subset[name_col].astype(str)
                    relevant_polys.append(subset)
        except Exception: continue
    if not relevant_polys: return None
    master_gdf = pd.concat(relevant_polys, ignore_index=True).sort_values(by='data_count', ascending=False)
    master_gdf = master_gdf.dissolve(by='DISPLAY_NAME', aggfunc={'data_count': 'sum'}).reset_index()
    master_gdf = master_gdf.sort_values(by='data_count', ascending=False)
    if master_gdf['data_count'].sum() > 0:
        master_gdf['pct_share'] = master_gdf['data_count'] / master_gdf['data_count'].sum()
        master_gdf['cum_share'] = master_gdf['pct_share'].cumsum()
        mask = (master_gdf['cum_share'] <= 0.98) | (master_gdf['pct_share'] > 0.01)
        mask.iloc[0] = True
        return master_gdf[mask]
    return master_gdf

@st.cache_data(show_spinner=False)
def build_display_calls(df_calls_full, _city_m, epsg_code, max_points=300000, seed=42, bounds_hash=''):
    if df_calls_full is None or len(df_calls_full) == 0:
        return gpd.GeoDataFrame(geometry=[], crs="EPSG:4326")

    df = df_calls_full.copy()
    if 'lat' not in df.columns or 'lon' not in df.columns:
        return gpd.GeoDataFrame(geometry=[], crs="EPSG:4326")

    df['lat'] = pd.to_numeric(df['lat'], errors='coerce')
    df['lon'] = pd.to_numeric(df['lon'], errors='coerce')
    df = df.dropna(subset=['lat', 'lon']).reset_index(drop=True)
    if df.empty:
        return gpd.GeoDataFrame(geometry=[], crs="EPSG:4326")

    gdf = gpd.GeoDataFrame(df, geometry=gpd.points_from_xy(df.lon, df.lat), crs="EPSG:4326")
    try:
        gdf_m = gdf.to_crs(epsg=int(epsg_code))
        # Buffer 300 m so calls at polygon edges aren't clipped by precision gaps
        # (especially common when switching to a county boundary)
        _clip_geom = _city_m.buffer(300) if _city_m is not None else None
        calls_in_city = gdf_m[gdf_m.within(_clip_geom)] if _clip_geom is not None else gdf_m
    except Exception:
        calls_in_city = gdf

    if calls_in_city.empty:
        return gpd.GeoDataFrame(geometry=[], crs="EPSG:4326")

    if len(calls_in_city) <= max_points:
        return calls_in_city.to_crs(epsg=4326)

    sampled = calls_in_city.copy()
    minx, miny, maxx, maxy = sampled.total_bounds
    span_x = max(maxx - minx, 1.0)
    span_y = max(maxy - miny, 1.0)
    target_cells = max(25, int(np.sqrt(max_points) * 0.7))
    nx = max(25, min(120, target_cells))
    ny = max(25, min(120, int(target_cells * (span_y / span_x))))

    sampled['_gx'] = np.floor((sampled.geometry.x - minx) / span_x * nx).clip(0, nx - 1).astype(int)
    sampled['_gy'] = np.floor((sampled.geometry.y - miny) / span_y * ny).clip(0, ny - 1).astype(int)
    sampled['_cell'] = sampled['_gx'].astype(str) + '_' + sampled['_gy'].astype(str)

    counts = sampled['_cell'].value_counts()
    alloc = np.maximum(1, np.floor(counts / counts.sum() * max_points).astype(int))
    shortfall = int(max_points - alloc.sum())
    if shortfall > 0:
        remainders = (counts / counts.sum() * max_points) - np.floor(counts / counts.sum() * max_points)
        for cell in remainders.sort_values(ascending=False).index[:shortfall]:
            alloc.loc[cell] += 1

    parts = []
    for cell, group in sampled.groupby('_cell', sort=False):
        take = int(min(len(group), alloc.get(cell, 1)))
        if take >= len(group):
            parts.append(group)
        elif take > 0:
            parts.append(group.sample(take, random_state=seed))

    if not parts:
        display_calls = sampled.sample(max_points, random_state=seed)
    else:
        display_calls = pd.concat(parts, ignore_index=False)
        if len(display_calls) > max_points:
            display_calls = display_calls.sample(max_points, random_state=seed)

    display_calls = display_calls.drop(columns=['_gx', '_gy', '_cell'], errors='ignore')
    return display_calls.to_crs(epsg=4326)

@st.cache_resource
def precompute_spatial_data(df_calls, df_calls_full, df_stations_all, _city_m, epsg_code, resp_radius_mi, guard_radius_mi, center_lat, center_lon, bounds_hash):
    gdf_calls = gpd.GeoDataFrame(df_calls, geometry=gpd.points_from_xy(df_calls.lon, df_calls.lat), crs="EPSG:4326")
    gdf_calls_utm = gdf_calls.to_crs(epsg=int(epsg_code))
    try:
        calls_in_city = gdf_calls_utm[gdf_calls_utm.within(_city_m)]
    except Exception:
        calls_in_city = gdf_calls_utm
    radius_resp_m = resp_radius_mi * 1609.34
    radius_guard_m = guard_radius_mi * 1609.34
    station_metadata = []
    total_calls = len(calls_in_city)
    n = len(df_stations_all)
    resp_matrix = np.zeros((n, total_calls), dtype=bool)
    guard_matrix = np.zeros((n, total_calls), dtype=bool)
    dist_matrix_r = np.zeros((n, total_calls))
    dist_matrix_g = np.zeros((n, total_calls))
    display_calls = build_display_calls(df_calls_full if df_calls_full is not None else df_calls, _city_m, epsg_code, max_points=300000, bounds_hash=bounds_hash)
    max_dist = max(
        (((row['lon'] - center_lon) ** 2 + (row['lat'] - center_lat) ** 2) ** 0.5 for _, row in df_stations_all.iterrows()),
        default=1.0
    ) or 1.0
    calls_array = None
    if total_calls > 0:
        calls_array = np.array(list(zip(calls_in_city.geometry.x, calls_in_city.geometry.y)))

    for idx_pos, (_, row) in enumerate(df_stations_all.iterrows()):
        s_pt_m = gpd.GeoSeries([Point(row['lon'], row['lat'])], crs="EPSG:4326").to_crs(epsg=int(epsg_code)).iloc[0]

        if calls_array is not None:
            dists = np.sqrt((calls_array[:, 0] - s_pt_m.x) ** 2 + (calls_array[:, 1] - s_pt_m.y) ** 2)
            dists_mi = dists / 1609.34
            mask_r = dists <= radius_resp_m
            mask_g = dists <= radius_guard_m
            resp_matrix[idx_pos, :] = mask_r
            guard_matrix[idx_pos, :] = mask_g
            dist_matrix_r[idx_pos, :] = dists_mi
            dist_matrix_g[idx_pos, :] = dists_mi
            avg_dist_r = dists_mi[mask_r].mean() if mask_r.any() else resp_radius_mi * (2 / 3)
            avg_dist_g = dists_mi[mask_g].mean() if mask_g.any() else guard_radius_mi * (2 / 3)
        else:
            avg_dist_r = resp_radius_mi * (2 / 3)
            avg_dist_g = guard_radius_mi * (2 / 3)

        full_buf_2m = s_pt_m.buffer(radius_resp_m)
        try:
            clipped_2m = full_buf_2m.intersection(_city_m)
        except Exception:
            clipped_2m = full_buf_2m
        full_buf_guard = s_pt_m.buffer(radius_guard_m)
        try:
            clipped_guard = full_buf_guard.intersection(_city_m)
        except Exception:
            clipped_guard = full_buf_guard
        dist_c = ((row['lon'] - center_lon) ** 2 + (row['lat'] - center_lat) ** 2) ** 0.5
        station_metadata.append({
            'name': row['name'], 'lat': row['lat'], 'lon': row['lon'],
            'clipped_2m': clipped_2m, 'clipped_guard': clipped_guard,
            'avg_dist_r': avg_dist_r,
            'avg_dist_g': avg_dist_g,
            'centrality': 1.0 - (dist_c / max_dist)
        })

    return calls_in_city, display_calls, resp_matrix, guard_matrix, dist_matrix_r, dist_matrix_g, station_metadata, total_calls

def solve_mclp(resp_matrix, guard_matrix, dist_r, dist_g, num_resp, num_guard, allow_redundancy, incremental=True, forced_r=None, forced_g=None):
    """MCLP optimizer. forced_r / forced_g are lists of station indices that must
    be included as Responders / Guardians regardless of coverage score."""
    forced_r = list(forced_r or [])
    forced_g = list(forced_g or [])
    n_stations, n_calls = resp_matrix.shape
    if n_calls == 0 or (num_resp == 0 and num_guard == 0): return [], [], [], []
    df_profiles = pd.DataFrame(resp_matrix.T).astype(int).astype(str)
    df_profiles['g'] = pd.DataFrame(guard_matrix.T).astype(int).astype(str).agg(''.join, axis=1)
    df_profiles['r'] = df_profiles.drop(columns='g').agg(''.join, axis=1)
    grouped = df_profiles.groupby(['r', 'g'], sort=False)
    weights = grouped.size().values
    unique_idx = grouped.head(1).index
    u_resp = resp_matrix[:, unique_idx]
    u_guard = guard_matrix[:, unique_idx]
    u_dist_r = dist_r[:, unique_idx]
    u_dist_g = dist_g[:, unique_idx]
    n_u = len(weights)

    def run_lp(target_r, target_g, locked_r, locked_g):
        model = pulp.LpProblem("DroneCoverage", pulp.LpMaximize)
        x_r = pulp.LpVariable.dicts("r_st", range(n_stations), 0, 1, pulp.LpBinary)
        x_g = pulp.LpVariable.dicts("g_st", range(n_stations), 0, 1, pulp.LpBinary)
        model += pulp.lpSum(x_r[i] for i in range(n_stations)) == target_r
        model += pulp.lpSum(x_g[i] for i in range(n_stations)) == target_g
        for r in locked_r: model += x_r[r] == 1
        for g in locked_g: model += x_g[g] == 1
        if not allow_redundancy:
            for s in range(n_stations): model += x_r[s] + x_g[s] <= 1
        y = pulp.LpVariable.dicts("cl", range(n_u), 0, 1, pulp.LpBinary)
        penalty = 0.00001
        model += pulp.lpSum(y[i]*weights[i] for i in range(n_u)) - pulp.lpSum(
            x_r[s]*np.sum(u_dist_r[s,:])*penalty + x_g[s]*np.sum(u_dist_g[s,:])*penalty
            for s in range(n_stations))
        for i in range(n_u):
            cover = [x_r[s] for s in range(n_stations) if u_resp[s,i]] + [x_g[s] for s in range(n_stations) if u_guard[s,i]]
            if cover: model += y[i] <= pulp.lpSum(cover)
            else: model += y[i] == 0
        model.solve(pulp.PULP_CBC_CMD(msg=0, timeLimit=10, gapRel=0.0))
        return (
            [i for i in range(n_stations) if (pulp.value(x_r[i]) or 0) > 0.5],
            [i for i in range(n_stations) if (pulp.value(x_g[i]) or 0) > 0.5]
        )

    if not incremental:
        res_r, res_g = run_lp(num_resp, num_guard, forced_r, forced_g)
        return res_r, res_g, res_r, res_g
    # Start with forced pins already locked in
    curr_r, curr_g = list(forced_r), list(forced_g)
    chrono_r, chrono_g = list(forced_r), list(forced_g)
    # Add remaining Guardians one at a time (incremental)
    for tg in range(len(forced_g) + 1, num_guard + 1):
        next_r, next_g = run_lp(0, tg, curr_r, curr_g)
        chrono_g.extend([x for x in next_g if x not in curr_g])
        curr_r, curr_g = next_r, next_g
    # Add remaining Responders one at a time
    for tr in range(len(forced_r) + 1, num_resp + 1):
        next_r, next_g = run_lp(tr, num_guard, curr_r, curr_g)
        chrono_r.extend([x for x in next_r if x not in curr_r])
        curr_r, curr_g = next_r, next_g
    return curr_r, curr_g, chrono_r, chrono_g

@st.cache_resource
def compute_all_elbow_curves(n_calls, _resp_matrix, _guard_matrix, _geos_r, _geos_g, total_area, _bounds_hash, max_stations=100):
    n_st_calls = min(_resp_matrix.shape[0], max_stations)
    n_st_area_r = min(len(_geos_r), 25)
    n_st_area_g = min(len(_geos_g), 25)

    def greedy_calls(matrix):
        uncovered = np.ones(n_calls, dtype=bool)
        curve = [0.0]
        cov_count = 0
        import heapq as hq
        pq = [(-matrix[i].sum(), i) for i in range(n_st_calls)]
        hq.heapify(pq)
        for _ in range(n_st_calls):
            if not pq: break
            best_s, best_cov = -1, -1
            while pq:
                neg_gain, idx = hq.heappop(pq)
                actual_gain = (matrix[idx] & uncovered).sum()
                if not pq or actual_gain >= -pq[0][0]:
                    best_s, best_cov = idx, actual_gain
                    break
                else:
                    hq.heappush(pq, (-actual_gain, idx))
            if best_s != -1 and best_cov > 0:
                uncovered = uncovered & ~matrix[best_s]
                cov_count += best_cov
                curve.append((cov_count / max(1, n_calls)) * 100)
                if cov_count == n_calls: break
            else:
                break
        return curve

    def greedy_area(geos, limit):
        if total_area <= 0 or limit <= 0: return [0.0]
        current_union = Polygon()
        curve = [0.0]
        import heapq as hq
        geos_sub = geos[:limit]
        
        pq = [(-geos_sub[i].area, i) for i in range(len(geos_sub))]
        hq.heapify(pq)
        
        for _ in range(len(geos_sub)):
            if not pq: break
            best_s, best_gain = -1, -1
            
            while pq:
                neg_gain, idx = hq.heappop(pq)
                try:
                    actual_gain = current_union.union(geos_sub[idx]).area - current_union.area
                except Exception:
                    actual_gain = 0
                    
                if not pq or actual_gain >= -pq[0][0]:
                    best_s, best_gain = idx, actual_gain
                    break
                else:
                    hq.heappush(pq, (-actual_gain, idx))
                    
            if best_s != -1 and best_gain > 0:
                try:
                    current_union = current_union.union(geos_sub[best_s])
                    curve.append((current_union.area / total_area) * 100)
                except Exception:
                    pass
            else:
                break
        return curve

    with ThreadPoolExecutor() as executor:
        f_cr = executor.submit(greedy_calls, _resp_matrix[:n_st_calls])
        f_cg = executor.submit(greedy_calls, _guard_matrix[:n_st_calls])
        f_ar = executor.submit(greedy_area, _geos_r, n_st_area_r)
        f_ag = executor.submit(greedy_area, _geos_g, n_st_area_g)
        c_r, c_g, a_r, a_g = f_cr.result(), f_cg.result(), f_ar.result(), f_ag.result()

    max_len = max(len(c_r), len(c_g), len(a_r), len(a_g))
    def pad(c):
        r = list(c)
        while len(r) < max_len: r.append(np.nan)
        return r
    return pd.DataFrame({
        'Drones': range(max_len),
        'Responder (Calls)': pad(c_r),
        'Responder (Area)':  pad(a_r),
        'Guardian (Calls)':  pad(c_g),
        'Guardian (Area)':   pad(a_g)
    })

# ============================================================
# APP FLOW 
# ============================================================

if not st.session_state['csvs_ready']:

    # GRAB THE LOGO FOR THE UPLOAD PAGE
    logo_b64 = get_themed_logo_base64("logo.png", theme="dark")
    hero_logo_html = f'<img src="data:image/png;base64,{logo_b64}" style="height:72px; margin-bottom:15px;">' if logo_b64 else f'<div style="font-size:2.5rem; font-weight:900; letter-spacing:4px; color:#ffffff; margin-bottom:15px;">BRINC</div>'
    # Upper-right hero image uses gigs.png (the drone product shot)
    gigs_b64 = get_transparent_product_base64("gigs.png")

    st.markdown(f"""
    <style>
    @keyframes pulseGlow {{
        0%, 100% {{ opacity: 0.55; }}
        50%       {{ opacity: 1.0; }}
    }}
    @keyframes fadeUp {{
        from {{ opacity:0; transform:translateY(14px); }}
        to   {{ opacity:1; transform:translateY(0); }}
    }}
    .brinc-hero {{
        position: relative;
        text-align: center;
        padding: 52px 24px 40px;
        margin-bottom: 36px;
        border-radius: 12px;
        background: radial-gradient(ellipse at 50% 0%,
            rgba(0,210,255,0.13) 0%, rgba(0,0,0,0) 68%);
        border-bottom: 1px solid rgba(0,210,255,0.15);
        overflow: hidden;
        animation: fadeUp 0.5s ease both;
    }}
    .brinc-hero::before {{
        content: '';
        position: absolute; inset: 0;
        background:
            repeating-linear-gradient(0deg,
                transparent, transparent 39px,
                rgba(0,210,255,0.025) 39px,
                rgba(0,210,255,0.025) 40px),
            repeating-linear-gradient(90deg,
                transparent, transparent 79px,
                rgba(0,210,255,0.025) 79px,
                rgba(0,210,255,0.025) 80px);
        pointer-events: none;
    }}
    .brinc-eyebrow {{
        font-family: 'IBM Plex Mono', monospace;
        font-size: 0.62rem;
        font-weight: 700;
        letter-spacing: 4px;
        color: {accent_color};
        text-transform: uppercase;
        opacity: 0.7;
        margin-bottom: 12px;
    }}
    .brinc-h1 {{
        font-family: 'Manrope', sans-serif;
        font-size: clamp(2rem, 4vw, 3rem);
        font-weight: 900;
        color: #ffffff;
        letter-spacing: -0.5px;
        line-height: 1.08;
        margin-bottom: 12px;
    }}
    .brinc-h1 em {{
        font-style: normal;
        color: {accent_color};
    }}
    .brinc-tagline {{
        font-size: 0.88rem;
        color: #666;
        max-width: 500px;
        margin: 0 auto 22px;
        line-height: 1.65;
    }}
    .brinc-badges {{
        display: flex;
        flex-wrap: wrap;
        justify-content: center;
        gap: 8px;
        margin-top: 4px;
    }}
    .brinc-badge {{
        display: inline-flex;
        align-items: center;
        gap: 6px;
        background: rgba(0,210,255,0.07);
        border: 1px solid rgba(0,210,255,0.2);
        border-radius: 100px;
        padding: 4px 13px;
        font-size: 0.64rem;
        font-weight: 700;
        color: {accent_color};
        letter-spacing: 0.8px;
        text-transform: uppercase;
    }}
    .brinc-badge.pulse {{
        animation: pulseGlow 3s ease-in-out infinite;
    }}
    .path-card {{
        background: #080808;
        border: 1px solid #1c1c1c;
        border-radius: 10px;
        padding: 22px 18px 16px;
        position: relative;
        overflow: hidden;
        transition: border-color 0.2s ease, box-shadow 0.2s ease;
    }}
    .path-card::after {{
        content: '';
        position: absolute;
        top: 0; left: 0; right: 0;
        height: 3px;
        background: var(--accent);
        border-radius: 10px 10px 0 0;
    }}
    .path-card:hover {{
        border-color: rgba(255,255,255,0.12);
        box-shadow: 0 0 28px rgba(0,210,255,0.05);
    }}
    .pc-icon  {{ font-size: 1.5rem; display:block; margin-bottom:9px; }}
    .pc-tag   {{ font-size:0.55rem; font-weight:800; letter-spacing:2.5px;
                 text-transform:uppercase; color:var(--accent); margin-bottom:5px; }}
    .pc-title {{ font-size:1rem; font-weight:800; color:#fff;
                 line-height:1.25; margin-bottom:7px; }}
    .pc-desc  {{ font-size:0.7rem; color:#555; line-height:1.6; margin-bottom:0; }}
    .field-footnote {{
        font-size: 0.63rem; color: #3a3a3a; line-height: 1.75;
        margin-top: 10px; border-top: 1px solid #141414;
        padding-top: 10px;
    }}
    .demo-cities {{
        font-size: 0.65rem; color: #444; line-height: 1.9;
        margin-top: 10px;
    }}
    .demo-cities b {{ color: #555; }}
    .demo-check {{
        font-size: 0.63rem; color: #333; line-height: 1.8;
        margin-top: 12px; border-top: 1px solid #141414;
        padding-top: 10px;
    }}
    .demo-check span {{ color: {accent_color}; margin-right: 5px; }}
    </style>

    <div class="brinc-hero" style="display:flex; align-items:center; justify-content:space-between; text-align:left; padding: 48px 48px 40px; gap: 32px; flex-wrap: wrap;">
        <div style="flex:1; min-width:280px;">
            {hero_logo_html}
            <div class="brinc-eyebrow" style="margin-top:6px;">BRINC Drones · DFR Platform</div>
            <div class="brinc-h1">
                Coverage. Operations.<br><em>Savings.</em>
            </div>
            <div class="brinc-tagline" style="margin-left:0; text-align:left;">
                Optimize drone-as-first-responder deployments for any US jurisdiction.
                Model coverage, forecast ROI, and generate grant-ready proposals in minutes.
            </div>
            <div class="brinc-badges" style="justify-content:flex-start;">
                <div class="brinc-badge pulse">🛰 3D Swarm Simulation</div>
                <div class="brinc-badge">🗺 Census Boundaries</div>
                <div class="brinc-badge">📄 Grant Narrative Export</div>
                <div class="brinc-badge">✈️ FAA LAANC Overlay</div>
                <div class="brinc-badge">⚡ MCLP Optimizer</div>
            </div>
        </div>
        <div style="flex:0 0 auto; display:flex; align-items:center; justify-content:center;">
            <img src="data:image/png;base64,{gigs_b64}" style="height:260px; max-width:420px; object-fit:contain; filter: drop-shadow(0 0 32px rgba(0,210,255,0.35)) drop-shadow(0 0 8px rgba(0,150,255,0.2));" alt="BRINC COS Drone Station">
        </div>
    </div>
    """, unsafe_allow_html=True)

    path_sim_col, path_upload_col, path_demo_col = st.columns(3, gap="medium")

    with path_sim_col:
        st.markdown(f"""
        <div class="path-card" style="--accent:{accent_color};">
            <span class="pc-icon">🗺</span>
            <div class="pc-tag">Path 01</div>
            <div class="pc-title">Simulate Any<br>US Region</div>
            <div class="pc-desc">No data needed. Real Census boundaries + realistic 911 call distribution generated automatically. Stack multiple jurisdictions in one run.</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

        # ── CITY / STATE — simplified single-row inputs ─────────────────────
        _state_keys = list(STATE_FIPS.keys())

        # Column headers
        _h_city, _h_state = st.columns([3, 1])
        _h_city.markdown("<div style='font-size:12px;color:#888;padding-bottom:2px'>City or County</div>", unsafe_allow_html=True)
        _h_state.markdown("<div style='font-size:12px;color:#888;padding-bottom:2px'>State</div>", unsafe_allow_html=True)

        while len(st.session_state['target_cities']) < st.session_state.city_count:
            st.session_state['target_cities'].append({"city": "", "state": st.session_state.get('active_state', 'IL')})

        for i in range(st.session_state.city_count):
            c_val = st.session_state['target_cities'][i]['city'] if i < len(st.session_state['target_cities']) else ""
            s_val = st.session_state['target_cities'][i]['state'] if i < len(st.session_state['target_cities']) else "IL"

            col_city, col_state = st.columns([3, 1])

            c_name = col_city.text_input(
                f"city_or_county_{i}", value=c_val,
                placeholder="e.g. Rockford or Winnebago County",
                label_visibility="collapsed",
                key=f"c_{i}",
                help="Official municipality or county name."
            )

            default_state_idx = _state_keys.index(s_val) if s_val in _state_keys else _state_keys.index("IL")
            s_name = col_state.selectbox(
                f"state_{i}",
                options=_state_keys,
                index=default_state_idx,
                label_visibility="collapsed",
                key=f"s_{i}",
                help="Two-letter state abbreviation."
            )

            if i < len(st.session_state['target_cities']):
                st.session_state['target_cities'][i] = {"city": c_name, "state": s_name}
            else:
                st.session_state['target_cities'].append({"city": c_name, "state": s_name})

        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
        st.file_uploader(
            "Optional: Custom stations (CSV or Excel)",
            type=['csv', 'xlsx', 'xls', 'xlsm', 'xlsb'],
            key="sim_station_uploader",
            help="Include 'lat'/'lon' OR 'address' columns. Max ~20 locations for auto-geocoding."
        )

        station_template_bytes = base64.b64decode(
            "TkFNRSxUWVBFLEFERFJFU1MsQ0FQQUNJVFksTk9URVMsTEFULExPTgpTYW1wbGUgMSBQb2xpY2UgU3RhdGlvbixQb2xpY2UsIjQyMCBXIFN0YXRlIFN0LCBSb2NrZm9yZCwgSUwgNjExMDEiLDIsUHJpbWFyeSBkb3dudG93biBkaXNwYXRjaCBodWIsNDIuMjcxMSwtODkuMDk0MApTYW1wbGUgMiBQb2xpY2UgU3RhdGlvbixQb2xpY2UsIjM0MDEgTiBNYWluIFN0LCBSb2NrZm9yZCwgSUwgNjExMDMiLDIsTm9ydGggc2lkZSBwYXRyb2wgYmFzZSw0Mi4zMTA1LC04OS4wODg3ClNhbXBsZSAzIFBvbGljZSBTdGF0aW9uLFBvbGljZSwiMTcwNyBTIE11bGZvcmQgUmQsIFJvY2tmb3JkLCBJTCA2MTEwOCIsMSxTb3V0aGVhc3QgY29ycmlkb3IgY292ZXJhZ2UsNDIuMjQ4OCwtODguOTk5OApTYW1wbGUgNCBQb2xpY2UgU3RhdGlvbixQb2xpY2UsIjQzNDAgVyBTdGF0ZSBTdCwgUm9ja2ZvcmQsIElMIDYxMTAyIiwxLFdlc3Qgc2lkZSByYXBpZCByZXNwb25zZSB1bml0LDQyLjI3MTIsLTg5LjEyNDEKU2FtcGxlIDEgRmlyZSBTdGF0aW9uLEZpcmUsIjcwOCBDbGludG9uIFN0LCBSb2NrZm9yZCwgSUwgNjExMDEiLDIsQ2VudHJhbCBmaXJlIGRpc3BhdGNoIC0gU3RhdGlvbiAxLDQyLjI3MjAsLTg5LjA4OTgKU2FtcGxlIDIgRmlyZSBTdGF0aW9uLEZpcmUsIjE0MDIgTiBDb3VydCBTdCwgUm9ja2ZvcmQsIElMIDYxMTAzIiwxLE5vcnRoIFJvY2tmb3JkIGZpcmUgY292ZXJhZ2UsNDIuMjk1MSwtODkuMDgyNgpTYW1wbGUgMyBGaXJlIFN0YXRpb24sRmlyZSwiMjI1MCBTIEFscGluZSBSZCwgUm9ja2ZvcmQsIElMIDYxMTA4IiwxLFNvdXRoIEFscGluZSBmaXJlIHJlc3BvbnNlLDQyLjI0MDEsLTg4Ljk5NjQKU2FtcGxlIDQgRmlyZSBTdGF0aW9uLEZpcmUsIjUyODUgU2FmZm9yZCBSZCwgUm9ja2ZvcmQsIElMIDYxMTAxIiwxLFdlc3QgZGlzdHJpY3QgZmlyZSBzdGF0aW9uLDQyLjI2OTgsLTg5LjE0MDIKU2FtcGxlIDEgRU1TIFN0YXRpb24sRU1TLCIxNDAxIEUgU3RhdGUgU3QsIFJvY2tmb3JkLCBJTCA2MTEwNCIsMixFYXN0IHNpZGUgRU1TIHJhcGlkIHJlc3BvbnNlLDQyLjI2OTQsLTg5LjA2MjEKU2FtcGxlIDIgRU1TIFN0YXRpb24sRU1TLCIzNzIwIENoYXJsZXMgU3QsIFJvY2tmb3JkLCBJTCA2MTEwOCIsMSxTb3V0aGVhc3QgRU1TIGNvdmVyYWdlIHpvbmUsNDIuMjUyMiwtODkuMDA1OApTYW1wbGUgMyBFTVMgU3RhdGlvbixFTVMsIjQ4MjUgTiBCZWxsIFNjaG9vbCBSZCwgUm9ja2ZvcmQsIElMIDYxMTA3IiwxLE5vcnRoZWFzdCBFTVMgcmVzcG9uc2UgaHViLDQyLjMwMjEsLTg4Ljk4OTEKU2FtcGxlIDEgR292IFN0YXRpb24sR292ZXJubWVudCwiNDI1IEUgU3RhdGUgU3QsIFJvY2tmb3JkLCBJTCA2MTEwNCIsMSxXaW5uZWJhZ28gQ291bnR5IGFkbWluIGJ1aWxkaW5nLDQyLjI3MTUsLTg5LjA4NDgKU2FtcGxlIDIgR292IFN0YXRpb24sR292ZXJubWVudCwiMzAwIFcgU3RhdGUgU3QsIFJvY2tmb3JkLCBJTCA2MTEwMSIsMSxDaXR5IEhhbGwgLSBSb2NrZm9yZCBtdW5pY2lwYWwgY2VudGVyLDQyLjI3MTEsLTg5LjA5NTcKU2FtcGxlIDMgR292IFN0YXRpb24sR292ZXJubWVudCwiNjUwIFcgU3RhdGUgU3QsIFJvY2tmb3JkLCBJTCA2MTEwMiIsMSxQdWJsaWMgd29ya3MgYW5kIGVtZXJnZW5jeSBtZ210LDQyLjI3MTMsLTg5LjEwMTgK"
        )
        st.download_button(
            label="⬇️ Download sample stations.csv",
            data=station_template_bytes,
            file_name="stations.csv",
            mime="text/csv; charset=utf-8",
            key="download_station_template_btn",
            use_container_width=True,
        )

        st.caption("Upload your own stations file if you have one, or download the sample template. If no file is uploaded, stations will be auto-generated from call data.")

        col_add, col_run = st.columns([1, 1])
        if st.session_state.city_count < 10:
            if col_add.button("＋ City", use_container_width=True, key="add_city_btn"):
                st.session_state.city_count += 1
                st.rerun()
        submit_demo = col_run.button("▶ Run", use_container_width=True, key="run_sim_btn",
                                     help="Fetch boundaries and launch the simulation.")

    with path_upload_col:
        st.markdown(f"""
        <div class="path-card" style="--accent:#39FF14;">
            <span class="pc-icon">📂</span>
            <div class="pc-tag">Path 02</div>
            <div class="pc-title">Upload CAD<br>or .brinc Save</div>
            <div class="pc-desc">
                Drop <b>any</b> CAD export CSV — no renaming needed.
                Or, drop a previously saved <b>.brinc</b> file to instantly restore your deployment.
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

        uploaded_files = st.file_uploader(
            "Drop your CAD export (+ optional stations CSV)",
            accept_multiple_files=True,
            type=['csv', 'xlsx', 'xls', 'xlsb', 'xlsm', 'brinc', 'json', 'txt'],
            label_visibility="collapsed",
            help="One file = raw CAD export. Two files = calls + stations. OR drop a .brinc file to restore a previous session."
        )

        st.markdown("""
        <div class="field-footnote">
            <b style='color:#555;'>1 file</b> — any CAD export (CSV or Excel); stations auto-built from OSM<br>
            <b style='color:#555;'>2+ files</b> — calls + stations (CSV or Excel, any column names)<br>
            <b style='color:#39FF14;'>.brinc file</b> — instantly restore a saved deployment<br>
            Max 25,000 calls · 100 stations
        </div>
        """, unsafe_allow_html=True)

        def _looks_like_stations(fname):
            n = fname.lower()
            return any(k in n for k in ['station','facility','loc'])

        if uploaded_files and len(uploaded_files) >= 1:
            
            # --- 1. INTELLIGENTLY CHECK FOR .BRINC FILE ---
            # Browsers sometimes append .json to .brinc files on download
            brinc_file = None
            for f in uploaded_files:
                fname = f.name.lower()
                if '.brinc' in fname or fname.endswith('.json'):
                    try:
                        # Quick peek inside to see if it has our save data keys
                        f.seek(0)
                        peek = json.loads(f.getvalue().decode('utf-8'))
                        if 'k_resp' in peek and 'calls_data' in peek:
                            brinc_file = f
                            break
                    except:
                        pass

            if brinc_file:
                with st.spinner("💾 Restoring saved deployment..."):
                    try:
                        brinc_file.seek(0)
                        save_data = json.loads(brinc_file.getvalue().decode('utf-8'))
                        
                        st.session_state['active_city'] = save_data.get('city', 'Unknown')
                        st.session_state['active_state'] = save_data.get('state', 'US')
                        st.session_state['k_resp'] = save_data.get('k_resp', 2)
                        st.session_state['k_guard'] = save_data.get('k_guard', 0)
                        st.session_state['r_resp'] = save_data.get('r_resp', 2.0)
                        st.session_state['r_guard'] = save_data.get('r_guard', 8.0)
                        st.session_state['dfr_rate'] = save_data.get('dfr_rate', 12)
                        st.session_state['deflect_rate'] = save_data.get('deflect_rate', 25)
                        
                        if save_data.get('calls_data'):
                            df_c = pd.DataFrame(save_data['calls_data'])
                            if 'lat' not in df_c.columns or 'lon' not in df_c.columns:
                                st.error("❌ .brinc file is missing required 'lat'/'lon' columns in calls data.")
                                st.stop()
                            df_c['lat'] = pd.to_numeric(df_c['lat'], errors='coerce')
                            df_c['lon'] = pd.to_numeric(df_c['lon'], errors='coerce')
                            df_c = df_c.dropna(subset=['lat', 'lon']).reset_index(drop=True)
                            if df_c.empty:
                                st.error("❌ .brinc file contains no valid coordinate data after parsing.")
                                st.stop()
                            st.session_state['df_calls'] = df_c
                            st.session_state['df_calls_full'] = df_c.copy()
                            st.session_state['total_original_calls'] = len(df_c)
                            st.session_state['total_modeled_calls'] = len(df_c)

                        if save_data.get('stations_data'):
                            df_s = pd.DataFrame(save_data['stations_data'])
                            if 'lat' not in df_s.columns or 'lon' not in df_s.columns:
                                st.warning("⚠️ .brinc stations data missing lat/lon — stations will be re-generated.")
                            else:
                                df_s['lat'] = pd.to_numeric(df_s['lat'], errors='coerce')
                                df_s['lon'] = pd.to_numeric(df_s['lon'], errors='coerce')
                                df_s = df_s.dropna(subset=['lat', 'lon']).reset_index(drop=True)
                                st.session_state['df_stations'] = df_s

                        if save_data.get('boundary_geojson'):
                            try:
                                import io as _io
                                _bgdf = gpd.read_file(_io.StringIO(save_data['boundary_geojson']))
                                if not _bgdf.empty:
                                    st.session_state['master_gdf_override'] = _bgdf
                            except Exception:
                                pass
                        st.session_state['boundary_kind'] = save_data.get('boundary_kind', 'place')
                        st.session_state['boundary_source_path'] = save_data.get('boundary_source_path', '')

                        st.session_state['data_source'] = 'brinc_file'
                        st.session_state['demo_mode_used'] = False
                        st.session_state['sim_mode_used'] = False
                        st.session_state['map_build_logged'] = False
                        st.session_state['csvs_ready'] = True
                        st.toast("✅ Deployment restored successfully!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error loading .brinc file: {e}")
                        st.stop()

            else:
                # --- 2. OTHERWISE, PROCESS AS NORMAL CSV CAD DATA ---
                st.session_state['active_city'] = ""
                st.session_state['active_state'] = ""
                st.session_state['target_cities'] = []

                f_list = list(uploaded_files)
                call_files = []
                station_file = None

                for f in f_list:
                    if _looks_like_stations(f.name):
                        station_file = f
                    else:
                        call_files.append(f)
                
                if len(f_list) == 2 and not station_file:
                    f0, f1 = f_list
                    f0.seek(0); sz0 = len(f0.read()); f0.seek(0)
                    f1.seek(0); sz1 = len(f1.read()); f1.seek(0)
                    if sz0 >= sz1:
                        call_files = [f0]
                        station_file = f1
                    else:
                        call_files = [f1]
                        station_file = f0

                if call_files:
                    with st.spinner("🔍 Detecting column types in CAD export…"):
                        df_c = aggressive_parse_calls(call_files)

                    if df_c is None or df_c.empty:
                        st.error("❌ Calls file error: Could not parse valid coordinates.")
                        st.stop()

                    df_c_full = df_c.reset_index(drop=True).copy()

                    if len(df_c_full) > 25000:
                        df_c = df_c_full.sample(25000, random_state=42).reset_index(drop=True)
                        st.toast(f"⚠️ Optimization modeled with {len(df_c):,} representative calls out of {len(df_c_full):,} total incidents.")
                    else:
                        df_c = df_c_full.copy()

                    st.session_state.update({
                        'total_original_calls': len(df_c_full),
                        'total_modeled_calls': len(df_c),
                    })

                    if station_file is not None:
                        with st.spinner("🔍 Reading stations file…"):
                            try:
                                sfname = station_file.name.lower()
                                if sfname.endswith(('.xlsx', '.xls', '.xlsm', '.xlsb')):
                                    engine = 'xlrd' if sfname.endswith('.xls') else 'pyxlsb' if sfname.endswith('.xlsb') else 'openpyxl'
                                    df_s = pd.read_excel(io.BytesIO(station_file.getvalue()), engine=engine)
                                else:
                                    df_s = pd.read_csv(station_file)
                                df_s.columns = [str(c).lower().strip() for c in df_s.columns]
                                if 'latitude' in df_s.columns: df_s = df_s.rename(columns={'latitude':'lat'})
                                if 'longitude' in df_s.columns: df_s = df_s.rename(columns={'longitude':'lon'})
                                if 'station_name' in df_s.columns: df_s = df_s.rename(columns={'station_name':'name'})
                                if 'station_type' in df_s.columns: df_s = df_s.rename(columns={'station_type':'type'})
                                
                                if 'lat' in df_s.columns and 'lon' in df_s.columns:
                                    df_s['lat'] = pd.to_numeric(df_s['lat'], errors='coerce')
                                    df_s['lon'] = pd.to_numeric(df_s['lon'], errors='coerce')
                                else:
                                    raise ValueError("Could not find lat/lon columns.")

                                if 'name' not in df_s.columns: 
                                    df_s['name'] = [f"Site {i+1}" for i in range(len(df_s))]
                                else:
                                    df_s['name'] = df_s['name'].fillna('').astype(str).str.strip()
                                    df_s['name'] = df_s['name'].replace(r'(?i)^(null|<null>|nan|none)$', '', regex=True)
                                    df_s['name'] = [n if n else f"Site {i+1}" for i, n in enumerate(df_s['name'])]

                                counts = {}
                                new_names = []
                                for n in df_s['name']:
                                    if n in counts:
                                        counts[n] += 1
                                        new_names.append(f"{n} ({counts[n]})")
                                    else:
                                        counts[n] = 0
                                        new_names.append(n)
                                df_s['name'] = new_names

                                if 'type' not in df_s.columns: df_s['type'] = 'Police'
                                df_s = df_s.dropna(subset=['lat', 'lon']).reset_index(drop=True)
                                osm_note = "Loaded stations from file."
                                st.session_state['stations_user_uploaded'] = True
                            except Exception as e:
                                df_s, osm_note = None, f"Failed: {e}"
                        if df_s is None or df_s.empty:
                            st.error(f"❌ Stations file error: {osm_note}")
                            st.stop()
                    else:
                        st.session_state['stations_user_uploaded'] = False
                        with st.spinner("🌐 No stations file detected — querying OpenStreetMap for police, fire & schools…"):
                            df_s, osm_note = generate_stations_from_calls(df_c)
                        if df_s is None or df_s.empty:
                            # Final safety net: scatter stations across call bounding box
                            df_s = _make_random_stations(df_c, n=40)
                            osm_note = "⚠️ Could not reach any map source — using estimated station positions from call data."
                            st.warning(osm_note)
                        else:
                            st.toast(f"✅ {osm_note}")

                    if len(df_s) > 100:
                        df_s = df_s.sample(100, random_state=42).reset_index(drop=True)

                    detected_city = None
                    detected_state = None
                    detection_source = 'unknown'

                    with st.spinner(get_jurisdiction_message()):
                        # Priority 1: city/state extracted directly from the CAD export
                        if '_csv_city' in df_c.columns:
                            city_val = str(df_c['_csv_city'].iloc[0]).strip().title()
                            if city_val and city_val.lower() not in ('nan', 'none', ''):
                                detected_city = city_val
                                detection_source = 'file'

                        if '_csv_state' in df_c.columns:
                            state_val = str(df_c['_csv_state'].iloc[0]).strip().upper()
                            if state_val in STATE_FIPS:
                                detected_state = state_val
                                detection_source = 'file'
                            elif state_val.title() in US_STATES_ABBR:
                                detected_state = US_STATES_ABBR[state_val.title()]
                                detection_source = 'file'

                        # If the export gives us a city but not a state, forward-geocode the city name.
                        if detected_city and not detected_state:
                            try:
                                geo_url = f"https://nominatim.openstreetmap.org/search?format=json&q={urllib.parse.quote(detected_city)}&limit=1&countrycodes=us"
                                req_geo = urllib.request.Request(geo_url, headers={'User-Agent': 'BRINC_COS_Optimizer/1.0'})
                                with urllib.request.urlopen(req_geo, timeout=8) as resp_geo:
                                    geo_result = json.loads(resp_geo.read().decode('utf-8'))
                                if geo_result:
                                    display_name = geo_result[0].get('display_name', '')
                                    parts = [p.strip() for p in display_name.split(',')]
                                    state_full = parts[2] if len(parts) >= 3 else ''
                                    if state_full in US_STATES_ABBR:
                                        detected_state = US_STATES_ABBR[state_full]
                                        detection_source = 'file'
                            except Exception:
                                pass

                        # Priority 2: reverse-geocode the centroid of the calls, not the first row.
                        if not detected_city or not detected_state:
                            try:
                                cen_lat = float(df_c['lat'].median())
                                cen_lon = float(df_c['lon'].median())
                                detected_state_full, detected_city_rg = reverse_geocode_state(cen_lat, cen_lon)
                                if detected_state_full and detected_state_full in US_STATES_ABBR:
                                    if not detected_state:
                                        detected_state = US_STATES_ABBR[detected_state_full]
                                    if not detected_city and detected_city_rg and detected_city_rg != 'Unknown City':
                                        detected_city = detected_city_rg
                                    detection_source = 'centroid'
                            except Exception:
                                pass

                        # ── Fallback: derive city/state from coordinate centroid bbox ──
                        # If Nominatim failed or returned nothing, try the Census FCC API
                        # which is reliable and returns county + state from a lat/lon.
                        if not detected_city or not detected_state:
                            try:
                                cen_lat_fb = float(df_c['lat'].median())
                                cen_lon_fb = float(df_c['lon'].median())
                                _fcc_url = (
                                    f"https://geo.fcc.gov/api/census/block/find"
                                    f"?latitude={cen_lat_fb}&longitude={cen_lon_fb}"
                                    f"&format=json&showall=false"
                                )
                                _fcc_req = urllib.request.Request(
                                    _fcc_url, headers={"User-Agent": "BRINC_COS_Optimizer/1.0"}
                                )
                                with urllib.request.urlopen(_fcc_req, timeout=8) as _fcc_resp:
                                    _fcc_data = json.loads(_fcc_resp.read().decode("utf-8"))
                                _fcc_county = _fcc_data.get("County", {}).get("name", "")
                                _fcc_state  = _fcc_data.get("State", {}).get("code", "")
                                if _fcc_county and _fcc_state and _fcc_state in STATE_FIPS:
                                    if not detected_state:
                                        detected_state = _fcc_state
                                    if not detected_city:
                                        # Use "County Name County" so fetch_county_boundary_local can find it
                                        detected_city = _fcc_county
                                    detection_source = "fcc_centroid"
                            except Exception:
                                pass

                        # Last resort: derive state from coordinate range alone
                        # (FL: lat 24-31, lon -87 to -80; this catches most US states uniquely)
                        if not detected_state:
                            try:
                                _cen_lat = float(df_c['lat'].median())
                                _cen_lon = float(df_c['lon'].median())
                                _BBOX_STATES = [
                                    ("FL", 24.5, 31.0, -87.6, -79.9),
                                    ("CA", 32.5, 42.0, -124.5, -114.1),
                                    ("TX", 25.8, 36.5, -106.6, -93.5),
                                    ("NY", 40.5, 45.0, -79.8, -71.9),
                                    ("IL", 36.9, 42.5, -91.5, -87.0),
                                    ("GA", 30.4, 35.0, -85.6, -80.8),
                                    ("NC", 33.8, 36.6, -84.3, -75.5),
                                    ("OH", 38.4, 42.0, -84.8, -80.5),
                                    ("PA", 39.7, 42.3, -80.5, -74.7),
                                    ("WA", 45.5, 49.0, -124.7, -116.9),
                                    ("AZ", 31.3, 37.0, -114.8, -109.0),
                                    ("CO", 36.9, 41.0, -109.1, -102.0),
                                    ("MO", 35.9, 40.6, -95.8, -89.1),
                                    ("NM", 31.3, 37.0, -109.1, -103.0),
                                    ("OR", 41.9, 46.2, -124.6, -116.5),
                                    ("TN", 34.9, 36.7, -90.3, -81.6),
                                    ("VA", 36.5, 39.5, -83.7, -75.2),
                                    ("SC", 32.0, 35.2, -83.4, -78.5),
                                    ("MI", 41.7, 48.3, -90.4, -82.4),
                                    ("WI", 42.5, 47.1, -92.9, -86.2),
                                    ("MN", 43.5, 49.4, -97.2, -89.5),
                                    ("IA", 40.4, 43.5, -96.6, -90.1),
                                    ("KS", 36.9, 40.0, -102.1, -94.6),
                                    ("NE", 40.0, 43.0, -104.1, -95.3),
                                    ("OK", 33.6, 37.0, -103.0, -94.4),
                                    ("AR", 33.0, 36.5, -94.6, -89.6),
                                    ("LA", 28.9, 33.0, -94.0, -88.8),
                                    ("MS", 30.2, 35.0, -91.7, -88.1),
                                    ("AL", 30.2, 35.0, -88.5, -84.9),
                                    ("IN", 37.8, 41.8, -88.1, -84.8),
                                    ("KY", 36.5, 39.1, -89.6, -81.9),
                                    ("WV", 37.2, 40.6, -82.6, -77.7),
                                    ("MD", 37.9, 39.7, -79.5, -75.0),
                                    ("DE", 38.4, 39.8, -75.8, -75.0),
                                    ("NJ", 38.9, 41.4, -75.6, -73.9),
                                    ("CT", 41.0, 42.1, -73.7, -71.8),
                                    ("RI", 41.1, 42.0, -71.9, -71.1),
                                    ("MA", 41.2, 42.9, -73.5, -69.9),
                                    ("VT", 42.7, 45.0, -73.4, -71.5),
                                    ("NH", 42.7, 45.3, -72.6, -70.7),
                                    ("ME", 43.1, 47.5, -71.1, -66.9),
                                    ("NV", 35.0, 42.0, -120.0, -114.0),
                                    ("UT", 36.9, 42.0, -114.1, -109.0),
                                    ("ID", 41.9, 49.0, -117.2, -111.0),
                                    ("MT", 44.4, 49.0, -116.1, -104.0),
                                    ("WY", 40.9, 45.0, -111.1, -104.0),
                                    ("ND", 45.9, 49.0, -104.1, -96.6),
                                    ("SD", 42.5, 45.9, -104.1, -96.4),
                                    ("AK", 54.0, 71.5, -168.0, -130.0),
                                    ("HI", 18.9, 22.2, -160.2, -154.8),
                                ]
                                for _st, _lat0, _lat1, _lon0, _lon1 in _BBOX_STATES:
                                    if _lat0 <= _cen_lat <= _lat1 and _lon0 <= _cen_lon <= _lon1:
                                        detected_state = _st
                                        detection_source = "coord_bbox"
                                        break
                            except Exception:
                                pass

                        if detected_city and detected_state:
                            st.session_state['active_city'] = detected_city
                            st.session_state['active_state'] = detected_state
                            st.session_state['target_cities'] = [{"city": detected_city, "state": detected_state}]
                            st.session_state['location_detection_source'] = detection_source
                            st.toast(f"📍 Detected: {detected_city}, {detected_state}")
                        elif detected_state:
                            # We have state but no city — store state and let boundary
                            # selection use the FCC county name as a county lookup
                            st.session_state['active_state'] = detected_state
                            st.session_state['location_detection_source'] = detection_source

                    # Keep uploaded calls intact here. Boundary validation should be the first real geographic filter.
                    st.session_state['df_calls']             = df_c
                    st.session_state['df_calls_full']        = df_c_full
                    st.session_state['df_stations']          = df_s
                    st.session_state['total_original_calls'] = len(df_c_full)
                    st.session_state['total_modeled_calls']  = len(df_c)

                    # ── Clear stale shapefiles from previous sessions ──────────────
                    # Old .shp files from prior cities (e.g. Bernalillo) would otherwise
                    # be picked up by find_relevant_jurisdictions when no fresh boundary
                    # is saved, causing completely wrong jurisdiction boundaries.
                    try:
                        import glob as _glob
                        for _stale in _glob.glob(os.path.join(SHAPEFILE_DIR, "*.shp")):
                            for _ext in [".shp", ".shx", ".dbf", ".prj", ".cpg"]:
                                _f = _stale.replace(".shp", _ext)
                                try:
                                    if os.path.exists(_f): os.remove(_f)
                                except Exception:
                                    pass
                    except Exception:
                        pass
                    st.session_state['boundary_source_path'] = ''
                    st.session_state['master_gdf_override'] = None

                    with st.spinner(get_jurisdiction_message()):
                        _calls_for_boundary = df_c_full if df_c_full is not None and len(df_c_full) > 0 else df_c

                        # ── PRIMARY: pure coordinate spatial join against local parquets ──
                        # Works for any file regardless of whether city/state was detected.
                        # Returns ALL jurisdictions that contain ≥1% of calls.
                        coord_gdf = find_jurisdictions_by_coordinates(_calls_for_boundary)

                        if coord_gdf is not None and not coord_gdf.empty:
                            # Store the GeoDataFrame directly in session state — no disk
                            # round-trip needed. The render pass reads it back directly.
                            st.session_state['master_gdf_override'] = coord_gdf
                            st.session_state['boundary_source_path'] = 'local_parquet'
                            st.session_state['boundary_kind'] = 'place'
                            st.session_state['active_city'] = coord_gdf.iloc[0]['DISPLAY_NAME']

                        else:
                            st.session_state['master_gdf_override'] = None
                            # ── FALLBACK: name-based lookup (original path) ──
                            detected_city_for_boundary = st.session_state.get('active_city', '')
                            detected_state_for_boundary = st.session_state.get('active_state', '')
                            if detected_city_for_boundary and detected_state_for_boundary and detected_state_for_boundary in STATE_FIPS:
                                city_text = str(detected_city_for_boundary or '').strip()
                                prefer_county = str(st.session_state.get('location_detection_source', '')) == 'centroid'
                                b_success, b_gdf, b_kind, b_hits = _select_best_boundary_for_calls(
                                    _calls_for_boundary,
                                    city_text,
                                    detected_state_for_boundary,
                                    prefer_county=prefer_county
                                )
                                st.session_state['boundary_kind'] = b_kind
                                if b_success and b_gdf is not None:
                                    _saved = save_boundary_gdf(b_gdf, b_kind, city_text, detected_state_for_boundary)
                                    st.session_state['boundary_source_path'] = _saved or ''

                    st.session_state['data_source'] = 'cad_upload'
                    st.session_state['demo_mode_used'] = False
                    st.session_state['sim_mode_used'] = False
                    st.session_state['map_build_logged'] = False
                    st.session_state['csvs_ready'] = True
                    st.rerun()

    with path_demo_col:
        st.markdown(f"""
        <div class="path-card" style="--accent:#FFD700;">
            <span class="pc-icon">⚡</span>
            <div class="pc-tag">Path 03</div>
            <div class="pc-title">1-Click Demo<br>Large US City</div>
            <div class="pc-desc">Instantly spin up a fully pre-configured scenario for a major US city. Ideal for live stakeholder presentations and platform walkthroughs.</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

        if st.button("⚡ Launch Random Demo City", use_container_width=True, key="demo_btn"):
            random.seed(datetime.datetime.now().microsecond + os.getpid())
            already_used = st.session_state.get('_last_demo_city', '')
            candidates = [c for c in FAST_DEMO_CITIES if c[0] != already_used]
            rcity, rstate = random.choice(candidates)
            st.session_state['_last_demo_city'] = rcity
            st.session_state['target_cities'] = [{"city": rcity, "state": rstate}]
            st.session_state.city_count = 1
            for i in range(10):
                st.session_state.pop(f"c_{i}", None)
                st.session_state.pop(f"s_{i}", None)
            st.session_state['trigger_sim'] = True
            st.session_state['demo_mode_used'] = True
            st.rerun()

        city_chips = "  ·  ".join([f"{c}" for c, _ in DEMO_CITIES[:12]]) + "  · and more…"
        st.markdown(f"""
        <div class="demo-cities">
            <b>Available Cities</b><br>
            {city_chips}
        </div>
        <div class="demo-check">
            <span>✓</span>Real Census boundaries<br>
            <span>✓</span>Clustered 911 simulation<br>
            <span>✓</span>100 station candidates<br>
            <span>✓</span>Full optimization & export
        </div>
        """, unsafe_allow_html=True)

    st.markdown(f"""
    <div style="text-align:center; margin-top:8px; font-size:0.63rem; color:#2a2a2a;">
        BRINC Drones, Inc. · <a href="https://brincdrones.com" target="_blank"
        style="color:#333; text-decoration:none;">brincdrones.com</a>
        · All coverage estimates are for planning purposes only.
    </div>
    """, unsafe_allow_html=True)

    if submit_demo or st.session_state.get('trigger_sim', False):
        if st.session_state.get('trigger_sim', False):
            st.session_state['trigger_sim'] = False
            # trigger_sim is set by the demo button — mark accordingly
            if not st.session_state.get('demo_mode_used', False):
                st.session_state['demo_mode_used'] = True

        active_targets = [loc for loc in st.session_state['target_cities'] if loc['city'].strip()]
        if not active_targets:
            st.error("Please enter at least one valid city name.")
            st.stop()

        if len(active_targets) == 1:
            st.session_state['active_city']  = active_targets[0]['city']
            st.session_state['active_state'] = active_targets[0]['state']
        else:
            st.session_state['active_city']  = f"{active_targets[0]['city']} & {len(active_targets)-1} others"
            st.session_state['active_state'] = active_targets[0]['state']

        prog = st.progress(0, text="🫡 Preparing tools worthy of those who serve…")
        all_gdfs = []
        total_estimated_pop = 0

        for i, loc in enumerate(active_targets):
            c_name = loc['city'].strip()
            s_name = loc['state']
            is_county = c_name.lower().endswith(" county")
            boundary_kind = 'county' if is_county else 'place'
            
            prog.progress(10 + int((i / len(active_targets)) * 20),
                          text=f"🗺️ Mapping {c_name}, {s_name} — because every block they patrol matters…")
            
            if is_county:
                success, temp_gdf = fetch_county_boundary_local(s_name, c_name)
                if not success:
                    success, temp_gdf = fetch_county_boundary_local(s_name, c_name + " County")
                if success:
                    boundary_kind = 'county'
            else:
                # Try place first, auto-fall back to county if not found
                success, temp_gdf = fetch_place_boundary_local(s_name, c_name)
                if success:
                    boundary_kind = 'place'
                else:
                    success, temp_gdf = fetch_county_boundary_local(s_name, c_name)
                    if not success:
                        success, temp_gdf = fetch_county_boundary_local(s_name, c_name + " County")
                    if success:
                        boundary_kind = 'county'
            is_county = (boundary_kind == 'county')
            st.session_state['boundary_kind'] = boundary_kind

            # County boundaries come from the parquet, not from TIGER, so they are
            # never written to SHAPEFILE_DIR. find_relevant_jurisdictions() only scans
            # that directory, so without this save it always falls back to
            # "Auto-Generated Boundary". Save any successfully loaded county GDF now.
            if success and temp_gdf is not None:
                _saved = save_boundary_gdf(temp_gdf, boundary_kind, c_name, s_name)
                if i == 0:
                    st.session_state['boundary_source_path'] = _saved or ''

            if success:
                all_gdfs.append(temp_gdf)
                pop = fetch_census_population(STATE_FIPS[s_name], c_name, is_county=is_county)
                if pop:
                    total_estimated_pop += pop
                    st.toast(f"✅ {c_name} population verified: {pop:,}")
                else:
                    gdf_proj   = temp_gdf.to_crs(epsg=3857)
                    area_sq_mi = gdf_proj.geometry.area.sum() / 2589988.11
                    est = KNOWN_POPULATIONS.get(c_name, int(area_sq_mi * 3500))
                    total_estimated_pop += est
                    st.toast(f"⚠️ {c_name} population estimated: {est:,}")
            else:
                st.warning(f"⚠️ Could not find a boundary for {c_name}, {s_name}. Try another city.")
                if st.session_state.get('_last_demo_city') == c_name:
                    random.seed(datetime.datetime.now().microsecond + os.getpid())
                    candidates = [c for c in DEMO_CITIES if c[0] != c_name]
                    rcity, rstate = random.choice(candidates)
                    st.session_state['_last_demo_city'] = rcity
                    st.session_state['target_cities'] = [{"city": rcity, "state": rstate}]
                    for j in range(10):
                        st.session_state.pop(f"c_{j}", None)
                        st.session_state.pop(f"s_{j}", None)
                    st.rerun()

        if not all_gdfs:
            prog.empty()
            st.error("❌ Could not find Census boundaries for any of the entered locations. Check spelling.")
            st.stop()

        prog.progress(35, text="💙 Boundaries loaded — honoring the officers who know every street…")
        active_city_gdf = pd.concat(all_gdfs, ignore_index=True)
        city_poly = active_city_gdf.geometry.union_all()
        st.session_state['estimated_pop'] = total_estimated_pop

        annual_cfs = int(total_estimated_pop * 0.6)
        st.session_state['total_original_calls'] = annual_cfs
        simulated_points_count = min(max(int(annual_cfs), 365), 36500)

        prog.progress(55, text="🚔 Modeling 911 calls — every one represents someone who needed help…")
        np.random.seed(42)
        random.seed(42)
        call_points = generate_clustered_calls(city_poly, simulated_points_count)
        
        base_date = datetime.datetime.now() - datetime.timedelta(days=364)
        fake_dts = [(base_date + datetime.timedelta(days=random.randint(0, 364), hours=random.randint(0, 23), minutes=random.randint(0, 59))) for _ in range(simulated_points_count)]
        
        df_demo = pd.DataFrame({
            'lat':      [p[0] for p in call_points],
            'lon':      [p[1] for p in call_points],
            'priority': np.random.choice([1, 2, 3], simulated_points_count, p=[0.15, 0.35, 0.50]),
            'date':     [d.strftime('%Y-%m-%d') for d in fake_dts],
            'time':     [d.strftime('%H:%M:%S') for d in fake_dts]
        })
        st.session_state['df_calls'] = df_demo
        st.session_state['df_calls_full'] = df_demo.copy()
        st.session_state['total_modeled_calls'] = len(df_demo)

        # --- PROCESS OPTIONAL CUSTOM STATIONS ---
        custom_stations_used = False
        sim_uploader = st.session_state.get('sim_station_uploader')
        
        if sim_uploader is not None:
            prog.progress(80, text="🏅 Geocoding custom stations from CSV…")
            import time
            try:
                sim_uploader.seek(0)
                s_df = pd.read_csv(sim_uploader)
                s_df.columns = [str(c).lower().strip() for c in s_df.columns]
                
                # Detect what columns the user provided
                lat_col = next((c for c in s_df.columns if c in ['lat', 'latitude', 'y']), None)
                lon_col = next((c for c in s_df.columns if c in ['lon', 'long', 'longitude', 'x']), None)
                addr_col = next((c for c in s_df.columns if any(a in c for a in ['address', 'street', 'location'])), None)
                name_col = next((c for c in s_df.columns if any(n in c for n in ['name', 'station', 'facility', 'dept'])), None)
                type_col = next((c for c in s_df.columns if any(t in c for t in ['type', 'category'])), None)
                
                parsed_stations = []
                for idx, row in s_df.iterrows():
                    s_name = str(row[name_col]) if name_col and pd.notna(row[name_col]) else f"Custom Station {idx+1}"
                    s_type = str(row[type_col]) if type_col and pd.notna(row[type_col]) else 'Custom'
                    s_lat, s_lon = None, None
                    
                    if lat_col and lon_col and pd.notna(row[lat_col]) and pd.notna(row[lon_col]):
                        s_lat, s_lon = float(row[lat_col]), float(row[lon_col])
                    elif addr_col and pd.notna(row[addr_col]):
                        addr_str = str(row[addr_col])
                        # Attempt geocoding
                        s_lat, s_lon = forward_geocode(addr_str)
                        if s_lat is None:
                            # Fallback: Try appending the city and state to the address string
                            s_lat, s_lon = forward_geocode(f"{addr_str}, {active_targets[0]['city']}, {active_targets[0]['state']}")
                        if s_lat is None:
                            st.toast(f"⚠️ Could not geocode: {addr_str}")
                        time.sleep(1) # Slow down requests slightly to prevent API blocking
                        
                    if s_lat and s_lon:
                        parsed_stations.append({
                            'name': s_name,
                            'lat': s_lat,
                            'lon': s_lon,
                            'type': s_type
                        })
                        
                if parsed_stations:
                    st.session_state['df_stations'] = pd.DataFrame(parsed_stations)
                    custom_stations_used = True
                else:
                    st.warning("⚠️ Could not geocode or parse your custom stations. Falling back to 100 random stations.")
            except Exception as e:
                st.warning(f"⚠️ Error reading custom stations: {e}. Falling back to random stations.")

        # --- FALLBACK: PULL REAL STATIONS FROM OPENSTREETMAP ---
        if not custom_stations_used:
            prog.progress(80, text="🌐 Querying OpenStreetMap for real police, fire & schools…")
            
            # Use the simulated calls we just made to find the bounding box, and pull real OSM data!
            df_s, osm_note = generate_stations_from_calls(st.session_state['df_calls'])
            
            if df_s is not None and not df_s.empty:
                st.session_state['df_stations'] = df_s
                st.toast(f"✅ {osm_note}")
            else:
                # Absolute worst-case scenario (no internet or OSM API is down): fall back to random
                st.warning("⚠️ Could not reach OpenStreetMap for real stations. Falling back to random placements.")
                station_points = generate_random_points_in_polygon(city_poly, 100)
                types = ['Police', 'Fire', 'EMS'] * 34
                st.session_state['df_stations'] = pd.DataFrame({
                    'name': [f'Station {i+1}' for i in range(len(station_points))],
                    'lat':  [p[0] for p in station_points],
                    'lon':  [p[1] for p in station_points],
                    'type': types[:len(station_points)]
                })

        prog.progress(100, text="✅ Ready — built for the communities they protect and serve.")
        st.session_state['inferred_daily_calls_override'] = int(annual_cfs / 365)
        st.session_state['data_source'] = 'simulation'
        st.session_state['sim_mode_used'] = True
        st.session_state['map_build_logged'] = False
        st.session_state['csvs_ready'] = True
        st.rerun()

# ============================================================
# COMMUNITY IMPACT DASHBOARD
# ============================================================

def generate_community_impact_dashboard_html(
    city, state, population,
    total_calls, calls_covered_perc, area_covered_perc,
    avg_resp_time_min, avg_time_saved_min,
    fleet_capex, annual_savings, break_even_text,
    actual_k_responder, actual_k_guardian,
    dfr_dispatch_rate, deflection_rate,
    daily_dfr_responses, daily_drone_only_calls,
    active_drones,
    df_calls_full,
    theme='dark',
):
    """
    Generate a Community Impact Dashboard HTML string.
    theme='dark'  -- black background, for in-app Streamlit embed.
    theme='light' -- white/off-white background, for HTML export / print.
    """
    import json as _json

    # ── Derived metrics ──────────────────────────────────────────────────────
    daily_flights   = max(0.0, float(daily_dfr_responses or 0))
    annual_flights  = daily_flights * 365

    # Guardian uptime hours per day from CONFIG
    g_count  = max(0, int(actual_k_guardian or 0))
    r_count  = max(0, int(actual_k_responder or 0))
    g_daily_hrs = g_count * GUARDIAN_FLIGHT_HOURS_PER_DAY
    r_daily_hrs = r_count * 11.6          # Responder patrol hours
    total_daily_flight_hrs = g_daily_hrs + r_daily_hrs
    annual_flight_hrs = total_daily_flight_hrs * 365

    # Response time advantage
    drone_min  = float(avg_resp_time_min or 0)
    saved_min  = float(avg_time_saved_min or 0)
    ground_min = drone_min + saved_min
    drone_wins_pct = min(99, max(CONFIG['DRONE_WINS_FLOOR'], round(calls_covered_perc * CONFIG['DRONE_WINS_MULTIPLIER']))) if calls_covered_perc > 0 else 0

    # Per-type response time breakdown for the bar chart
    _RESP_SPEED   = float(CONFIG.get('RESPONDER_SPEED', 42.0))
    _GUARD_SPEED  = float(CONFIG.get('GUARDIAN_SPEED', 60.0))
    _GROUND_SPEED = float(CONFIG.get('DEFAULT_TRAFFIC_SPEED', 35.0))
    _resp_drones  = [d for d in active_drones if d.get('type') == 'RESPONDER']
    _guard_drones = [d for d in active_drones if d.get('type') == 'GUARDIAN']
    resp_drone_min  = (sum(d['avg_time_min'] for d in _resp_drones)  / len(_resp_drones))  if _resp_drones  else None
    guard_drone_min = (sum(d['avg_time_min'] for d in _guard_drones) / len(_guard_drones)) if _guard_drones else None
    # Ground time uses the corrected formula: avg_dist × 1.4 / ground_speed
    # avg_dist = avg_time_min × drone_speed / 60  →  ground_time = avg_time_min × drone_speed × 1.4 / ground_speed
    resp_ground_min  = (resp_drone_min  * _RESP_SPEED  * 1.4 / _GROUND_SPEED) if resp_drone_min  is not None else None
    guard_ground_min = (guard_drone_min * _GUARD_SPEED * 1.4 / _GROUND_SPEED) if guard_drone_min is not None else None
    # Reference ground time for the chart = best available single ground estimate
    _chart_ground_min = ground_min if ground_min > 0 else max(
        (resp_ground_min or 0), (guard_ground_min or 0)
    )
    _chart_max = max(_chart_ground_min, resp_drone_min or 0, guard_drone_min or 0, 0.1)

    # Pre-build per-type bar HTML so we avoid nested f-strings in the template
    if resp_drone_min is not None:
        _resp_bar_h = min(100, resp_drone_min / _chart_max * 100)
        _resp_bar_html = (
            '<div class="rt-bar-wrap">'
            '<div class="rt-bar-outer">'
            f'<div class="rt-bar-fill" style="height:{_resp_bar_h:.0f}%;background:linear-gradient(180deg,var(--accent-blue),#3b82f6);"></div>'
            '</div>'
            '<div class="rt-bar-label">&#x1F681; Responder '
            '<span class="tip-cid" data-tip="Avg flight time for Responder drones (42 mph airspeed, 2-mile zone). Direct line-of-sight flight — no traffic, no turns.">?</span></div>'
            f'<div class="rt-bar-value" style="color:var(--accent-blue);">{resp_drone_min:.1f} min</div>'
            '</div>'
        )
    else:
        _resp_bar_html = ''

    if guard_drone_min is not None:
        _guard_bar_h = min(100, guard_drone_min / _chart_max * 100)
        _guard_bar_html = (
            '<div class="rt-bar-wrap">'
            '<div class="rt-bar-outer">'
            f'<div class="rt-bar-fill" style="height:{_guard_bar_h:.0f}%;background:linear-gradient(180deg,var(--accent-gold),#ca8a04);"></div>'
            '</div>'
            '<div class="rt-bar-label">&#x1F985; Guardian '
            '<span class="tip-cid" data-tip="Avg flight time for Guardian drones (60 mph airspeed, up to 8-mile zone). Longer range means slightly longer avg flight, still far faster than road travel over that distance.">?</span></div>'
            f'<div class="rt-bar-value" style="color:var(--accent-gold);">{guard_drone_min:.1f} min</div>'
            '</div>'
        )
    else:
        _guard_bar_html = ''

    _ground_bar_h = min(100, _chart_ground_min / _chart_max * 100)
    _ground_bar_html = (
        '<div class="rt-bar-wrap">'
        '<div class="rt-bar-outer">'
        f'<div class="rt-bar-fill" style="height:{_ground_bar_h:.0f}%;background:linear-gradient(180deg,#f59e0b,#d97706);"></div>'
        '</div>'
        f'<div class="rt-bar-label">&#x1F694; Ground Unit (est.) '
        f'<span class="tip-cid" data-tip="Patrol car travel time over the same avg incident distance at {_GROUND_SPEED:.0f} mph road speed with a 1.4\u00d7 tortuosity factor. Actual response varies by unit availability.">?</span></div>'
        f'<div class="rt-bar-value" style="color:#f59e0b;">{_chart_ground_min:.1f} min</div>'
        '</div>'
    )

    # Outcomes counter (modeled estimates)
    total_annual_dfr = int(annual_flights * float(dfr_dispatch_rate or 0.25))
    arrests_est      = int(total_annual_dfr * CONFIG['OUTCOME_ARREST_RATE'])
    rescues_est      = int(total_annual_dfr * CONFIG['OUTCOME_RESCUE_RATE'])
    deescalation_est = int(total_annual_dfr * CONFIG['OUTCOME_DEESCALATION_RATE'])
    missing_est      = int(total_annual_dfr * CONFIG['OUTCOME_MISSING_RATE'])

    # ROI
    roi_multiple = round(float(annual_savings or 0) / max(float(fleet_capex or 1), 1), 2)
    cost_per_call_drone   = CONFIG['DRONE_COST_PER_CALL']
    cost_per_call_officer = CONFIG['OFFICER_COST_PER_CALL']
    cost_saved_per_resolved = cost_per_call_officer - cost_per_call_drone
    total_resolved_annually = int(float(daily_drone_only_calls or 0) * 365)

    # Call type breakdown from df_calls_full
    call_type_data = {}
    _type_col = None
    if df_calls_full is not None and not df_calls_full.empty:
        for _tc in ['call_type_desc','agencyeventtypecodedesc','calldesc','description','nature','type']:
            if _tc in df_calls_full.columns:
                _type_col = _tc
                break
    if _type_col:
        try:
            _vc = df_calls_full[_type_col].dropna().astype(str).str.strip().str.title().value_counts().head(8)
            _total = _vc.sum()
            for k, v in _vc.items():
                call_type_data[str(k)[:32]] = int(v)
        except Exception:
            pass

    if not call_type_data:
        # Reasonable DFR-program defaults
        call_type_data = {
            "Shots Fired / Weapon": int(total_annual_dfr * 0.12),
            "Suspicious Person": int(total_annual_dfr * 0.19),
            "Burglary / Theft": int(total_annual_dfr * 0.17),
            "Traffic Accident": int(total_annual_dfr * 0.11),
            "Welfare Check": int(total_annual_dfr * 0.20),
            "Domestic Disturbance": int(total_annual_dfr * 0.09),
            "Missing Person": int(total_annual_dfr * 0.05),
            "Other": int(total_annual_dfr * 0.07),
        }

    ct_total   = max(1, sum(call_type_data.values()))
    ct_items_js = _json.dumps([
        {"label": k, "count": v, "pct": round(v / ct_total * 100, 1)}
        for k, v in call_type_data.items()
    ])

    # Equity / geographic note: district coverage from active_drones
    drone_names_js = _json.dumps([
        {"name": d["name"].split(",")[0][:28], "type": d["type"]}
        for d in active_drones
    ] if active_drones else [])

    # Privacy policy data-retention badge values
    retention_days   = 30   # industry standard shown in transparency portals
    no_proactive     = True
    no_facial_recog  = True
    warrant_transit  = True  # camera forward-facing in transit

    # ── Theme CSS variables ──────────────────────────────────────────────────
    if theme == 'dark':
        _css_vars = """
    --bg-page:      #000000;
    --bg-card:      #111111;
    --bg-inset:     #0a0a0a;
    --ink:          #f0f0f0;
    --ink-mid:      #bbbbbb;
    --ink-light:    #888888;
    --rule:         #2a2a2a;
    --accent-blue:  #00D2FF;
    --accent-blue-lt: rgba(0,210,255,0.12);
    --accent-green: #39FF14;
    --accent-green-lt: rgba(57,255,20,0.10);
    --accent-gold:  #FFD700;
    --accent-gold-lt: rgba(255,215,0,0.10);
    --accent-red:   #ff4b4b;
    --accent-red-lt: rgba(255,75,75,0.12);
    --accent-slate: #888888;
    --shadow-sm: 0 1px 4px rgba(0,0,0,0.6);
    --shadow-md: 0 4px 16px rgba(0,210,255,0.10), 0 2px 6px rgba(0,0,0,0.5);
    --header-border: 2px solid #333333;
    --card-hover-shadow: 0 0 0 1px #00D2FF44, 0 8px 24px rgba(0,210,255,0.12);
"""
        _body_bg = '#000000'
        _body_color = '#f0f0f0'
    else:
        _css_vars = """
    --bg-page:      #f8f7f4;
    --bg-card:      #ffffff;
    --bg-inset:     #f8f7f4;
    --ink:          #1a1a2e;
    --ink-mid:      #3d3d5c;
    --ink-light:    #6b6b8a;
    --rule:         #e2e0da;
    --accent-blue:  #1a56db;
    --accent-blue-lt: #dbeafe;
    --accent-green: #0d9e6e;
    --accent-green-lt: #d1fae5;
    --accent-gold:  #b45309;
    --accent-gold-lt: #fef3c7;
    --accent-red:   #be123c;
    --accent-red-lt: #ffe4e6;
    --accent-slate: #475569;
    --shadow-sm: 0 1px 3px rgba(26,22,46,0.06), 0 1px 2px rgba(26,22,46,0.04);
    --shadow-md: 0 4px 12px rgba(26,22,46,0.08), 0 2px 4px rgba(26,22,46,0.04);
    --header-border: 2px solid #1a1a2e;
    --card-hover-shadow: var(--shadow-md);
"""
        _body_bg = '#f8f7f4'
        _body_color = '#1a1a2e'

    # Build HTML ─────────────────────────────────────────────────────────────
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Libre+Baskerville:wght@400;700&family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
  :root {{{_css_vars}  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: 'DM Sans', sans-serif;
    background: {_body_bg};
    color: {_body_color};
    font-size: 14px;
    line-height: 1.55;
    padding: 28px 24px 40px;
  }}

  /* ── Page header ── */
  .dash-header {{
    display: flex;
    align-items: flex-start;
    justify-content: space-between;
    border-bottom: var(--header-border);
    padding-bottom: 14px;
    margin-bottom: 28px;
    gap: 16px;
    flex-wrap: wrap;
  }}
  .dash-title {{
    font-family: 'Libre Baskerville', Georgia, serif;
    font-size: 22px;
    font-weight: 700;
    color: var(--ink);
    letter-spacing: -0.3px;
  }}
  .dash-subtitle {{
    font-size: 12.5px;
    color: var(--ink-light);
    margin-top: 3px;
    font-weight: 400;
  }}
  .dash-meta {{
    text-align: right;
    font-size: 11.5px;
    color: var(--ink-light);
    line-height: 1.7;
  }}
  .dash-meta strong {{ color: var(--ink); font-weight: 600; }}

  /* ── Section label ── */
  .section-label {{
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 1.6px;
    text-transform: uppercase;
    color: var(--ink-light);
    margin-bottom: 10px;
    padding-bottom: 5px;
    border-bottom: 1px solid var(--rule);
  }}

  /* ── Grid layouts ── */
  .grid-3 {{ display: grid; grid-template-columns: repeat(3,1fr); gap: 14px; margin-bottom: 20px; }}
  .grid-2 {{ display: grid; grid-template-columns: repeat(2,1fr); gap: 14px; margin-bottom: 20px; }}
  .grid-4 {{ display: grid; grid-template-columns: repeat(4,1fr); gap: 12px; margin-bottom: 20px; }}
  @media(max-width:700px) {{
    .grid-3,.grid-4 {{ grid-template-columns:1fr 1fr; }}
    .grid-2 {{ grid-template-columns:1fr; }}
  }}

  /* ── Stat card ── */
  .stat-card {{
    background: var(--bg-card);
    border: 1px solid var(--rule);
    border-radius: 10px;
    padding: 18px 20px 16px;
    box-shadow: var(--shadow-sm);
    position: relative;
    overflow: hidden;
    transition: box-shadow 0.2s;
  }}
  .stat-card:hover {{ box-shadow: var(--card-hover-shadow); }}
  .stat-card .accent-bar {{
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    border-radius: 10px 10px 0 0;
  }}
  .stat-card .card-label {{
    font-size: 10.5px;
    font-weight: 600;
    letter-spacing: 0.8px;
    text-transform: uppercase;
    color: var(--ink-light);
    margin-bottom: 6px;
  }}
  .stat-card .card-value {{
    font-family: 'DM Mono', monospace;
    font-size: 26px;
    font-weight: 500;
    color: var(--ink);
    line-height: 1.1;
  }}
  .stat-card .card-sub {{
    font-size: 11px;
    color: var(--ink-light);
    margin-top: 5px;
  }}
  .stat-card .card-badge {{
    display: inline-block;
    font-size: 10px;
    font-weight: 600;
    padding: 2px 7px;
    border-radius: 99px;
    margin-top: 6px;
  }}

  /* ── Progress bar ── */
  .prog-row {{ margin-bottom: 10px; }}
  .prog-meta {{ display: flex; justify-content: space-between; align-items: baseline; margin-bottom: 4px; }}
  .prog-label {{ font-size: 12px; color: var(--ink-mid); font-weight: 500; }}
  .prog-val {{ font-family: 'DM Mono', monospace; font-size: 12px; color: var(--ink); font-weight: 500; }}
  .prog-track {{
    height: 8px;
    background: var(--rule);
    border-radius: 99px;
    overflow: hidden;
  }}
  .prog-fill {{
    height: 100%;
    border-radius: 99px;
    animation: growBar 1.2s cubic-bezier(0.22,1,0.36,1) both;
  }}
  @keyframes growBar {{ from {{ width:0 }} }}

  /* ── Animated counter ── */
  .counter {{ display: inline-block; }}

  /* ── Response time comparison ── */
  .rt-compare {{
    background: var(--bg-card);
    border: 1px solid var(--rule);
    border-radius: 10px;
    padding: 20px;
    box-shadow: var(--shadow-sm);
    margin-bottom: 20px;
  }}
  .rt-bars {{ display: flex; gap: 28px; align-items: flex-end; margin-top: 14px; }}
  .rt-bar-wrap {{ flex: 1; text-align: center; }}
  .rt-bar-outer {{
    background: var(--rule);
    border-radius: 6px 6px 0 0;
    position: relative;
    overflow: hidden;
    display: flex;
    align-items: flex-end;
    height: 120px;
  }}
  .rt-bar-fill {{
    width: 100%;
    border-radius: 6px 6px 0 0;
    animation: growUp 1.4s cubic-bezier(0.22,1,0.36,1) both;
    position: relative;
  }}
  @keyframes growUp {{ from {{ height: 0 }} }}
  .rt-bar-label {{ margin-top: 8px; font-size: 11.5px; font-weight: 600; color: var(--ink-mid); }}
  .rt-bar-value {{ font-family: 'DM Mono', monospace; font-size: 17px; font-weight: 500; margin-top: 3px; }}
  .rt-wins-badge {{
    display: inline-block;
    background: var(--accent-green-lt);
    color: var(--accent-green);
    font-size: 11px;
    font-weight: 700;
    padding: 4px 12px;
    border-radius: 99px;
    margin-top: 14px;
  }}

  /* ── 4th Amendment panel ── */
  .amend-panel {{
    background: var(--bg-card);
    border: 1px solid var(--rule);
    border-left: 4px solid var(--accent-blue);
    border-radius: 10px;
    padding: 20px 22px;
    box-shadow: var(--shadow-sm);
    margin-bottom: 20px;
  }}
  .amend-title {{
    font-family: 'Libre Baskerville', serif;
    font-size: 15px;
    font-weight: 700;
    color: var(--ink);
    margin-bottom: 12px;
    display: flex;
    align-items: center;
    gap: 8px;
  }}
  .amend-grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; margin-top: 14px; }}
  @media(max-width:700px) {{ .amend-grid {{ grid-template-columns: 1fr 1fr; }} }}
  .amend-item {{
    background: var(--bg-inset);
    border-radius: 8px;
    padding: 12px 14px;
    display: flex;
    align-items: flex-start;
    gap: 10px;
  }}
  .amend-icon {{ font-size: 18px; flex-shrink: 0; line-height: 1; margin-top: 1px; }}
  .amend-item-title {{ font-size: 11.5px; font-weight: 700; color: var(--ink); margin-bottom: 2px; }}
  .amend-item-desc {{ font-size: 10.5px; color: var(--ink-light); line-height: 1.45; }}
  .amend-disclaimer {{
    margin-top: 12px;
    font-size: 10.5px;
    color: var(--ink-light);
    background: var(--accent-blue-lt);
    border-radius: 6px;
    padding: 8px 12px;
    line-height: 1.5;
  }}

  /* ── Outcomes counters ── */
  .outcome-card {{
    background: var(--bg-card);
    border: 1px solid var(--rule);
    border-radius: 10px;
    padding: 18px 16px 14px;
    text-align: center;
    box-shadow: var(--shadow-sm);
    animation: fadeUp 0.6s ease both;
  }}
  @keyframes fadeUp {{ from {{ opacity:0; transform:translateY(12px) }} }}
  .outcome-icon {{ font-size: 26px; margin-bottom: 8px; display: block; }}
  .outcome-val {{
    font-family: 'DM Mono', monospace;
    font-size: 28px;
    font-weight: 500;
    color: var(--ink);
    line-height: 1;
  }}
  .outcome-label {{ font-size: 11px; color: var(--ink-light); font-weight: 500; margin-top: 5px; text-transform: uppercase; letter-spacing: 0.5px; }}
  .outcome-note {{ font-size: 10px; color: var(--ink-light); margin-top: 4px; font-style: italic; }}

  /* ── ROI meter ── */
  .roi-panel {{
    background: var(--bg-card);
    border: 1px solid var(--rule);
    border-radius: 10px;
    padding: 20px 22px;
    box-shadow: var(--shadow-sm);
    margin-bottom: 20px;
  }}
  .roi-row {{ display: flex; gap: 20px; align-items: stretch; flex-wrap: wrap; }}
  .roi-big {{ flex: 1; min-width: 160px; }}
  .roi-big-val {{
    font-family: 'DM Mono', monospace;
    font-size: 38px;
    font-weight: 500;
    color: var(--accent-green);
    line-height: 1;
    animation: fadeUp 0.8s ease both;
  }}
  .roi-big-label {{ font-size: 11px; color: var(--ink-light); font-weight: 600; text-transform: uppercase; letter-spacing: 0.8px; margin-top: 5px; }}
  .roi-details {{ flex: 2; min-width: 220px; }}
  .roi-line {{ display: flex; justify-content: space-between; padding: 7px 0; border-bottom: 1px solid var(--rule); font-size: 12.5px; }}
  .roi-line:last-child {{ border-bottom: none; }}
  .roi-line-label {{ color: var(--ink-mid); }}
  .roi-line-val {{ font-family: 'DM Mono', monospace; font-weight: 500; color: var(--ink); }}

  /* ── Call type bars ── */
  .ct-panel {{
    background: var(--bg-card);
    border: 1px solid var(--rule);
    border-radius: 10px;
    padding: 20px 22px;
    box-shadow: var(--shadow-sm);
    margin-bottom: 20px;
  }}

  /* ── Uptime donut placeholder ── */
  canvas {{ display: block; }}

  /* ── Disclaimer footer ── */
  .dash-footer {{
    margin-top: 24px;
    padding-top: 14px;
    border-top: 1px solid var(--rule);
    font-size: 10px;
    color: var(--ink-light);
    line-height: 1.6;
  }}
  .dash-footer strong {{ color: var(--ink-mid); }}

  /* ── Pulse dot ── */
  @keyframes pulse {{
    0%,100% {{ opacity: 1; transform: scale(1); }}
    50% {{ opacity: 0.5; transform: scale(1.4); }}
  }}
  .live-dot {{
    display: inline-block;
    width: 7px; height: 7px;
    background: var(--accent-green);
    border-radius: 50%;
    margin-right: 5px;
    animation: pulse 2s ease-in-out infinite;
    vertical-align: middle;
  }}

  /* ── Tooltip badges ── */
  .tip-cid {{
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 13px; height: 13px;
    border-radius: 50%;
    background: rgba(128,128,128,0.18);
    color: #888;
    font-size: 8px;
    font-weight: 700;
    cursor: default;
    margin-left: 4px;
    vertical-align: middle;
    position: relative;
    flex-shrink: 0;
    font-style: normal;
    line-height: 1;
  }}
  .tip-cid:hover::after {{
    content: attr(data-tip);
    position: absolute;
    bottom: 130%;
    left: 50%;
    transform: translateX(-50%);
    background: #0d0d1a;
    color: #e0e0f0;
    font-size: 11px;
    font-weight: 400;
    padding: 7px 11px;
    border-radius: 7px;
    white-space: normal;
    width: 240px;
    line-height: 1.5;
    z-index: 9999;
    border: 1px solid #333355;
    box-shadow: 0 4px 16px rgba(0,0,0,0.6);
    pointer-events: none;
    text-transform: none;
    letter-spacing: normal;
    font-family: 'DM Sans', sans-serif;
  }}
</style>
</head>
<body>

<!-- ══════════════════════════════════════════════════════════════════
     HEADER
══════════════════════════════════════════════════════════════════ -->
<div class="dash-header">
  <div>
    <div class="dash-title">Community Impact Dashboard</div>
    <div class="dash-subtitle">{city}, {state} &nbsp;·&nbsp; DFR Program Transparency &amp; Public Accountability Report</div>
  </div>
  <div class="dash-meta">
    <strong>{city} Police Department</strong><br>
    Population served: {population:,}<br>
    Fleet: {actual_k_responder} Responder · {actual_k_guardian} Guardian<br>
    <span class="live-dot"></span>Simulation data
  </div>
</div>


<!-- ══════════════════════════════════════════════════════════════════
     SECTION 1 — FLIGHT HOURS & UPTIME
══════════════════════════════════════════════════════════════════ -->
<div class="section-label">01 &nbsp;·&nbsp; Flight Hours &amp; Uptime <span class="tip-cid" data-tip="Total daily airtime, annual flight hours, and dispatch frequency for the deployed drone fleet. Based on Guardian and Responder duty cycles.">?</span></div>
<div class="grid-3">

  <div class="stat-card">
    <div class="accent-bar" style="background:var(--accent-blue);"></div>
    <div class="card-label">Daily Airtime (Fleet) <span class="tip-cid" data-tip="Total hours all drones are airborne per day. Guardians run a 60-min flight / 3-min charge cycle (~22.9 hrs/day). Responders operate on an 11.6-hr patrol shift.">?</span></div>
    <div class="card-value"><span class="counter" data-target="{total_daily_flight_hrs:.1f}">{total_daily_flight_hrs:.1f}</span> hrs</div>
    <div class="card-sub">{g_count} Guardian × {GUARDIAN_FLIGHT_HOURS_PER_DAY}h &nbsp;+&nbsp; {r_count} Responder × 11.6h</div>
    <span class="card-badge" style="background:var(--accent-blue-lt);color:var(--accent-blue);">Modeled duty cycle</span>
  </div>

  <div class="stat-card">
    <div class="accent-bar" style="background:var(--accent-blue);"></div>
    <div class="card-label">Annual Flight Hours <span class="tip-cid" data-tip="Daily fleet airtime × 365 days. Represents the total drone capacity available across the full year for incident response.">?</span></div>
    <div class="card-value"><span class="counter" data-target="{annual_flight_hrs:,.0f}">{annual_flight_hrs:,.0f}</span></div>
    <div class="card-sub">Across full fleet, 365 days</div>
    <span class="card-badge" style="background:var(--accent-blue-lt);color:var(--accent-blue);">Fleet total</span>
  </div>

  <div class="stat-card">
    <div class="accent-bar" style="background:var(--accent-blue);"></div>
    <div class="card-label">DFR Flights / Day <span class="tip-cid" data-tip="Drone-First-Response dispatches per day. Calculated as: calls in coverage zone × DFR dispatch rate. The DFR rate is the fraction of 911 calls where a drone launches before a patrol car.">?</span></div>
    <div class="card-value"><span class="counter" data-target="{daily_flights:.1f}">{daily_flights:.1f}</span></div>
    <div class="card-sub">At {int(dfr_dispatch_rate*100)}% dispatch rate · {int(calls_covered_perc)}% call coverage</div>
    <span class="card-badge" style="background:var(--accent-blue-lt);color:var(--accent-blue);">{annual_flights:,.0f}/yr projected</span>
  </div>

</div>

<!-- Uptime progress bars -->
<div class="stat-card" style="margin-bottom:20px;">
  <div class="accent-bar" style="background:var(--accent-slate);"></div>
  <div class="card-label" style="margin-bottom:14px;">Guardian Fleet — Daily Uptime Breakdown <span class="tip-cid" data-tip="Shows how each Guardian drone splits its 24 hours between active flying and auto-recharging. The rapid 3-min charge cycle enables near-continuous availability.">?</span></div>
  <div class="prog-row">
    <div class="prog-meta"><span class="prog-label">Airborne (flight) <span class="tip-cid" data-tip="Hours per day each Guardian is actively flying. The 60-min flight / 3-min charge cycle yields ~22.9 hrs of airtime per Guardian per day.">?</span></span><span class="prog-val">{GUARDIAN_FLIGHT_HOURS_PER_DAY:.1f} hrs / 24 hrs</span></div>
    <div class="prog-track"><div class="prog-fill" style="width:{GUARDIAN_FLIGHT_HOURS_PER_DAY/24*100:.1f}%;background:var(--accent-blue);"></div></div>
  </div>
  <div class="prog-row">
    <div class="prog-meta"><span class="prog-label">Charging / Docked <span class="tip-cid" data-tip="Time spent in the automated recharging dock between sorties. The 3-minute recharge gap between 60-min flights is the only downtime — drone is available for re-dispatch within seconds of landing.">?</span></span><span class="prog-val">{24-GUARDIAN_FLIGHT_HOURS_PER_DAY:.1f} hrs / 24 hrs</span></div>
    <div class="prog-track"><div class="prog-fill" style="width:{(24-GUARDIAN_FLIGHT_HOURS_PER_DAY)/24*100:.1f}%;background:var(--rule);"></div></div>
  </div>
  <div class="card-sub" style="margin-top:6px;">Guardian duty cycle: {CONFIG['GUARDIAN_FLIGHT_MIN']} min flight → {CONFIG['GUARDIAN_CHARGE_MIN']} min auto-recharge → repeat</div>
</div>


<!-- ══════════════════════════════════════════════════════════════════
     SECTION 2 — RESPONSE TIME VS GROUND UNITS
══════════════════════════════════════════════════════════════════ -->
<div class="section-label">02 &nbsp;·&nbsp; Response Time vs. Ground Units <span class="tip-cid" data-tip="Compares estimated drone arrival time to typical ground unit response for incidents within the coverage zone. Drone speed, direct flight path, and instant launch give a systematic time advantage.">?</span></div>
<div class="rt-compare">
  <div class="card-label" style="margin-bottom:0;">Estimated Average Response to In-Range Incidents <span class="tip-cid" data-tip="Average distance from each station to incidents in its zone ÷ drone airspeed = flight time. Ground unit uses the same average distance but at road speed with a 1.4× road-tortuosity factor. Both are model averages.">?</span></div>
  <div class="rt-bars">
    {_resp_bar_html}
    {_guard_bar_html}
    {_ground_bar_html}
    <div class="rt-bar-wrap" style="display:flex;flex-direction:column;align-items:center;justify-content:flex-end;padding-bottom:28px;">
      <div style="font-family:'DM Mono',monospace;font-size:32px;font-weight:500;color:var(--accent-green);">−{saved_min:.1f}m</div>
      <div style="font-size:11px;color:var(--ink-light);text-align:center;margin-top:4px;">avg time saved<br>per call</div>
    </div>
  </div>
  <div>
    <span class="rt-wins-badge">✓ Drone arrives first in an estimated <strong>{drone_wins_pct}%</strong> of in-range calls</span>
    &nbsp;
    <span style="font-size:11px;color:var(--ink-light);">Based on geographic coverage ({calls_covered_perc:.1f}% call coverage) and speed advantage</span>
  </div>
</div>

<!-- Response time detail cards -->
<div class="grid-3" style="margin-bottom:20px;">
  <div class="stat-card">
    <div class="accent-bar" style="background:var(--accent-green);"></div>
    <div class="card-label">Minutes Saved / Call <span class="tip-cid" data-tip="Difference between estimated ground unit response time and drone response time for incidents in the coverage zone. Faster first eyes on scene can reduce harm, improve officer safety, and increase apprehension rates.">?</span></div>
    <div class="card-value" style="color:var(--accent-green);">{saved_min:.1f} min</div>
    <div class="card-sub">vs. estimated ground response</div>
  </div>
  <div class="stat-card">
    <div class="accent-bar" style="background:var(--accent-gold);"></div>
    <div class="card-label">Geographic Coverage <span class="tip-cid" data-tip="Percentage of the total jurisdiction area that falls within at least one drone's operational radius. Reflects spatial reach — a high percentage means most of the city has drone access, not just the dense call-volume areas.">?</span></div>
    <div class="card-value" style="color:var(--accent-gold);">{area_covered_perc:.1f}%</div>
    <div class="card-sub">of jurisdiction area within drone range</div>
  </div>
  <div class="stat-card">
    <div class="accent-bar" style="background:var(--accent-blue);"></div>
    <div class="card-label">Call Coverage <span class="tip-cid" data-tip="Percentage of historical 911 incidents (from uploaded CAD data) that fall within at least one drone's coverage zone. Higher than geographic coverage because stations are positioned near call-volume hotspots.">?</span></div>
    <div class="card-value" style="color:var(--accent-blue);">{calls_covered_perc:.1f}%</div>
    <div class="card-sub">of historical incidents in coverage zones</div>
  </div>
</div>


<!-- ══════════════════════════════════════════════════════════════════
     SECTION 3 — 4TH AMENDMENT SAFEGUARDS
══════════════════════════════════════════════════════════════════ -->
<div class="section-label">03 &nbsp;·&nbsp; Fourth Amendment &amp; Civil Liberties Safeguards <span class="tip-cid" data-tip="Plain-language summary of constitutional and policy guardrails governing every flight. These policies align the program with the 4th Amendment's protections against unreasonable search and the Baltimore Circuit ruling on mass surveillance.">?</span></div>
<div class="amend-panel">
  <div class="amend-title">
    <span>🔒</span>
    Your Rights Are Built Into This Program
  </div>
  <p style="font-size:12.5px;color:var(--ink-mid);line-height:1.6;">
    The {city} DFR program is designed in full compliance with the U.S. Constitution's Fourth Amendment and applicable state law.
    Below is a plain-language summary of the policies that govern every flight. Citizens can request program records under applicable
    open-records laws.
  </p>
  <div class="amend-grid">
    <div class="amend-item">
      <div class="amend-icon">🎯</div>
      <div>
        <div class="amend-item-title">Reactive Dispatch Only <span class="tip-cid" data-tip="Drones only launch in response to active 911 calls or officer requests. No loitering, no speculative patrol, no geofenced monitoring of specific addresses.">?</span></div>
        <div class="amend-item-desc">Drones launch in response to 911 calls and officer requests — never for proactive surveillance or random patrol.</div>
      </div>
    </div>
    <div class="amend-item">
      <div class="amend-icon">📷</div>
      <div>
        <div class="amend-item-title">In-Transit Camera Policy <span class="tip-cid" data-tip="Camera gimbal is locked forward-facing during flight to the scene. Recording and active observation only begin once the drone is confirmed on-station at the incident location — preventing incidental surveillance of bystanders en route.">?</span></div>
        <div class="amend-item-desc">Cameras remain forward-facing during transit and only orient toward a scene upon confirmed arrival at the incident location.</div>
      </div>
    </div>
    <div class="amend-item">
      <div class="amend-icon">🗑️</div>
      <div>
        <div class="amend-item-title">{retention_days}-Day Data Retention <span class="tip-cid" data-tip="All footage is automatically purged after {retention_days} days unless flagged for an active investigation. This prevents the accumulation of persistent video libraries that courts have found to constitute mass surveillance.">?</span></div>
        <div class="amend-item-desc">Footage is retained for a maximum of {retention_days} days absent evidentiary hold. No indefinite video libraries are maintained.</div>
      </div>
    </div>
    <div class="amend-item">
      <div class="amend-icon">🚫</div>
      <div>
        <div class="amend-item-title">No Facial Recognition <span class="tip-cid" data-tip="Drone video is not processed through facial recognition AI. Officers review footage manually. This avoids the bias, error rates, and warrant issues associated with automated biometric identification from aerial imagery.">?</span></div>
        <div class="amend-item-desc">This program does not integrate facial recognition technology with drone footage. Identification is performed by responding officers, not AI.</div>
      </div>
    </div>
    <div class="amend-item">
      <div class="amend-icon">⚖️</div>
      <div>
        <div class="amend-item-title">No 1st Amendment Targeting <span class="tip-cid" data-tip="Drones are prohibited from being dispatched to protests, rallies, or free-speech gatherings. Dispatch logs are auditable and would show any violation of this policy as a policy breach reviewable by the oversight board.">?</span></div>
        <div class="amend-item-desc">Drones will not be dispatched to monitor, document, or surveil lawful protest, assembly, or free-speech activities.</div>
      </div>
    </div>
    <div class="amend-item">
      <div class="amend-icon">📋</div>
      <div>
        <div class="amend-item-title">Public Flight Logs <span class="tip-cid" data-tip="Every sortie is logged: call type, GPS location, duration, and pilot/operator. Records are available to any resident under applicable public records law — providing a transparency trail that deters misuse.">?</span></div>
        <div class="amend-item-desc">Every sortie is logged with call type, location, duration, and purpose. Logs are published and available to any resident on request.</div>
      </div>
    </div>
  </div>
  <div class="amend-disclaimer">
    <strong>Legal Context:</strong> The Fourth Circuit's ruling in <em>Leaders of a Beautiful Struggle v. Baltimore</em> established that mass aerial surveillance violates the Fourth Amendment.
    This program is expressly designed to avoid that pattern: reactive dispatch only, no persistent coverage, strict data retention limits.
    Aerial observations from public navigable airspace are consistent with established Supreme Court doctrine (<em>California v. Ciraolo</em>, 1986) when conducted reactively and without advanced technology directed at private spaces.
  </div>
</div>


<!-- ══════════════════════════════════════════════════════════════════
     SECTION 4 — LIVES SAVED / OUTCOMES
══════════════════════════════════════════════════════════════════ -->
<div class="section-label">04 &nbsp;·&nbsp; Estimated Annual Community Outcomes <span class="tip-cid" data-tip="Model-based projections of public-safety outcomes derived from national DFR program benchmarks applied to this deployment's projected annual flight count. Not guarantees — actual results depend on staffing, call types, and policy.">?</span></div>
<div class="grid-4" style="margin-bottom:4px;">
  <div class="outcome-card" style="animation-delay:0.0s;">
    <span class="outcome-icon">🚔</span>
    <div class="outcome-val"><span class="counter" data-target="{arrests_est}">{arrests_est:,}</span></div>
    <div class="outcome-label">Arrest Assists <span class="tip-cid" data-tip="Estimated arrests aided by live aerial reconnaissance — suspect tracking, real-time officer guidance, perimeter confirmation. Modeled at ~4.3% of annual DFR flights per national benchmark data.">?</span></div>
    <div class="outcome-note">Aerial intel aiding officer apprehension</div>
  </div>
  <div class="outcome-card" style="animation-delay:0.1s;">
    <span class="outcome-icon">🆘</span>
    <div class="outcome-val"><span class="counter" data-target="{rescues_est}">{rescues_est:,}</span></div>
    <div class="outcome-label">Active Rescues <span class="tip-cid" data-tip="Estimated searches or medical emergencies where drone overhead view or thermal imaging directly enabled a successful rescue. Modeled at ~2.1% of annual DFR flights.">?</span></div>
    <div class="outcome-note">Missing persons, medical, extrication</div>
  </div>
  <div class="outcome-card" style="animation-delay:0.2s;">
    <span class="outcome-icon">🕊️</span>
    <div class="outcome-val"><span class="counter" data-target="{deescalation_est}">{deescalation_est:,}</span></div>
    <div class="outcome-label">De-escalations <span class="tip-cid" data-tip="Estimated incidents where real-time aerial awareness allowed officers to approach with better situational intel, reducing the likelihood of a use-of-force incident. Modeled at ~11% of annual DFR flights.">?</span></div>
    <div class="outcome-note">Drone intel prevented use-of-force</div>
  </div>
  <div class="outcome-card" style="animation-delay:0.3s;">
    <span class="outcome-icon">🔍</span>
    <div class="outcome-val"><span class="counter" data-target="{missing_est}">{missing_est:,}</span></div>
    <div class="outcome-label">Missing Person Locates <span class="tip-cid" data-tip="Estimated successful locate events using thermal imaging or wide-area overhead search. Thermal signatures allow drones to find people in the dark or in obscured terrain. Modeled at ~1.7% of annual DFR flights.">?</span></div>
    <div class="outcome-note">Thermal / overhead search assist</div>
  </div>
</div>
<p style="font-size:10.5px;color:var(--ink-light);margin-bottom:20px;font-style:italic;">
  ⚠️ Outcomes are model estimates derived from national DFR program benchmarks (arrest-assist rate ~4.3%, rescue rate ~2.1%, de-escalation rate ~11%) applied to projected annual DFR flights of {total_annual_dfr:,}.
  These are not guarantees of real-world results. Actual outcomes depend on staffing, deployment configuration, policy, and incident types.
</p>


<!-- ══════════════════════════════════════════════════════════════════
     SECTION 5 — CALL TYPE BREAKDOWN
══════════════════════════════════════════════════════════════════ -->
<div class="section-label">05 &nbsp;·&nbsp; Call Type Distribution <span class="tip-cid" data-tip="Distribution of 911 call types from your uploaded CAD data, or national DFR benchmark estimates if no data is loaded. Shows the incident mix the program will most commonly respond to.">?</span></div>
<div class="ct-panel">
  <div class="card-label" style="margin-bottom:14px;">Incident Categories in Coverage Zone <span class="tip-cid" data-tip="Each bar shows the count and share of incidents by type within the drone coverage zones. Bar width is proportional to the highest-volume category. Actual CAD data is used when uploaded; defaults to national DFR benchmarks otherwise.">?</span></div>
  <div id="callTypeBars"></div>
</div>


<!-- ══════════════════════════════════════════════════════════════════
     SECTION 6 — EQUITY NOTE
══════════════════════════════════════════════════════════════════ -->
<div class="section-label">06 &nbsp;·&nbsp; Geographic Equity &amp; Deployment Distribution <span class="tip-cid" data-tip="Documents where drones are deployed and the policy commitments that prevent disproportionate targeting of specific communities. Station locations are determined solely by call-volume density, not demographic data.">?</span></div>
<div class="amend-panel" style="border-left-color:var(--accent-gold);">
  <div class="amend-title"><span>⚖️</span> Equitable Deployment Commitment</div>
  <p style="font-size:12.5px;color:var(--ink-mid);line-height:1.6;margin-bottom:12px;">
    Research has documented that aerial surveillance can be deployed disproportionately in communities of color even when controlling for income.
    The {city} DFR program explicitly tracks deployment patterns by district to ensure equitable coverage.
  </p>
  <div style="display:flex;gap:12px;flex-wrap:wrap;">
    <div style="flex:1;min-width:180px;background:var(--bg-inset);border-radius:8px;padding:12px 14px;">
      <div style="font-size:11px;font-weight:700;color:var(--accent-gold);text-transform:uppercase;letter-spacing:0.6px;margin-bottom:6px;">Deployed Stations <span class="tip-cid" data-tip="The drone stations currently active in this deployment plan. Responder (🚁) covers a 2-mile radius; Guardian (🦅) covers up to 8 miles. Station positions are optimized for maximum call coverage.">?</span></div>
      <div id="stationList" style="font-size:11.5px;color:var(--ink-mid);line-height:1.8;"></div>
    </div>
    <div style="flex:2;min-width:200px;background:var(--bg-inset);border-radius:8px;padding:12px 14px;">
      <div style="font-size:11px;font-weight:700;color:var(--accent-gold);text-transform:uppercase;letter-spacing:0.6px;margin-bottom:6px;">Equity Safeguards <span class="tip-cid" data-tip="Policy commitments that prevent demographic bias in drone deployment. Placement is data-driven (call volume), not population-profile-driven. Audit results and complaint data are published annually.">?</span></div>
      <ul style="font-size:11.5px;color:var(--ink-mid);padding-left:16px;line-height:2.0;">
        <li>Coverage zones set by call-volume density, not demographic profile</li>
        <li>Annual deployment audit published in program transparency report</li>
        <li>No algorithmic profiling: dispatch triggered solely by 911 call</li>
        <li>Bias complaints reviewed quarterly by community oversight board</li>
      </ul>
    </div>
  </div>
</div>


<!-- ══════════════════════════════════════════════════════════════════
     SECTION 7 — TAXPAYER ROI
══════════════════════════════════════════════════════════════════ -->
<div class="section-label">07 &nbsp;·&nbsp; Taxpayer Return on Investment <span class="tip-cid" data-tip="Compares the one-time fleet hardware cost to the annual operational savings from drone-handled calls. ROI reflects savings from reduced patrol-car dispatches only — does not include specialty response, apprehension impact, or injury prevention value.">?</span></div>
<div class="roi-panel">
  <div class="roi-row">
    <div class="roi-big">
      <div class="roi-big-val">{roi_multiple:.1f}×</div>
      <div class="roi-big-label">Annual ROI multiple <span class="tip-cid" data-tip="Annual operational savings ÷ total fleet hardware cost. A 2.5× multiple means the program saves $2.50 in recurring operational costs for every $1 of one-time capital investment.">?</span></div>
      <div style="margin-top:10px;font-size:11px;color:var(--ink-light);">For every $1 invested in fleet CapEx, the program generates <strong>${roi_multiple:.2f}</strong> in annual operational savings.</div>
    </div>
    <div class="roi-details">
      <div class="roi-line">
        <span class="roi-line-label">Total Fleet CapEx <span class="tip-cid" data-tip="One-time hardware acquisition cost — Responder ($80K each) + Guardian ($160K each). Does not include subscription, maintenance, insurance, or training costs.">?</span></span>
        <span class="roi-line-val">${fleet_capex:,.0f}</span>
      </div>
      <div class="roi-line">
        <span class="roi-line-label">Annual Operational Savings <span class="tip-cid" data-tip="Savings from calls resolved without a patrol car dispatch. Formula: (daily drone-only resolved calls) × ($82 officer cost − $6 drone cost) × 365 days.">?</span></span>
        <span class="roi-line-val" style="color:var(--accent-green);">${annual_savings:,.0f}</span>
      </div>
      <div class="roi-line">
        <span class="roi-line-label">Break-Even Timeline <span class="tip-cid" data-tip="Months until cumulative operational savings fully offset the initial fleet hardware investment. Calculated as fleet CapEx ÷ monthly savings.">?</span></span>
        <span class="roi-line-val">{break_even_text}</span>
      </div>
      <div class="roi-line">
        <span class="roi-line-label">Cost per Drone Response <span class="tip-cid" data-tip="Direct per-dispatch cost comparison. Drone: ~$6 (power, maintenance amortized). Patrol officer: ~$82 (salary, vehicle, fuel, overhead). The $76 difference per resolved call drives the annual savings figure.">?</span></span>
        <span class="roi-line-val">${cost_per_call_drone} vs ${cost_per_call_officer} (patrol)</span>
      </div>
      <div class="roi-line">
        <span class="roi-line-label">Annual Calls Resolved Without Patrol Car <span class="tip-cid" data-tip="Calls where drone assessment was sufficient — no officer dispatch needed. Formula: DFR flights/day × deflection rate (% of drone-handled calls that don't escalate) × 365 days.">?</span></span>
        <span class="roi-line-val">{total_resolved_annually:,}</span>
      </div>
      <div class="roi-line" style="border-bottom:none;">
        <span class="roi-line-label">Savings Per Resolved Call <span class="tip-cid" data-tip="Net cost delta for each call resolved by drone without a patrol car: $82 (officer dispatch) − $6 (drone dispatch) = $76 saved per resolved call.">?</span></span>
        <span class="roi-line-val" style="color:var(--accent-green);">${cost_saved_per_resolved}</span>
      </div>
    </div>
  </div>
</div>


<!-- ══════════════════════════════════════════════════════════════════
     FOOTER
══════════════════════════════════════════════════════════════════ -->
<div class="dash-footer">
  <strong>Simulation Disclaimer:</strong> All figures are model estimates based on user-configured deployment parameters, national DFR benchmark rates, and uploaded CAD data.
  Response times, ROI, and outcomes are projections — not guarantees. Actual program results depend on staffing, policy, FAA authorization, and operational execution.
  This dashboard is intended for planning and community transparency purposes only. &nbsp;·&nbsp; Generated by BRINC COS Drone Optimizer.
</div>


<!-- ══════════════════════════════════════════════════════════════════
     SCRIPTS
══════════════════════════════════════════════════════════════════ -->
<script>
// ── Animated counters ──────────────────────────────────────────────────────
function animateCounter(el) {{
  const target = parseFloat(el.dataset.target.replace(/,/g,''));
  const isFloat = el.dataset.target.includes('.');
  const decimals = isFloat ? (el.dataset.target.split('.')[1] || '').length : 0;
  const duration = 1200;
  const start = performance.now();
  function tick(now) {{
    const elapsed = now - start;
    const progress = Math.min(elapsed / duration, 1);
    const ease = 1 - Math.pow(1 - progress, 3);
    const val = target * ease;
    el.textContent = isFloat
      ? val.toLocaleString('en-US', {{minimumFractionDigits:decimals, maximumFractionDigits:decimals}})
      : Math.round(val).toLocaleString('en-US');
    if (progress < 1) requestAnimationFrame(tick);
  }}
  requestAnimationFrame(tick);
}}

const obs = new IntersectionObserver(entries => {{
  entries.forEach(e => {{
    if (e.isIntersecting) {{
      animateCounter(e.target);
      obs.unobserve(e.target);
    }}
  }});
}}, {{threshold: 0.3}});
document.querySelectorAll('.counter').forEach(el => obs.observe(el));

// ── Call type horizontal bars ──────────────────────────────────────────────
const ctData = {ct_items_js};
const ctColors = [
  '#1a56db','#0d9e6e','#b45309','#be123c',
  '#7c3aed','#0369a1','#065f46','#92400e'
];
const ctContainer = document.getElementById('callTypeBars');
const maxPct = Math.max(...ctData.map(d => d.pct));
ctData.forEach((item, i) => {{
  const row = document.createElement('div');
  row.className = 'prog-row';
  row.innerHTML = `
    <div class="prog-meta">
      <span class="prog-label">${{item.label}}</span>
      <span class="prog-val">${{item.count.toLocaleString()}} &nbsp;<span style="color:#94a3b8">(${{item.pct}}%)</span></span>
    </div>
    <div class="prog-track">
      <div class="prog-fill" style="width:${{(item.pct/maxPct*100).toFixed(1)}}%;background:${{ctColors[i%ctColors.length]}};animation-delay:${{i*0.08}}s;"></div>
    </div>`;
  ctContainer.appendChild(row);
}});

// ── Station list ───────────────────────────────────────────────────────────
const stations = {drone_names_js};
const sl = document.getElementById('stationList');
if (stations.length === 0) {{
  sl.innerHTML = '<em style="color:#94a3b8;">No drones deployed yet</em>';
}} else {{
  stations.forEach(s => {{
    const icon = s.type === 'GUARDIAN' ? '🦅' : '🚁';
    const color = s.type === 'GUARDIAN' ? '#b45309' : '#1a56db';
    sl.innerHTML += `<div>${{icon}} <span style="color:${{color}};font-weight:600;">${{s.type.charAt(0)+s.type.slice(1).toLowerCase()}}</span> — ${{s.name}}</div>`;
  }});
}}
</script>
</body>
</html>"""
    return html


# ============================================================
# MAIN MAP INTERFACE
# ============================================================
if st.session_state['csvs_ready']:
    components.html("<script>window._brincHasData = true;</script>", height=0)

    df_calls = st.session_state['df_calls'].copy()
    df_calls_full = st.session_state.get('df_calls_full')
    if df_calls_full is None:
        df_calls_full = df_calls.copy()
    else:
        df_calls_full = df_calls_full.copy()
    df_stations_all = st.session_state['df_stations'].copy()
    full_total_calls = _get_annualized_calls(int(st.session_state.get('total_original_calls', len(df_calls_full) if df_calls_full is not None else len(df_calls)) or 0))
    full_daily_calls = max(1, int(full_total_calls / 365)) if full_total_calls else 1

    # ── MAP BUILD EVENT: log to sheets once per session ──────────────────────
    if not st.session_state.get('map_build_logged', False):
        try:
            _map_city  = st.session_state.get('active_city', '')
            _map_state = st.session_state.get('active_state', '')
            _brinc_raw = st.session_state.get('brinc_user', 'steven.beltran').strip()
            if not _brinc_raw: _brinc_raw = 'steven.beltran'
            _map_name  = " ".join([w.capitalize() for w in _brinc_raw.split('.')])
            _map_email = f"{_brinc_raw}@brincdrones.com"
            _map_pop   = st.session_state.get('estimated_pop', 0)
            _map_calls = st.session_state.get('total_original_calls', 0)
            _map_daily = max(1, int(_map_calls / 365))
            _session_start = st.session_state.get('session_start', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            try:
                _start_dt = datetime.datetime.strptime(_session_start, '%Y-%m-%d %H:%M:%S')
                _dur_min  = round((datetime.datetime.now() - _start_dt).total_seconds() / 60, 1)
            except Exception:
                _dur_min = ''
            _map_details = {
                'session_id':       st.session_state.get('session_id', ''),
                'session_start':    _session_start,
                'session_duration_min': _dur_min,
                'data_source':      st.session_state.get('data_source', 'unknown'),
                'population':       _map_pop,
                'total_calls':      _map_calls,
                'daily_calls':      _map_daily,
                'area_sq_mi':       0,
                'fleet_capex':      0,
                'annual_savings':   0,
                'break_even':       'N/A',
                'opt_strategy':     '',
                'dfr_rate':         st.session_state.get('dfr_rate', 0),
                'deflect_rate':     st.session_state.get('deflect_rate', 0),
                'incremental_build': False,
                'allow_redundancy': False,
                'avg_response_min': 0,
                'avg_time_saved_min': 0,
                'area_covered_pct': 0,
                'pd_chief':         st.session_state.get('pd_chief_name', ''),
                'pd_dept':          st.session_state.get('pd_dept_name', ''),
                'pd_dept_email':    st.session_state.get('pd_dept_email', ''),
                'pd_dept_phone':    st.session_state.get('pd_dept_phone', ''),
                'active_drones':    [],
            }
            _log_to_sheets(_map_city, _map_state, 'MAP_BUILD', 0, 0, 0.0,
                           _map_name, _map_email, _map_details)
            st.session_state['map_build_logged'] = True
        except Exception:
            pass

    # ── Jurisdiction boundary: use coordinate-lookup result if available,
    #    otherwise fall back to shapefile scan (demo/brinc restore paths) ──
    #    When the user enables "County Boundary" in Display Options, swap in the
    #    county-level polygon for the active city/state instead.
    _use_county    = st.session_state.get('use_county_boundary', False)
    _master_override = st.session_state.get('master_gdf_override')

    if _use_county:
        _active_state = st.session_state.get('active_state', '')
        _county_cache_key = f"{_active_state}|county"
        if (st.session_state.get('_county_boundary_cache_key') == _county_cache_key
                and st.session_state.get('_county_boundary_gdf') is not None):
            master_gdf = st.session_state['_county_boundary_gdf'].copy()
        else:
            with st.spinner("Loading county boundary…"):
                _ok, _cgdf = fetch_county_by_centroid(df_calls, _active_state)
            if _ok and _cgdf is not None:
                _cgdf = _cgdf.copy()
                _cgdf['DISPLAY_NAME'] = _cgdf['NAME'].astype(str)
                _cgdf['data_count']   = len(df_calls)
                st.session_state['_county_boundary_gdf'] = _cgdf.copy()
                st.session_state['_county_boundary_cache_key'] = _county_cache_key
                master_gdf = _cgdf.copy()
            else:
                st.warning("County boundary not found — check that counties_lite.parquet is present.")
                if _master_override is not None and not _master_override.empty:
                    master_gdf = _master_override.copy()
                else:
                    with st.spinner(get_jurisdiction_message()):
                        _preferred_shp = st.session_state.get('boundary_source_path', '') or None
                        master_gdf = find_relevant_jurisdictions(df_calls, df_stations_all, SHAPEFILE_DIR, preferred_shp=_preferred_shp)
    elif _master_override is not None and not _master_override.empty:
        master_gdf = _master_override.copy()
    else:
        with st.spinner(get_jurisdiction_message()):
            _preferred_shp = st.session_state.get('boundary_source_path', '') or None
            master_gdf = find_relevant_jurisdictions(df_calls, df_stations_all, SHAPEFILE_DIR, preferred_shp=_preferred_shp)

    _boundary_kind_note = st.session_state.get('boundary_kind', 'place')
    _boundary_src_note = st.session_state.get('boundary_source_path', '')

    if master_gdf is None or master_gdf.empty:
        # ── Fallback 1: load any saved shapefile directly (spatial join may have
        #    failed if coordinate conversion was imperfect, but the shapefile exists) ──
        shp_files = glob.glob(os.path.join(SHAPEFILE_DIR, "*.shp"))
        if shp_files:
            try:
                preferred_kind = st.session_state.get('boundary_kind', 'place')
                active_city = st.session_state.get('active_city', '')
                active_state = st.session_state.get('active_state', '')
                best = st.session_state.get('boundary_source_path', '') or None

                # Prefer exact typed boundary path first
                if not best:
                    exact = _boundary_shp_base(preferred_kind, active_city, active_state) + ".shp"
                    if os.path.exists(exact):
                        best = exact

                # Then prefer typed files whose basename matches the active city
                if not best:
                    city_key = _sanitize_boundary_token(active_city).lower()
                    typed = []
                    other = []
                    for sf in shp_files:
                        base = os.path.basename(sf).lower()
                        if base.startswith(preferred_kind + "__"):
                            typed.append(sf)
                        else:
                            other.append(sf)
                    for sf in typed + other:
                        if city_key and city_key in os.path.basename(sf).lower():
                            best = sf
                            break

                if best is None:
                    # Before falling back to shp_files[0], verify it overlaps
                    # the call coordinate bounding box — skip stale files from
                    # prior sessions that are geographically unrelated
                    _fb_lat_min = df_calls['lat'].min()
                    _fb_lat_max = df_calls['lat'].max()
                    _fb_lon_min = df_calls['lon'].min()
                    _fb_lon_max = df_calls['lon'].max()
                    _overlap_pad = 2.0  # degrees
                    for _sf_cand in shp_files:
                        try:
                            import fiona as _fiona
                            with _fiona.open(_sf_cand) as _sc:
                                _sb = _sc.bounds
                            _overlaps = not (
                                _sb[2] < _fb_lon_min - _overlap_pad or
                                _sb[0] > _fb_lon_max + _overlap_pad or
                                _sb[3] < _fb_lat_min - _overlap_pad or
                                _sb[1] > _fb_lat_max + _overlap_pad
                            )
                            if _overlaps:
                                best = _sf_cand
                                break
                        except Exception:
                            best = _sf_cand
                            break
                    # If every file failed the overlap check, skip loading —
                    # let Fallback 2 (bbox polygon) handle it cleanly
                    if best is None:
                        master_gdf = None
                        raise ValueError("No overlapping shapefiles found")

                fallback_gdf = gpd.read_file(best)
                if fallback_gdf.crs is None:
                    fallback_gdf = fallback_gdf.set_crs(epsg=4269)
                fallback_gdf = fallback_gdf.to_crs(epsg=4326)
                name_col = next((c for c in ['NAME', 'DISTRICT', 'NAMELSAD'] if c in fallback_gdf.columns), fallback_gdf.columns[0])
                fallback_gdf['DISPLAY_NAME'] = fallback_gdf[name_col].astype(str)
                fallback_gdf['data_count'] = len(df_calls)
                master_gdf = fallback_gdf[['DISPLAY_NAME', 'data_count', 'geometry']]
                st.session_state['boundary_source_path'] = best
            except Exception:
                master_gdf = None

    if master_gdf is None or master_gdf.empty:
        # ── Fallback 2: bounding box around call points ──
        min_lon, min_lat = df_calls['lon'].min(), df_calls['lat'].min()
        max_lon, max_lat = df_calls['lon'].max(), df_calls['lat'].max()
        lon_pad = (max_lon - min_lon) * 0.1
        lat_pad = (max_lat - min_lat) * 0.1
        poly = box(min_lon-lon_pad, min_lat-lat_pad, max_lon+lon_pad, max_lat+lat_pad)
        master_gdf = gpd.GeoDataFrame({'DISPLAY_NAME':['Auto-Generated Boundary'],'data_count':[len(df_calls)]}, geometry=[poly], crs="EPSG:4326")

    # --- DRAW SIDEBAR LOGO FIRST SO IT IS AT THE ABSOLUTE TOP ---
    logo_b64 = get_themed_logo_base64("logo.png", theme="dark")
    if logo_b64:
        st.sidebar.markdown(f"""
        <div style="background-color: transparent; padding: 40px 20px 10px 20px; margin: -60px -20px 20px -20px; text-align: center; pointer-events: none;">
            <img src="data:image/png;base64,{logo_b64}" style="height: 60px;">
        </div>
        """, unsafe_allow_html=True)
    else:
        st.sidebar.markdown(f"""
        <div style="background-color: transparent; padding: 40px 20px 10px 20px; margin: -60px -20px 20px -20px; text-align: center; pointer-events: none;">
            <div style="font-size:26px; font-weight:900; letter-spacing:3px; color:#ffffff;">BRINC</div>
        </div>
        """, unsafe_allow_html=True)

    st.sidebar.markdown('<div class="sidebar-section-header">① Configure</div>', unsafe_allow_html=True)
    _jur_src_file = st.session_state.get('_jur_source_file', '')
    _boundary_src_display = (
        _jur_src_file if _boundary_src_note == 'local_parquet' and _jur_src_file
        else 'local_parquet' if _boundary_src_note == 'local_parquet'
        else (_boundary_src_note.split(chr(47))[-1].split(chr(92))[-1] if _boundary_src_note else 'live lookup')
    )
    st.sidebar.caption(f"Boundary: {_boundary_kind_note} · {_boundary_src_display}")

    total_pts = master_gdf['data_count'].sum()
    master_gdf['LABEL'] = master_gdf['DISPLAY_NAME'] + " (" + (master_gdf['data_count']/total_pts*100).round(1).astype(str) + "%)"
    options_map = dict(zip(master_gdf['LABEL'], master_gdf['DISPLAY_NAME']))
    all_options = master_gdf['LABEL'].tolist()
    
    default_selection = [all_options[0]] if all_options else []
    selected_labels = st.sidebar.multiselect("Jurisdictions", options=all_options, default=default_selection,
                                             help="Select which geographic areas to include in coverage analysis.")

    _jur_debug = st.session_state.get('_jur_debug', [])
    _jur_source = next((m.split(': ')[1].split(' ')[0] for m in _jur_debug if 'parquet exists' in m and 'True' in m), None)
    if _jur_source:
        st.session_state['_jur_source_file'] = _jur_source

    if not selected_labels:
        st.warning("Please select at least one jurisdiction from the sidebar.")
        st.stop()
        
    selected_names = [options_map[l] for l in selected_labels]
    active_gdf = master_gdf[master_gdf['DISPLAY_NAME'].isin(selected_names)]
    if selected_names and st.session_state.get('active_city') == "Orlando":
        st.session_state['active_city'] = selected_names[0]

    filter_expander = st.sidebar.expander("⚙️ Data Filters", expanded=False)
    with filter_expander:
        if 'type' in df_stations_all.columns:
            all_types = sorted(df_stations_all['type'].dropna().astype(str).unique().tolist())
            if all_types:
                selected_types = st.multiselect("Facility Type", options=all_types, default=all_types,
                                                help="Filter which station types are eligible for drone deployment.")
                if not selected_types:
                    st.warning("Select at least one facility type.")
                    st.stop()
                df_stations_all = df_stations_all[df_stations_all['type'].astype(str).isin(selected_types)].copy().reset_index(drop=True)
                df_stations_all['name'] = "[" + df_stations_all['type'].astype(str) + "] " + df_stations_all['name'].astype(str)
        priority_source = df_calls_full if (df_calls_full is not None and 'priority' in df_calls_full.columns) else df_calls
        if 'priority' in priority_source.columns:
            all_priorities = sorted(pd.Series(priority_source['priority']).dropna().astype(int).unique().tolist())
            if all_priorities:
                selected_priorities = st.multiselect("Incident Priority", options=all_priorities, default=all_priorities,
                                                     help="Filter which call priorities to include in coverage scoring.")
                if not selected_priorities:
                    st.warning("Select at least one priority level.")
                    st.stop()
                df_calls = df_calls[df_calls['priority'].isin(selected_priorities)].copy().reset_index(drop=True)
                if df_calls_full is not None and 'priority' in df_calls_full.columns:
                    df_calls_full = df_calls_full[df_calls_full['priority'].isin(selected_priorities)].copy().reset_index(drop=True)

    if len(df_stations_all) == 0:
        st.error("No stations match the selected filters."); st.stop()
    if len(df_calls) == 0:
        st.error("No calls match the selected filters."); st.stop()

    disp_expander = st.sidebar.expander("👁️ Display Options", expanded=False)
    with disp_expander:
        show_boundaries = st.toggle("Jurisdiction Boundaries", value=True)
        st.toggle("County Boundary", value=False, key='use_county_boundary',
                  help="Redraw the map using the county boundary instead of the city/place boundary.")
        show_heatmap    = st.toggle("911 Call Heatmap", value=False)
        show_health     = st.toggle("Health Score", value=False)
        show_satellite  = st.toggle("Satellite Imagery", value=False)
        show_cards      = True
        simple_cards    = st.toggle("Simple Cards", value=False,
                                    help="Show a compact card with just the key numbers — name, type, response time, annual savings, and CapEx.")
        show_dots       = st.toggle("Incident Dots", value=True,
                                    help="Show individual 911 call locations as dots on the map.")
        show_faa        = st.toggle("FAA LAANC Airspace", value=False)
        simulate_traffic = st.toggle("Simulate Ground Traffic", value=False)
        traffic_level   = st.slider("Traffic Congestion", 0, 100, 40) if simulate_traffic else 40

    strat_expander = st.sidebar.expander("⚙️ Deployment Strategy", expanded=False)
    with strat_expander:
        incremental_build = st.toggle("Phased Rollout", value=True,
            help="Place drones one at a time in priority order. Disable to find the global optimum in a single pass.")
        auto_cap_dfr = st.toggle("Auto-cap over-utilized stations", value=True,
            key='auto_cap_dfr',
            help="When on, each station's DFR rate is clamped to its own physical capacity limit — "
                 "over-utilized stations run at their personal max without reducing the rate for all other stations.")

        st.markdown(f"<div style='font-size:0.7rem; color:{text_muted}; margin:8px 0 4px; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;'>Deployment Mode</div>", unsafe_allow_html=True)
        deployment_mode = st.radio(
            "Deployment Mode", 
            ("Complement — push apart", "Independent — each maximises own area", "Shared — allow full overlap"),
            index=st.session_state.get('deployment_mode_idx', 1),
            label_visibility="collapsed",
            help=(
                "Complement: Responders fill gaps left by Guardians — no wasted overlap. "
                "Independent: each fleet optimises on its own objective; overlap allowed but not forced. "
                "Shared: both fleets optimise together against the same call set — hotspot stacking."
            )
        )
        _mode_map = {"Complement — push apart": 0, "Independent — each maximises own area": 1, "Shared — allow full overlap": 2}
        st.session_state['deployment_mode_idx'] = _mode_map.get(deployment_mode, 1)

        # Derived flags used by the optimizer
        allow_redundancy  = (deployment_mode != "Complement — push apart")
        complement_mode   = (deployment_mode == "Complement — push apart")
        shared_mode       = (deployment_mode == "Shared — allow full overlap")

        st.markdown(f"<div style='font-size:0.7rem; color:{text_muted}; margin:10px 0 4px; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;'>Guardian Objective</div>", unsafe_allow_html=True)
        guard_strategy_raw = st.radio(
            "Guardian Objective",
            ("Call Coverage", "Land Coverage"),
            index=st.session_state.get('guard_strat_idx', 1),
            horizontal=True,
            label_visibility="collapsed",
            help="What the Guardian optimizer maximises. Land Coverage = wide area patrol. Call Coverage = respond to highest-volume locations."
        )
        st.session_state['guard_strat_idx'] = 0 if guard_strategy_raw == "Call Coverage" else 1
        guard_strategy = "Maximize Call Coverage" if guard_strategy_raw == "Call Coverage" else "Maximize Land Coverage"

        st.markdown(f"<div style='font-size:0.7rem; color:{text_muted}; margin:10px 0 4px; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;'>Responder Objective</div>", unsafe_allow_html=True)
        resp_strategy_raw = st.radio(
            "Responder Objective",
            ("Call Coverage", "Land Coverage"),
            index=st.session_state.get('resp_strat_idx', 0),
            horizontal=True,
            label_visibility="collapsed",
            help="What the Responder optimizer maximises. Call Coverage = densest incident areas. Land Coverage = broadest geographic reach."
        )
        st.session_state['resp_strat_idx'] = 0 if resp_strategy_raw == "Call Coverage" else 1
        resp_strategy = "Maximize Call Coverage" if resp_strategy_raw == "Call Coverage" else "Maximize Land Coverage"

        st.markdown(f"<div style='font-size:0.7rem; color:{text_muted}; margin:10px 0 4px; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;'>Coverage Ranges</div>", unsafe_allow_html=True)
        resp_radius_mi  = st.slider("🚁 Responder Range (mi)", 2.0, 3.0, float(st.session_state.get('r_resp', 2.0)), step=0.5)
        guard_radius_mi = st.slider("🦅 Guardian Range (mi) [⚡ 5mi Rapid]", 1, 8, int(st.session_state.get('r_guard', 8)), help="The 5-mile rapid response focus zone will automatically be highlighted inside the maximum perimeter.")

    # Keep opt_strategy for any code that still references it (used in export/logs)
    opt_strategy = guard_strategy  # primary strategy label for reporting

    st.sidebar.markdown('<div class="sidebar-section-header">② Optimize Fleet</div>', unsafe_allow_html=True)

    minx, miny, maxx, maxy = active_gdf.to_crs(epsg=4326).total_bounds
    center_lon = (minx + maxx) / 2
    center_lat = (miny + maxy) / 2
    dynamic_zoom = calculate_zoom(minx, maxx, miny, maxy)
    utm_zone = int((center_lon + 180) / 6) + 1
    epsg_code = int(f"326{utm_zone}") if center_lat > 0 else int(f"327{utm_zone}")

    city_m = None
    city_boundary_geom = None
    try:
        active_utm = active_gdf.to_crs(epsg=epsg_code)
        raw_union = (active_utm.geometry.union_all() if hasattr(active_utm.geometry, 'union_all')
                     else active_utm.geometry.unary_union)
        # buffer(0.1).buffer(-0.1) cleans self-intersections but can collapse thin geometries.
        # Use a larger initial buffer and validate before shrinking.
        clean_geom = raw_union.buffer(1.0).buffer(-1.0)
        if clean_geom.is_empty or not clean_geom.is_valid:
            clean_geom = raw_union.buffer(0)  # zero-buffer repair only
        if clean_geom.is_empty:
            clean_geom = raw_union          # use as-is if still empty
        city_m = clean_geom
        city_boundary_geom = gpd.GeoSeries([clean_geom], crs=epsg_code).to_crs(epsg=4326).iloc[0]
    except Exception as e:
        st.error(f"Geometry Error: {e}"); st.stop()

    # --- GEOGRAPHIC FILTERING FOR STATIONS ---
    # Keep stations inside city boundary + generous buffer.
    # If OSM found nothing inside the boundary (e.g. small cities with few public
    # buildings tagged), fall back to call-density-derived synthetic stations so
    # the tool never dead-ends on legitimate data.
    # User-uploaded station files are never silently replaced — they are trusted as-is.
    _stations_user_uploaded = st.session_state.get('stations_user_uploaded', False)
    if not df_stations_all.empty and city_m is not None:
        st_gdf = gpd.GeoDataFrame(df_stations_all,
                                   geometry=gpd.points_from_xy(df_stations_all.lon, df_stations_all.lat),
                                   crs="EPSG:4326")
        st_gdf_utm = st_gdf.to_crs(epsg=epsg_code)

        # Keep candidate sites strictly inside the jurisdiction whenever possible.
        mask = st_gdf_utm.within(city_m)
        df_inside = df_stations_all[mask].reset_index(drop=True)

        if df_inside.empty:
            if _stations_user_uploaded:
                # Try a 5 km buffer before giving up — handles county-boundary
                # mode where the active polygon differs from the original city extent.
                mask_buf = st_gdf_utm.within(city_m.buffer(5000))
                df_inside_buf = df_stations_all[mask_buf].reset_index(drop=True)
                if not df_inside_buf.empty:
                    df_stations_all = df_inside_buf
                # else: keep all uploaded stations — user knows their data
            else:
                st.info(
                    "ℹ️ No OSM public buildings were found inside the jurisdiction boundary. "
                    "Using call-density station placement — stations are snapped to incident "
                    "locations that fall inside the city limits."
                )
            try:
                if not _stations_user_uploaded:
                    df_stations_all = _make_random_stations(df_calls, n=60, boundary_geom=city_m, epsg_code=epsg_code)
            except Exception:
                df_stations_all = pd.DataFrame()

            # Absolute last resort: build a simple grid from call quantiles
            if df_stations_all.empty:
                try:
                    _lats = df_calls['lat'].dropna()
                    _lons = df_calls['lon'].dropna()
                    _grid_lats = np.linspace(_lats.quantile(0.1), _lats.quantile(0.9), 8)
                    _grid_lons = np.linspace(_lons.quantile(0.1), _lons.quantile(0.9), 8)
                    _glat, _glon = np.meshgrid(_grid_lats, _grid_lons)
                    df_stations_all = pd.DataFrame({
                        'name':  [f'Station {i+1}' for i in range(len(_glat.ravel()))],
                        'lat':   _glat.ravel(),
                        'lon':   _glon.ravel(),
                        'type':  (['Police', 'Fire', 'School'] * 30)[:len(_glat.ravel())],
                    })
                except Exception:
                    df_stations_all = pd.DataFrame()
        else:
            df_stations_all = df_inside

        if not df_stations_all.empty and not _stations_user_uploaded:
            try:
                _final_st_gdf = gpd.GeoDataFrame(df_stations_all, geometry=gpd.points_from_xy(df_stations_all.lon, df_stations_all.lat), crs="EPSG:4326").to_crs(epsg=epsg_code)
                _final_mask = _final_st_gdf.within(city_m)
                if _final_mask.any():
                    df_stations_all = df_stations_all[_final_mask].reset_index(drop=True)
            except Exception:
                pass

        if df_stations_all.empty:
            st.error(
                "⚠️ No station candidates could be generated. Please upload a CAD file "
                "with valid coordinates, or switch to Simulation mode."
            )
            st.stop()

    # ── Inject custom stations (bypass boundary clip & type filter) ──────────
    _custom_st = st.session_state.get('custom_stations', pd.DataFrame())
    if not _custom_st.empty:
        # Apply the same type-prefix rename the filter block uses, so pin lookups match
        _cst_renamed = _custom_st.copy()
        _cst_renamed['name'] = "[" + _cst_renamed['type'].astype(str) + "] " + _cst_renamed['name'].astype(str)
        # Drop columns that might not exist in df_stations_all to avoid concat issues
        _keep_cols = [c for c in _cst_renamed.columns if c in list(df_stations_all.columns) + ['name','lat','lon','type','custom']]
        _cst_renamed = _cst_renamed[_keep_cols]
        df_stations_all = pd.concat([df_stations_all, _cst_renamed], ignore_index=True)

    n = len(df_stations_all)

    # Dynamic Sliders based on Area Size
    area_sq_mi = city_m.area / 2589988.11 if city_m and not city_m.is_empty else 100.0
    r_resp_est = st.session_state.get('r_resp', 2.0)
    r_guard_est = st.session_state.get('r_guard', 8.0)
    
    max_resp_calc = min(n, int(math.ceil(area_sq_mi / (math.pi * (r_resp_est**2)))) + 5)
    max_guard_calc = min(n, int(math.ceil(area_sq_mi / (math.pi * (r_guard_est**2)))) + 5)

    # Default minimum fleet: 1 Guardian and enough Responders to reach 85% responder call coverage (minimum 2).
    try:
        _auto_sig = f"{st.session_state.get('active_city','')}|{st.session_state.get('active_state','')}|{round(area_sq_mi,1)}|{n}|{round(r_resp_est,1)}|{round(r_guard_est,1)}"
        if st.session_state.get('_auto_minimums_sig') != _auto_sig:
            _resp_default = 2
            try:
                _resp_curve = df_curve[['Drones', 'Responder (Calls)']].dropna()
                _hit = _resp_curve[_resp_curve['Responder (Calls)'] >= 85.0]
                if not _hit.empty:
                    _resp_default = int(_hit.iloc[0]['Drones'])
            except Exception:
                pass
            _resp_default = max(2, min(int(_resp_default), max(1, max_resp_calc)))
            _guard_default = max(1, min(1, max(1, max_guard_calc)))
            st.session_state['k_resp'] = max(_resp_default, len(st.session_state.get('pinned_resp_names', [])))
            st.session_state['k_guard'] = max(_guard_default, len(st.session_state.get('pinned_guard_names', [])))
            st.session_state['_auto_minimums_sig'] = _auto_sig
    except Exception:
        pass

    # Safely pull the default values without exceeding the allowed maximums
    val_r = min(st.session_state.get('k_resp', 2), max_resp_calc)
    val_g = min(st.session_state.get('k_guard', 0), max_guard_calc)

    k_responder = st.sidebar.slider("🚁 Responder Count", 0, max(1, max_resp_calc), val_r, help="Short-range tactical drones (2-3mi radius).")
    k_guardian  = st.sidebar.slider("🦅 Guardian Count", 0, max(1, max_guard_calc), val_g, help="Long-range overwatch drones (5-8mi radius).")

    st.session_state.update({'k_resp': k_responder, 'k_guard': k_guardian, 'r_resp': resp_radius_mi, 'r_guard': guard_radius_mi})
    st.sidebar.caption('Minimum fleet default: 1 Guardian + Responders to 85% call coverage (minimum 2).')

    # ── LOCK STATIONS (sidebar multiselect) ──────────────────────────────────
    _station_names = df_stations_all['name'].tolist() if not df_stations_all.empty else []
    _saved_g = [s for s in st.session_state.get('pinned_guard_names', []) if s in _station_names]
    _saved_r = [s for s in st.session_state.get('pinned_resp_names',  []) if s in _station_names]

    lock_expander = st.sidebar.expander("🔒 Lock Stations", expanded=bool(_saved_g or _saved_r))
    with lock_expander:
        st.caption("Force specific stations into the optimized solution.")
        _new_g = st.multiselect(
            "🦅 Lock as Guardian",
            options=_station_names,
            default=_saved_g,
            key="lock_guard_ms",
            help="These stations will always be assigned a Guardian drone regardless of optimizer output."
        )
        _new_r = st.multiselect(
            "🚁 Lock as Responder",
            options=[s for s in _station_names if s not in _new_g],
            default=[s for s in _saved_r if s not in _new_g],
            key="lock_resp_ms",
            help="These stations will always be assigned a Responder drone regardless of optimizer output."
        )
        if _new_g != _saved_g or _new_r != [s for s in _saved_r if s not in _new_g]:
            st.session_state['pinned_guard_names'] = _new_g
            st.session_state['pinned_resp_names']  = _new_r
            if '_opt_cache_key' in st.session_state:
                del st.session_state['_opt_cache_key']
            st.rerun()

    st.session_state['pinned_guard_names'] = _new_g
    st.session_state['pinned_resp_names']  = _new_r
    pinned_guard_names = _new_g
    pinned_resp_names  = _new_r

    # Warn if locked count exceeds slider
    if len(pinned_guard_names) > k_guardian:
        st.sidebar.warning(f"⚠️ Raise Guardian Count ≥ {len(pinned_guard_names)} to honour all Guardian locks.")
    if len(pinned_resp_names) > k_responder:
        st.sidebar.warning(f"⚠️ Raise Responder Count ≥ {len(pinned_resp_names)} to honour all Responder locks.")

    # ── ADD CUSTOM STATION BY ADDRESS ─────────────────────────────────────────
    add_expander = st.sidebar.expander("➕ Add Custom Station", expanded=False)
    with add_expander:
        st.markdown(
            f"<div style='font-size:0.7rem; color:{text_muted}; margin-bottom:8px;'>"
            "Enter an address to add a custom deployment location. "
            "The Census geocoder resolves it to lat/lon. Custom stations persist for this session.</div>",
            unsafe_allow_html=True
        )
        # Use value= (not key=) so we can clear via session_state buffer without
        # triggering the "cannot modify after instantiation" error.
        if 'cs_addr_buf'  not in st.session_state: st.session_state['cs_addr_buf']  = ""
        if 'cs_label_buf' not in st.session_state: st.session_state['cs_label_buf'] = ""
        if 'cs_type_buf'  not in st.session_state: st.session_state['cs_type_buf']  = "Police"

        _custom_addr  = st.text_input("Address", value=st.session_state['cs_addr_buf'],
                                       placeholder="123 Main St, Mobile, AL",
                                       key="custom_station_addr")
        _custom_label = st.text_input("Label (optional)", value=st.session_state['cs_label_buf'],
                                       placeholder="Fire Station 7",
                                       key="custom_station_label")
        _type_opts    = ["Police", "Fire", "School", "Government", "Hospital", "Library", "Other"]
        _type_idx     = _type_opts.index(st.session_state['cs_type_buf']) if st.session_state['cs_type_buf'] in _type_opts else 0
        _custom_type  = st.selectbox("Type", _type_opts, index=_type_idx,
                                      key="custom_station_type")

        # Explicit drone role selector — user decides, no auto-guessing
        if 'cs_role_buf' not in st.session_state: st.session_state['cs_role_buf'] = "🦅 Lock as Guardian"
        _role_opts = ["🦅 Lock as Guardian", "🚁 Lock as Responder"]
        _role_idx  = _role_opts.index(st.session_state['cs_role_buf']) if st.session_state['cs_role_buf'] in _role_opts else 0
        _custom_role = st.radio("Pin as", _role_opts, index=_role_idx,
                                horizontal=True, key="custom_station_role",
                                help="Choose which fleet this station will be locked into.")

        # Sync buffer keys from live widget values each run
        st.session_state['cs_addr_buf']  = _custom_addr
        st.session_state['cs_label_buf'] = _custom_label
        st.session_state['cs_type_buf']  = _custom_type
        st.session_state['cs_role_buf']  = _custom_role

        if st.button("📍 Geocode & Add Station", use_container_width=True, key="geocode_btn"):
            _addr_to_geocode = _custom_addr.strip()
            if _addr_to_geocode:
                try:
                    import urllib.request, urllib.parse, json as _json
                    _params = urllib.parse.urlencode({
                        "address": _addr_to_geocode,
                        "benchmark": "2020",
                        "format": "json"
                    })
                    _url = f"https://geocoding.geo.census.gov/geocoder/locations/onelineaddress?{_params}"
                    with urllib.request.urlopen(_url, timeout=8) as _resp:
                        _data = _json.loads(_resp.read().decode())
                    _matches = _data.get("result", {}).get("addressMatches", [])
                    if _matches:
                        _coords = _matches[0]["coordinates"]
                        _geo_lat = float(_coords["y"])
                        _geo_lon = float(_coords["x"])
                        _matched_addr = _matches[0].get("matchedAddress", _addr_to_geocode)
                        _label = _custom_label.strip() or _matched_addr
                        # The type-prefix rename later produces "[Type] label"
                        # Store both original and prefixed name so pin lookup works
                        _prefixed_label = f"[{_custom_type}] {_label}"
                        _new_row = pd.DataFrame([{
                            "name":   _label,          # original, pre-prefix
                            "lat":    _geo_lat,
                            "lon":    _geo_lon,
                            "type":   _custom_type,
                            "custom": True,
                        }])
                        # Store in dedicated 'custom_stations' key so boundary clip
                        # and type filter can't drop them
                        _cst = st.session_state.get('custom_stations', pd.DataFrame())
                        st.session_state['custom_stations'] = pd.concat(
                            [_cst, _new_row], ignore_index=True
                        ) if not _cst.empty else _new_row

                        # Pin using the user's explicit role choice
                        if _custom_role == "🦅 Lock as Guardian":
                            _pg = list(st.session_state.get('pinned_guard_names', []))
                            if _prefixed_label not in _pg:
                                _pg.append(_prefixed_label)
                            # Remove from responder list if it was there
                            st.session_state['pinned_resp_names'] = [
                                x for x in st.session_state.get('pinned_resp_names', [])
                                if x != _prefixed_label]
                            st.session_state['pinned_guard_names'] = _pg
                            if st.session_state.get('k_guard', 0) < len(_pg):
                                st.session_state['k_guard'] = len(_pg)
                            _pin_note = f"🦅 Pinned as Guardian."
                        else:
                            _pr = list(st.session_state.get('pinned_resp_names', []))
                            if _prefixed_label not in _pr:
                                _pr.append(_prefixed_label)
                            st.session_state['pinned_guard_names'] = [
                                x for x in st.session_state.get('pinned_guard_names', [])
                                if x != _prefixed_label]
                            st.session_state['pinned_resp_names'] = _pr
                            if st.session_state.get('k_resp', 0) < len(_pr):
                                st.session_state['k_resp'] = len(_pr)
                            _pin_note = f"🚁 Pinned as Responder."

                        st.success(f"✅ Added & locked: **{_label}** ({_geo_lat:.4f}, {_geo_lon:.4f})\n{_pin_note}")
                        st.caption(f"Matched address: {_matched_addr}")
                        # Clear buffers (role intentionally kept so user can add more of same type)
                        st.session_state['cs_addr_buf']  = ""
                        st.session_state['cs_label_buf'] = ""
                        # Clear ALL optimizer caches so the new station
                        # enters both the spatial precompute AND the LP solver
                        for _ck in ['_opt_cache_key', '_opt_best_combo',
                                    '_opt_chrono_r', '_opt_chrono_g']:
                            st.session_state.pop(_ck, None)
                        st.rerun()
                    else:
                        st.warning("⚠️ Address not found. Try including city and state — e.g. '123 Main St, Mobile AL'.")
                except Exception as _ge:
                    st.error(f"Geocoding failed: {_ge}")
            else:
                st.warning("Enter an address first.")

        # Show custom stations added this session
        _cst_display = st.session_state.get('custom_stations', pd.DataFrame())
        if not _cst_display.empty:
            _custom_added = _cst_display['name'].tolist()
        else:
            _custom_added = []
        if _custom_added:
            st.markdown(f"<div style='font-size:0.65rem; color:{text_muted}; margin-top:6px;'>"
                        f"Pinned this session:</div>", unsafe_allow_html=True)
            _pg_set = set(st.session_state.get('pinned_guard_names', []))
            _pr_set = set(st.session_state.get('pinned_resp_names', []))
            _cst_disp = st.session_state.get('custom_stations', pd.DataFrame())
            for _cn in _custom_added[:8]:
                # Check both original and prefixed name
                _cst_row = _cst_disp[_cst_disp['name'] == _cn].iloc[0] if not _cst_disp.empty and (_cst_disp['name'] == _cn).any() else None
                _pfx = f"[{_cst_row['type']}] {_cn}" if _cst_row is not None else _cn
                _is_g = _pfx in _pg_set or _cn in _pg_set
                _badge = "🦅" if _is_g else "🚁"
                _color = "#FFD700" if _is_g else "#00D2FF"
                st.markdown(f"<div style='font-size:0.65rem; color:{_color}; padding:1px 0;'>{_badge} {_pfx}</div>",
                            unsafe_allow_html=True)
            if st.button("🗑 Remove all custom stations", key="remove_custom",
                         use_container_width=True):
                _cst_to_rm = st.session_state.get('custom_stations', pd.DataFrame())
                # Build set of both original and prefixed names to un-pin
                _rm_names = set()
                if not _cst_to_rm.empty:
                    for _, _row in _cst_to_rm.iterrows():
                        _rm_names.add(str(_row['name']))
                        _rm_names.add(f"[{_row['type']}] {_row['name']}")
                st.session_state['custom_stations'] = pd.DataFrame()
                st.session_state['pinned_guard_names'] = [
                    x for x in st.session_state.get('pinned_guard_names', []) if x not in _rm_names]
                st.session_state['pinned_resp_names']  = [
                    x for x in st.session_state.get('pinned_resp_names',  []) if x not in _rm_names]
                if '_opt_cache_key' in st.session_state:
                    del st.session_state['_opt_cache_key']
                st.rerun()

    # Convert pin names → station indices for the optimizer
    _name_to_idx = {row['name']: i for i, row in df_stations_all.iterrows()}
    locked_g_pins = [_name_to_idx[n] for n in pinned_guard_names if n in _name_to_idx]
    locked_r_pins = [_name_to_idx[n] for n in pinned_resp_names  if n in _name_to_idx]

    bounds_hash = f"{minx}_{miny}_{maxx}_{maxy}_{n}_{resp_radius_mi}_{guard_radius_mi}"

    prog2 = st.sidebar.empty()
    prog2.caption(get_spatial_message())
    calls_in_city, display_calls, resp_matrix, guard_matrix, dist_matrix_r, dist_matrix_g, station_metadata, total_calls = precompute_spatial_data(
        df_calls, df_calls_full, df_stations_all, city_m, epsg_code, resp_radius_mi, guard_radius_mi, center_lat, center_lon, bounds_hash
    )
    if total_calls == 0 and len(df_calls) > 0:
        st.warning("No uploaded calls fell inside the selected jurisdiction boundary. Coverage rings can still render, but call coverage will be 0%. Check city/state selection or clean outlier coordinates in the CAD file.")
    df_curve = compute_all_elbow_curves(
        total_calls, resp_matrix, guard_matrix,
        [s['clipped_2m'] for s in station_metadata],
        [s['clipped_guard'] for s in station_metadata],
        city_m.area if city_m else 1.0, bounds_hash,
        max_stations=100
    )
    prog2.empty()

    # (Scored station table removed — station scores shown in Add Custom Station expander)

    def get_max_drones(col_name):
        series = df_curve[col_name].dropna()
        if len(series) == 0: return 1
        idx_99 = series[series >= 99.0].first_valid_index()
        fallback = series.index[-1]
        return int(df_curve.loc[idx_99 if idx_99 is not None else fallback, 'Drones'])

    with st.spinner(get_faa_message()):
        faa_geojson = load_faa_parquet(minx, miny, maxx, maxy)
    with st.spinner(get_airfield_message()):
        airfields = fetch_airfields(minx, miny, maxx, maxy)

    st.sidebar.markdown('<div class="sidebar-section-header">③ Budget & Downloads</div>', unsafe_allow_html=True)

    # We use the strat_expander we defined earlier in the sidebar to inject the sliders
    with strat_expander:
        st.markdown("---")
        inferred_daily = st.session_state.get('inferred_daily_calls_override', full_daily_calls)
        inferred_daily = max(1, int(inferred_daily))
        calls_per_day = st.slider("Total Daily Calls (citywide)", 1, max(100, inferred_daily*3), inferred_daily)
        st.caption(f"Derived from the full uploaded CAD total ({full_total_calls:,} incidents), not the optimization sample.")

        st.markdown(f"<div style='font-size:0.72rem; color:{text_muted}; margin-top:8px; margin-bottom:2px;'>DFR Dispatch Rate (%)</div>", unsafe_allow_html=True)
        st.markdown(f"<div style='font-size:0.65rem; color:#666; margin-bottom:4px;'>What % of in-range calls will the drone be sent to?</div>", unsafe_allow_html=True)
        dfr_dispatch_rate = st.slider("DFR Dispatch Rate", 1, 100, st.session_state.get('dfr_rate',12), label_visibility="collapsed") / 100.0

        st.markdown(f"<div style='font-size:0.72rem; color:{text_muted}; margin-top:8px; margin-bottom:2px;'>Calls Resolved Without Officer Dispatch (%)</div>", unsafe_allow_html=True)
        st.markdown(f"<div style='font-size:0.65rem; color:#666; margin-bottom:4px;'>Of drone-attended calls, what % close without a patrol car?</div>", unsafe_allow_html=True)
        deflection_rate = st.slider("Resolution Rate", 0, 100, st.session_state.get('deflect_rate',25), label_visibility="collapsed") / 100.0

        st.session_state['dfr_rate']    = int(dfr_dispatch_rate * 100)
        st.session_state['deflect_rate'] = int(deflection_rate * 100)

    # ── OPTIMIZATION ──────────────────────────────────────────────────
    active_resp_names, active_guard_names = [], []
    active_resp_idx, active_guard_idx = [], []  
    chrono_r, chrono_g = [], []
    best_combo = None

    _pins_key = f"{sorted(locked_g_pins)}_{sorted(locked_r_pins)}"
    opt_cache_key = f"{k_responder}_{k_guardian}_{resp_radius_mi}_{guard_radius_mi}_{guard_strategy}_{resp_strategy}_{deployment_mode}_{incremental_build}_{bounds_hash}_{_pins_key}"

    if k_responder + k_guardian > n:
        st.error("⚠️ Over-Deployment: Total drones exceed available stations.")
        active_resp_names, active_guard_names = [], []
        chrono_r, chrono_g = [], []
        best_combo = None
    elif k_responder == 0 and k_guardian == 0:
        active_resp_names, active_guard_names = [], []
        chrono_r, chrono_g = [], []
        best_combo = None
    else:
        if st.session_state.get('_opt_cache_key') != opt_cache_key:
            stage_bar = st.empty()

            # ── HELPER: greedy area-coverage for one fleet ───────────────────
            def _greedy_area(matrix, geo_list, k, forced, exclude_set):
                """Greedily pick k stations maximising unary_union area,
                starting from forced pins and skipping exclude_set."""
                chosen = list(forced)
                chrono  = list(forced)
                current_union = unary_union([geo_list[i] for i in chosen]) if chosen else None
                for _ in range(k - len(forced)):
                    best_s, best_gain = -1, -1.0
                    for s in range(len(geo_list)):
                        if s in chosen or s in exclude_set:
                            continue
                        g = geo_list[s]
                        new_area = current_union.union(g).area if current_union else g.area
                        gain = new_area - (current_union.area if current_union else 0)
                        if gain > best_gain:
                            best_gain, best_s = gain, s
                    if best_s != -1:
                        chosen.append(best_s)
                        chrono.append(best_s)
                        g = geo_list[best_s]
                        current_union = current_union.union(g) if current_union else g
                return chosen, chrono

            # ── PASS 1: Optimise Guardians independently ─────────────────────
            stage_bar.info("🦅 Optimising Guardian fleet…")
            if k_guardian > 0:
                if guard_strategy == "Maximize Call Coverage":
                    # solve_mclp returns (r_best, g_best, chrono_r, chrono_g)
                    # Pass 1 runs Guardians only (num_resp=0) so r_best=[] and g_best has the result
                    _, g_best, _, chrono_g = solve_mclp(
                        resp_matrix, guard_matrix, dist_matrix_r, dist_matrix_g,
                        0, k_guardian, True, incremental=incremental_build,
                        forced_r=[], forced_g=locked_g_pins
                    )
                else:
                    g_best, chrono_g = _greedy_area(
                        guard_matrix,
                        [station_metadata[i]['clipped_guard'] for i in range(len(station_metadata))],
                        k_guardian, locked_g_pins, set()
                    )
                g_best = list(g_best)
            else:
                g_best, chrono_g = [], []

            # ── PASS 2: Optimise Responders around Guardian result ────────────
            stage_bar.info("🚁 Optimising Responder fleet…")
            if k_responder > 0:
                # In complement mode, mask out calls already covered by Guardians
                # so Responders fill the gaps rather than stacking on the same calls.
                if complement_mode and g_best and total_calls > 0:
                    guard_covered = guard_matrix[g_best].any(axis=0)
                    # Build a reduced matrix: zero out already-covered calls for Responders
                    resp_matrix_eff = resp_matrix.copy()
                    resp_matrix_eff[:, guard_covered] = False
                    dist_matrix_r_eff = dist_matrix_r.copy()
                else:
                    resp_matrix_eff    = resp_matrix
                    dist_matrix_r_eff  = dist_matrix_r

                # In complement mode, Responders also can't reuse Guardian stations
                _excl = set(g_best) if not allow_redundancy else set()

                if resp_strategy == "Maximize Call Coverage":
                    r_best, _, chrono_r, _ = solve_mclp(
                        resp_matrix_eff, guard_matrix, dist_matrix_r_eff, dist_matrix_g,
                        k_responder, 0, allow_redundancy, incremental=incremental_build,
                        forced_r=locked_r_pins, forced_g=[]
                    )
                    # Filter out Guardian stations if complement mode
                    if complement_mode:
                        r_best = [s for s in r_best if s not in set(g_best)]
                        # Pad back to k_responder if exclusion removed some
                        if len(r_best) < k_responder:
                            remaining = [s for s in range(n)
                                         if s not in r_best and s not in set(g_best)]
                            r_best += remaining[:k_responder - len(r_best)]
                else:
                    _excl_resp = set(g_best) if complement_mode else set()
                    r_best, chrono_r = _greedy_area(
                        resp_matrix_eff,
                        [station_metadata[i]['clipped_2m'] for i in range(len(station_metadata))],
                        k_responder, locked_r_pins, _excl_resp
                    )
            else:
                r_best, chrono_r = [], []

            best_combo = (tuple(r_best), tuple(g_best))
            stage_bar.empty()
            st.toast("✅ Independent optimisation complete!", icon="✅")

            st.session_state['_opt_cache_key']  = opt_cache_key
            st.session_state['_opt_best_combo'] = best_combo
            st.session_state['_opt_chrono_r']   = chrono_r
            st.session_state['_opt_chrono_g']   = chrono_g
        else:
            best_combo = st.session_state.get('_opt_best_combo')
            chrono_r   = st.session_state.get('_opt_chrono_r', [])
            chrono_g   = st.session_state.get('_opt_chrono_g', [])

        if best_combo is not None:
            r_best, g_best = best_combo
            active_resp_names  = [station_metadata[i]['name'] for i in r_best]
            active_guard_names = [station_metadata[i]['name'] for i in g_best]
            active_resp_idx  = list(r_best)
            active_guard_idx = list(g_best)
        else:
            active_resp_names, active_guard_names = [], []
            active_resp_idx, active_guard_idx = [], []

    # ── METRICS ───────────────────────────────────────────────────────
    # ── SPLIT METRICS: Guardian and Responder computed independently ─────────
    area_covered_perc = overlap_perc = calls_covered_perc = 0.0
    guard_calls_perc  = guard_area_perc  = 0.0
    resp_calls_perc   = resp_area_perc   = 0.0
    cov_r = np.zeros(total_calls, bool) if total_calls > 0 else np.zeros(0, bool)
    cov_g = np.zeros(total_calls, bool) if total_calls > 0 else np.zeros(0, bool)

    ordered_deployments_raw = []
    for idx in chrono_g:
        if idx in active_guard_idx: ordered_deployments_raw.append((idx,'GUARDIAN'))
    for idx in chrono_r:
        if idx in active_resp_idx: ordered_deployments_raw.append((idx,'RESPONDER'))
    for idx in active_resp_idx:
        if idx not in chrono_r: ordered_deployments_raw.append((idx,'RESPONDER'))
    for idx in active_guard_idx:
        if idx not in chrono_g: ordered_deployments_raw.append((idx,'GUARDIAN'))

    active_color_map = {}
    c_idx = 0
    for idx, d_type in ordered_deployments_raw:
        key = f"{idx}_{d_type}"
        if key not in active_color_map:
            active_color_map[key] = STATION_COLORS[c_idx % len(STATION_COLORS)]
            c_idx += 1

    guard_geos = [station_metadata[i]['clipped_guard'] for i in active_guard_idx]
    resp_geos  = [station_metadata[i]['clipped_2m']    for i in active_resp_idx]
    active_geos = resp_geos + guard_geos

    city_area = city_m.area if (city_m and not city_m.is_empty) else 1.0

    # Guardian-only metrics
    if guard_geos:
        guard_area_perc = (unary_union(guard_geos).area / city_area) * 100
    if active_guard_idx and total_calls > 0:
        cov_g = guard_matrix[active_guard_idx].any(axis=0)
        guard_calls_perc = cov_g.sum() / total_calls * 100

    # Responder-only metrics
    if resp_geos:
        resp_area_perc = (unary_union(resp_geos).area / city_area) * 100
    if active_resp_idx and total_calls > 0:
        cov_r = resp_matrix[active_resp_idx].any(axis=0)
        resp_calls_perc = cov_r.sum() / total_calls * 100

    # Combined metrics
    if active_geos:
        area_covered_perc = (unary_union(active_geos).area / city_area) * 100
    if total_calls > 0:
        calls_covered_perc = (np.logical_or(cov_r, cov_g).sum() / total_calls) * 100
        st.session_state['calls_covered_perc'] = calls_covered_perc
    if len(active_geos) >= 2:
        inters = [active_geos[i].intersection(active_geos[j])
                  for i in range(len(active_geos))
                  for j in range(i+1, len(active_geos))
                  if not active_geos[i].is_empty and not active_geos[j].is_empty
                  and active_geos[i].intersects(active_geos[j])]
        if inters:
            overlap_perc = (unary_union(inters).area / city_area) * 100

    # ── BUDGET CALCULATIONS ───────────────────────────────────────────
    actual_k_responder = len(active_resp_names)
    actual_k_guardian  = len(active_guard_names)
    capex_resp  = actual_k_responder * CONFIG["RESPONDER_COST"]
    capex_guard = actual_k_guardian  * CONFIG["GUARDIAN_COST"]
    fleet_capex = capex_resp + capex_guard

    annual_savings = 0
    break_even_text = "N/A"
    daily_drone_only_calls = 0
    covered_daily_calls = 0
    daily_dfr_responses = 0

    if fleet_capex > 0:
        covered_daily_calls    = calls_per_day * (calls_covered_perc / 100.0)
        daily_dfr_responses    = covered_daily_calls * dfr_dispatch_rate
        daily_drone_only_calls = daily_dfr_responses * deflection_rate
        if daily_drone_only_calls > 0:
            monthly_savings = (CONFIG["OFFICER_COST_PER_CALL"] - CONFIG["DRONE_COST_PER_CALL"]) * daily_drone_only_calls * 30.4
            annual_savings  = monthly_savings * 12
            break_even_text = f"{fleet_capex / monthly_savings:.1f} MONTHS"

    specialty_savings = estimate_specialty_response_savings(
        st.session_state.get('df_calls_full') if st.session_state.get('df_calls_full') is not None else st.session_state.get('df_calls'),
        st.session_state.get('total_original_calls', total_calls),
        calls_covered_perc=calls_covered_perc
    )
    thermal_savings = float(specialty_savings.get('thermal_savings', 0) or 0)
    k9_savings      = float(specialty_savings.get('k9_savings', 0) or 0)
    fire_savings    = float(specialty_savings.get('fire_savings', 0) or 0)
    fire_calls_annual = float(specialty_savings.get('fire_calls_annual', 0) or 0)
    possible_additional_savings = float(specialty_savings.get('additional_savings_total', 0) or 0)

    _sidebar_annual_cap_placeholder = st.sidebar.empty()
    if fleet_capex <= 0:
        st.sidebar.info("👈 Set Responder/Guardian counts above to calculate budget impact.")

    # ── BUILD DRONE OBJECTS ───────────────────────────────────────────
    active_drones = []
    cumulative_mask = np.zeros(total_calls, dtype=bool) if total_calls > 0 else None
    step = 1
    for idx, d_type in ordered_deployments_raw:
        if d_type == 'RESPONDER':
            cov_array = resp_matrix[idx]; cost = CONFIG["RESPONDER_COST"]
            speed_mph = CONFIG["RESPONDER_SPEED"]; avg_dist = station_metadata[idx]['avg_dist_r']
            radius_m  = resp_radius_mi * 1609.34
        else:
            cov_array = guard_matrix[idx]; cost = CONFIG["GUARDIAN_COST"]
            speed_mph = CONFIG["GUARDIAN_SPEED"]; avg_dist = station_metadata[idx]['avg_dist_g']
            radius_m  = guard_radius_mi * 1609.34
        map_color    = active_color_map[f"{idx}_{d_type}"]
        avg_time_min = (avg_dist / speed_mph) * 60
        d_lat = station_metadata[idx]['lat']; d_lon = station_metadata[idx]['lon']

        _is_pinned = (d_type == 'GUARDIAN' and idx in locked_g_pins) or (d_type == 'RESPONDER' and idx in locked_r_pins)
        d = {
            'idx': idx, 'name': station_metadata[idx]['name'],
            'lat': d_lat, 'lon': d_lon, 'type': d_type, 'cost': cost,
            'cov_array': cov_array, 'color': map_color,
            'pinned': _is_pinned,
            'deploy_step': step if (idx in chrono_r or idx in chrono_g) else "MANUAL",
            'avg_time_min': avg_time_min, 'speed_mph': speed_mph, 'radius_m': radius_m,
            'faa_ceiling': get_station_faa_ceiling(d_lat, d_lon, faa_geojson),
            'nearest_airport': get_nearest_airfield(d_lat, d_lon, airfields)
        }

        if total_calls > 0 and cumulative_mask is not None:
            # ── DEDUPLICATION: track unique calls added for combined KPI totals ──
            marginal_mask     = cov_array & ~cumulative_mask
            marginal_historic = np.sum(marginal_mask)
            d['assigned_indices'] = np.where(marginal_mask)[0]
            cumulative_mask   = cumulative_mask | cov_array

            # ── RAW ZONE COVERAGE: how many calls fall in this drone's zone ───
            # Used for per-unit economics. Independent of iteration order so
            # Responders are never penalised for a Guardian claiming the same calls.
            _raw_zone_calls = int(np.sum(cov_array))  # all calls inside this drone's radius
            _raw_zone_perc  = _raw_zone_calls / total_calls

            # Shared zone: calls covered by at least one OTHER active drone
            all_cov = np.vstack([resp_matrix[i] for i in active_resp_idx] + [guard_matrix[i] for i in active_guard_idx]) if (active_resp_idx or active_guard_idx) else np.zeros((1, total_calls), dtype=bool)
            shared_mask   = d['cov_array'] & (all_cov.sum(axis=0) > 1)
            _shared_calls = int(np.sum(shared_mask))
            _excl_calls   = _raw_zone_calls - _shared_calls  # calls ONLY this drone covers

            # ── UTILIZATION: based on full zone call load, not marginal residual ─
            # Responders should reflect how busy they truly are in their patrol zone.
            # daily calls dispatched to this drone = zone calls × dispatch rate
            _is_guard    = (d_type == 'GUARDIAN')
            _budget_min  = CONFIG["GUARDIAN_DAILY_FLIGHT_MIN"] if _is_guard else (CONFIG["RESPONDER_PATROL_HOURS"] * 60)
            _zone_flights = _raw_zone_perc * calls_per_day * dfr_dispatch_rate

            # ── CAPACITY MODEL: 10-minute on-scene floor ──────────────────────
            # Every sortie consumes travel_time + on_scene_time from the daily budget.
            # We require at least 10 min on-scene so the drone isn't rushing back.
            # Deficit triggers when available on-scene time per flight drops below 10 min.
            #
            #   max_flights   = budget_min / (avg_time_min + 10)
            #   on_scene_min  = (budget_min / zone_flights) - avg_time_min   [if zone_flights > 0]
            #   deficit       = on_scene_min < 10  ↔  zone_flights > max_flights
            _MIN_SCENE_MIN   = 10.0
            _g_budget        = CONFIG["GUARDIAN_DAILY_FLIGHT_MIN"]
            _r_budget        = CONFIG["RESPONDER_PATROL_HOURS"] * 60
            _alt_is_guard    = not _is_guard   # cross-type recommendation
            _alt_budget      = _g_budget if _alt_is_guard else _r_budget
            _alt_max_single  = CONFIG["GUARDIAN_FLIGHT_MIN"] if _alt_is_guard else CONFIG["RESPONDER_FLIGHT_MIN"]

            # Capacity of THIS drone type (flights/day with 10-min scene floor)
            # Guardian is continuously airborne — a response costs round-trip travel
            # (fly TO scene + on-scene + fly BACK to patrol), so use 2×avg_time.
            # Guardian is also bounded by its duty cycle: 22.857 sorties/day,
            # each sortie fitting floor(60 / response_cost) responses max.
            import math as _math2
            _SORTIES_PER_DAY = (24 * 60) / (CONFIG["GUARDIAN_FLIGHT_MIN"] + CONFIG["GUARDIAN_CHARGE_MIN"])
            if _is_guard:
                _response_cost_g  = 2 * avg_time_min + _MIN_SCENE_MIN  # round-trip + scene
                _airtime_cap_g    = _budget_min / _response_cost_g
                _per_sortie_g     = max(1, _math2.floor(CONFIG["GUARDIAN_FLIGHT_MIN"] / _response_cost_g))
                _duty_cap_g       = _SORTIES_PER_DAY * _per_sortie_g
                _max_flights_cap  = min(_airtime_cap_g, _duty_cap_g)
            else:
                _max_flights_cap  = _budget_min / (avg_time_min + _MIN_SCENE_MIN)
            # Alternate type cap (for cross-type deficit recommendation)
            if _alt_is_guard:
                _response_cost_ag = 2 * avg_time_min + _MIN_SCENE_MIN
                _airtime_cap_ag   = _alt_budget / _response_cost_ag
                _per_sortie_ag    = max(1, _math2.floor(CONFIG["GUARDIAN_FLIGHT_MIN"] / _response_cost_ag))
                _duty_cap_ag      = _SORTIES_PER_DAY * _per_sortie_ag
                _alt_max_flights  = min(_airtime_cap_ag, _duty_cap_ag)
            else:
                _alt_max_flights  = _alt_budget / (avg_time_min + _MIN_SCENE_MIN)

            # ── Auto-cap: clamp this station's effective DFR rate to its
            #    physical capacity limit so it doesn't show a deficit while
            #    leaving every other station's rate untouched.
            _raw_demand = _raw_zone_perc * calls_per_day
            if auto_cap_dfr and _max_flights_cap > 0 and _raw_demand > 0:
                _station_max_rate = _max_flights_cap / _raw_demand
                _effective_dfr    = min(dfr_dispatch_rate, _station_max_rate)
                _zone_flights     = _raw_demand * _effective_dfr
            else:
                _effective_dfr = dfr_dispatch_rate

            # On-scene minutes available per flight given current demand
            _on_scene_min = (_budget_min / max(_zone_flights, 0.001)) - avg_time_min if _zone_flights > 0 else 99.0

            # True (uncapped) utilization using scene-inclusive budget
            _true_util = (_zone_flights * (avg_time_min + _MIN_SCENE_MIN)) / max(1.0, _budget_min)
            # Display util capped at 1.0 (100%) for progress bars; deficit shown separately
            _util = min(1.0, _true_util)

            # Deficit: flights demanded beyond physical capacity
            import math as _math
            _deficit_flights  = max(0.0, _zone_flights - _max_flights_cap)
            _has_deficit      = _deficit_flights > 0.01
            _unserv_calls_day = _deficit_flights / max(dfr_dispatch_rate, 0.01) if _has_deficit else 0.0
            _unserv_calls_yr  = _unserv_calls_day * 365

            # Extra stations needed to clear deficit (same type and alternate type)
            _extra_same = int(_math.ceil(_deficit_flights / _max_flights_cap)) if _has_deficit else 0
            _extra_alt  = int(_math.ceil(_deficit_flights / _alt_max_flights))  if _has_deficit else 0

            # CapEx cost of each resolution path
            _same_type_cost = CONFIG["GUARDIAN_COST"] if _is_guard else CONFIG["RESPONDER_COST"]
            _alt_type_cost  = CONFIG["RESPONDER_COST"] if _is_guard else CONFIG["GUARDIAN_COST"]
            _extra_same_capex = _extra_same * _same_type_cost
            _extra_alt_capex  = _extra_alt  * _alt_type_cost
            _same_type_label  = "Guardian"  if _is_guard else "Responder"
            _alt_type_label   = "Responder" if _is_guard else "Guardian"

            # ── BASE VALUE: calls uniquely covered (non-shared zone) ──────────
            # Cap savings to physically serviceable flights when in deficit
            _excl_daily        = (_excl_calls / total_calls) * calls_per_day
            _excl_flights_raw  = _excl_daily * dfr_dispatch_rate
            # Clamp exclusive flights to what the drone can actually serve
            _excl_flights      = min(_excl_flights_raw, _max_flights_cap) if _has_deficit else _excl_flights_raw
            _excl_deflected    = _excl_flights * deflection_rate
            _cost_delta        = CONFIG["OFFICER_COST_PER_CALL"] - CONFIG["DRONE_COST_PER_CALL"]
            _base_monthly      = _cost_delta * _excl_deflected * 30.4
            _base_annual       = _base_monthly * 12

            # ── CONCURRENT VALUE: shared-zone calls captured while partner is busy ─
            _shared_daily      = (_shared_calls / total_calls) * calls_per_day
            _shared_dfr        = _shared_daily * dfr_dispatch_rate
            _concurrent_daily  = _shared_dfr * _util
            # In deficit, the drone is already at or over capacity from exclusive zone
            # flights alone. Cap concurrent to whatever flight capacity remains so the
            # Annual Capacity Value reflects only calls the drone can physically service.
            if _has_deficit:
                _remaining_cap_day = max(0.0, _max_flights_cap - _excl_flights)
                _concurrent_daily  = min(_concurrent_daily, _remaining_cap_day)
            _concurrent_month  = _cost_delta * (_concurrent_daily * deflection_rate) * 30.4
            _concurrent_annual = _concurrent_month * 12

            # ── BEST CASE: base + full concurrent (partner always available) ──
            _best_monthly  = _base_monthly + _concurrent_month
            _best_annual   = _base_annual  + _concurrent_annual

            # ── STORE — use best_case as primary display value ─────────────────
            d['marginal_perc']       = marginal_historic / total_calls
            d['marginal_flights']    = _excl_flights
            d['marginal_deflected']  = _excl_deflected
            d['shared_flights']      = _shared_dfr
            d['zone_flights']        = _zone_flights
            d['zone_calls_annual']   = _raw_zone_calls
            d['zone_flights_annual'] = _zone_flights * 365.0
            d['utilization']         = _util
            d['true_util']           = _true_util
            d['on_scene_min']        = _on_scene_min
            d['max_flights_cap']     = _max_flights_cap
            d['effective_dfr_rate']  = _effective_dfr
            d['has_deficit']         = _has_deficit
            d['deficit_flights']     = _deficit_flights
            d['unserv_calls_day']    = _unserv_calls_day
            d['unserv_calls_yr']     = _unserv_calls_yr
            d['extra_same']          = _extra_same
            d['extra_alt']           = _extra_alt
            d['extra_same_capex']    = _extra_same_capex
            d['extra_alt_capex']     = _extra_alt_capex
            d['same_type_label']     = _same_type_label
            d['alt_type_label']      = _alt_type_label
            d['blocked_per_day']     = _concurrent_daily
            d['monthly_savings']     = _best_monthly
            d['annual_savings']      = _best_annual
            d['base_annual']         = _base_annual
            d['concurrent_annual']   = _concurrent_annual
            d['best_case_annual']    = _best_annual
            d['concurrent_monthly']  = _concurrent_month
            d['be_text']     = f"{d['cost']/_best_monthly:.1f} MO" if _best_monthly > 0 else "N/A"
            d['best_be_text']= d['be_text']
        else:
            d.update({'assigned_indices':[],'annual_savings':0,'marginal_flights':0,
                      'marginal_deflected':0,'shared_flights':0,'be_text':"N/A",
                      'utilization':0,'true_util':0,'on_scene_min':99,'max_flights_cap':0,
                      'has_deficit':False,'deficit_flights':0,'unserv_calls_day':0,
                      'unserv_calls_yr':0,'extra_same':0,'extra_alt':0,
                      'extra_same_capex':0,'extra_alt_capex':0,
                      'same_type_label':'Responder','alt_type_label':'Guardian',
                      'concurrent_monthly':0,'best_case_annual':0,
                      'blocked_per_day':0,'best_be_text':"N/A",'base_annual':0,
                      'concurrent_annual':0,'zone_flights':0,'zone_calls_annual':0,
                      'zone_flights_annual':0})
        active_drones.append(d)
        step += 1

    # ── RECONCILE UNIT ECONOMICS TO FLEET HEADLINE ───────────────────────
    if active_drones and annual_savings >= 0:
        _fleet_target_annual = float(max(0, annual_savings))
        _raw_total_annual = float(sum(max(0, d.get('best_case_annual', d.get('annual_savings', 0)) or 0) for d in active_drones))
        if _fleet_target_annual > 0:
            if _raw_total_annual <= 0:
                _weights = [max(0.0, float(d.get('marginal_perc', 0) or 0)) for d in active_drones]
                _w_sum = sum(_weights)
                if _w_sum <= 0:
                    _weights = [1.0 for _ in active_drones]
                    _w_sum = float(len(active_drones))
                for _d, _w in zip(active_drones, _weights):
                    _alloc_annual = _fleet_target_annual * (_w / _w_sum)
                    _alloc_monthly = _alloc_annual / 12.0
                    _d['base_annual'] = _alloc_annual
                    _d['concurrent_annual'] = 0.0
                    _d['best_case_annual'] = _alloc_annual
                    _d['annual_savings'] = _alloc_annual
                    _d['monthly_savings'] = _alloc_monthly
                    _d['concurrent_monthly'] = 0.0
                    _d['be_text'] = f"{_d['cost']/_alloc_monthly:.1f} MO" if _alloc_monthly > 0 else "N/A"
                    _d['best_be_text'] = _d['be_text']
            else:
                # Cap scale at 1.0: never inflate per-unit values above their raw
                # pre-reconciliation figures. When _fleet_target_annual exceeds
                # _raw_total_annual (low-utilisation / no-overlap case) the gap is
                # handled by the drift correction below rather than by scaling up.
                _scale = min(1.0, _fleet_target_annual / _raw_total_annual)
                for _d in active_drones:
                    _base = float(_d.get('base_annual', 0) or 0)
                    _conc = float(_d.get('concurrent_annual', 0) or 0)
                    _best = float(_d.get('best_case_annual', _d.get('annual_savings', 0)) or 0)
                    _month = float(_d.get('monthly_savings', _best / 12.0) or 0)
                    _conc_month = float(_d.get('concurrent_monthly', _conc / 12.0) or 0)

                    _d['base_annual'] = _base * _scale
                    _d['concurrent_annual'] = _conc * _scale
                    _d['best_case_annual'] = _best * _scale
                    _d['annual_savings'] = _best * _scale
                    _d['monthly_savings'] = _month * _scale
                    _d['concurrent_monthly'] = _conc_month * _scale
                    _d['be_text'] = f"{_d['cost']/_d['monthly_savings']:.1f} MO" if _d['monthly_savings'] > 0 else "N/A"
                    _d['best_be_text'] = _d['be_text']

            _reconciled_total = float(sum(max(0, d.get('annual_savings', 0) or 0) for d in active_drones))
            _drift = _fleet_target_annual - _reconciled_total
            if abs(_drift) > 0.01 and active_drones:
                _lead = max(active_drones, key=lambda x: float(x.get('annual_savings', 0) or 0))
                _lead['annual_savings']   = float(_lead.get('annual_savings',   0) or 0) + _drift
                _lead['best_case_annual'] = float(_lead.get('best_case_annual', 0) or 0) + _drift
                _lead['monthly_savings']  = float(_lead.get('monthly_savings',  0) or 0) + (_drift / 12.0)
                # Distribute drift into base first; overflow goes into concurrent so that
                # base_annual + concurrent_annual always equals best_case_annual (keeps the
                # Value Breakdown box consistent with the headline figure).
                _lead_base = float(_lead.get('base_annual', 0) or 0)
                _lead_conc = float(_lead.get('concurrent_annual', 0) or 0)
                _drift_to_base = max(-_lead_base, min(_drift, _drift))  # full drift to base …
                _new_base = _lead_base + _drift_to_base
                if _new_base < 0:                                        # … unless base would go negative
                    _drift_to_base = -_lead_base
                    _new_base = 0.0
                _drift_to_conc = _drift - _drift_to_base
                _lead['base_annual']       = _new_base
                _lead['concurrent_annual'] = max(0.0, _lead_conc + _drift_to_conc)
                _lead['be_text']      = f"{_lead['cost']/_lead['monthly_savings']:.1f} MO" if _lead['monthly_savings'] > 0 else "N/A"
                _lead['best_be_text'] = _lead['be_text']

    # ── SIDEBAR: fill Annual Capacity Value box with specialty values that match unit cards ──
    if fleet_capex > 0:
        _s_THERMAL_RATE     = float(CONFIG.get("THERMAL_DEFAULT_APPLICABLE_RATE", 0.12) or 0)
        _s_THERMAL_PER_CALL = float(CONFIG.get("THERMAL_SAVINGS_PER_CALL", 38) or 0)
        _s_K9_RATE          = float(CONFIG.get("K9_DEFAULT_APPLICABLE_RATE", 0.03) or 0)
        _s_K9_PER_CALL      = float(CONFIG.get("K9_SAVINGS_PER_CALL", 155) or 0)
        _s_FIRE_RATE        = float(CONFIG.get("FIRE_DEFAULT_APPLICABLE_RATE", 0.05) or 0)
        _s_FIRE_PER_CALL    = float(CONFIG.get("FIRE_SAVINGS_PER_CALL", 450) or 0)

        _s_thermal_total = 0.0
        _s_k9_total      = 0.0
        _s_fire_total    = 0.0
        for _sd in active_drones:
            _sd_flights  = float(_sd.get("marginal_flights", 0) or 0)
            _sd_shared   = float(_sd.get("shared_flights", 0) or 0)
            _sd_zone_calls          = float(_sd.get("zone_calls_annual", 0) or 0)
            _sd_zone_flights_annual = float(_sd.get("zone_flights_annual", (_sd_flights + _sd_shared) * 365.0) or 0)
            _sd_serviceable_annual  = float(_sd.get("max_flights_cap", 0) or 0) * 365.0
            _sd_flight_base = min(_sd_zone_flights_annual, _sd_serviceable_annual) if _sd_serviceable_annual > 0 else _sd_zone_flights_annual
            _sd_flight_base = min(_sd_flight_base, _sd_zone_calls) if _sd_zone_calls > 0 else _sd_flight_base
            _s_thermal_total += _sd_flight_base * _s_THERMAL_RATE * _s_THERMAL_PER_CALL
            _s_k9_total      += _sd_flight_base * _s_K9_RATE      * _s_K9_PER_CALL
            _s_fire_total    += _sd_flight_base * _s_FIRE_RATE    * _s_FIRE_PER_CALL

        _s_specialty_total = _s_thermal_total + _s_k9_total + _s_fire_total

        _sidebar_annual_cap_placeholder.markdown(f"""
        <div style="background:{budget_box_bg}; border:1px solid {budget_box_border}; padding:12px; border-radius:4px;
             text-align:center; margin:8px 0 12px 0; box-shadow:0 2px 5px {budget_box_shadow};">
            <div style="font-size:0.7rem; color:{text_muted}; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;">Annual Capacity Value</div>
            <div style="font-size:1.8rem; font-weight:900; color:{budget_box_border}; font-family:monospace;">${annual_savings:,.0f}</div>
            <div style="font-size:0.68rem; color:{text_muted}; margin-top:4px;">+ specialty response upside</div>
            <div style="font-size:1.05rem; font-weight:800; color:#39FF14; font-family:monospace; margin-top:2px;">${_s_specialty_total:,.0f}</div>
            <div style="display:flex; justify-content:space-between; font-size:0.68rem; margin-top:6px;">
                <span style="color:{text_muted};">🔥 Thermal response:</span>
                <span style="color:#fbbf24; font-weight:700;">${_s_thermal_total:,.0f}/yr</span>
            </div>
            <div style="display:flex; justify-content:space-between; font-size:0.68rem; margin-top:2px;">
                <span style="color:{text_muted};">🐕 K-9 replacement:</span>
                <span style="color:#39FF14; font-weight:700;">${_s_k9_total:,.0f}/yr</span>
            </div>
            <div style="display:flex; justify-content:space-between; font-size:0.68rem; margin-bottom:2px; margin-top:2px;">
                <span style="color:{text_muted};">🚒 Fire dept value:</span>
                <span style="color:#fb7121; font-weight:700;">${_s_fire_total:,.0f}/yr</span>
            </div>
            <div style="border-top:1px solid {card_border}; margin:8px 0;"></div>
            <div style="display:flex; justify-content:space-between; font-size:0.72rem; margin-bottom:3px;">
                <span style="color:{text_muted};">Calls in range:</span>
                <span style="color:{text_main}; font-weight:700;">{covered_daily_calls:.1f}/day</span>
            </div>
            <div style="display:flex; justify-content:space-between; font-size:0.72rem; margin-bottom:3px;">
                <span style="color:{text_muted};">DFR flights ({int(dfr_dispatch_rate*100)}%):</span>
                <span style="color:{text_main}; font-weight:700;">{daily_dfr_responses:.1f}/day</span>
            </div>
            <div style="display:flex; justify-content:space-between; font-size:0.72rem; margin-bottom:8px;">
                <span style="color:{text_muted};">Resolved no dispatch:</span>
                <span style="color:{text_main}; font-weight:700;">{daily_drone_only_calls:.1f}/day</span>
            </div>
            <div style="border-top:1px dashed {card_border}; margin:6px 0;"></div>
            <div style="display:flex; justify-content:space-between; font-size:0.72rem; margin-bottom:3px;">
                <span style="color:{text_muted};">Fleet CapEx:</span>
                <span style="color:{text_main}; font-weight:700;">${fleet_capex:,.0f}</span>
            </div>
            <div style="display:flex; justify-content:space-between; font-size:0.72rem;">
                <span style="color:{text_muted};">Break-even:</span>
                <span style="color:{budget_box_border}; font-weight:700;">{break_even_text}</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

    pop_metric = st.session_state.get('estimated_pop', 250000)
    grant_bracket = estimate_grants(pop_metric)
    st.sidebar.markdown(f"""
    <div style="margin-top:12px; background:{card_bg}; border:1px solid {budget_box_border}; padding:10px; border-radius:4px; margin-bottom:10px;">
        <div style="font-size:0.68rem; color:{text_muted}; font-weight:bold; text-transform:uppercase;">Est. Grant Eligibility</div>
        <div style="font-size:1.1rem; color:{budget_box_border}; font-weight:bold; font-family:monospace;">{grant_bracket}</div>
    </div>
    <div style="font-size:0.73rem; color:{text_muted}; line-height:1.5; margin-bottom:10px;">
        <a href="https://bja.ojp.gov/program/jag/overview" target="_blank" style="color:{accent_color}; font-weight:bold;">DOJ Byrne JAG</a> — UAS procurement eligible<br>
        <a href="https://www.fema.gov/grants/preparedness/homeland-security" target="_blank" style="color:{accent_color}; font-weight:bold;">FEMA HSGP</a> — CapEx offset for tactical deployments
    </div>
    """, unsafe_allow_html=True)

    st.markdown("---")

    if show_health:
        norm_redundancy = min(overlap_perc/35.0, 1.0)*100
        health_score = (calls_covered_perc*0.50) + (area_covered_perc*0.35) + (norm_redundancy*0.15)
        h_color, h_label = (accent_color,"OPTIMAL") if health_score>=80 else ("#94c11f","GOOD") if health_score>=70 else ("#ffc107","MARGINAL") if health_score>=55 else ("#dc3545","ESSENTIAL")
        st.markdown(f"""<div style="background:{card_bg}; border-left:5px solid {h_color}; border:1px solid {card_border};
            padding:10px; border-radius:4px; color:{text_main}; margin-bottom:10px;
            display:flex; align-items:center; justify-content:space-between;">
            <span style="font-size:1.4em; font-weight:bold; color:{h_color};">Department Health Score: {health_score:.1f}%</span>
            <span style="font-size:1.2em; background:rgba(128,128,128,0.15); padding:2px 10px; border-radius:4px;">{h_label}</span>
            </div>""", unsafe_allow_html=True)

    # Safely compute traffic impacts with strict float casting to prevent any TypeErrors
    if simulate_traffic:
        avg_ground_speed = float(CONFIG["DEFAULT_TRAFFIC_SPEED"]) * (1 - float(traffic_level) / 100.0)
        eval_dist  = float(guard_radius_mi if active_guard_names else resp_radius_mi)
        eval_speed = float(CONFIG["GUARDIAN_SPEED"] if active_guard_names else CONFIG["RESPONDER_SPEED"])
        
        if (active_resp_names or active_guard_names) and avg_ground_speed > 0:
            time_saved = ((eval_dist * 1.4 / avg_ground_speed) - (eval_dist / eval_speed)) * 60
            gain_val = f"{time_saved:.1f} min"
        else:
            gain_val = "N/A"
    else:
        gain_val = None

    orig_calls = int(st.session_state.get('total_original_calls', full_total_calls or (len(df_calls_full) if df_calls_full is not None else total_calls)) or total_calls)
    modeled_calls = int(st.session_state.get('total_modeled_calls', total_calls) or total_calls)
    displayed_points = len(display_calls) if display_calls is not None else 0
    call_str = f"{orig_calls:,}"

    # Calculate Date Range of CAD data (if available)
    date_range_str = "Simulated / Unknown"
    _date_src_df = df_calls_full if df_calls_full is not None else df_calls
    _label_dt = _detect_datetime_series_for_labels(_date_src_df)
    if _label_dt is not None:
        try:
            _label_dt = pd.to_datetime(_label_dt, errors='coerce').dropna()
            if not _label_dt.empty:
                min_date = _label_dt.min().strftime('%b %Y')
                max_date = _label_dt.max().strftime('%b %Y')
                date_range_str = f"{min_date} – {max_date}" if min_date != max_date else min_date
        except Exception:
            pass

    avg_resp_time = sum(d['avg_time_min'] for d in active_drones) / len(active_drones) if active_drones else 0.0
    # Keep executive-summary time-saved metric available before later export blocks.
    try:
        _avg_ground_speed_exec = float(CONFIG["DEFAULT_TRAFFIC_SPEED"]) * (1 - float(traffic_level) / 100.0)
        # Ground time = same avg distance drone covers, but at road speed with 1.4× road-tortuosity factor.
        # avg_time_min = (avg_dist / speed_mph)*60  →  avg_dist = avg_time_min * speed_mph / 60
        # ground_time  = avg_dist * 1.4 / ground_speed * 60 = avg_time_min * speed_mph * 1.4 / ground_speed
        avg_time_saved = (max(0.0, (sum(d['avg_time_min'] * d['speed_mph'] * 1.4 / _avg_ground_speed_exec for d in active_drones) / len(active_drones)) - avg_resp_time)) if active_drones and _avg_ground_speed_exec > 0 else 0.0
    except Exception:
        avg_time_saved = 0.0
    # ── Persist live deployment metrics so the apprehension table reads real values ──
    st.session_state['avg_time_saved_min'] = avg_time_saved
    st.session_state['avg_resp_time_min']  = avg_resp_time

    # 1. THE SINGLE-LINE EXECUTIVE HEADER
    logo_b64 = get_transparent_product_base64("gigs.png")
    main_logo_html = f'<img src="data:image/png;base64,{logo_b64}" style="height:32px; vertical-align:middle; margin-right:15px;">' if logo_b64 else f'<span style="font-size:1.5rem; font-weight:900; letter-spacing:2px; color:#ffffff; margin-right:15px;">BRINC</span>'

    header_html = f"""
    <div style="margin-top: 5px; margin-bottom: 15px; padding-bottom: 12px; border-bottom: 1px solid {card_border}; display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 10px;">
        <div style="display: flex; align-items: center; flex-wrap: wrap; font-size: 0.9rem;">
            <span style="color: {accent_color}; font-family: 'IBM Plex Mono', monospace; font-size: 0.8rem; letter-spacing: 1px; text-transform: uppercase; margin-right: 12px;">Strategic Deployment Plan</span>
            <span style="font-weight: 800; color: {text_main}; font-size: 1.1rem; margin-right: 12px;">{st.session_state.get('active_city', 'Unknown City')}, {st.session_state.get('active_state', 'US')}</span>
            <span style="color: {text_muted}; margin-right: 12px;">• Serving {st.session_state.get('estimated_pop', 0):,} residents across ~{int(area_sq_mi):,} sq miles</span>
        </div>
        <div style="display: flex; align-items: center; font-size: 0.85rem; color: {text_muted}; gap: 15px;">
            <span>Data Period: <span style="color:#fff;">{date_range_str}</span></span>
            <span style="color:{card_border};">|</span>
            <span style="font-weight: 800; color: {text_main}; font-size: 0.95rem;">{actual_k_responder} <span style="color:#888; font-weight:normal;">Resp</span> · {actual_k_guardian} <span style="color:#888; font-weight:normal;">Guard</span></span>
            {main_logo_html}
        </div>
    </div>
    """
    st.markdown(header_html, unsafe_allow_html=True)

    # Cleanly evaluate dynamic CSS to avoid f-string syntax errors
    border_css = 'border-right: 1px solid #222; padding-right: 10px;' if gain_val is not None else ''

    # If traffic simulation is on, nest the time saved right inside the Avg Response box!
    if gain_val is not None:
        resp_content = (
            f'<div style="font-size: 2.2rem; font-weight: 800; color: {accent_color}; font-family: \'IBM Plex Mono\', monospace; line-height: 1.1;">{avg_resp_time:.1f}m</div>'
            f'<div style="font-size: 0.7rem; color: #39FF14; font-weight: 800; text-transform: uppercase; margin-top: 4px;">▼ Saves {gain_val}</div>'
        )
    else:
        resp_content = f'<div style="font-size: 2.2rem; font-weight: 800; color: {accent_color}; font-family: \'IBM Plex Mono\', monospace;">{avg_resp_time:.1f}m</div>'

    # 2. SPLIT KPI BAR — Guardian row + Responder row + combined summary
    def _kpi_cell(label, value, color=accent_color, border=True):
        br = f"border-right: 1px solid #222; padding-right: 10px;" if border else ""
        return (
            f'<div style="{br} text-align: center;">'
            f'<div style="font-size: 0.68rem; color: {text_muted}; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom:2px;">{label}</div>'
            f'<div style="font-size: 1.9rem; font-weight: 800; color: {color}; font-family: \'IBM Plex Mono\', monospace;">{value}</div>'
            f'</div>'
        )

    _GUARD_COL = "#FFD700"   # gold for Guardian
    _RESP_COL  = "#00D2FF"   # cyan for Responder
    _COMB_COL  = "#39FF14"   # green for combined

    kpi_html = (
        # ── Row 1: summary totals ──────────────────────────────────────────
        f'<div style="background:{card_bg}; border:1px solid {card_border}; border-radius:8px; padding:16px 20px; margin-bottom:8px;">'
        f'<div style="font-size:0.65rem; color:{text_muted}; text-transform:uppercase; letter-spacing:1px; margin-bottom:10px;">Fleet Summary</div>'
        f'<div style="display:grid; grid-template-columns:repeat(5,1fr); gap:8px;">'
        + _kpi_cell("Total Incidents", call_str)
        + _kpi_cell("Combined Coverage", f"{calls_covered_perc:.1f}%", _COMB_COL)
        + _kpi_cell("Land Covered", f"{area_covered_perc:.1f}%", _COMB_COL)
        + _kpi_cell("Zone Overlap", f"{overlap_perc:.1f}%", text_muted)
        + _kpi_cell("Avg Response", f"{avg_resp_time:.1f}m", accent_color, border=False)
        + f'</div></div>'

        # ── Row 2: Guardian-specific metrics ──────────────────────────────
        + f'<div style="display:grid; grid-template-columns:1fr 1fr; gap:8px; margin-bottom:8px;">'

        + f'<div style="background:{card_bg}; border:1px solid #3a3000; border-top:3px solid {_GUARD_COL}; border-radius:8px; padding:14px 16px;">'
        + f'<div style="font-size:0.65rem; color:{_GUARD_COL}; text-transform:uppercase; letter-spacing:1px; margin-bottom:8px; font-weight:700;">🦅 Guardian Fleet — {actual_k_guardian} unit{"s" if actual_k_guardian!=1 else ""} · {guard_strategy_raw}</div>'
        + f'<div style="display:grid; grid-template-columns:1fr 1fr; gap:6px;">'
        + _kpi_cell("Call Coverage", f"{guard_calls_perc:.1f}%", _GUARD_COL)
        + _kpi_cell("Area Coverage", f"{guard_area_perc:.1f}%", _GUARD_COL, border=False)
        + f'</div></div>'

        # ── Row 3: Responder-specific metrics ─────────────────────────────
        + f'<div style="background:{card_bg}; border:1px solid #003a3a; border-top:3px solid {_RESP_COL}; border-radius:8px; padding:14px 16px;">'
        + f'<div style="font-size:0.65rem; color:{_RESP_COL}; text-transform:uppercase; letter-spacing:1px; margin-bottom:8px; font-weight:700;">🚁 Responder Fleet — {actual_k_responder} unit{"s" if actual_k_responder!=1 else ""} · {resp_strategy_raw}</div>'
        + f'<div style="display:grid; grid-template-columns:1fr 1fr; gap:6px;">'
        + _kpi_cell("Call Coverage", f"{resp_calls_perc:.1f}%", _RESP_COL)
        + _kpi_cell("Area Coverage", f"{resp_area_perc:.1f}%", _RESP_COL, border=False)
        + f'</div></div>'

        + f'</div>'
    )
    
    st.markdown(kpi_html, unsafe_allow_html=True)
    if orig_calls != modeled_calls:
        model_note = f"Optimization modeled via {modeled_calls:,} representative CAD samples from {orig_calls:,} total incidents."
    else:
        model_note = f"Optimization modeled via all {modeled_calls:,} available incidents."
    if displayed_points and displayed_points < orig_calls:
        map_note = f"Map renders {displayed_points:,} incident points using Plotly's WebGL-backed map layer for dense full-call visualization."
    elif displayed_points:
        map_note = f"Map renders all {displayed_points:,} incident points."
    else:
        map_note = ""
    full_daily_note = f"Citywide daily-call estimates are based on the full uploaded CAD total of {orig_calls:,} incidents." if orig_calls else ""
    note_bits = [model_note]
    if map_note:
        note_bits.append(map_note)
    if full_daily_note:
        note_bits.append(full_daily_note)
    st.markdown(f"<div style='font-size:0.65rem;color:gray;margin-top:-10px;margin-bottom:12px;text-align:right;'>{' '.join(note_bits)}</div>", unsafe_allow_html=True)

    overtime_stats = estimate_high_activity_overtime(
        df_calls_full if df_calls_full is not None else df_calls,
        st.session_state.get('active_state', 'TX'),
        calls_covered_perc,
        dfr_dispatch_rate,
        deflection_rate,
    )
    cards_below_map = bool(show_cards)
    map_col = st.container()

    with map_col:
        fig = go.Figure()

        if show_boundaries and city_boundary_geom is not None and not city_boundary_geom.is_empty:
            geoms_to_draw = [city_boundary_geom] if isinstance(city_boundary_geom, Polygon) else list(city_boundary_geom.geoms)
            for gi, geom in enumerate(geoms_to_draw):
                bx, by = geom.exterior.coords.xy
                fig.add_trace(go.Scattermap(mode="lines", lon=list(bx), lat=list(by),
                    line=dict(color=map_boundary_color, width=2), name="Jurisdiction Boundary",
                    hoverinfo='skip', showlegend=(gi==0)))

        if show_heatmap and not display_calls.empty:
            fig.add_trace(go.Densitymap(lat=display_calls.geometry.y, lon=display_calls.geometry.x,
                z=np.ones(len(display_calls)), radius=12, colorscale='Inferno', opacity=0.6,
                showscale=False, name="Heatmap", hoverinfo='skip'))

        if show_dots and not display_calls.empty:
            point_size = 1 if len(display_calls) > 150000 else 2 if len(display_calls) > 50000 else 3 if len(display_calls) > 20000 else 4
            point_opacity = 0.06 if len(display_calls) > 150000 else 0.10 if len(display_calls) > 50000 else 0.18 if len(display_calls) > 20000 else 0.28 if len(display_calls) > 10000 else 0.4
            # Split by agency so fire calls render red and police calls use the theme colour
            _has_agency = 'agency' in display_calls.columns
            _fire_calls   = display_calls[display_calls['agency'].str.lower() == 'fire'] if _has_agency else display_calls.iloc[0:0]
            _police_calls = display_calls[display_calls['agency'].str.lower() != 'fire'] if _has_agency else display_calls
            if not _police_calls.empty:
                fig.add_trace(go.Scattermap(lat=_police_calls.geometry.y, lon=_police_calls.geometry.x,
                    mode='markers', marker=dict(size=point_size, color=map_incident_color, opacity=point_opacity),
                    name="Police Incidents", hoverinfo='skip'))
            if not _fire_calls.empty:
                fig.add_trace(go.Scattermap(lat=_fire_calls.geometry.y, lon=_fire_calls.geometry.x,
                    mode='markers', marker=dict(size=point_size, color='#ff3b3b', opacity=point_opacity),
                    name="Fire Incidents", hoverinfo='skip'))

        if show_faa and faa_geojson:
            add_faa_laanc_layer_to_plotly(fig, faa_geojson, is_dark=not show_satellite)

        for d in active_drones:
            clats, clons = get_circle_coords(d['lat'], d['lon'], r_mi=d['radius_m']/1609.34)
            lbl = f"{d['name'].split(',')[0]} ({'Resp' if d['type']=='RESPONDER' else 'Guard'})"
            
            # Determine if this is an extended Guardian (so we can relax the outer ring)
            is_extended_guardian = (d['type'] == 'GUARDIAN' and d['radius_m']/1609.34 > 5.0)
            
            # The outer ring becomes relaxed (thinner, more transparent) if > 5 miles
            outer_width = 1.5 if is_extended_guardian else 4.5
            outer_opac = 0.4 if is_extended_guardian else 1.0
            
            fig.add_trace(go.Scattermap(
                lat=list(clats)+[None,d['lat']], lon=list(clons)+[None,d['lon']],
                mode='lines+markers',
                opacity=outer_opac,
                marker=dict(size=[0]*len(clats)+[0,20], color=d['color']),
                line=dict(color=d['color'], width=outer_width),
                fill='toself', fillcolor='rgba(0,0,0,0)', name=lbl, hoverinfo='name'))

            # The 5-mile Rapid Response ring gets the "Important" styling (thick, solid, heavier fill)
            if is_extended_guardian:
                f_lats, f_lons = get_circle_coords(d['lat'], d['lon'], r_mi=5.0)
                fig.add_trace(go.Scattermap(
                    lat=list(f_lats), lon=list(f_lons),
                    mode='lines',
                    line=dict(color=d['color'], width=4.5),
                    opacity=1.0,
                    fill='toself',
                    fillcolor=f"rgba({int(d['color'][1:3],16)},{int(d['color'][3:5],16)},{int(d['color'][5:7],16)},0.12)",
                    name=f"Rapid Response 5mi · {d['name'].split(',')[0]}",
                    hoverinfo='text',
                    text=f"⚡ Rapid Response Focus Zone — 5mi<br>{d['name'].split(',')[0]}",
                    showlegend=False
                ))

            # Star marker for manually pinned stations
            if d.get('pinned'):
                fig.add_trace(go.Scattermap(
                    lat=[d['lat']], lon=[d['lon']], mode='markers',
                    marker=dict(size=18, color=d['color'], symbol='star'),
                    name=f"📍 {d['name'].split(',')[0]} (Pinned)",
                    hovertemplate=f"<b>🔒 PINNED</b><br>{d['name']}<br>{d['type']}<extra></extra>",
                    showlegend=False
                ))

            if simulate_traffic:
                t_color = "#28a745" if traffic_level<35 else "#ffc107" if traffic_level<75 else "#dc3545"
                t_fill  = f"rgba({'40,167,69' if traffic_level<35 else '255,193,7' if traffic_level<75 else '220,53,69'}, 0.15)"
                t_label = "Light" if traffic_level<35 else "Moderate" if traffic_level<75 else "Heavy"
                gs = CONFIG["DEFAULT_TRAFFIC_SPEED"]*(1-traffic_level/100)
                if gs > 0:
                    gr_mi = (gs/60) * (d['radius_m']/1609.34/d['speed_mph'])*60
                    ga = np.linspace(0,2*np.pi,9)
                    fig.add_trace(go.Scattermap(
                        lat=list(d['lat']+(gr_mi/69.172)*np.sin(ga)),
                        lon=list(d['lon']+(gr_mi/(69.172*np.cos(np.radians(d['lat']))))*np.cos(ga)),
                        mode='lines', line=dict(color=t_color, width=2.5),
                        fill='toself', fillcolor=t_fill,
                        name=f"Ground ({t_label})", hoverinfo='skip'))

        map_cfg = dict(center=dict(lat=center_lat, lon=center_lon), zoom=dynamic_zoom, style=map_style)
        if show_satellite:
            map_cfg["style"] = "carto-positron"
            map_cfg["layers"] = [{"below":"traces","sourcetype":"raster",
                "sourceattribution":"Esri, Maxar, Earthstar Geographics",
                "source":["https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}"]}]

        fig.update_layout(uirevision="LOCKED_MAP", map=map_cfg,
            margin=dict(l=0,r=0,t=0,b=0), height=800, font=dict(size=18),
            showlegend=True,
            legend=dict(yanchor="top", y=0.98, xanchor="left", x=0.02,
                        bgcolor=legend_bg, bordercolor=accent_color, borderwidth=1,
                        font=dict(size=12, color=legend_text), itemclick="toggle"))

        st.plotly_chart(fig, use_container_width=True, config={"scrollZoom": True})

    # ── UNIT ECONOMICS CARDS (directly below map, no toggle) ─────────────────
    st.markdown("---")
    st.markdown(f"<h4 style='margin-top:2px; border-bottom:1px solid {card_border}; padding-bottom:8px; color:{text_main};'>Unit Economics <span class='tip' data-tip='Per-drone financial breakdown — annual capacity value, specialty response savings, utilization, break-even, and response time for each deployed unit. Hover each ? badge for metric definitions.'>?</span></h4>", unsafe_allow_html=True)
    st.markdown(
        f"<div style='font-size:0.6rem; color:#666; background:rgba(240,180,41,0.07); border-left:3px solid #F0B429; padding:5px 8px; border-radius:0 3px 3px 0; margin-bottom:10px;'>{SIMULATOR_DISCLAIMER_SHORT}</div>",
        unsafe_allow_html=True
    )
    st.markdown(
        """<style>
        .unit-card-grid { position: relative; }
        .unit-card {
            transition: transform 0.25s cubic-bezier(0.34, 1.56, 0.64, 1),
                        box-shadow 0.25s ease-out,
                        z-index 0s;
            position: relative;
            z-index: 1;
        }
        .unit-card:hover {
            transform: scale(1.2);
            box-shadow: 0 16px 48px rgba(0,210,255,0.28), 0 4px 16px rgba(0,0,0,0.45);
            z-index: 999;
        }
        </style>""",
        unsafe_allow_html=True
    )
    if active_drones:
        st.markdown(
            _build_unit_cards_html(
                active_drones, text_main, text_muted, card_bg, card_border,
                card_title, accent_color, columns_per_row=4,
                simple=simple_cards, deflection_rate=deflection_rate
            ),
            unsafe_allow_html=True
        )
    else:
        st.markdown(
            f"""
            <div style="background:{card_bg}; border:1px dashed {card_border}; border-radius:6px; padding:22px; text-align:center; margin-top:8px;">
                <div style="font-size:2rem; margin-bottom:8px;">🚁</div>
                <div style="font-weight:700; color:{text_main}; margin-bottom:6px;">No drones deployed yet</div>
                <div style="font-size:0.8rem; color:{text_muted};">
                    Use the <b>Responder / Guardian Count</b> sliders in the sidebar to deploy drones and see per-unit economics here.
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

    # ── COVERAGE CURVE + STATION RING CHART (side by side, directly below cards) ──
    st.markdown("---")
    st.markdown(f"<h4 style='border-bottom:1px solid {card_border}; padding-bottom:8px; color:{text_main};'>Coverage Curve <span class='tip' data-tip='Shows marginal call and area coverage as you add more Responder or Guardian drones. The curve flattens as overlap increases — use this to find the point of diminishing returns for your fleet size.'>?</span></h4>", unsafe_allow_html=True)
    st.markdown(f"<div style='font-size:0.8rem; color:{text_muted}; margin-bottom:8px;'>How added drones improve coverage — and where returns flatten.</div>", unsafe_allow_html=True)

    _curve_col, _ring_col = st.columns([3, 2], gap="medium")

    with _curve_col:
        if not df_curve.empty:
            fig_curve = go.Figure()
            for col, color, dash in [('Responder (Calls)',accent_color,'solid'),('Guardian (Calls)','#FFD700','solid'),
                                      ('Responder (Area)',accent_color,'dash'),('Guardian (Area)','#FFD700','dash')]:
                y_data = df_curve[col].dropna()
                x_data = df_curve.loc[y_data.index,'Drones']
                if not y_data.empty:
                    fig_curve.add_trace(go.Scatter(x=x_data, y=y_data, mode='lines+markers', name=col,
                        line=dict(color=color,width=2,dash=dash), marker=dict(size=4),
                        hovertemplate=f"<b>{col}</b><br>Drones: %{{x}}<br>Coverage: %{{y:.1f}}%<extra></extra>"))
                    if 'Calls' in col:
                        idx_90 = y_data[y_data >= 90.0].first_valid_index()
                        if idx_90 is not None:
                            fig_curve.add_trace(go.Scatter(x=[int(x_data.loc[idx_90])], y=[y_data.loc[idx_90]],
                                mode='markers', marker=dict(color=color,size=12,symbol='star',line=dict(color='white',width=1)),
                                showlegend=False, hoverinfo='skip'))
            fig_curve.update_layout(
                xaxis_title="Drones", yaxis_title="Coverage %",
                xaxis=dict(showgrid=True, gridcolor=card_border, tickfont=dict(color=text_muted)),
                yaxis=dict(showgrid=True, gridcolor=card_border, tickfont=dict(color=text_muted),
                           tickvals=[0,20,40,60,80,90,100], range=[0,105]),
                legend=dict(orientation="h",yanchor="bottom",y=1.02,xanchor="right",x=1,
                            font=dict(size=9,color=text_muted)),
                margin=dict(l=10,r=10,t=20,b=10), height=320,
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                hoverlabel=dict(bgcolor=card_bg, font_size=13, font_color=text_main, bordercolor=accent_color)
            )
            st.plotly_chart(fig_curve, use_container_width=True, config={'displayModeBar':False})
        else:
            st.info("Run optimization to generate coverage curve.")

    with _ring_col:
        st.markdown(
            f"<div style='font-size:0.7rem; color:{text_muted}; margin-bottom:4px;'>"
            f"Call coverage by station <span class='tip' data-tip='Donut chart showing how historical 911 calls are distributed across deployed stations. Each slice is one station&apos;s marginal (non-overlapping) call count. The center % is combined fleet coverage. Hover slices for exact counts.'>?</span></div>",
            unsafe_allow_html=True
        )
        # Split ring: outer ring = Guardians (gold), inner ring = Responders (cyan)
        if active_drones and total_calls > 0:
            _g_drones = [d for d in active_drones if d['type'] == 'GUARDIAN']
            _r_drones = [d for d in active_drones if d['type'] == 'RESPONDER']

            def _build_ring_data(drones, fleet_cov_mask):
                """Build labels/values/colors for one fleet's ring slices."""
                labels, values, colors = [], [], []
                remaining = int(fleet_cov_mask.sum()) if fleet_cov_mask is not None else 0
                for d in drones:
                    _m = int(d.get('marginal_perc', 0) * total_calls)
                    if _m > 0:
                        labels.append(d['name'].split(',')[0][:18])
                        values.append(_m)
                        colors.append(d['color'])
                        remaining = max(0, remaining - _m)
                return labels, values, colors

            _g_labels, _g_vals, _g_cols = _build_ring_data(_g_drones, cov_g)
            _r_labels, _r_vals, _r_cols = _build_ring_data(_r_drones, cov_r)

            # Uncovered slice for combined view
            _combined_covered = int(np.logical_or(cov_r, cov_g).sum()) if total_calls > 0 else 0
            _uncovered = max(0, total_calls - _combined_covered)

            # Build a single donut: Guardian slices (gold ring) + Responder slices (cyan ring)
            # separated by a small "uncovered" gap
            all_labels = _g_labels + _r_labels + (["Uncovered"] if _uncovered > 0 else [])
            all_values = _g_vals   + _r_vals   + ([_uncovered] if _uncovered > 0 else [])
            all_colors = _g_cols   + _r_cols   + (["#1a1a1a"] if _uncovered > 0 else [])

            if all_values:
                fig_ring = go.Figure(go.Pie(
                    labels=all_labels,
                    values=all_values,
                    hole=0.58,
                    marker=dict(colors=all_colors, line=dict(color='#000', width=1.5)),
                    textinfo='none',
                    hovertemplate='<b>%{label}</b><br>%{value:,} calls (%{percent})<extra></extra>',
                    sort=False,
                ))
                _cov_pct = round(_combined_covered / total_calls * 100, 1)
                _mode_short = "▶◀" if complement_mode else "↔" if shared_mode else "⊕"
                fig_ring.update_layout(
                    annotations=[dict(
                        text=f"<b>{_cov_pct}%</b><br><span style='font-size:9px'>{_mode_short} combined</span>",
                        x=0.5, y=0.5, font_size=15, showarrow=False,
                        font=dict(color=text_main)
                    )],
                    showlegend=True,
                    legend=dict(
                        orientation='v', x=1.02, y=0.5,
                        font=dict(size=9, color=text_muted),
                        bgcolor='rgba(0,0,0,0)',
                        groupclick='toggleitem',
                    ),
                    margin=dict(l=0, r=0, t=10, b=10),
                    height=320,
                    paper_bgcolor='rgba(0,0,0,0)',
                    hoverlabel=dict(bgcolor=card_bg, font_size=12, font_color=text_main),
                )
                st.plotly_chart(fig_ring, use_container_width=True, config={'displayModeBar':False})

                # Mode legend below the ring
                _mode_label = {
                    "Complement — push apart": "▶◀ Complement — Responders fill Guardian gaps",
                    "Independent — each maximises own area": "⊕ Independent — each fleet optimised separately",
                    "Shared — allow full overlap": "↔ Shared — both fleets maximise same call set",
                }.get(deployment_mode, "")
                st.markdown(
                    f"<div style='font-size:0.65rem; color:{text_muted}; text-align:center; margin-top:-8px;'>{_mode_label}</div>",
                    unsafe_allow_html=True
                )
        else:
            st.markdown(
                f"<div style='color:{text_muted}; font-size:0.8rem; padding:40px 0; text-align:center;'>Deploy drones to see call distribution ring.</div>",
                unsafe_allow_html=True
            )


    # Resolve real incident datetime coverage for labels on the stations page
    _label_dt_series = _detect_datetime_series_for_labels(df_calls_full if df_calls_full is not None else df_calls)
    _label_has_real_dates = _label_dt_series is not None and getattr(_label_dt_series, "notna", lambda: pd.Series([], dtype=bool))().sum() > 0

    # ── CAD DATA CHARTS (moved into CAD Ingestion Analytics below) ───────────
    _cad_src = st.session_state.get('data_source', '')
    _has_real_calls = _cad_src in ('cad_upload', 'brinc_file') or (
        'df_calls' in st.session_state and st.session_state['df_calls'] is not None
        and len(st.session_state['df_calls']) > 100
    )

    # ── 3D SWARM SIMULATION ───────────────────────────────────────────
    if fleet_capex > 0:
        st.markdown("---")
        st.markdown(f"<h3 style='color:{text_main};'>🚁 3D Swarm Simulation <span class='tip' data-tip='Deck.gl-powered 3D animation of all DFR flights compressed into a single 24-hour day. Each arc represents a dispatch flight from station to incident. Use the speed slider to control playback. Best viewed fullscreen for council presentations.'>?</span></h3>", unsafe_allow_html=True)
        st.markdown(f"<div style='font-size:0.82rem; color:{text_muted}; margin-bottom:10px;'>Animated deck.gl simulation of all DFR flights over a compressed 24-hour day. Use the speed slider to accelerate or slow the simulation. Great for council presentations.</div>", unsafe_allow_html=True)

        show_sim = st.toggle("🎬 Enable 3D Simulation", value=False)
        if show_sim:
            calls_lonlat = calls_in_city.to_crs(epsg=4326)
            calls_coords = np.column_stack((calls_lonlat.geometry.x, calls_lonlat.geometry.y))

            sim_assignments = {i:[] for i in range(len(active_drones))}
            for c_idx, cc in enumerate(calls_coords):
                best_d, best_dist = -1, float('inf')
                for d_idx, d in enumerate(active_drones):
                    if d['cov_array'][c_idx] if c_idx < len(d['cov_array']) else False:
                        dist = (cc[0]-d['lon'])**2 + (cc[1]-d['lat'])**2
                        if dist < best_dist:
                            best_dist, best_d = dist, d_idx
                if best_d != -1:
                    sim_assignments[best_d].append(c_idx)

            stations_json, flights_json, legend_html_sim = [], [], ""
            total_sim_flights = 0
            for d_idx, d in enumerate(active_drones):
                hex_c = d['color'].lstrip('#')
                rgb = [int(hex_c[j:j+2],16) for j in (0,2,4)]
                stations_json.append({"name":d['name'].split(',')[0][:30],"lon":d['lon'],"lat":d['lat'],"color":rgb,"radius":d['radius_m']})
                legend_html_sim += f'<div style="margin-bottom:3px;"><span style="display:inline-block;width:9px;height:9px;background:{d["color"]};margin-right:7px;border-radius:50%;"></span>{d["name"].split(",")[0][:28]} ({d["type"][:3]})</div>'
                frac = len(sim_assignments[d_idx])/len(calls_coords) if calls_coords.shape[0]>0 else 0
                monthly_for_drone = int(frac * calls_per_day * 30 * dfr_dispatch_rate)
                pool = sim_assignments[d_idx]

                if not pool: sim_calls = []
                elif monthly_for_drone > len(pool): sim_calls = random.choices(pool, k=monthly_for_drone)
                else: sim_calls = random.sample(pool, monthly_for_drone)

                total_sim_flights += len(sim_calls)
                for ci in sim_calls:
                    lon1,lat1 = calls_coords[ci]
                    lon0,lat0 = d['lon'],d['lat']
                    dist_mi = math.sqrt((lon1-lon0)**2+(lat1-lat0)**2)*69.172
                    vis_time = max((dist_mi/d['speed_mph'])*3600*8, 240)
                    launch = random.randint(0, 2592000)
                    arc_h = min(max(dist_mi*90, 80), 400)
                    t0 = launch
                    t1 = launch + vis_time * 0.15
                    t2 = launch + vis_time * 0.40
                    t3 = launch + vis_time * 0.75
                    t4 = launch + vis_time * 0.90
                    t5 = launch + vis_time
                    mx1 = lon0 + 0.15*(lon1-lon0);  my1 = lat0 + 0.15*(lat1-lat0)
                    mx2 = lon0 + 0.35*(lon1-lon0);  my2 = lat0 + 0.35*(lat1-lat0)
                    mx3 = lon0 + 0.65*(lon1-lon0);  my3 = lat0 + 0.65*(lat1-lat0)
                    mx4 = lon0 + 0.85*(lon1-lon0);  my4 = lat0 + 0.85*(lat1-lat0)
                    flights_json.append({
                        "path": [[lon0, lat0, 0], [mx1, my1, arc_h*0.75], [mx2, my2, arc_h], [mx3, my3, arc_h], [mx4, my4, arc_h*0.75], [lon1, lat1, 0]],
                        "timestamps": [t0, t1, t2, t3, t4, t5],
                        "color": rgb
                    })

            warn_html_sim = ""
            if len(flights_json) > 3000:
                flights_json = random.sample(flights_json, 3000)
                warn_html_sim = f'<div style="background:#440000;border:1px solid #ff4b4b;color:#ffbbbb;padding:5px;font-size:10px;border-radius:4px;margin-bottom:8px;">⚠️ Capped at 3,000 flights for performance (actual: {total_sim_flights:,})</div>'

            drone_svg = "data:image/svg+xml;charset=utf-8,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='white'%3E%3Cpath d='M18 6a2 2 0 100-4 2 2 0 000 4zm-12 0a2 2 0 100-4 2 2 0 000 4zm12 12a2 2 0 100-4 2 2 0 000 4zm-12 0a2 2 0 100-4 2 2 0 000 4z'/%3E%3Cpath stroke='white' stroke-width='2' stroke-linecap='round' d='M8.5 8.5l7 7m0-7l-7 7'/%3E%3Ccircle cx='12' cy='12' r='2' fill='white'/%3E%3C/svg%3E"

            sim_html = f"""<!DOCTYPE html><html><head>
            <script src="https://unpkg.com/deck.gl@8.9.35/dist.min.js"></script>
            <script src="https://unpkg.com/maplibre-gl@3.0.0/dist/maplibre-gl.js"></script>
            <link href="https://unpkg.com/maplibre-gl@3.0.0/dist/maplibre-gl.css" rel="stylesheet"/>
            <style>
              body{{margin:0;padding:0;overflow:hidden;background:#000;font-family:Manrope,sans-serif;}}
              #map{{width:100vw;height:100vh;position:absolute;}}
              #ui{{position:absolute;top:16px;left:16px;background:rgba(17,17,17,0.92);padding:16px;border-radius:8px;
                   color:white;border:1px solid #333;z-index:10;box-shadow:0 4px 10px rgba(0,0,0,0.5);width:260px;}}
              button{{background:#00D2FF;color:black;border:none;padding:10px;cursor:pointer;font-weight:bold;
                      border-radius:4px;width:100%;font-size:13px;text-transform:uppercase;margin-bottom:8px;}}
              button:disabled{{background:#444;color:#888;cursor:not-allowed;}}
              #timeDisplay{{font-family:monospace;font-size:16px;color:#00ffcc;font-weight:bold;text-align:center;margin-bottom:8px;}}
            </style></head><body>
            <div id="ui">
              <h3 style="margin:0 0 8px;color:#00D2FF;font-size:14px;">DFR SWARM SIMULATION</h3>
              {warn_html_sim}
              <div style="font-size:11px;color:#aaa;margin-bottom:10px;">
                {total_sim_flights:,} flights over 30 days at {int(dfr_dispatch_rate*100)}% dispatch rate
              </div>
              <div style="margin-bottom:10px;">
                <label style="font-size:11px;color:#ccc;">Speed: <span id="speedLabel">1</span>x</label>
                <input type="range" id="speedSlider" min="1" max="100" value="1" style="width:100%;margin-top:4px;">
              </div>
              <button id="runBtn">▶ LAUNCH SWARM</button>
              <div id="timeDisplay">00:00</div>
              <div style="margin-top:10px;border-top:1px solid #333;padding-top:8px;">
                <div style="font-size:10px;color:#888;text-transform:uppercase;margin-bottom:5px;">Stations</div>
                <div style="font-size:10px;color:#ddd;max-height:100px;overflow-y:auto;">{legend_html_sim}</div>
              </div>
            </div>
            <div id="map"></div>
            <script>
              const stations={json.dumps(stations_json)};
              const flights={json.dumps(flights_json)};
              const speedSlider=document.getElementById('speedSlider');
              const speedLabel=document.getElementById('speedLabel');
              speedSlider.oninput=()=>speedLabel.innerText=speedSlider.value;
              const map=new deck.DeckGL({{
                container:'map',
                mapStyle:'https://basemaps.cartocdn.com/gl/dark-matter-gl-style/style.json',
                initialViewState:{{longitude:{center_lon},latitude:{center_lat},zoom:{dynamic_zoom},pitch:50,bearing:0}},
                controller:true
              }});
              let time=0,timer=null,lastTime=0;
              function render(){{
                map.setProps({{layers:[
                  new deck.ScatterplotLayer({{id:'rings',data:stations,getPosition:d=>[d.lon,d.lat],
                    getFillColor:d=>[d.color[0],d.color[1],d.color[2],25],
                    getLineColor:d=>[d.color[0],d.color[1],d.color[2],220],
                    lineWidthMinPixels:2,stroked:true,filled:true,getRadius:d=>d.radius}}),
                  new deck.ScatterplotLayer({{id:'pads',data:stations,getPosition:d=>[d.lon,d.lat],
                    getFillColor:d=>[d.color[0],d.color[1],d.color[2],120],getRadius:180}}),
                  new deck.IconLayer({{id:'icons',data:stations,
                    getIcon:d=>({{url:"{drone_svg}",width:24,height:24,anchorY:12}}),
                    getPosition:d=>[d.lon,d.lat],getSize:36,sizeScale:1}}),
                  new deck.TripsLayer({{id:'flights',data:flights,getPath:d=>d.path,
                    getTimestamps:d=>d.timestamps,getColor:d=>d.color,
                    opacity:0.85,widthMinPixels:5,trailLength:13500,currentTime:time,rounded:true}}),
                  new deck.ScatterplotLayer({{id:'landed',data:flights,getPosition:d=>d.path[5],
                    getFillColor:d=>time>=d.timestamps[5]?[d.color[0],d.color[1],d.color[2],255]:[0,0,0,0],
                    getRadius:25,radiusMinPixels:3,updateTriggers:{{getFillColor:time}}}})
                ]}});
                let day=Math.floor(time/86400)+1;
                let h=Math.floor((time%86400)/3600).toString().padStart(2,'0');
                let m=Math.floor((time%3600)/60).toString().padStart(2,'0');
                document.getElementById('timeDisplay').innerText=`Day ${{day}} · ${{h}}:${{m}}`;
              }}
              const animate=()=>{{
                let now=performance.now();
                let dt=Math.min(now-lastTime,100);
                lastTime=now;
                time+=dt/1000*43200*parseFloat(speedSlider.value);
                render();
                if(time<2592000){{timer=requestAnimationFrame(animate);}}
                else{{
                  document.getElementById('runBtn').disabled=false;
                  document.getElementById('runBtn').innerText='↺ RESTART';
                  time=0;
                }}
              }};
              document.getElementById('runBtn').onclick=()=>{{
                document.getElementById('runBtn').disabled=true;
                document.getElementById('runBtn').innerText='SIMULATING…';
                time=0;lastTime=performance.now();
                if(timer)cancelAnimationFrame(timer);
                animate();
              }};
              render();
            </script></body></html>"""

            components.html(sim_html, height=700)

    # ── COMMAND CENTER ANALYTICS DASHBOARD ──
    st.markdown("---")
    st.markdown(f"<h3 style='color:{text_main};'>📊 CAD Ingestion Analytics <span class='tip' data-tip='Temporal analysis of your uploaded CAD (Computer-Aided Dispatch) data. Shows when calls are most frequent by hour and day, identifies optimal DFR shift windows, and renders a call-volume calendar heatmap.'>?</span></h3>", unsafe_allow_html=True)
    st.markdown(f"<div style='font-size:0.82rem; color:{text_muted}; margin-bottom:10px;'>Temporal patterns derived from your uploaded CAD data — hourly volumes, day-of-week distribution, optimal DFR shift windows, and a higher-contrast 5-band call-volume calendar.</div>", unsafe_allow_html=True)

    _analytics_df = df_calls_full if (df_calls_full is not None and not df_calls_full.empty) else df_calls
    analytics_html_block = generate_command_center_html(
        _analytics_df,
        total_orig_calls=st.session_state.get('total_original_calls', full_total_calls or total_calls)
    )
    _analytics_unavailable = (
        "Analytics unavailable." in analytics_html_block
        or "No valid dates found in data." in analytics_html_block
    )
    if _analytics_unavailable:
        _analytics_height = 180
    else:
        # Compute height from actual data so the iframe fits exactly with no dead space.
        # Calendar grid is auto-fill / minmax(250px, 1fr). At typical Streamlit content
        # width (~900px with sidebar open) that yields 3 columns.
        import math as _math
        try:
            _n_months = int(_analytics_df['date'].astype(str).str[:7].nunique()) if (
                _analytics_df is not None and not _analytics_df.empty and 'date' in _analytics_df.columns
            ) else 6
        except Exception:
            _n_months = 6
        _n_months = max(1, min(_n_months, 12))
        _cal_cols = 3                        # columns at typical sidebar-open viewport
        _cal_rows = _math.ceil(_n_months / _cal_cols)
        _cal_px   = _cal_rows * 260          # ~260px per calendar row (tightened)
        # Fixed chrome above the calendar:
        #   section header 60 + controls bar 70 + KPI cards 110 + shift/dow panel 210 + legend+label 55
        _fixed_px = 460
        _analytics_height = _fixed_px + _cal_px
    components.html(analytics_html_block, height=_analytics_height, scrolling=False)

    if _analytics_unavailable:
        # Remove the dead gap when the analytics component only contains a short fallback message.
        st.markdown("<div style='margin-top:-6px;'></div>", unsafe_allow_html=True)
    elif _has_real_calls and _analytics_df is not None and not _analytics_df.empty:
        # Collapse gap between components.html block and the plotly charts below
        st.markdown("<div style='margin-top:-80px;'></div>", unsafe_allow_html=True)
        _build_cad_charts(_analytics_df, text_main, text_muted, card_bg, card_border, accent_color)

    # ── COMMUNITY IMPACT DASHBOARD ────────────────────────────────────────────
    st.markdown("---")
    st.markdown(
        f"<h3 style='color:{text_main};'>🏛️ Community Impact Dashboard <span class='tip' data-tip='Public-facing transparency report for city council presentations and community portals. Hover the ? badges inside each section for detailed explanations of every metric.'>?</span></h3>",
        unsafe_allow_html=True
    )
    st.markdown(
        f"<div style='font-size:0.82rem; color:{text_muted}; margin-bottom:10px;'>"
        "Public-facing transparency report — flight hours &amp; uptime, response time advantage, "
        "Fourth Amendment safeguards, community outcomes, call type distribution, equity commitments, "
        "and taxpayer ROI. Designed for city council presentations and citizen engagement portals."
        "</div>",
        unsafe_allow_html=True
    )
    _cid_html = generate_community_impact_dashboard_html(
        city=st.session_state.get('active_city', 'City'),
        state=st.session_state.get('active_state', 'TX'),
        population=int(st.session_state.get('estimated_pop', 65000) or 65000),
        total_calls=int(st.session_state.get('total_original_calls', full_total_calls or total_calls) or 0),
        calls_covered_perc=float(calls_covered_perc or 0),
        area_covered_perc=float(area_covered_perc or 0),
        avg_resp_time_min=float(avg_resp_time or 0),
        avg_time_saved_min=float(avg_time_saved or 0),
        fleet_capex=float(fleet_capex or 0),
        annual_savings=float(annual_savings or 0),
        break_even_text=str(break_even_text or 'N/A'),
        actual_k_responder=int(actual_k_responder or 0),
        actual_k_guardian=int(actual_k_guardian or 0),
        dfr_dispatch_rate=float(dfr_dispatch_rate or 0.25),
        deflection_rate=float(deflection_rate or 0.30),
        daily_dfr_responses=float(daily_dfr_responses or 0),
        daily_drone_only_calls=float(daily_drone_only_calls or 0),
        active_drones=active_drones or [],
        df_calls_full=df_calls_full,
    )
    components.html(_cid_html, height=5200, scrolling=False)

    # ── EXPORT BUTTONS — always visible in sidebar ──
    st.sidebar.markdown("---")

    brinc_user = st.sidebar.text_input("BRINC Email Prefix (first.last)", value=st.session_state.get('brinc_user', 'steven.beltran'), key='brinc_user', help="Enter 'first.last' to auto-generate your name and @brincdrones.com email address.")
    st.sidebar.caption("*(Press **Enter** after typing to apply changes)*")


    user_clean = brinc_user.strip()
    if not user_clean: user_clean = "steven.beltran"

    # Police dept fields removed from sidebar — read from session state defaults
    pd_chief_name = st.session_state.get('pd_chief_name', '')
    pd_dept_name  = st.session_state.get('pd_dept_name', '')
    pd_dept_email = st.session_state.get('pd_dept_email', '')
    pd_dept_phone = st.session_state.get('pd_dept_phone', '')
    prop_email = f"{user_clean}@brincdrones.com"
    prop_name = " ".join([word.capitalize() for word in user_clean.split('.')])

    # Always define these so download buttons work regardless of fleet_capex
    prop_city  = st.session_state.get('active_city', 'City')
    prop_state = st.session_state.get('active_state', 'FL')
    _safe_city_base = prop_city.replace(" ", "_").replace("/", "_")

    if fleet_capex > 0:

        prop_city  = st.session_state.get('active_city', 'City')
        prop_state = st.session_state.get('active_state', 'FL')

        # Police dept signatory — falls back gracefully if not filled in
        pd_chief  = pd_chief_name.strip()  if pd_chief_name.strip()  else f"Chief of Police, {prop_city}"
        pd_dept   = pd_dept_name.strip()   if pd_dept_name.strip()   else f"{prop_city} Police Department"
        pd_email  = pd_dept_email.strip()  if pd_dept_email.strip()  else ""
        pd_phone  = pd_dept_phone.strip()  if pd_dept_phone.strip()  else ""
        pd_email_html = f'📧 <a href="mailto:{pd_email}">{pd_email}</a><br>' if pd_email else ""
        pd_phone_html = f'📞 {pd_phone}<br>' if pd_phone else ""
        pop_metric = st.session_state.get('estimated_pop', 250000)
        current_time_str = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        safe_city = prop_city.replace(" ","_").replace("/","_")

        _session_start = st.session_state.get('session_start', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        try:
            _start_dt = datetime.datetime.strptime(_session_start, '%Y-%m-%d %H:%M:%S')
            _dur_min  = round((datetime.datetime.now() - _start_dt).total_seconds() / 60, 1)
        except Exception:
            _dur_min = ''

        avg_resp_time    = sum(d['avg_time_min'] for d in active_drones) / len(active_drones) if active_drones else 0.0
        avg_ground_speed = CONFIG["DEFAULT_TRAFFIC_SPEED"] * (1 - traffic_level / 100)
        avg_time_saved   = (max(0.0, (sum(d['avg_time_min'] * d['speed_mph'] * 1.4 / avg_ground_speed for d in active_drones) / len(active_drones)) - avg_resp_time)) if active_drones and avg_ground_speed > 0 else 0.0
        _calls_lons = (df_calls_full if df_calls_full is not None else df_calls)['lon'].dropna()
        _calls_lats = (df_calls_full if df_calls_full is not None else df_calls)['lat'].dropna()
        minx = float(_calls_lons.min()) if len(_calls_lons) else 0
        maxx = float(_calls_lons.max()) if len(_calls_lons) else 0
        miny = float(_calls_lats.min()) if len(_calls_lats) else 0
        maxy = float(_calls_lats.max()) if len(_calls_lats) else 0
        area_sq_mi_est   = max(1, int((maxx - minx) * (maxy - miny) * 3280))
        export_details = {
            # Session
            "session_id":            st.session_state.get('session_id', ''),
            "session_start":         _session_start,
            "session_duration_min":  _dur_min,
            "data_source":           st.session_state.get('data_source', 'unknown'),
            # Jurisdiction (user-selected)
            "population":            pop_metric,
            "total_calls":           st.session_state.get('total_original_calls', 0),
            "daily_calls":           max(1, int(st.session_state.get('total_original_calls', 0) / 365)),
            "area_sq_mi":            area_sq_mi_est,
            # City/state enrichment
            "city_confirmed_match":  (
                st.session_state.get('file_meta', {}).get('file_inferred_city', '').lower().strip() ==
                prop_city.lower().strip()
                if st.session_state.get('file_meta', {}).get('file_inferred_city', '') else ''
            ),
            "multi_city_targets":    json.dumps(st.session_state.get('target_cities', [])),
            "num_cities_targeted":   len(st.session_state.get('target_cities', [])),
            # Financials
            "opt_strategy":          opt_strategy,
            "incremental_build":     incremental_build,
            "allow_redundancy":      allow_redundancy,
            "dfr_rate":              int(dfr_dispatch_rate*100),
            "deflect_rate":          int(deflection_rate*100),
            "fleet_capex":           fleet_capex,
            "annual_savings":        annual_savings,
            "thermal_savings":       thermal_savings,
            "k9_savings":            k9_savings,
            "possible_additional_savings": possible_additional_savings,
            "break_even":            break_even_text,
            # Operational
            "avg_response_min":      round(avg_resp_time, 2),
            "avg_time_saved_min":    round(avg_time_saved, 2),
            "area_covered_pct":      round(area_covered_perc, 1),
            "calls_per_capita":      round(st.session_state.get('total_original_calls', 0) / max(pop_metric, 1), 4),
            # Engagement / depth signals
            "r_resp_radius":         st.session_state.get('r_resp', ''),
            "r_guard_radius":        st.session_state.get('r_guard', ''),
            "estimated_pop_input":   pop_metric,
            "boundary_kind":         st.session_state.get('boundary_kind', ''),
            "boundary_source_path":  st.session_state.get('boundary_source_path', ''),
            "sim_or_upload":         st.session_state.get('data_source', 'unknown'),
            "onboarding_completed":  st.session_state.get('onboarding_done', False),
            "demo_mode_used":        st.session_state.get('demo_mode_used', False),
            "map_viewed":            st.session_state.get('map_build_logged', False),
            "export_type_sequence":  json.dumps(st.session_state.get('export_event_log', [])),
            "total_exports_in_session": st.session_state.get('export_count', 0),
            # File data matrix (populated by aggressive_parse_calls → _extract_file_meta)
            "file_meta":             st.session_state.get('file_meta', {}),
            # PD Contact
            "pd_chief":              st.session_state.get('pd_chief_name', ''),
            "pd_dept":               st.session_state.get('pd_dept_name', ''),
            "pd_dept_email":         st.session_state.get('pd_dept_email', ''),
            "pd_dept_phone":         st.session_state.get('pd_dept_phone', ''),
            # Drones
            "active_drones": [{
                "name": d['name'], "type": d['type'],
                "lat": d['lat'],   "lon": d['lon'],
                "avg_time_min":    d.get('avg_time_min', 0),
                "faa_ceiling":     d.get('faa_ceiling', ''),
                "annual_savings":  d.get('annual_savings', 0),
            } for d in active_drones],
        }

        _mgdf_export = st.session_state.get('master_gdf_override')
        _boundary_geojson_export = None
        if _mgdf_export is not None and not _mgdf_export.empty:
            try:
                _boundary_geojson_export = _mgdf_export.to_crs(epsg=4326).to_json()
            except Exception:
                _boundary_geojson_export = None

        export_dict = {
            "city": prop_city, "state": prop_state,
            "_disclaimer": (
                "SIMULATION TOOL: All figures in this file are model estimates based on user-provided inputs. "
                "Real-world results will vary. This is not a legal recommendation, binding proposal, contract, "
                "or guarantee of any product, service, or financial outcome."
            ),
            "k_resp": k_responder, "k_guard": k_guardian,
            "r_resp": resp_radius_mi, "r_guard": guard_radius_mi,
            "dfr_rate": int(dfr_dispatch_rate*100), "deflect_rate": int(deflection_rate*100),
            "calls_data": _safe_df_to_records(
                st.session_state.get('df_calls_full') if st.session_state.get('df_calls_full') is not None
                else st.session_state.get('df_calls')
            ),
            "stations_data": _safe_df_to_records(st.session_state.get('df_stations')),
            "faa_geojson": faa_geojson,
            "boundary_geojson": _boundary_geojson_export,
            "boundary_kind": st.session_state.get('boundary_kind', 'place'),
            "boundary_source_path": st.session_state.get('boundary_source_path', ''),
        }

        fig_for_export = go.Figure()
        for d in active_drones:
            clats, clons = get_circle_coords(d['lat'], d['lon'], r_mi=d['radius_m']/1609.34)
            fig_for_export.add_trace(go.Scattermap(
                lat=list(clats)+[None,d['lat']], lon=list(clons)+[None,d['lon']],
                mode='lines+markers', line=dict(color=d['color'], width=3),
                marker=dict(size=[0]*len(clats)+[0,16], color=d['color']),
                fill='toself', fillcolor='rgba(0,0,0,0)', name=d['name'][:30]
            ))
        fig_for_export.update_layout(
            map=dict(center=dict(lat=center_lat, lon=center_lon), zoom=dynamic_zoom, style="carto-darkmatter"),
            margin=dict(l=0,r=0,t=0,b=0), height=500, showlegend=True,
            legend=dict(
                yanchor="top", y=0.98, xanchor="left", x=0.02,
                bgcolor=legend_bg, bordercolor="#444444", borderwidth=1,
                font=dict(color=legend_text, size=11)
            )
        )
        map_html_str = fig_for_export.to_html(full_html=False, include_plotlyjs='cdn', default_height='500px', default_width='100%')
        station_rows = "".join(f"<tr><td>{d['name']}</td><td>{d['type']}</td><td>{d['avg_time_min']:.1f} min</td><td>{d['faa_ceiling']}</td><td>${d['cost']:,}</td></tr>" for d in active_drones)

        all_bldgs_rows = ""
        _type_colors = {
            "Police": ("rgba(0,210,255,0.1)","#0066aa"),
            "Fire": ("rgba(220,53,69,0.1)","#aa0022"),
            "School": ("rgba(255,215,0,0.12)","#7a6000"),
            "Hospital": ("rgba(34,197,94,0.1)","#006622"),
            "Government": ("rgba(139,92,246,0.1)","#4b0082"),
            "Library": ("rgba(249,115,22,0.1)","#8a3300"),
        }
        for _, row in df_stations_all.iterrows():
            gmaps_link = f"https://www.google.com/maps/search/?api=1&query={row['lat']},{row['lon']}"
            _rtype = str(row.get('type', 'Facility'))
            _tc = _type_colors.get(_rtype, ("rgba(0,0,0,0.04)","#555"))
            _short_name = str(row['name'])[:45] + ("…" if len(str(row['name']))>45 else "")
            all_bldgs_rows += (
                f'''<div class="infra-item">
                  <span class="i-name" title="{row['name']}">{_short_name}</span>
                  <span class="i-type" style="background:{_tc[0]};color:{_tc[1]}">{_rtype}</span>
                  <a class="i-link" href="{gmaps_link}" target="_blank">↗</a>
                </div>'''
            )

        logo_b64 = get_themed_logo_base64("logo.png", theme="dark")
        logo_html_str = f'<img src="data:image/png;base64,{logo_b64}" style="height:40px;">' if logo_b64 else '<div style="font-size:24px;font-weight:900;letter-spacing:3px;color:#fff;">BRINC</div>'

        jurisdiction_list = ", ".join(selected_names) if selected_names else prop_city
        all_station_types = df_stations_all['type'].dropna().unique().tolist() if 'type' in df_stations_all.columns else []
        police_dept_names = [d['name'] for d in active_drones if '[Police]' in d['name']]
        fire_dept_names   = [d['name'] for d in active_drones if '[Fire]' in d['name']]
        ems_dept_names    = [d['name'] for d in active_drones if '[EMS]' in d['name']]

        police_stations = [d['name'] for d in active_drones if 'Police' in d.get('name','') or (
            'type' in df_stations_all.columns and
            'Police' in str(df_stations_all[df_stations_all['name'].str.contains(
                d['name'].split(']')[-1].strip(), na=False, regex=False
            )]['type'].values[:1])
        )]

        dept_summary_parts = []
        if police_dept_names: dept_summary_parts.append(f"{len(police_dept_names)} Police station{'s' if len(police_dept_names)>1 else ''}")
        if fire_dept_names:   dept_summary_parts.append(f"{len(fire_dept_names)} Fire station{'s' if len(fire_dept_names)>1 else ''}")
        if ems_dept_names:    dept_summary_parts.append(f"{len(ems_dept_names)} EMS station{'s' if len(ems_dept_names)>1 else ''}")
        dept_summary = ", ".join(dept_summary_parts) if dept_summary_parts else f"{len(active_drones)} municipal stations"
        police_names_str = (", ".join([n.replace('[Police] ','') for n in police_dept_names[:6]]) + ("..." if len(police_dept_names)>6 else "")) if police_dept_names else "municipal facilities"
        total_fleet = actual_k_responder + actual_k_guardian
        analytics_html_export = generate_command_center_html(df_calls_full if df_calls_full is not None else df_calls, total_orig_calls=st.session_state.get('total_original_calls', full_total_calls or total_calls), export_mode=True)
        cad_charts_html_export = _build_cad_charts_html(df_calls_full if df_calls_full is not None else df_calls)
        staffing_pressure_html_export = ""

        prepared_for_city = st.session_state.get('active_city', prop_city) or prop_city
        prepared_by_name = prop_name
        # ── Export personalization: extract call-type and priority stats ────────
        _exp_df = df_calls_full if (df_calls_full is not None and not df_calls_full.empty) else df_calls
        _exp_top_calls, _exp_pri_str, _exp_date_range = [], "mixed priorities", "recent period"
        try:
            for _cc in ['call_type_desc','agencyeventtypecodedesc','calldesc','description']:
                if _cc in _exp_df.columns:
                    _tc = _exp_df[_cc].dropna().str.strip().value_counts().head(5)
                    if not _tc.empty:
                        _exp_top_calls = list(zip(_tc.index.tolist(), _tc.values.tolist()))
                    break
            if 'priority' in _exp_df.columns:
                _pc = _exp_df['priority'].dropna().value_counts().sort_index()
                _pri_parts = [f"Priority {k}: {v:,} ({v/max(len(_exp_df),1)*100:.0f}%)" for k,v in _pc.items() if str(k).strip().isdigit()]
                _exp_pri_str = " · ".join(_pri_parts[:4]) if _pri_parts else "mixed priorities"
            if 'date' in _exp_df.columns:
                _dd = pd.to_datetime(_exp_df['date'], errors='coerce').dropna()
                if not _dd.empty:
                    _exp_date_range = f"{_dd.min().strftime('%b %Y')} – {_dd.max().strftime('%b %Y')}"
        except Exception:
            pass
        _top5_html = ""
        for _cname, _ccnt in (_exp_top_calls or []):
            _cpct = f"{_ccnt/max(len(_exp_df),1)*100:.1f}%"
            _is_prop = any(w in str(_cname).upper() for w in ['THEFT','BURGLAR','VANDAL','ROBBERY','TRESPASS','LARCEN','AUTO','VEHICLE','BREAK','SHOPLI'])
            _cflag = ' <span style="color:#f59e0b;font-size:10px">property-related</span>' if _is_prop else ''
            _top5_html += f'<tr><td><strong>{_cname}</strong>{_cflag}</td><td style="font-family:monospace;color:#374151">{_ccnt:,}</td><td style="color:#6b7280">{_cpct}</td></tr>'
        _guardian_img_b64 = get_transparent_product_base64("gigs.png") or ""
        logo_b64_dark = get_themed_logo_base64("logo.png", theme="dark")
        logo_b64_light = get_themed_logo_base64("logo.png", theme="light")
        export_html = f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>BRINC DFR Proposal — {prop_city}, {prop_state}</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=IBM+Plex+Mono:wght@400;600&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
:root{{
  --black:#000000;--white:#ffffff;--cyan:#00D2FF;--gold:#FFD700;
  --ink:#0a0a0f;--surface:#f7f8fa;--border:#e4e6ea;
  --text:#111118;--muted:#6b7280;--light:#f0f2f6;
  --resp:#00D2FF;--guard:#FFD700;--green:#22c55e;--amber:#f59e0b;
}}
html{{scroll-behavior:smooth}}
body{{font-family:'Inter',sans-serif;background:var(--surface);color:var(--text);line-height:1.6;display:flex;min-height:100vh}}

/* ── LEFT SIDEBAR INDEX ──────────────────────────────────────── */
.doc-sidebar{{
  position:fixed;top:0;left:0;width:220px;height:100vh;
  background:var(--ink);overflow-y:auto;z-index:100;
  display:flex;flex-direction:column;
  border-right:1px solid #1a1a2a;
}}
.sidebar-logo{{
  padding:28px 24px 20px;
  border-bottom:1px solid #1a1a2a;
}}
.sidebar-logo img{{height:40px;display:block}}
.sidebar-logo .brand{{font-size:22px;font-weight:900;letter-spacing:3px;color:#fff}}
.sidebar-city{{
  padding:16px 24px;
  border-bottom:1px solid #1a1a2a;
}}
.sidebar-city .city-name{{font-size:13px;font-weight:700;color:#fff;letter-spacing:0.3px}}
.sidebar-city .city-sub{{font-size:11px;color:#666;margin-top:2px}}
.sidebar-nav{{padding:20px 0;flex:1}}
.sidebar-nav a{{
  display:flex;align-items:center;gap:10px;
  padding:9px 24px;font-size:12px;font-weight:500;
  color:#888;text-decoration:none;
  border-left:2px solid transparent;
  transition:all 0.15s;
}}
.sidebar-nav a:hover,.sidebar-nav a.active{{
  color:#fff;background:rgba(0,210,255,0.06);
  border-left-color:var(--cyan);
}}
.sidebar-nav .nav-num{{
  font-size:10px;font-weight:700;color:#333;
  width:18px;text-align:center;flex-shrink:0;
}}
.sidebar-nav a:hover .nav-num,.sidebar-nav a.active .nav-num{{color:var(--cyan)}}
.sidebar-footer{{
  padding:20px 24px;border-top:1px solid #1a1a2a;
  font-size:10px;color:#444;line-height:1.6;
}}
.sidebar-footer a{{color:#555;text-decoration:none}}

/* ── MAIN CONTENT ────────────────────────────────────────────── */
.doc-main{{margin-left:220px;flex:1;min-width:0}}

/* ── PAGE SECTIONS (each prints as independent page) ─────────── */
.doc-section{{
  background:#fff;
  border-bottom:6px solid var(--surface);
  padding:52px 60px;
  position:relative;
}}
.doc-section:first-child{{padding-top:64px}}

/* Section header bar */
.section-eyebrow{{
  display:flex;align-items:center;gap:10px;
  margin-bottom:32px;
}}
.copy-section-btn{{
  margin-left:auto;
  display:inline-flex;align-items:center;gap:6px;
  font-size:10px;font-weight:600;letter-spacing:0.8px;text-transform:uppercase;
  color:var(--cyan);border:1px solid rgba(0,210,255,0.35);
  background:rgba(0,210,255,0.06);
  padding:4px 12px;border-radius:100px;cursor:pointer;
  font-family:'IBM Plex Mono',monospace;
  transition:background 0.15s,border-color 0.15s;
  user-select:none;
}}
.copy-section-btn:hover{{background:rgba(0,210,255,0.14);border-color:rgba(0,210,255,0.6);}}
.copy-section-btn.copied{{color:#39FF14;border-color:rgba(57,255,20,0.5);background:rgba(57,255,20,0.07);}}
.section-eyebrow .pg-num{{
  font-size:10px;font-weight:700;letter-spacing:2px;text-transform:uppercase;
  color:var(--cyan);border:1px solid rgba(0,210,255,0.3);
  padding:3px 10px;border-radius:100px;font-family:'IBM Plex Mono',monospace;
}}
.section-eyebrow .pg-title{{
  font-size:10px;font-weight:600;letter-spacing:2px;text-transform:uppercase;
  color:var(--muted);
}}

/* ── COVER PAGE ──────────────────────────────────────────────── */
.cover-page{{
  background:var(--ink);color:#fff;min-height:100vh;
  display:flex;flex-direction:column;justify-content:space-between;
  padding:60px;position:relative;overflow:hidden;
}}
.cover-page::before{{
  content:'';position:absolute;inset:0;
  background:radial-gradient(ellipse 80% 60% at 70% 30%,rgba(0,210,255,0.08) 0%,transparent 70%);
  pointer-events:none;
}}
.cover-logo{{margin-bottom:auto}}
.cover-logo img{{height:52px}}
.cover-logo .brand{{font-size:28px;font-weight:900;letter-spacing:4px;color:#fff}}
.cover-headline{{margin:60px 0 40px}}
.cover-tag{{
  font-size:11px;font-weight:700;letter-spacing:3px;text-transform:uppercase;
  color:var(--cyan);margin-bottom:20px;
}}
.cover-headline h1{{
  font-size:52px;font-weight:900;line-height:1.05;letter-spacing:-1px;
  color:#fff;margin-bottom:16px;
}}
.cover-headline h1 span{{color:var(--cyan)}}
.cover-headline p{{font-size:16px;color:#888;max-width:480px;line-height:1.7}}
.cover-meta{{
  display:grid;grid-template-columns:repeat(4,1fr);gap:1px;
  background:#1a1a2a;border:1px solid #1a1a2a;border-radius:10px;overflow:hidden;
  margin-top:4px;
}}
.cover-meta-cell{{
  background:var(--ink);padding:16px 14px;
}}
.cover-meta-cell .label{{font-size:9px;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:#555;margin-bottom:5px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.cover-meta-cell .value{{font-size:clamp(11px,1vw,14px);font-weight:800;color:#fff;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;font-family:'IBM Plex Mono',monospace;letter-spacing:-0.3px}}
.cover-meta-cell .value.accent{{color:var(--cyan)}}
.cover-meta-cell .value.gold{{color:var(--gold)}}
.cover-bottom{{margin-top:40px;font-size:12px;color:#444;border-top:1px solid #1a1a2a;padding-top:24px;display:flex;justify-content:space-between}}

/* ── METRICS SECTION ─────────────────────────────────────────── */
.metrics-hero{{
  display:grid;grid-template-columns:repeat(4,1fr);gap:1px;
  background:var(--border);border-radius:12px;overflow:hidden;
  margin-bottom:40px;box-shadow:0 1px 3px rgba(0,0,0,0.04);
}}
.metric-cell{{
  background:#fff;padding:24px 16px;text-align:center;min-width:0;
}}
.metric-cell .m-label{{font-size:9px;font-weight:600;letter-spacing:1.2px;text-transform:uppercase;color:var(--muted);margin-bottom:8px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.metric-cell .m-value{{font-size:clamp(16px,1.8vw,28px);font-weight:900;font-family:'IBM Plex Mono',monospace;line-height:1.1;color:var(--text);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.metric-cell .m-value.cyan{{color:var(--cyan)}}
.metric-cell .m-value.gold{{color:var(--gold)}}
.metric-cell .m-value.green{{color:var(--green)}}
.metric-cell .m-sub{{font-size:11px;color:var(--muted);margin-top:6px}}

/* Fleet split row */
.fleet-split{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:32px}}
.fleet-card{{border-radius:10px;padding:24px;position:relative;overflow:hidden}}
.fleet-card.guardian{{background:#0a0800;border:1px solid #2a2400;color:#fff}}
.fleet-card.responder{{background:#00111a;border:1px solid #002a3a;color:#fff}}
.fleet-card .fc-icon{{font-size:28px;margin-bottom:12px}}
.fleet-card .fc-type{{font-size:10px;font-weight:700;letter-spacing:2px;text-transform:uppercase;margin-bottom:8px}}
.fleet-card.guardian .fc-type{{color:var(--gold)}}
.fleet-card.responder .fc-type{{color:var(--resp)}}
.fleet-card .fc-val{{font-size:32px;font-weight:900;font-family:'IBM Plex Mono',monospace;color:#fff}}
.fleet-card .fc-sub{{font-size:12px;color:#888;margin-top:4px}}
.fleet-card .fc-row{{display:flex;justify-content:space-between;font-size:12px;padding:6px 0;border-bottom:1px solid rgba(255,255,255,0.06)}}
.fleet-card .fc-row:last-child{{border-bottom:none}}
.fleet-card .fc-row .k{{color:#666}}.fleet-card .fc-row .v{{font-weight:600;color:#ccc}}

/* ── SECTION HEADINGS ────────────────────────────────────────── */
.sh{{
  font-size:22px;font-weight:800;color:var(--text);
  margin-bottom:20px;padding-bottom:12px;
  border-bottom:2px solid var(--border);
  display:flex;align-items:center;gap:10px;
}}
.sh .sh-accent{{color:var(--cyan)}}

/* ── TABLES ──────────────────────────────────────────────────── */
table{{width:100%;border-collapse:collapse;font-size:13px;margin-bottom:28px}}
thead tr{{background:var(--ink);color:#fff}}
thead th{{padding:12px 16px;text-align:left;font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase}}
tbody tr:nth-child(even){{background:#fafbfc}}
tbody tr:hover{{background:#f0f8ff}}
td{{padding:12px 16px;border-bottom:1px solid var(--border);color:var(--text)}}
.tag-resp{{background:rgba(0,210,255,0.1);color:#0066aa;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600}}
.tag-guard{{background:rgba(255,215,0,0.15);color:#8a6a00;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600}}

/* ── MAP ─────────────────────────────────────────────────────── */
.map-wrap{{border-radius:12px;overflow:hidden;border:1px solid var(--border);box-shadow:0 4px 20px rgba(0,0,0,0.06);margin-bottom:32px}}

/* ── DISCLAIMER ──────────────────────────────────────────────── */
.disc{{background:#fffbeb;border-left:4px solid var(--amber);padding:16px 20px;border-radius:0 8px 8px 0;font-size:12px;color:#7a5a00;margin-bottom:24px}}

/* ── FOOTER ──────────────────────────────────────────────────── */
.doc-footer{{background:var(--ink);color:#555;padding:36px 60px;font-size:11px;display:flex;justify-content:space-between;align-items:center;gap:20px}}
.doc-footer a{{color:#666;text-decoration:none}}
.doc-footer .brand-mark{{font-size:16px;font-weight:900;letter-spacing:3px;color:#fff;flex-shrink:0}}

/* ── PRODUCT IMAGE ───────────────────────────────────────────── */
.cover-body{{
  display:flex;align-items:flex-start;gap:40px;flex:1;margin:40px 0;
}}
.cover-left{{flex:1;min-width:0}}
.cover-right{{
  flex-shrink:0;width:420px;display:flex;align-items:center;
  justify-content:flex-end;
}}
.product-img{{
  width:100%;max-width:420px;
  filter:drop-shadow(0 20px 60px rgba(0,210,255,0.18));
  pointer-events:none;
}}
@media (max-width:900px){{.cover-right{{display:none}}}}

/* ── TWO-COL INFRA TABLE ─────────────────────────────────────── */
.infra-grid{{
  display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:28px;
}}
.infra-item{{
  background:#fafbfc;border:1px solid var(--border);border-radius:6px;
  padding:10px 12px;font-size:11px;display:flex;justify-content:space-between;
  align-items:center;gap:8px;
}}
.infra-item .i-name{{font-weight:600;color:var(--text);flex:1;min-width:0;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}}
.infra-item .i-type{{font-size:10px;font-weight:600;padding:2px 6px;border-radius:4px;flex-shrink:0}}
.infra-item .i-link{{color:var(--cyan);text-decoration:none;font-weight:700;font-size:11px;flex-shrink:0}}

/* ── GRANT STATS SIDEBAR ─────────────────────────────────────── */
.grant-layout{{display:grid;grid-template-columns:1fr 280px;gap:40px;align-items:start}}
.grant-sidebar{{display:flex;flex-direction:column;gap:12px}}
.grant-stat{{
  background:#fafbfc;border:1px solid var(--border);border-radius:8px;
  padding:16px;text-align:center;
}}
.grant-stat .gs-label{{font-size:10px;font-weight:600;letter-spacing:1px;
  text-transform:uppercase;color:var(--muted);margin-bottom:6px}}
.grant-stat .gs-val{{font-size:24px;font-weight:900;font-family:'IBM Plex Mono',monospace;color:var(--cyan)}}
.grant-stat .gs-sub{{font-size:11px;color:var(--muted);margin-top:4px}}
.grant-stat.gold .gs-val{{color:var(--gold)}}
.grant-stat.green .gs-val{{color:var(--green)}}

/* ── CRIME STATS BOX ─────────────────────────────────────────── */
.crime-box{{
  background:linear-gradient(135deg,#0a0800 0%,#120e00 100%);
  border:1px solid #2a2000;border-radius:10px;
  padding:24px;margin-bottom:24px;
}}
.crime-box h4{{color:var(--gold);font-size:12px;font-weight:700;
  letter-spacing:1.5px;text-transform:uppercase;margin-bottom:16px}}
.crime-stat-row{{
  display:flex;justify-content:space-between;align-items:baseline;
  padding:8px 0;border-bottom:1px solid #1a1600;font-size:13px;
}}
.crime-stat-row:last-child{{border-bottom:none}}
.crime-stat-row .csk{{color:#888}}
.crime-stat-row .csv{{font-weight:700;color:#fff;font-family:'IBM Plex Mono',monospace}}
.crime-stat-row .csv.accent{{color:var(--gold)}}

/* ── PRINT ───────────────────────────────────────────────────── */
@media print{{
  @page{{size:auto;margin:0.45in 0.5in}}
  html,body{{background:#fff !important;-webkit-print-color-adjust:exact;print-color-adjust:exact}}
  .doc-sidebar{{display:none !important}}
  .doc-main{{margin-left:0 !important}}
  .doc-section,.cover-page{{
    page-break-after:always;
    page-break-inside:avoid;
    break-after:page;
    border-bottom:none;
    min-height:0;
    box-shadow:none !important;
  }}
  .doc-section{{padding:30px 34px !important}}
  .cover-page{{padding:38px 40px !important;min-height:auto !important}}
  .doc-footer{{padding:20px 34px !important}}
  .doc-section:last-child,.cover-page:last-child{{page-break-after:auto;break-after:auto}}
  .section-eyebrow,.cover-headline,h1,h2,h3,h4{{break-after:avoid-page;page-break-after:avoid}}
  .metrics-hero,.roi-strip,.grant-layout,.infra-grid,.map-wrap,table,img,svg,canvas{{
    break-inside:avoid-page;
    page-break-inside:avoid;
  }}
  a{{color:inherit !important;text-decoration:none !important}}
}}
</style>
</head>
<body>

<!-- ── TOP-RIGHT CORNER LOGO ─────────────────────────────────── -->


<!-- ── SIDEBAR INDEX ─────────────────────────────────────────── -->
<nav class="doc-sidebar">
  <div class="sidebar-logo">
    {"<img src='data:image/png;base64," + logo_b64_dark + "' style='height:40px;display:block'>" if logo_b64_dark else '<div class="brand">BRINC</div>'}
  </div>
  <div class="sidebar-city">
    <div class="city-name">{prop_city}, {prop_state}</div>
    <div class="city-sub">DFR Deployment Proposal</div>
  </div>
  <nav class="sidebar-nav">
    <a href="#cover"><span class="nav-num">00</span>Cover</a>
    <a href="#executive"><span class="nav-num">01</span>Executive Summary</a>
    <a href="#fleet"><span class="nav-num">02</span>Fleet &amp; Coverage</a>
    <a href="#map"><span class="nav-num">03</span>Coverage Map</a>
    <a href="#incident-data"><span class="nav-num">04</span>Incident Analysis</a>
    <a href="#deployment"><span class="nav-num">05</span>Deployment Locations</a>
    <a href="#grant"><span class="nav-num">06</span>Grant Narrative</a>
    <a href="#infrastructure"><span class="nav-num">07</span>Infrastructure Directory</a>
    <a href="#community"><span class="nav-num">08</span>Community Partnership</a>
    <a href="#analytics"><span class="nav-num">09</span>Analytics Dashboard</a>
    <a href="#community-impact"><span class="nav-num">10</span>Community Impact</a>
  </nav>
  <div class="sidebar-footer">
    Prepared {datetime.datetime.now().strftime("%b %d, %Y")}<br>
    {prop_name}<br>
    <a href="mailto:{prop_email}">{prop_email}</a>
  </div>
</nav>

<main class="doc-main">

<!-- ── 00: COVER ─────────────────────────────────────────────── -->
<section class="cover-page" id="cover">
  <div class="cover-logo">
    {"<img src='data:image/png;base64," + logo_b64_dark + "' style='height:48px;display:block'>" if logo_b64_dark else '<div class="brand">BRINC</div>'}
  </div>
  <div class="cover-body">
    <div class="cover-left">
      <div class="cover-headline">
        <div class="cover-tag">Drone as a First Responder · Deployment Proposal</div>
        <h1>Protecting <span>{prop_city}</span>,<br>{prop_state}</h1>
        <p>A data-driven deployment plan for {actual_k_responder + actual_k_guardian} BRINC aerial units covering {calls_covered_perc:.1f}% of incidents across {prop_city}.</p>
      </div>
      <div class="cover-meta">
        <div class="cover-meta-cell"><div class="label">Fleet CapEx</div><div class="value accent">${fleet_capex:,.0f}</div></div>
        <div class="cover-meta-cell"><div class="label">Annual Savings</div><div class="value gold">${annual_savings:,.0f}</div></div>
        <div class="cover-meta-cell"><div class="label">Add'l Thermal + K-9</div><div class="value accent">${possible_additional_savings:,.0f}</div></div>
        <div class="cover-meta-cell"><div class="label">Break-Even</div><div class="value">{break_even_text}</div></div>
        <div class="cover-meta-cell"><div class="label">Call Coverage</div><div class="value accent">{calls_covered_perc:.1f}%</div></div>
        <div class="cover-meta-cell"><div class="label">Avg Response</div><div class="value">{avg_resp_time:.1f} min</div></div>
        <div class="cover-meta-cell"><div class="label">Time Saved</div><div class="value gold">{avg_time_saved:.1f} min</div></div>
        <div class="cover-meta-cell" style="background:rgba(0,210,255,0.04)"><div class="label" style="color:#1a4a5a">Fleet Size</div><div class="value" style="color:#00a0bf">{actual_k_responder + actual_k_guardian} Units</div></div>
      </div>
    </div>
    <div class="cover-right">
      <img class="product-img" src="data:image/jpeg;base64,{_guardian_img_b64}" alt="BRINC Guardian Station">
    </div>
  </div>
  <div class="cover-bottom">
    <span>Prepared for <strong style="color:#fff">{prepared_for_city}</strong> by <strong style="color:#fff">{prepared_by_name}</strong></span>
    <span>{datetime.datetime.now().strftime("%B %d, %Y")}</span>
  </div>
</section>

<!-- ── 01: EXECUTIVE SUMMARY ──────────────────────────────────── -->
<section class="doc-section" id="executive">
  <div class="section-eyebrow"><span class="pg-num">01</span><span class="pg-title">Executive Summary</span></div>
  <div class="metrics-hero">
    <div class="metric-cell"><div class="m-label">Fleet Capital Expenditure</div><div class="m-value cyan">${fleet_capex:,.0f}</div><div class="m-sub">{actual_k_responder} Responder · {actual_k_guardian} Guardian</div></div>
    <div class="metric-cell"><div class="m-label">Annual Savings Capacity</div><div class="m-value gold">${annual_savings:,.0f}</div><div class="m-sub">At {int(dfr_dispatch_rate*100)}% dispatch · {int(deflection_rate*100)}% resolution</div></div>
    <div class="metric-cell"><div class="m-label">Specialty Response Value</div><div class="m-value green">${possible_additional_savings:,.0f}</div><div class="m-sub">Thermal ${thermal_savings:,.0f} · K-9 ${k9_savings:,.0f} · Fire ${fire_savings:,.0f}</div></div>
    <div class="metric-cell"><div class="m-label">Program Break-Even</div><div class="m-value">{break_even_text}</div><div class="m-sub">Full cost recovery timeline</div></div>
    <div class="metric-cell"><div class="m-label">911 Call Coverage</div><div class="m-value cyan">{calls_covered_perc:.1f}%</div><div class="m-sub">of {st.session_state.get('total_original_calls', total_calls):,} annual incidents</div></div>
    <div class="metric-cell"><div class="m-label">Avg Aerial Response</div><div class="m-value">{avg_resp_time:.1f} min</div><div class="m-sub">vs. ground patrol baseline</div></div>
    <div class="metric-cell"><div class="m-label">Time Saved vs Patrol</div><div class="m-value green">{avg_time_saved:.1f} min</div><div class="m-sub">per incident, on average</div></div>
    <div class="metric-cell" style="background:#fafbfc"><div class="m-label">Total Fleet Units</div><div class="m-value" style="color:#374151">{actual_k_responder + actual_k_guardian}</div><div class="m-sub">{actual_k_responder} Responder · {actual_k_guardian} Guardian</div></div>
  </div>
  <p style="font-size:15px;color:#444;line-height:1.8;max-width:680px">
    The {jurisdiction_list} proposes a BRINC Drones Drone as a First Responder (DFR) program deploying
    <strong>{actual_k_responder + actual_k_guardian} aerial units</strong> — {actual_k_responder} BRINC Responders
    and {actual_k_guardian} BRINC Guardians — across {dept_summary}. The system is projected to cover
    <strong>{calls_covered_perc:.1f}% of historical incidents</strong>, reach scenes
    <strong>{avg_time_saved:.1f} minutes faster</strong> than ground patrol, and deliver
    <strong>${annual_savings:,.0f} in annual operational savings</strong> with a break-even horizon of {break_even_text.lower()}.
  </p>
</section>

<!-- ── 02: FLEET & COVERAGE ───────────────────────────────────── -->
<section class="doc-section" id="fleet">
  <div class="section-eyebrow"><span class="pg-num">02</span><span class="pg-title">Fleet &amp; Coverage</span></div>
  <div class="fleet-split">
    <div class="fleet-card guardian">
      <div class="fc-icon">🦅</div>
      <div class="fc-type">BRINC Guardian</div>
      <div class="fc-val">{actual_k_guardian} Unit{"s" if actual_k_guardian != 1 else ""}</div>
      <div class="fc-sub">{guard_radius_mi}-mile operational radius · {guard_strategy_raw}</div>
      <div style="margin-top:16px">
        <div class="fc-row"><span class="k">Unit CapEx</span><span class="v">${CONFIG['GUARDIAN_COST']:,}</span></div>
        <div class="fc-row"><span class="k">Call Coverage</span><span class="v">{guard_calls_perc:.1f}%</span></div>
        <div class="fc-row"><span class="k">Area Coverage</span><span class="v">{guard_area_perc:.1f}%</span></div>
      </div>
    </div>
    <div class="fleet-card responder">
      <div class="fc-icon">🚁</div>
      <div class="fc-type">BRINC Responder</div>
      <div class="fc-val">{actual_k_responder} Unit{"s" if actual_k_responder != 1 else ""}</div>
      <div class="fc-sub">{resp_radius_mi}-mile operational radius · {resp_strategy_raw}</div>
      <div style="margin-top:16px">
        <div class="fc-row"><span class="k">Unit CapEx</span><span class="v">${CONFIG['RESPONDER_COST']:,}</span></div>
        <div class="fc-row"><span class="k">Call Coverage</span><span class="v">{resp_calls_perc:.1f}%</span></div>
        <div class="fc-row"><span class="k">Area Coverage</span><span class="v">{resp_area_perc:.1f}%</span></div>
      </div>
    </div>
  </div>
</section>

<!-- ── 03: COVERAGE MAP ───────────────────────────────────────── -->
<section class="doc-section" id="map">
  <div class="section-eyebrow"><span class="pg-num">03</span><span class="pg-title">Coverage Map</span></div>
  <div class="map-wrap">{map_html_str}</div>
</section>

<!-- ── 04: INCIDENT ANALYSIS ─────────────────────────────────── -->
<section class="doc-section" id="incident-data">
  <div class="section-eyebrow"><span class="pg-num">04</span><span class="pg-title">Incident Data Analysis</span></div>
  {cad_charts_html_export}
  {staffing_pressure_html_export}
</section>

<!-- ── 05: DEPLOYMENT LOCATIONS ──────────────────────────────── -->
<section class="doc-section" id="deployment">
  <div class="section-eyebrow"><span class="pg-num">05</span><span class="pg-title">Deployment Locations</span></div>
  <table>
    <thead><tr><th>Station</th><th>Type</th><th>Avg Response</th><th>FAA Ceiling</th><th>CapEx</th></tr></thead>
    <tbody>{station_rows}</tbody>
  </table>
</section>

<!-- ── 06: GRANT NARRATIVE ───────────────────────────────────── -->
<section class="doc-section" id="grant">
  <div class="section-eyebrow"><span class="pg-num">06</span><span class="pg-title">Grant Narrative</span></div>
  <div class="disc"><strong>DISCLAIMER:</strong> AI-generated draft. Must be reviewed, localized, and fact-checked by your grants administrator before submission.</div>

  <script>
  function copyGrantText(id, btn) {{
    var el = document.getElementById(id);
    var text = el.innerText;
    var origLabel = btn.innerHTML;
    function confirm() {{
      btn.innerHTML = '&#10003; Copied!';
      btn.style.background = '#22c55e';
      btn.style.color = '#fff';
      setTimeout(function() {{ btn.innerHTML = origLabel; btn.style.background = ''; btn.style.color = ''; }}, 2200);
    }}
    if (navigator.clipboard && window.isSecureContext) {{
      navigator.clipboard.writeText(text).then(confirm).catch(function() {{ fallback(text, confirm); }});
    }} else {{ fallback(text, confirm); }}
  }}
  function fallback(text, cb) {{
    var ta = document.createElement('textarea');
    ta.value = text;
    ta.style.cssText = 'position:fixed;top:0;left:0;opacity:0;pointer-events:none';
    document.body.appendChild(ta);
    ta.focus(); ta.select();
    try {{ document.execCommand('copy'); cb(); }} catch(e) {{ alert('Press Ctrl+C / Cmd+C to copy.'); }}
    document.body.removeChild(ta);
  }}
  </script>

  <div style="margin-bottom:10px;">
    <button onclick="copyGrantText('grant-body-law', this)"
      style="background:#1e40af;color:#fff;border:none;border-radius:6px;padding:8px 18px;
             font-size:13px;font-weight:600;cursor:pointer;display:inline-flex;align-items:center;gap:6px;">
      &#128203; Copy Grant Text
    </button>
    <span style="font-size:11px;color:#888;margin-left:10px;">Paste directly into Word or Google Docs</span>
  </div>

  <div class="grant-layout">
  <div class="grant-body" id="grant-body-law">
    <p><strong>Project Title:</strong> BRINC Drones Drone as a First Responder (DFR) Program — {prepared_for_city}, {prop_state}</p>

    <p><strong>Executive Summary:</strong> The {jurisdiction_list} respectfully requests funding to establish a Drone as a First Responder (DFR) program deploying {actual_k_responder + actual_k_guardian} purpose-built BRINC aerial units — {actual_k_responder} BRINC Responder and {actual_k_guardian} BRINC Guardian — across {dept_summary} serving {pop_metric:,} residents. Modeled against {st.session_state.get('total_original_calls', total_calls):,} historical incidents from {_exp_date_range}, the program is projected to cover <strong>{calls_covered_perc:.1f}%</strong> of calls for service, arrive an average of <strong>{avg_time_saved:.1f} minutes faster</strong> than ground patrol, and generate <strong>${annual_savings:,.0f} in annual operational savings</strong>, with an additional modeled specialty-response upside of <strong>${possible_additional_savings:,.0f}</strong> from thermal-supported searches and avoided K-9 deployments, reaching full cost recovery in <strong>{break_even_text.lower()}</strong>.</p>

    <p><strong>Statement of Need:</strong> {jurisdiction_list} currently responds to an estimated {st.session_state.get('total_original_calls', total_calls):,} calls for service annually — approximately {max(1,int(st.session_state.get('total_original_calls',total_calls)/365)):,} calls per day. Incident prioritization ({_exp_pri_str}) demonstrates sustained demand across all severity levels. Ground-based patrol response is constrained by traffic, unit availability, and geographic coverage gaps. First-arriving aerial units with live HD/thermal video enable officers to assess scenes, coordinate response, and in many cases resolve incidents without physical dispatch — compressing the critical gap between call receipt and situational awareness from minutes to seconds. BRINC Drones is the only DFR platform purpose-designed for law enforcement, with deployments across hundreds of US agencies.</p>

    <p><strong>Geographic Scope &amp; Participating Agencies:</strong> The proposed network covers <strong>{jurisdiction_list}</strong> ({prop_state}), hosted at {dept_summary} — including facilities operated by <em>{police_names_str}</em>. The deployment area encompasses approximately <strong>{area_sq_mi_est:,} square miles</strong>, achieving <strong>{calls_covered_perc:.1f}%</strong> historical incident coverage and <strong>{area_covered_perc:.1f}%</strong> geographic area coverage. All sites have been pre-screened against FAA LAANC UAS Facility Maps; no controlled-airspace conflicts were identified in the current configuration.</p>

    <p><strong>Fleet Architecture &amp; Program Design:</strong> The fleet consists of <strong>{actual_k_responder} BRINC Responder</strong> units ({resp_radius_mi}-mile operational radius, {CONFIG["RESPONDER_SPEED"]:.0f} mph, 2-minute average response, optimized for <em>{resp_strategy_raw.lower()}</em>) and <strong>{actual_k_guardian} BRINC Guardian</strong> units ({guard_radius_mi}-mile wide-area radius, {CONFIG["GUARDIAN_SPEED"]:.0f} mph, optimized for <em>{guard_strategy_raw.lower()}</em>, {CONFIG["GUARDIAN_FLIGHT_MIN"]}-minute flight cycles with {CONFIG["GUARDIAN_CHARGE_MIN"]}-minute auto-recharge). Guardians provide continuous wide-area patrol at {round(CONFIG["GUARDIAN_PATROL_HOURS"],1)} hours of daily airtime; Responders deliver rapid tactical response within dense call-volume zones. The two-fleet architecture ensures that when a Guardian is engaged on a call, Responders maintain independent coverage of their patrol areas — eliminating single-point-of-failure gaps in aerial response. Deployment mode: <strong>{deployment_mode}</strong>.</p>

    <p><strong>Technology Platform:</strong> BRINC Drones provides fully automated launch-on-dispatch, live-streaming HD and thermal video to dispatch and responding officers, FAA-compliant Beyond Visual Line of Sight (BVLOS) operations, chain-of-custody flight logging, and integrated data analytics. All hardware is manufactured in the United States. BRINC provides full agency onboarding, FAA coordination support, Part 107 pilot training, and ongoing operational guidance at no additional cost.</p>

    <p><strong>Fiscal Impact &amp; Return on Investment:</strong> Total program capital expenditure is <strong>${fleet_capex:,.0f}</strong> ({actual_k_responder} Responder × ${CONFIG["RESPONDER_COST"]:,} + {actual_k_guardian} Guardian × ${CONFIG["GUARDIAN_COST"]:,}). At a <strong>{int(dfr_dispatch_rate*100)}% DFR dispatch rate</strong> and <strong>{int(deflection_rate*100)}% call resolution rate</strong> (no officer dispatch required), the program is projected to generate <strong>${annual_savings:,.0f} per year</strong> in operational savings, plus a conservative <strong>${possible_additional_savings:,.0f}</strong> in possible additional specialty response savings (thermal imaging, K-9 replacement, and fire department aerial support), reaching break-even in <strong>{break_even_text.lower()}</strong>. Cost per drone response is ${CONFIG["DRONE_COST_PER_CALL"]} versus ${CONFIG["OFFICER_COST_PER_CALL"]} for a ground patrol dispatch — a <strong>{int((1-CONFIG["DRONE_COST_PER_CALL"]/CONFIG["OFFICER_COST_PER_CALL"])*100)}% cost reduction</strong> per incident. The program also reduces officer exposure to unknown-risk calls, decreasing liability and improving officer retention outcomes.</p>

    <p><strong>Evaluation Plan:</strong> Program outcomes will be tracked quarterly across four dimensions: (1) response time comparison vs. pre-deployment baseline, (2) incident resolution rate for drone-attended calls, (3) officer injury rate reduction in drone-supported zones, and (4) community satisfaction via annual resident survey. All flight data, incident assignments, and outcome records are retained in BRINC's cloud platform and available for agency reporting.</p>

    <p><strong>10-Year Program Value Model:</strong> The following projections assume consistent call volume and constant dispatch/resolution rates. Actual results may vary.</p>
    <table style="font-size:12px;margin-bottom:16px">
      <thead><tr><th>Metric</th><th>Year 1</th><th>Year 3</th><th>Year 5</th><th>Year 10</th></tr></thead>
      <tbody>
        <tr><td>Annual Savings</td><td>${annual_savings:,.0f}</td><td>${annual_savings*1.05:,.0f}</td><td>${annual_savings*1.1:,.0f}</td><td>${annual_savings*1.22:,.0f}</td></tr>
        <tr><td>Cumulative Savings</td><td>${annual_savings:,.0f}</td><td>${annual_savings*3.15:,.0f}</td><td>${annual_savings*5.53:,.0f}</td><td>${annual_savings*12.58:,.0f}</td></tr>
        <tr><td>Drone-Attended Calls</td><td>{int(calls_per_day*dfr_dispatch_rate*(calls_covered_perc/100)*365):,}</td><td>{int(calls_per_day*dfr_dispatch_rate*(calls_covered_perc/100)*365*1.03):,}</td><td>{int(calls_per_day*dfr_dispatch_rate*(calls_covered_perc/100)*365*1.06):,}</td><td>{int(calls_per_day*dfr_dispatch_rate*(calls_covered_perc/100)*365*1.13):,}</td></tr>
        <tr><td>Calls Resolved w/o Dispatch</td><td>{int(calls_per_day*dfr_dispatch_rate*(calls_covered_perc/100)*deflection_rate*365):,}</td><td>{int(calls_per_day*dfr_dispatch_rate*(calls_covered_perc/100)*deflection_rate*365*1.03):,}</td><td>{int(calls_per_day*dfr_dispatch_rate*(calls_covered_perc/100)*deflection_rate*365*1.06):,}</td><td>{int(calls_per_day*dfr_dispatch_rate*(calls_covered_perc/100)*deflection_rate*365*1.13):,}</td></tr>
        <tr><td>Net Program ROI</td><td style="color:#dc3545">({int((fleet_capex-annual_savings)/1000)}K deficit)</td><td>{f"+${int((annual_savings*3.15-fleet_capex)/1000)}K" if annual_savings*3.15>fleet_capex else f"({int((fleet_capex-annual_savings*3.15)/1000)}K)"}</td><td style="color:#22c55e">+${int((annual_savings*5.53-fleet_capex)/1000):,}K</td><td style="color:#22c55e">+${int((annual_savings*12.58-fleet_capex)/1000):,}K</td></tr>
      </tbody>
    </table>

    <p><strong>Officer Safety &amp; Liability Reduction:</strong> Beyond direct operational savings, DFR programs measurably reduce officer exposure to unknown-risk call scenarios. First-arriving drones perform scene reconnaissance before ground units arrive, enabling officers to approach with full situational awareness. Documented outcomes across peer agencies include: reduced officer injuries in drone-supported zones (avg. 18% reduction, per DOJ data), faster suspect identification improving apprehension rates, and reduced use-of-force incidents through earlier de-escalation intelligence. These outcomes reduce agency liability costs and workers' compensation claims — benefits not captured in the direct cost model above.</p>

    <p><strong>Community &amp; Economic Impact:</strong> Response time improvements of <strong>{avg_time_saved:.1f} minutes</strong> translate directly to better outcomes in time-sensitive incidents: cardiac events, structure fires, crimes in progress, and missing persons cases. Studies by the International Association of Chiefs of Police (IACP) document measurable improvements in case clearance rates, property crime deterrence (15–30% reduction in areas with visible DFR patrols), and community trust metrics in agencies with active drone programs. For {prop_city}'s business community, faster emergency response reduces property damage, shortens insurance claim cycles, and improves the commercial district safety perception that drives foot traffic and investment.</p>

    <p style="background:#f8f9fa;padding:15px;border-radius:8px;border:1px solid #eee;font-size:13px">
      <strong>Applicable Grant Funding Sources:</strong><br>
      <a href="https://bja.ojp.gov/program/jag/overview">DOJ Byrne JAG</a> — Technology and equipment procurement<br>
      <a href="https://www.fema.gov/grants/preparedness/homeland-security">FEMA HSGP</a> — Homeland security CapEx offset<br>
      <a href="https://cops.usdoj.gov/grants">DOJ COPS Office</a> — Law enforcement technology<br>
      <a href="https://www.transportation.gov/grants">DOT RAISE</a> — Regional infrastructure and safety<br>
      <a href="https://bja.ojp.gov/program/smart-policing-initiative/overview">DOJ Smart Policing Initiative</a> — Data-driven public safety
    </p>
  </div>
  <div class="grant-sidebar">
    <div class="grant-stat"><div class="gs-label">Annual Calls</div><div class="gs-val">{st.session_state.get('total_original_calls', total_calls):,}</div><div class="gs-sub">{_exp_date_range}</div></div>
    <div class="grant-stat"><div class="gs-label">Calls/Day</div><div class="gs-val">{max(1,int(st.session_state.get('total_original_calls',total_calls)/365)):,}</div><div class="gs-sub">citywide avg</div></div>
    <div class="grant-stat gold"><div class="gs-label">Call Coverage</div><div class="gs-val">{calls_covered_perc:.1f}%</div><div class="gs-sub">of historical incidents</div></div>
    <div class="grant-stat"><div class="gs-label">Avg Response</div><div class="gs-val">{avg_resp_time:.1f}m</div><div class="gs-sub">{avg_time_saved:.1f} min faster than patrol</div></div>
    <div class="grant-stat gold"><div class="gs-label">Annual Savings</div><div class="gs-val">${annual_savings:,.0f}</div><div class="gs-sub">break-even {break_even_text.lower()}</div></div>
    <div class="grant-stat green"><div class="gs-label">Cost Reduction</div><div class="gs-val">{int((1-CONFIG["DRONE_COST_PER_CALL"]/CONFIG["OFFICER_COST_PER_CALL"])*100)}%</div><div class="gs-sub">per incident vs. patrol dispatch</div></div>
  </div>
  </div>
</section>

<!-- ── 06B: FIRE DEPARTMENT VALUE ─────────────────────────────── -->
<section class="doc-section" id="fire-value">
  <div class="section-eyebrow"><span class="pg-num">06B</span><span class="pg-title">Fire Department Value</span></div>

  <div class="metrics-hero">
    <div class="metric-cell" style="background:rgba(251,113,33,0.07);border:1px solid rgba(251,113,33,0.3)">
      <div class="m-label">Fire Calls Assisted</div>
      <div class="m-value" style="color:#fb7121">{fire_calls_annual:,.0f}</div>
      <div class="m-sub">Est. {int(CONFIG["FIRE_DEFAULT_APPLICABLE_RATE"]*100)}% of covered incidents</div>
    </div>
    <div class="metric-cell" style="background:rgba(251,113,33,0.07);border:1px solid rgba(251,113,33,0.3)">
      <div class="m-label">Scene Size-Up Savings</div>
      <div class="m-value" style="color:#fb7121">${fire_savings * 0.8:,.0f}/yr</div>
      <div class="m-sub">Avoided premature aerial ladder deployment</div>
    </div>
    <div class="metric-cell" style="background:rgba(251,113,33,0.07);border:1px solid rgba(251,113,33,0.3)">
      <div class="m-label">Overhaul Hotspot Savings</div>
      <div class="m-value" style="color:#fb7121">${fire_savings * 0.2:,.0f}/yr</div>
      <div class="m-sub">Crew time saved via thermal detection</div>
    </div>
    <div class="metric-cell" style="background:rgba(251,113,33,0.12);border:1px solid rgba(251,113,33,0.4)">
      <div class="m-label">Total Fire Dept Value</div>
      <div class="m-value" style="color:#fb7121">${fire_savings:,.0f}/yr</div>
      <div class="m-sub">${CONFIG["FIRE_SAVINGS_PER_CALL"]} blended savings per fire call</div>
    </div>
  </div>

  <h3 style="font-size:15px;font-weight:700;margin:20px 0 10px;color:#1e293b">How Drones Deliver Fire Department Value</h3>
  <table style="font-size:12px;margin-bottom:20px">
    <thead><tr><th>Value Driver</th><th>Mechanism</th><th>Est. Savings</th><th>Source Basis</th></tr></thead>
    <tbody>
      <tr>
        <td><strong>Aerial Scene Size-Up</strong></td>
        <td>Drone arrives before engine company, streams live roof and exterior view. Incident commander makes informed entry decisions, avoids premature aerial ladder deployment (~15% of fire calls).</td>
        <td style="color:#fb7121;font-weight:700">${fire_savings * 0.8:,.0f}/yr</td>
        <td>NFPA: aerial ladder deployment $3,000–$8,000/call; 15% avoidance rate applied</td>
      </tr>
      <tr>
        <td><strong>Overhaul Hotspot Detection</strong></td>
        <td>Thermal imaging pinpoints hidden hotspots in walls and attic after knockdown. Reduces overhaul crew time by ~45 min (4-person crew) per fire call.</td>
        <td style="color:#fb7121;font-weight:700">${fire_savings * 0.2:,.0f}/yr</td>
        <td>IAFC: avg. overhaul crew cost ~$200/hr; 45-min reduction applied to 60% of fire calls</td>
      </tr>
      <tr>
        <td><strong>Structure Fire Reconnaissance</strong></td>
        <td>Pre-entry aerial survey identifies structural compromise, vent points, and victim locations — reducing secondary search risk and interior exposure time.</td>
        <td>Included above</td>
        <td>USFA: secondary collapses are #1 cause of on-duty firefighter fatalities</td>
      </tr>
      <tr>
        <td><strong>Wildfire / Brush Perimeter Monitoring</strong></td>
        <td>Thermal-equipped Guardian drones track fire perimeter in real time, replacing helicopter spotting at $5,000–$15,000/hr.</td>
        <td>Situational; not modeled</td>
        <td>CAL FIRE, NWCG: helicopter spotting costs $8,000–$15,000/flight hour</td>
      </tr>
    </tbody>
  </table>

  <h3 style="font-size:15px;font-weight:700;margin:20px 0 10px;color:#1e293b">🚒 Fire Department Grant Narrative</h3>
  <div class="disc"><strong>DISCLAIMER:</strong> AI-generated draft. Must be reviewed, localized, and fact-checked by your grants administrator before submission.</div>

  <div style="margin-bottom:10px;">
    <button onclick="copyGrantText('grant-body-fire', this)"
      style="background:#c2410c;color:#fff;border:none;border-radius:6px;padding:8px 18px;
             font-size:13px;font-weight:600;cursor:pointer;display:inline-flex;align-items:center;gap:6px;">
      &#128203; Copy Fire Grant Text
    </button>
    <span style="font-size:11px;color:#888;margin-left:10px;">Paste directly into Word or Google Docs</span>
  </div>

  <div class="grant-layout">
  <div class="grant-body" id="grant-body-fire">
    <p><strong>Project Title:</strong> BRINC Drones Drone as a First Responder (DFR) Program — Fire Department Aerial Operations Enhancement — {prepared_for_city}, {prop_state}</p>

    <p><strong>Executive Summary:</strong> The {jurisdiction_list} respectfully requests funding to extend its Drone as a First Responder (DFR) program to support fire department operations in {prepared_for_city}, {prop_state}. Deploying {actual_k_responder + actual_k_guardian} BRINC aerial units across {dept_summary}, the program is projected to assist an estimated <strong>{fire_calls_annual:,.0f} fire-related calls annually</strong> — delivering <strong>${fire_savings:,.0f} per year</strong> in fire department operational value through aerial scene size-up, avoided aerial ladder deployments, and thermal-guided overhaul support. The program reaches full capital cost recovery in <strong>{break_even_text.lower()}</strong> when combined with law enforcement operational savings.</p>

    <p><strong>Statement of Need — Fire Operations:</strong> {jurisdiction_list} fire personnel respond to an estimated {fire_calls_annual:,.0f} fire-related incidents annually within the proposed DFR coverage area, including structure fires, fire alarms, brush and vegetation fires, smoke investigations, carbon monoxide events, and hazardous materials calls. Current operations require engine and aerial apparatus to respond blind — without pre-arrival situational awareness of structural conditions, flame/smoke location, victim position, or access constraints. This information gap creates two operational costs: (1) unnecessary aerial ladder deployments when roof access is not required, and (2) extended overhaul operations when thermal hotspots cannot be rapidly located. A drone arriving 2–4 minutes ahead of ground apparatus eliminates this gap at a fraction of the apparatus cost.</p>

    <p><strong>Operational Application:</strong> Upon dispatch of a fire call within the coverage zone, the nearest BRINC drone auto-launches and arrives on scene in an average of <strong>{avg_resp_time:.1f} minutes</strong> — typically before the first engine company. The drone streams live HD and thermal video to the incident commander and apparatus en route, enabling: (1) real-time roof and exterior structural assessment to guide aerial ladder deployment decisions; (2) thermal identification of fire location and spread within the structure; (3) victim search support in smoke-filled environments; (4) post-knockdown hotspot detection to guide targeted overhaul and reduce crew exposure time; and (5) perimeter monitoring for exterior fires and brush incidents. All flight data is logged for after-action review and NFIRS documentation support.</p>

    <p><strong>Fiscal Impact — Fire Department:</strong> The modeled fire department value of <strong>${fire_savings:,.0f} per year</strong> is derived from two primary cost-avoidance mechanisms. First, aerial scene size-up enables incident commanders to defer or cancel aerial ladder deployment in approximately 15% of attended fire calls. At a cost of $3,000–$8,000 per aerial ladder deployment (NFPA), this represents substantial apparatus cost avoidance and equipment preservation. Second, thermal-guided overhaul reduces crew exposure time by an estimated 45 minutes per fire call (4-person crew at $200/hr equivalent labor cost), applied to 60% of attended fire incidents. These figures are intentionally conservative and do not capture reduced workers' compensation exposure, decreased vehicle wear, or avoided overtime from extended scene operations.</p>

    <p><strong>Officer and Firefighter Safety:</strong> The U.S. Fire Administration reports that structure collapses — frequently caused by delayed recognition of structural compromise — remain the leading cause of on-duty firefighter line-of-duty deaths. Pre-entry aerial reconnaissance directly addresses this risk by identifying compromised roof structures, concentrated fire loads, and unsafe entry points before personnel commit to interior positions. Additionally, real-time thermal monitoring during overhaul eliminates the need for crews to conduct repeated manual inspections of walls and ceilings — reducing both exposure time and the risk of delayed ignition injuries.</p>

    <p><strong>Applicable Fire Department Grant Sources:</strong><br>
    <strong>FEMA Assistance to Firefighters Grant (AFG)</strong> — Equipment and technology procurement for fire departments; drones qualify under the Operations &amp; Safety category.<br>
    <strong>FEMA Fire Prevention &amp; Safety (FP&amp;S)</strong> — Research and technology projects reducing firefighter fatalities and injuries.<br>
    <strong>FEMA BRIC (Building Resilient Infrastructure and Communities)</strong> — Mitigation technology including wildfire monitoring systems.<br>
    <strong>USDA Community Facilities Grant</strong> — Rural fire department equipment for communities under 20,000 population.<br>
    <strong>DHS Urban Areas Security Initiative (UASI)</strong> — Multi-agency technology for high-threat urban areas.<br>
    <strong>State Homeland Security Program (SHSP)</strong> — State-administered equipment grants for emergency response agencies.</p>

    <p><strong>10-Year Fire Department Value Projection:</strong></p>
    <table style="font-size:12px;margin-bottom:16px">
      <thead><tr><th>Metric</th><th>Year 1</th><th>Year 3</th><th>Year 5</th><th>Year 10</th></tr></thead>
      <tbody>
        <tr><td>Fire Calls Assisted</td><td>{fire_calls_annual:,.0f}</td><td>{fire_calls_annual*1.03:,.0f}</td><td>{fire_calls_annual*1.06:,.0f}</td><td>{fire_calls_annual*1.13:,.0f}</td></tr>
        <tr><td>Annual Fire Dept Value</td><td>${fire_savings:,.0f}</td><td>${fire_savings*1.05:,.0f}</td><td>${fire_savings*1.10:,.0f}</td><td>${fire_savings*1.22:,.0f}</td></tr>
        <tr><td>Cumulative Fire Value</td><td>${fire_savings:,.0f}</td><td>${fire_savings*3.15:,.0f}</td><td>${fire_savings*5.53:,.0f}</td><td>${fire_savings*12.58:,.0f}</td></tr>
        <tr><td>Scene Size-Up Savings</td><td>${fire_savings*0.8:,.0f}</td><td>${fire_savings*0.8*1.05:,.0f}</td><td>${fire_savings*0.8*1.10:,.0f}</td><td>${fire_savings*0.8*1.22:,.0f}</td></tr>
        <tr><td>Overhaul Crew Savings</td><td>${fire_savings*0.2:,.0f}</td><td>${fire_savings*0.2*1.05:,.0f}</td><td>${fire_savings*0.2*1.10:,.0f}</td><td>${fire_savings*0.2*1.22:,.0f}</td></tr>
      </tbody>
    </table>
  </div>
  <div class="grant-sidebar">
    <div class="grant-stat" style="border-color:rgba(251,113,33,0.4)"><div class="gs-label">Fire Calls/Year</div><div class="gs-val" style="color:#fb7121">{fire_calls_annual:,.0f}</div><div class="gs-sub">within coverage zone</div></div>
    <div class="grant-stat" style="border-color:rgba(251,113,33,0.4)"><div class="gs-label">Scene Size-Up Value</div><div class="gs-val" style="color:#fb7121">${fire_savings*0.8:,.0f}</div><div class="gs-sub">aerial ladder cost avoidance</div></div>
    <div class="grant-stat" style="border-color:rgba(251,113,33,0.4)"><div class="gs-label">Overhaul Value</div><div class="gs-val" style="color:#fb7121">${fire_savings*0.2:,.0f}</div><div class="gs-sub">crew time &amp; hotspot detection</div></div>
    <div class="grant-stat gold"><div class="gs-label">Total Fire Value</div><div class="gs-val">${fire_savings:,.0f}/yr</div><div class="gs-sub">${CONFIG["FIRE_SAVINGS_PER_CALL"]}/call blended</div></div>
    <div class="grant-stat green"><div class="gs-label">Avg Drone Response</div><div class="gs-val">{avg_resp_time:.1f} min</div><div class="gs-sub">{avg_time_saved:.1f} min faster than apparatus</div></div>
    <div class="grant-stat" style="border-color:rgba(251,113,33,0.4)"><div class="gs-label">10-Year Fire Value</div><div class="gs-val" style="color:#fb7121">${fire_savings*12.58:,.0f}</div><div class="gs-sub">cumulative projected value</div></div>
  </div>
  </div>
</section>

<!-- ── 07: INFRASTRUCTURE DIRECTORY ──────────────────────────── -->
<section class="doc-section" id="infrastructure">
  <div class="section-eyebrow"><span class="pg-num">07</span><span class="pg-title">Infrastructure Directory</span></div>
  <p style="color:var(--muted);font-size:12px;margin-bottom:16px">Public facilities evaluated as candidate deployment locations during optimization. All coordinates verified against FAA LAANC facility maps.</p>
  <div class="infra-grid">{all_bldgs_rows}</div>
</section>

<!-- ── 08: COMMUNITY PARTNERSHIP ─────────────────────────────── -->
<section class="doc-section" id="community">
  <div class="section-eyebrow">
    <span class="pg-num">08</span>
    <span class="pg-title">Community Business Partnership</span>
    <button class="copy-section-btn" id="copyPartnershipBtn" onclick="copyCommunitySection()">
      <svg width="11" height="11" viewBox="0 0 16 16" fill="currentColor" style="flex-shrink:0"><path d="M4 2a2 2 0 0 1 2-2h8a2 2 0 0 1 2 2v8a2 2 0 0 1-2 2H6a2 2 0 0 1-2-2V2zm2-1a1 1 0 0 0-1 1v8a1 1 0 0 0 1 1h8a1 1 0 0 0 1-1V2a1 1 0 0 0-1-1H6z"/><path d="M2 5a1 1 0 0 0-1 1v8a1 1 0 0 0 1 1h8a1 1 0 0 0 1-1v-1h-1v1H2V6h1V5H2z"/></svg>
      Copy Letter
    </button>
  </div>

  <!-- Letter header -->
  <div style="background:#f0f8ff;border-left:4px solid var(--cyan);padding:18px 22px;border-radius:0 8px 8px 0;margin-bottom:24px;font-size:13px;color:#333">
    <strong>To:</strong> Local Business Community of {prop_city}, {prop_state}<br>
    <strong>Re:</strong> Drone as a First Responder Program — Community Investment &amp; Partnership Opportunity<br>
    <strong>Date:</strong> {datetime.datetime.now().strftime("%B %d, %Y")}
  </div>

  <p>Dear {prop_city} Business Owner,</p>

  <p>The {prop_city} Police Department is proposing a transformational public safety initiative that will directly protect your business, your employees, and your customers. We are deploying a <strong>BRINC Drones Drone as a First Responder (DFR)</strong> network — {actual_k_responder + actual_k_guardian} purpose-built aerial units that launch automatically on 911 dispatch and arrive on scene in under two minutes, before any ground unit can respond.</p>

  <p>This is not a surveillance program. It is a rapid-response tool. When a burglar alarm trips at your storefront at 2 AM, a BRINC drone is airborne in seconds — streaming HD and thermal video to dispatch, enabling real-time decision-making, and in many cases capturing the suspect before they can flee. That footage becomes court-admissible evidence that dramatically improves prosecution outcomes.</p>

  <!-- Local incident context box -->
  <div class="crime-box">
    <h4>📊 {prop_city} Public Safety Context</h4>
    <div class="crime-stat-row"><span class="csk">Annual Calls for Service (citywide)</span><span class="csv accent">{st.session_state.get('total_original_calls', total_calls):,}</span></div>
    <div class="crime-stat-row"><span class="csk">Average Calls Per Day</span><span class="csv">{max(1,int(st.session_state.get('total_original_calls',total_calls)/365)):,}</span></div>
    <div class="crime-stat-row"><span class="csk">Incidents Covered by DFR Network</span><span class="csv accent">{calls_covered_perc:.1f}% of all calls</span></div>
    <div class="crime-stat-row"><span class="csk">DFR Avg Aerial Response Time</span><span class="csv accent">{avg_resp_time:.1f} minutes</span></div>
    <div class="crime-stat-row"><span class="csk">Time Saved vs. Ground Patrol</span><span class="csv accent">~{avg_time_saved:.1f} min faster per incident</span></div>
    <div class="crime-stat-row"><span class="csk">Estimated Calls Resolved Without Officer Dispatch</span><span class="csv">{int(calls_per_day * dfr_dispatch_rate * (calls_covered_perc/100) * deflection_rate * 365):,}/year</span></div>
    <div class="crime-stat-row"><span class="csk">Geographic Coverage Area</span><span class="csv">{area_sq_mi_est:,} sq mi · {area_covered_perc:.1f}% of city</span></div>
  </div>

  <h3 style="color:var(--text);font-size:16px;margin:24px 0 12px">Why This Matters to Your Business</h3>

  <p>Property crimes cost American businesses over <strong>$50 billion annually</strong> in direct losses — before accounting for insurance premium increases, business interruption, and the intangible cost of a less safe commercial environment. The businesses of {prop_city} are not immune. Consider what faster response means in practice:</p>

  <table style="margin-bottom:20px;font-size:13px">
    <thead><tr><th>Crime Type</th><th>Avg Cost per Incident</th><th>DFR Impact</th></tr></thead>
    <tbody>
      <tr><td><strong>Retail Theft / Shoplifting</strong></td><td>$559 per incident <em>(NRF 2023)</em></td><td>Real-time apprehension · evidence capture · deterrence</td></tr>
      <tr><td><strong>Commercial Burglary</strong></td><td>$3,000–$8,500 per incident</td><td>Scene arrival before suspect can exit · HD identification footage</td></tr>
      <tr><td><strong>Vandalism / Property Damage</strong></td><td>$1,000–$5,000 per incident</td><td>Suspect identification · reduced repeat incidents in monitored zones</td></tr>
      <tr><td><strong>Robbery / Armed Incidents</strong></td><td>$8,000+ in costs &amp; liability</td><td>Officer-safety pre-arrival intel · coordinated ground response</td></tr>
      <tr><td><strong>Trespass / Loitering</strong></td><td>$200–$1,500 in staff/facility cost</td><td>Non-confrontational aerial deterrence</td></tr>
    </tbody>
  </table>

  <p>Peer agencies with active DFR programs — including Chula Vista, CA; El Cajon, CA; and Westerville, OH — have documented <strong>15–30% reductions in property crime rates</strong> within covered zones, and average response times of under 90 seconds from dispatch. {prop_city}'s proposed deployment covers <strong>{calls_covered_perc:.1f}% of incidents</strong> and arrives an average of <strong>{avg_time_saved:.1f} minutes faster</strong> than current patrol response.</p>

  <h3 style="color:var(--text);font-size:16px;margin:24px 0 12px">The Investment We're Requesting</h3>

  <p>Total program CapEx is <strong>${fleet_capex:,.0f}</strong>. The {prop_city} Police Department is seeking community partnership contributions to offset a portion of this cost and accelerate deployment. Every dollar contributed directly funds equipment that protects your street, your block, your customers.</p>

  <table style="margin-bottom:20px">
    <thead><tr><th>Sponsorship Tier</th><th>Contribution</th><th>Recognition &amp; Benefits</th></tr></thead>
    <tbody>
      <tr style="background:rgba(255,215,0,0.04)"><td><strong>🥇 Founding Sponsor</strong></td><td><strong>$25,000+</strong></td><td>Named drone unit bearing your business name · press release · plaque at station · annual program briefing with command staff · tax-deductible receipt</td></tr>
      <tr><td><strong>🥈 Community Champion</strong></td><td><strong>$10,000–$24,999</strong></td><td>Logo on all program materials &amp; vehicle decals · certificate of recognition · annual impact report · public acknowledgment at City Council</td></tr>
      <tr><td><strong>🥉 Neighborhood Partner</strong></td><td><strong>$2,500–$9,999</strong></td><td>Listed on program website and department newsletter · certificate · public acknowledgment</td></tr>
      <tr><td><strong>🤝 Business Supporter</strong></td><td><strong>$500–$2,499</strong></td><td>Donor roll listing · formal thank-you letter on department letterhead</td></tr>
      <tr><td><strong>💙 Community Contributor</strong></td><td><strong>Any amount</strong></td><td>Recognized in annual program transparency report</td></tr>
    </tbody>
  </table>

  <p>For every <strong>$10,000</strong> contributed, the program is projected to generate approximately <strong>${int(annual_savings/max(fleet_capex,1)*10000):,}</strong> in annual operational savings and property crime cost avoidance for the {prop_city} business community — a {round(annual_savings/max(fleet_capex,1),1):.1f}x return on community investment.</p>

  <p>To make a contribution or learn more, please contact:</p>

  <div style="background:#f8f9fa;border:1px solid #eee;border-radius:8px;padding:20px;margin:20px 0;font-size:13px">
    <strong>{pd_chief}</strong><br>{pd_dept}<br>{pd_email_html}{pd_phone_html}
  </div>

  <p>Together, we can make {prop_city} safer, faster, and more resilient. Thank you for your commitment to this community.</p>
  <p>Respectfully,</p>
  <p><strong>{pd_chief}</strong><br>{pd_dept}</p>
</section>

<!-- ── 09: ANALYTICS DASHBOARD ───────────────────────────────── -->
<section class="doc-section" id="analytics">
  <div class="section-eyebrow"><span class="pg-num">09</span><span class="pg-title">Analytics Dashboard</span></div>
  [ANALYTICS_HTML_EXPORT]
</section>

<!-- ── 10: COMMUNITY IMPACT DASHBOARD ────────────────────────── -->
<section class="doc-section" id="community-impact">
  <div class="section-eyebrow"><span class="pg-num">10</span><span class="pg-title">Community Impact &amp; Transparency</span></div>
  [COMMUNITY_IMPACT_HTML_EXPORT]
</section>

<!-- ── DISCLAIMER ─────────────────────────────────────────────── -->
<div style="background:#fffbeb;border:1px solid #f59e0b;border-radius:8px;padding:20px 60px;margin:0;font-size:11px;color:#7a5a00;line-height:1.7">
  <strong>&#9888; SIMULATION TOOL DISCLAIMER</strong> — All figures are model estimates based on user inputs and publicly available data. Not a legal recommendation, binding proposal, contract, or guarantee. Deployments require FAA authorization and formal procurement.
</div>

<!-- ── FOOTER ─────────────────────────────────────────────────── -->
<footer class="doc-footer">
  <span class="brand-mark">BRINC</span>
  <span>{"<img src='data:image/png;base64," + logo_b64_light + "' style='height:24px;vertical-align:middle;'>" if logo_b64_light else ""} BRINC Drones, Inc. · <a href="https://brincdrones.com">brincdrones.com</a> · <a href="mailto:sales@brincdrones.com">sales@brincdrones.com</a> · +1 (855) 950-0226</span>
  <span>Prepared by {prop_name} · <a href="mailto:{prop_email}">{prop_email}</a></span>
</footer>

</main>
<script>
// Highlight active nav link on scroll
const sections=document.querySelectorAll('section[id],div[id="analytics"]');
const links=document.querySelectorAll('.sidebar-nav a');
const obs=new IntersectionObserver(entries=>{{
  entries.forEach(e=>{{
    if(e.isIntersecting){{
      links.forEach(l=>l.classList.remove('active'));
      const a=document.querySelector('.sidebar-nav a[href="#'+e.target.id+'"]');
      if(a)a.classList.add('active');
    }}
  }})
}},{{threshold:0.3}});
sections.forEach(s=>obs.observe(s));

// ── Copy Community Business Partnership letter ──────────────────────────────
function copyCommunitySection() {{
  const section = document.getElementById('community');
  const btn = document.getElementById('copyPartnershipBtn');
  if (!section || !btn) return;

  // Build clean plain-text version by walking the section's text nodes
  // but skipping the button itself
  function getTextContent(node) {{
    if (node === btn) return '';
    if (node.nodeType === Node.TEXT_NODE) return node.textContent;
    if (node.nodeType !== Node.ELEMENT_NODE) return '';
    const tag = node.tagName.toUpperCase();
    if (tag === 'STYLE' || tag === 'SCRIPT') return '';
    let out = '';
    for (const child of node.childNodes) out += getTextContent(child);
    // Block-level elements get newlines
    const block = ['DIV','P','H1','H2','H3','H4','H5','H6','LI','TR','BR','SECTION','ARTICLE'];
    if (block.includes(tag)) out = '\n' + out.trimEnd() + '\n';
    if (tag === 'TH' || tag === 'TD') out = out.trim() + '\t';
    return out;
  }}

  const rawText = getTextContent(section)
    .replace(/\n{{3,}}/g, '\n\n')   // collapse excess blank lines
    .trim();

  navigator.clipboard.writeText(rawText).then(() => {{
    const orig = btn.innerHTML;
    btn.classList.add('copied');
    btn.innerHTML = '<svg width="11" height="11" viewBox="0 0 16 16" fill="currentColor"><path d="M13.854 3.646a.5.5 0 0 1 0 .708l-7 7a.5.5 0 0 1-.708 0l-3.5-3.5a.5.5 0 1 1 .708-.708L6.5 10.293l6.646-6.647a.5.5 0 0 1 .708 0z"/></svg> Copied!';
    setTimeout(() => {{ btn.classList.remove('copied'); btn.innerHTML = orig; }}, 2000);
  }}).catch(() => {{
    // Fallback for non-https / older browsers
    const ta = document.createElement('textarea');
    ta.value = rawText;
    ta.style.cssText = 'position:fixed;opacity:0;top:0;left:0;';
    document.body.appendChild(ta);
    ta.select();
    document.execCommand('copy');
    document.body.removeChild(ta);
    const orig = btn.innerHTML;
    btn.classList.add('copied');
    btn.innerHTML = '<svg width="11" height="11" viewBox="0 0 16 16" fill="currentColor"><path d="M13.854 3.646a.5.5 0 0 1 0 .708l-7 7a.5.5 0 0 1-.708 0l-3.5-3.5a.5.5 0 1 1 .708-.708L6.5 10.293l6.646-6.647a.5.5 0 0 1 .708 0z"/></svg> Copied!';
    setTimeout(() => {{ btn.classList.remove('copied'); btn.innerHTML = orig; }}, 2000);
  }});
}}
</script>
</body></html>"""

        export_html = export_html.replace("[ANALYTICS_HTML_EXPORT]", analytics_html_export)

        # ── Community Impact Dashboard (light theme for print/export) ────────
        _cid_export_html = generate_community_impact_dashboard_html(
            city=prop_city,
            state=prop_state,
            population=int(pop_metric or 65000),
            total_calls=int(st.session_state.get('total_original_calls', 0) or 0),
            calls_covered_perc=float(calls_covered_perc or 0),
            area_covered_perc=float(area_covered_perc or 0),
            avg_resp_time_min=float(avg_resp_time or 0),
            avg_time_saved_min=float(avg_time_saved or 0),
            fleet_capex=float(fleet_capex or 0),
            annual_savings=float(annual_savings or 0),
            break_even_text=str(break_even_text or 'N/A'),
            actual_k_responder=int(actual_k_responder or 0),
            actual_k_guardian=int(actual_k_guardian or 0),
            dfr_dispatch_rate=float(dfr_dispatch_rate or 0.25),
            deflection_rate=float(deflection_rate or 0.30),
            daily_dfr_responses=float(daily_dfr_responses or 0),
            daily_drone_only_calls=float(daily_drone_only_calls or 0),
            active_drones=active_drones or [],
            df_calls_full=df_calls_full,
            theme='light',
        )
        # Extract <style> block and body content separately, then scope the styles
        # with a .cid-wrap prefix so they don't collide with the export document's CSS.
        import re as _re
        _style_match = _re.search(r'<style>(.*?)</style>', _cid_export_html, _re.DOTALL)
        _cid_style = _style_match.group(1) if _style_match else ''
        # Scope every CSS rule inside the style block by prefixing with .cid-wrap
        # Simple approach: wrap rules that start at column 0 (non-nested)
        def _scope_css(raw_css):
            # Replace :root { with .cid-wrap { so vars apply within scope
            raw_css = raw_css.replace(':root {', '.cid-wrap {')
            # Prepend .cid-wrap to each rule selector (lines that end with {)
            lines = raw_css.split('\n')
            scoped = []
            for line in lines:
                stripped = line.strip()
                # Skip empty, @-rules, closing braces, and already-scoped lines
                if (stripped.startswith('@') or stripped == '}' or stripped == ''
                        or stripped.startswith('/*') or stripped.startswith('*')
                        or stripped.startswith('.cid-wrap {')):
                    scoped.append(line)
                elif stripped.endswith('{') and not stripped.startswith('.cid-wrap'):
                    # It's a selector line — prefix it
                    selector = stripped[:-1].strip()
                    # Don't double-scope :root replacement or @keyframes internals
                    if selector and not selector.startswith('.cid-wrap') and not selector.startswith('from') and not selector.startswith('to') and not selector.startswith('0%') and not selector.startswith('50%') and not selector.startswith('100%'):
                        scoped.append(f'  .cid-wrap {selector} {{')
                    else:
                        scoped.append(line)
                else:
                    scoped.append(line)
            return '\n'.join(scoped)

        _scoped_style = _scope_css(_cid_style)
        # Extract body content (between <body> and </body>)
        _body_match = _re.search(r'<body[^>]*>(.*?)</body>', _cid_export_html, _re.DOTALL)
        _cid_body = _body_match.group(1).strip() if _body_match else _cid_export_html
        # Build the scoped embed: scoped <style> + wrapper div
        _cid_embed = f'<style>{_scoped_style}</style>\n<div class="cid-wrap" style="font-family:\'DM Sans\',sans-serif;background:#f8f7f4;border-radius:10px;overflow:hidden;">{_cid_body}</div>'
        export_html = export_html.replace("[COMMUNITY_IMPACT_HTML_EXPORT]", _cid_embed)

    # ── Download buttons — always rendered so they're visible in the sidebar ──
    _safe_city   = _safe_city_base
    _ts          = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    # 1. Save Deployment Plan (.brinc) — always available
    _brinc_data = json.dumps(export_dict) if fleet_capex > 0 else json.dumps({
        "city": st.session_state.get('active_city', ''), "state": st.session_state.get('active_state', ''),
        "_disclaimer": "No drones deployed yet.", "k_resp": 0, "k_guard": 0,
        "calls_data": _safe_df_to_records(
            st.session_state.get('df_calls_full') if st.session_state.get('df_calls_full') is not None
            else st.session_state.get('df_calls')
        ),
        "stations_data": _safe_df_to_records(st.session_state.get('df_stations')),
    })
    if st.sidebar.download_button("💾 Save Deployment Plan", data=_brinc_data,
                                  file_name=f"Brinc_{_safe_city}_{_ts}.brinc",
                                  mime="application/json", use_container_width=True):
        # ── Track export event ───────────────────────────────────────────────
        st.session_state['export_event_log'] = st.session_state.get('export_event_log', []) + ['BRINC']
        st.session_state['export_count'] = st.session_state.get('export_count', 0) + 1
        if fleet_capex > 0:
            _notify_email(st.session_state.get('active_city',''), st.session_state.get('active_state',''),
                          "BRINC", k_responder, k_guardian, calls_covered_perc,
                          prop_name, prop_email, details=export_details)
            _log_to_sheets(st.session_state.get('active_city',''), st.session_state.get('active_state',''),
                           "BRINC", k_responder, k_guardian, calls_covered_perc,
                           prop_name, prop_email, details=export_details)
    # 2. Executive Summary / proposal HTML export
    if fleet_capex > 0:
        if st.sidebar.download_button(f"📄 {prop_city}, {prop_state} — Executive Summary",
                                      data=export_html,
                                      file_name=f"BRINC_Executive_Summary_{_safe_city}_{_ts}.html",
                                      mime="text/html",
                                      use_container_width=True):
            # ── Track export event ───────────────────────────────────────────
            st.session_state['export_event_log'] = st.session_state.get('export_event_log', []) + ['HTML']
            st.session_state['export_count'] = st.session_state.get('export_count', 0) + 1
            _notify_email(st.session_state.get('active_city',''), st.session_state.get('active_state',''),
                          "HTML", k_responder, k_guardian, calls_covered_perc,
                          prop_name, prop_email, details=export_details)
            _log_to_sheets(st.session_state.get('active_city',''), st.session_state.get('active_state',''),
                           "HTML", k_responder, k_guardian, calls_covered_perc,
                           prop_name, prop_email, details=export_details)
        st.sidebar.caption("HTML export includes the latest executive summary, incident analysis, staffing pressure, and print-ready proposal sections.")

    # 3. Google Earth KML — only when drones are placed
    if active_drones:
        if st.sidebar.download_button("🌏 Google Earth Briefing File",
                                      data=generate_kml(active_gdf, active_drones, calls_in_city),
                                      file_name="drone_deployment.kml",
                                      mime="application/vnd.google-earth.kml+xml",
                                      use_container_width=True):
            # ── Track export event ───────────────────────────────────────────
            st.session_state['export_event_log'] = st.session_state.get('export_event_log', []) + ['KML']
            st.session_state['export_count'] = st.session_state.get('export_count', 0) + 1
            _notify_email(st.session_state.get('active_city',''), st.session_state.get('active_state',''),
                          "KML", k_responder, k_guardian, calls_covered_perc,
                          prop_name, prop_email, details=export_details)
            _log_to_sheets(st.session_state.get('active_city',''), st.session_state.get('active_state',''),
                           "KML", k_responder, k_guardian, calls_covered_perc,
                           prop_name, prop_email, details=export_details)
    else:
        st.sidebar.button("🌏 Google Earth Briefing File", disabled=True,
                          use_container_width=True,
                          help="Deploy at least one drone to generate the KML file.")
